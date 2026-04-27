#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║  📈 FMP STOCK SCREENER — Professional Fundamentals Edition     ║
║  17 AI specialists + Master Manager · Financial Modeling Prep  ║
║  No hardcoded data — everything from FMP API + AI analysis     ║
╚══════════════════════════════════════════════════════════════════╝

Separate from stockscreenerultra.py — this is the FMP-native screener.
Covers: NASDAQ + NYSE + AMEX (full US market ~3,500+ stocks)

Tabs:
  1. Overview          — AI market pulse + top picks + tab summaries
  2. IV Discount       — DCF intrinsic value + Lynch quality (rev consistency, FCF conv, buybacks), top 50
  2b. IV by Sector     — Same, 30 per sector
  3. Stalwarts         — PEG 1-2, rev 8-20%, >$2B (Lynch category)
  4. Fast Growers      — PEG <1.5, rev >20% (Lynch category)
  5. Slow Growers      — Dividend >2%, stable (Lynch category)
  6. Cyclicals         — HIGH/no P/E in cyclical sectors = trough earnings = Lynch BUY signal
  7. Turnarounds       — Down >40%, recovering (Lynch category)
  8. Asset Plays       — P/B <1, hidden value (Lynch category)
  9. Sector Valuations — 11 SPDR ETFs with PEG/FCF/growth
  10. Insider Buying   — FMP insider transaction data
  11. Picks Tracking   — Performance measurement

Requirements:
  - FMP API key (env var: FMP_API_KEY)
  - Anthropic API key (env var: ANTHROPIC_API_KEY) for AI overview
  - pip: openpyxl, requests

Usage:
  python stockscreener_fmp.py
"""

import os, sys, json, time, datetime, csv, math, pickle, subprocess
from collections import defaultdict

# ── Windows UTF-8 fix: force stdout/stderr to UTF-8 so emojis don't crash on redirect ──
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

def _load_dotenv():
    """Load .env file from the script's directory into os.environ (no extra deps)."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            # Overwrite if not set OR if set to empty string (e.g. blank PyCharm run config entry)
            if k and v and not os.environ.get(k):
                os.environ[k] = v

_load_dotenv()

# ─────────────────────────────────────────────
# DEPENDENCIES
# ─────────────────────────────────────────────
try:
    import requests
except ImportError:
    sys.exit("pip install requests")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("pip install openpyxl")

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
FMP_KEY = os.environ.get("FMP_API_KEY", "")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
NTFY_TOPIC    = os.environ.get("NTFY_TOPIC",    "")  # optional: phone push notifications
GITHUB_REPO   = os.environ.get("GITHUB_REPO",   "")  # "username/repo" for GitHub Pages hosting
GITHUB_TOKEN  = os.environ.get("GITHUB_TOKEN",  "")  # GitHub Personal Access Token
FMP_BASE = "https://financialmodelingprep.com/stable"

CACHE_FILE = "fmp_screener_cache.json"
CACHE_DAYS = 1
PICKS_LOG    = "fmp_picks_log.csv"
AI_PICKS_LOG = "fmp_ai_picks_log.csv"

# B6: Version stamp on every pick row — bump when prompts or scoring logic changes
# so backtest / attribution can group pre/post change comparisons correctly.
PROMPT_VERSION = "3.0.0"   # Phase A+B+C complete
PORTFOLIO_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fmp_portfolio.json")
OUTPUT_DIR = "."

# Portfolio manager constants
PORTFOLIO_INITIAL_CASH = 100_000.0   # paper money starting balance
PORTFOLIO_TARGET_POSITION = 10_000.0 # ~$10K per position (equal weight)
PORTFOLIO_MAX_POSITIONS = 10         # hard cap
PORTFOLIO_SLIPPAGE_RATE  = 0.001     # C3: 0.1% slippage per trade (bid/ask spread model)
PORTFOLIO_COMMISSION     = 1.00      # C3: $1 flat commission per trade

TOP_N = 50  # main tab results
SECTOR_N = 30  # per-sector results

# ── Fast Grower: small-cap scoring & filter constants ────────────────────
FG_MICRO_CAP_BONUS        = 12    # $100M–$500M: most underfollowed, highest 10x potential
FG_SMALL_CAP_BONUS        =  8    # $500M–$2B:   still neglected by institutional coverage
FG_SMALL_MID_BONUS        =  3    # $2B–$5B:     marginal recognition; slight nudge
FG_MICRO_CAP_MAX          = 500e6
FG_SMALL_CAP_MAX          =   2e9
FG_SMALL_MID_MAX          =   5e9
FG_RELAXED_FCF_REV_GROWTH = 0.25  # minimum rev growth to qualify for FCF exception

def _cap_size_label(mktcap: float) -> str:
    """Return a human-readable market cap tier label for display."""
    if mktcap >= 200e9: return "Mega"
    if mktcap >=  10e9: return "Large"
    if mktcap >=   2e9: return "Mid"
    if mktcap >= 500e6: return "Small"
    if mktcap >= 100e6: return "Micro"
    return "Nano"


# ─────────────────────────────────────────────
# RUN TIMER  (phase-by-phase elapsed + ETA)
# ─────────────────────────────────────────────
_run_start: float = 0.0
_phase_start: float = 0.0
_phase_name: str = ""

# Typical phase durations (seconds) for ETA — updated each run via exponential smoothing
# Order matches the actual run sequence in main()
_PHASE_WEIGHTS = {
    "macro":        3,
    "universe":     25,
    "enrichment":   40,
    "52w_ranges":   90,
    "tabs":         35,
    "ai_analysis":  210,
    "portfolio":    30,
    "save":         5,
}


def _fmt_elapsed(sec: float) -> str:
    """Format seconds as 'm:ss' or 'Xs'."""
    if sec >= 60:
        return f"{int(sec // 60)}m{int(sec % 60):02d}s"
    return f"{int(sec):.0f}s"


def phase_start(name: str, label: str = "") -> None:
    """Mark start of a new phase. Prints elapsed time since run start + ETA."""
    global _phase_start, _phase_name
    _phase_start = time.time()
    _phase_name  = name

    elapsed = _phase_start - _run_start if _run_start else 0
    # Compute ETA: sum of remaining phase weights
    phases = list(_PHASE_WEIGHTS.keys())
    try:
        idx = phases.index(name)
    except ValueError:
        idx = len(phases)
    remaining_est = sum(v for k, v in _PHASE_WEIGHTS.items()
                        if phases.index(k) > idx)
    eta_str = f" | ETA ~{_fmt_elapsed(remaining_est)} left" if remaining_est > 0 else ""
    print(f"\n  ⏱  [{_fmt_elapsed(elapsed)} elapsed{eta_str}]  {label or name.upper()}")


def phase_done(extra: str = "") -> None:
    """Print phase completion with duration."""
    dur = time.time() - _phase_start if _phase_start else 0
    suffix = f"  — {extra}" if extra else ""
    print(f"     ✔ done in {_fmt_elapsed(dur)}{suffix}")


def notify_phone(title: str, message: str, tags: str = "white_check_mark") -> None:
    """Send push notification via ntfy.sh (free, no account needed).
    Set NTFY_TOPIC env var to enable. Install ntfy app on phone and subscribe to the topic.
    Uses JSON body so emojis and Unicode in titles/messages work correctly.
    """
    if not NTFY_TOPIC:
        return
    try:
        requests.post(
            "https://ntfy.sh/",
            json={
                "topic":    NTFY_TOPIC,
                "title":    title,
                "message":  message,
                "tags":     tags.split(",") if tags else [],
                "priority": 3,
            },
            timeout=8,
        )
    except Exception:
        pass  # never crash the run over a notification


def push_html_to_github(html_content: str) -> str | None:
    """Upload HTML dashboard to GitHub via Contents API. Returns public Pages URL or None.
    Requires GITHUB_REPO='username/repo' and GITHUB_TOKEN in .env.
    No git clone or subprocess needed — pure HTTP.
    """
    if not GITHUB_REPO or not GITHUB_TOKEN:
        return None
    try:
        import base64
        encoded = base64.b64encode(html_content.encode("utf-8")).decode()
        hdrs = {"Authorization": f"token {GITHUB_TOKEN}",
                "Accept": "application/vnd.github.v3+json"}
        # Get current file SHA (required for updates; absent on first push)
        r = requests.get(f"https://api.github.com/repos/{GITHUB_REPO}/contents/index.html",
                         headers=hdrs, timeout=10)
        sha = r.json().get("sha") if r.status_code == 200 else None
        today = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        payload = {"message": f"Screener run {today}", "content": encoded, "branch": "main"}
        if sha:
            payload["sha"] = sha
        r2 = requests.put(f"https://api.github.com/repos/{GITHUB_REPO}/contents/index.html",
                          json=payload, headers=hdrs, timeout=20)
        if r2.status_code in (200, 201):
            owner, repo = GITHUB_REPO.split("/", 1)
            return f"https://{owner}.github.io/{repo}/"
        print(f"  ⚠️ GitHub push failed: {r2.status_code} {r2.text[:120]}")
    except Exception as e:
        print(f"  ⚠️ GitHub push error: {e}")
    return None


def commit_data_files():
    """Commit portfolio + picks logs to git after a local run.
    Skipped in GitHub Actions (which has its own 'Save data files' step).
    """
    if os.environ.get("GITHUB_ACTIONS"):
        return
    try:
        _cwd = os.path.dirname(os.path.abspath(__file__))
        _today = datetime.date.today().isoformat()
        # Sync any remote commits pushed by push_html_to_github() via API
        subprocess.run(["git", "pull", "--rebase", "--autostash"],
                       cwd=_cwd, capture_output=True, timeout=30)
        subprocess.run(["git", "add", "fmp_portfolio.json",
                        "fmp_picks_log.csv", "fmp_ai_picks_log.csv"],
                       cwd=_cwd, capture_output=True)
        staged = subprocess.run(["git", "diff", "--staged", "--quiet"],
                                cwd=_cwd, capture_output=True)
        if staged.returncode != 0:
            subprocess.run(["git", "commit", "-m", f"data: {_today}"],
                           cwd=_cwd, capture_output=True)
            subprocess.run(["git", "push"],
                           cwd=_cwd, capture_output=True, timeout=30)
            print("  💾 Data files committed to git")
    except Exception as _e:
        print(f"  ⚠️ Git data-file commit failed: {_e}")


# Excel styling
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALT_FILL = PatternFill("solid", fgColor="F5F5F5")
PLAIN_FILL = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
AMBER_FILL = PatternFill("solid", fgColor="FFF8E1")
RED_FILL = PatternFill("solid", fgColor="FFEBEE")

# Sector ETFs for sector analysis
SECTOR_ETFS = {
    "XLK": "Technology", "XLF": "Financial Services", "XLV": "Healthcare",
    "XLY": "Consumer Cyclical", "XLP": "Consumer Defensive", "XLE": "Energy",
    "XLI": "Industrials", "XLB": "Basic Materials", "XLRE": "Real Estate",
    "XLU": "Utilities", "XLC": "Communication Services",
}

CYCLICAL_SECTORS = {
    "Energy", "Basic Materials", "Industrials", "Consumer Cyclical",
    "Real Estate", "Financial Services",
}

_LYNCH_ASSET_HEAVY_SECTORS = {
    "Financial Services", "Real Estate", "Energy", "Basic Materials",
    "Utilities", "Industrials",
}


# ── A1 (Sprint 2): "Familiar Brand" / consumer-observable industries ──────────
# Companies in these GICS-style industries make products/services the user can
# directly experience as a consumer or workplace user.  This is the primary
# "do I know this?" cue for personal-knowledge-driven picking.
# Adobe → Software—Application 🛒 ;  small B2B chip designer → no badge.
_CONSUMER_OBSERVABLE_INDUSTRIES = {
    # Software / Internet (consumer-facing or workplace-observable)
    "Software—Application", "Software - Application",
    "Internet Content & Information",
    "Internet Retail",
    # Retail
    "Specialty Retail", "Apparel Retail", "Discount Stores", "Grocery Stores",
    "Department Stores", "Home Improvement Retail", "Pharmaceutical Retailers",
    "Auto & Truck Dealerships",
    # Apparel & footwear
    "Apparel Manufacturing", "Footwear & Accessories", "Luxury Goods",
    # Restaurants & travel
    "Restaurants", "Lodging", "Resorts & Casinos", "Travel Services",
    "Gambling", "Leisure",
    # Consumer staples
    "Beverages—Non-Alcoholic", "Beverages - Non-Alcoholic",
    "Beverages—Brewers", "Beverages—Wineries & Distilleries",
    "Beverages - Brewers", "Beverages - Wineries & Distilleries",
    "Confectioners", "Packaged Foods", "Personal Products",
    "Household & Personal Products", "Tobacco",
    # Consumer electronics & autos
    "Consumer Electronics", "Auto Manufacturers",
    # Media & entertainment
    "Entertainment", "Broadcasting", "Publishing", "Electronic Gaming & Multimedia",
    # Telecom & airlines
    "Telecom Services", "Airlines",
    # Toys
    "Toys & Hobbies",
    # Health / wellness consumer-touchable
    "Drug Manufacturers—General", "Drug Manufacturers - General",
    "Healthcare Plans",
}


def _classify_consumer_observable(s: dict) -> bool:
    """Return True if a stock's industry suggests a product/service the user
    can directly experience as a consumer or workplace user.

    Used for the 🛒 Familiar Brand tag — the primary signal for personal-
    knowledge-driven picking ("do I know this company?").
    """
    industry = (s.get("industry") or "").strip()
    return industry in _CONSUMER_OBSERVABLE_INDUSTRIES


def _classify_lynch(s: dict) -> str:
    """Classify a stock into one or more of Peter Lynch's six categories.

    Returns '+'-joined labels, e.g. "FastGrower+Stalwart".  Empty string
    if the stock matches no category (rare — but possible for early-stage
    shells or data-starved micro-caps).

    Each predicate is a SIMPLIFIED mirror of the corresponding tab filter
    in build_lynch_tabs — it trades a little precision for self-containment
    (no closures over tab-internal state) and is cheap enough to call on
    every stock in the universe.
    """
    labels = []
    rg      = s.get("revGrowth")
    rg5     = s.get("revGrowth5y")
    rg5h    = s.get("fiveYRevGrowth")
    eg5     = s.get("epsGrowth5y")
    pe      = s.get("pe")
    pb      = s.get("pb")
    tb      = s.get("tangibleBook")
    price   = s.get("price", 0) or 0
    pio     = s.get("piotroski")
    fcf     = s.get("fcfYield")
    dy      = s.get("divYield")
    eg      = s.get("epsGrowth")
    nig     = s.get("netIncomeGrowth")
    az      = s.get("altmanZ")
    mos     = s.get("mos")
    mc      = s.get("mktCap", 0) or 0
    sector  = s.get("sector") or ""
    ncr     = s.get("netCashRatio")
    cr      = s.get("currentRatio")
    de      = s.get("de")
    gnn     = s.get("grahamNetNet")
    rc      = s.get("revConsistency")

    # ── Stalwart: >$2B, 5–25% rev growth, FCF-positive, quality floor ────
    if (mc > 2e9
            and rg is not None and 0.05 <= rg <= 0.25
            and (fcf is None or fcf > 0)
            and sector != "Basic Materials"
            and (pio is None or pio >= 5)
            and (rc is None or rc >= 0.60)):
        labels.append("Stalwart")

    # ── Fast Grower: strong growth signal + size band + not commodity-led ─
    if (100e6 <= mc <= 150e9
            and sector not in ("Real Estate", "Basic Materials")):
        strong_current = (rg is not None and rg > 0.20)
        strong_5yr     = (rg5 is not None and rg5 > 0.15
                          and (rg is None or rg > 0.10))
        strong_5yr_h   = (rg5h is not None and rg5h > 0.15
                          and (rg is None or rg > 0.10))
        eps_led        = (eg5 is not None and eg5 > 0.15
                          and (rg is None or rg > 0.08))
        if strong_current or strong_5yr or strong_5yr_h or eps_led:
            labels.append("FastGrower")

    # ── Slow Grower: mature dividend-payer, low growth ──────────────────
    if (dy is not None and 0.02 < dy < 0.15
            and (rg is None or rg < 0.10)
            and (pio is None or pio >= 5)
            and (fcf is None or fcf > dy * 0.5)):
        labels.append("SlowGrower")

    # ── Cyclical: in cyclical sector, >$500M, not bankrupt, trough signal ─
    if (sector in CYCLICAL_SECTORS
            and mc > 500e6
            and (pio is None or pio >= 4)
            and (az is None or az >= 0.8)):
        trough_pe    = (pe is not None and 10 < pe < 65)
        loss_trough  = (pe is None and pb is not None and 0 < pb < 2.0)
        asset_trough = (pb is not None and 0 < pb < 0.9)
        if trough_pe or loss_trough or asset_trough:
            labels.append("Cyclical")

    # ── Turnaround: cheap, not bankrupt, showing recovery signals ───────
    if mc >= 300e6 and (pio is None or pio < 8) and (az is None or az >= 0.5):
        deep_discount = (mos is not None and mos > 0.20)
        asset_cheap   = (pb is not None and 0 < pb < 0.8)
        if deep_discount or asset_cheap:
            recovery = 0
            if rg is not None and rg > 0.03:  recovery += 1
            if eg is not None and eg > 0:     recovery += 1
            if nig is not None and nig > 0:   recovery += 1
            if fcf is not None and fcf > 0:   recovery += 1
            if rg5 is not None and rg5 > 0.05: recovery += 1
            if pio is not None and pio >= 5:  recovery += 1
            if recovery >= 2:
                labels.append("Turnaround")

    # ── Asset Play: hidden value in balance sheet ───────────────────────
    if mc >= 150e6 and (pio is None or pio >= 4):
        below_book     = (pb is not None and 0 < pb < 1.0)
        below_tangible = (tb is not None and tb > 0 and price > 0
                          and (tb / price) >= 0.95)
        cash_fortress  = (cr is not None and cr > 3.0
                          and de is not None and de < 0.3
                          and pb is not None and 0 < pb < 2.0)
        net_cash_play  = (ncr is not None and ncr > 0.25)
        graham_nn      = (gnn is not None and price > 0
                          and gnn > price * 0.5)
        if below_book or below_tangible or cash_fortress or net_cash_play or graham_nn:
            labels.append("AssetPlay")

    return "+".join(labels)


# ─────────────────────────────────────────────
# FMP API CLIENT
# ─────────────────────────────────────────────

_fmp_call_count = 0
FMP_V3 = "https://financialmodelingprep.com/api/v3"


def fetch_live_price(ticker: str) -> float | None:
    """Fetch real-time price via FMP stable quote endpoint."""
    global _fmp_call_count
    if not FMP_KEY:
        return None
    try:
        r = requests.get(f"{FMP_BASE}/quote",
                         params={"symbol": ticker, "apikey": FMP_KEY}, timeout=10)
        _fmp_call_count += 1
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list) and data and data[0].get("price"):
                return float(data[0]["price"])
            elif isinstance(data, dict) and data.get("price"):
                return float(data["price"])
    except Exception:
        pass
    return None


def fetch_spy_history(from_date: str) -> dict:
    """Fetch SPY historical EOD closes in a single batch call.
    Returns {date_str: close_price} dict. Empty dict on failure (graceful degradation).
    """
    global _fmp_call_count
    if not FMP_KEY:
        return {}
    try:
        r = requests.get(f"{FMP_BASE}/historical-price-eod/light",
                         params={"symbol": "SPY", "from": from_date,
                                 "to": datetime.date.today().isoformat(), "apikey": FMP_KEY},
                         timeout=20)
        _fmp_call_count += 1
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list):
                return {rec["date"]: float(rec.get("price") or rec.get("close") or 0)
                        for rec in data if rec.get("date") and (rec.get("price") or rec.get("close"))}
    except Exception:
        pass
    return {}


# B2: In-memory cache for historical price lookups (keyed ticker → {date→close})
# Populated lazily by fetch_price_on_date; survives for the lifetime of one run.
_hist_price_cache: dict = {}
_delisted_tickers: set  = set()   # C5: tickers with empty FMP history (presumed delisted)
_DELISTED_RETURN  = -1.0          # C5: conservative assumption when data is absent post-delist


def fetch_price_on_date(ticker: str, date_str: str) -> float | None:
    """Return the EOD close price for *ticker* on *date_str* (YYYY-MM-DD) or the
    nearest trading day up to 5 calendar days forward.

    Fetches FMP /historical-price-eod/light once per ticker (full range) and
    caches in module-level _hist_price_cache so repeated calls for the same
    ticker are free.  Returns None on any failure so callers degrade gracefully.
    """
    global _fmp_call_count, _hist_price_cache
    if not FMP_KEY or not ticker or not date_str:
        return None

    # Populate cache for this ticker if not already done
    if ticker not in _hist_price_cache:
        try:
            # Fetch ~2 years of history — enough for 180d checkpoints on any pick
            from_dt = (datetime.date.today() - datetime.timedelta(days=730)).isoformat()
            r = requests.get(
                f"{FMP_BASE}/historical-price-eod/light",
                params={"symbol": ticker, "from": from_dt,
                        "to": datetime.date.today().isoformat(), "apikey": FMP_KEY},
                timeout=20,
            )
            _fmp_call_count += 1
            if r.status_code == 200:
                data = r.json()
                if isinstance(data, list):
                    _hist_price_cache[ticker] = {
                        rec["date"]: float(rec.get("price") or rec.get("close") or 0)
                        for rec in data
                        if rec.get("date") and (rec.get("price") or rec.get("close"))
                    }
                else:
                    _hist_price_cache[ticker] = {}   # mark as attempted / empty
            else:
                _hist_price_cache[ticker] = {}       # mark as attempted / error
        except Exception:
            _hist_price_cache[ticker] = {}

    prices = _hist_price_cache.get(ticker, {})
    if not prices:
        return None

    # Find price on date_str or up to +5 trading-day tolerance
    try:
        target = datetime.date.fromisoformat(date_str)
    except ValueError:
        return None

    for delta in range(6):   # 0 = exact, 1-5 = next available trading day
        candidate = (target + datetime.timedelta(days=delta)).isoformat()
        price = prices.get(candidate)
        if price:
            return price
    return None


def fetch_52w_ranges(tickers: list) -> dict:
    """Fetch 52-week high/low for enrichment candidates using individual /stable/quote calls.
    Uses the same parallel threading as _parallel_fetch. Cached with other fundamentals.
    Returns {ticker: {"yearHigh": float, "yearLow": float}}
    Note: FMP /stable/quote batch (comma-separated) returns [] on Starter plan — must use
    individual calls. ~4000 calls run in parallel threads, adds ~60s to fresh runs only.
    """
    global _fmp_call_count
    cache_key = "52w_ranges"
    cached = _cache.get(cache_key, {})

    # Only use cache if it covers at least 80% of requested tickers (prevents debug run
    # (20 stocks) from permanently polluting the production cache for 4000 stocks)
    missing = [t for t in tickers if t not in cached]
    if len(missing) <= len(tickers) * 0.20:
        print(f"  📦 Using cached 52w ranges ({len(cached)} stocks, {len(missing)} missing)")
        return cached

    if cached:
        print(f"  📊 Fetching 52w ranges for {len(missing)} stocks (cached {len(cached)}, parallel)...")
    else:
        print(f"  📊 Fetching 52w ranges for {len(tickers)} stocks (parallel)...")

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    fetch_list = missing if cached else tickers
    results = dict(cached)  # start from what we have
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)

    def _fetch_one(t):
        with _throttle:
            try:
                r = requests.get(f"{FMP_BASE}/quote",
                                 params={"symbol": t, "apikey": FMP_KEY}, timeout=10)
                time.sleep(0.2)
                if r.status_code == 200:
                    data = r.json()
                    item = data[0] if isinstance(data, list) and data else (data if isinstance(data, dict) else {})
                    yh = item.get("yearHigh"); yl = item.get("yearLow")
                    if yh and yl and float(yh) > 0:
                        return t, {"yearHigh": float(yh), "yearLow": float(yl)}
            except Exception:
                pass
            return t, None

    done_count = 0
    with ThreadPoolExecutor(max_workers=12) as pool:
        futures = {pool.submit(_fetch_one, t): t for t in fetch_list}
        for future in as_completed(futures):
            t, data = future.result()
            if data:
                with _lock:
                    results[t] = data
            done_count += 1
            if done_count % 500 == 0:
                print(f"    [{done_count}/{len(fetch_list)}] 52w ranges fetched...")

    _cache[cache_key] = results
    print(f"  ✅ 52w ranges: {len(results)} stocks ({len(results)-len(cached)} new)")
    return results


def fetch_sparkline_data(tickers: list) -> dict:
    """Fetch 5-year monthly price history for AI pick sparkline mini-charts.

    Returns {ticker: [float, ...]} — 60 evenly-sampled close prices, oldest→newest.
    Cached for 7 days (sparklines only need weekly refresh for a 5Y view).
    Uses FMP /historical-price-full/{ticker}?serietype=line&from=5Y_AGO.
    """
    global _fmp_call_count
    SKEY = "_sparklines"
    DAYS = 7
    today_str = datetime.date.today().isoformat()
    five_yr_ago = (datetime.date.today() - datetime.timedelta(days=365 * 5)).strftime("%Y-%m-%d")

    cached = _cache.get(SKEY, {})

    def _fresh(entry):
        if not isinstance(entry, dict):
            return False
        try:
            age = (datetime.date.today() -
                   datetime.date.fromisoformat(entry.get("d", "2000-01-01"))).days
            return age < DAYS
        except Exception:
            return False

    missing = [t for t in tickers if t not in cached or not _fresh(cached[t])]
    if not missing:
        return {t: cached[t]["p"] for t in tickers if t in cached and _fresh(cached[t])}

    print(f"  📈 Fetching 5Y sparklines for {len(missing)} tickers"
          f"{' (+ ' + str(len(tickers) - len(missing)) + ' cached)' if len(tickers) > len(missing) else ''}...")

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    results = {t: cached[t]["p"] for t in tickers if t in cached and _fresh(cached[t])}
    _lock = threading.Lock()
    _throttle = threading.Semaphore(8)

    def _sample(prices, n=60):
        """Sample a list of floats down to n evenly-spaced points."""
        if len(prices) <= n:
            return prices
        step = (len(prices) - 1) / (n - 1)
        return [prices[round(i * step)] for i in range(n)]

    def _fetch_one(t):
        with _throttle:
            try:
                r = requests.get(
                    f"{FMP_BASE}/historical-price-eod/light",
                    params={"symbol": t, "from": five_yr_ago,
                            "to": datetime.date.today().isoformat(), "apikey": FMP_KEY},
                    timeout=15,
                )
                time.sleep(0.08)
                if r.status_code == 200:
                    data = r.json()
                    # Response is a list [{"date":"...","close":X}, ...] newest-first
                    if isinstance(data, list) and data:
                        prices = [float(h["close"]) for h in reversed(data)
                                  if "close" in h and h["close"]]
                        if prices:
                            return t, _sample(prices)
            except Exception:
                pass
            return t, None

    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_fetch_one, t): t for t in missing}
        for future in as_completed(futures):
            t, prices = future.result()
            if prices:
                with _lock:
                    results[t] = prices
                    cached[t] = {"p": prices, "d": today_str}

    _cache[SKEY] = cached
    _fmp_call_count += len(missing)
    return results


def fetch_macro_indicators() -> dict:
    """Fetch live macro indicators from FRED via the ivo-welch.info CSV gateway.

    Completely free — no API key needed. Same FRED data used by Fed economists.
    Cached for 1 day via the standard _cache system.
    Returns a dict with values + pre-computed signals (or empty dict on full failure).
    """
    cache_key = "macro_indicators"
    cached = _cache.get(cache_key)
    if cached and isinstance(cached, dict) and cached.get("as_of"):
        # Check age of the macro data itself using its as_of date (not the cache file
        # timestamp, which updates on every save_cache() call regardless of macro staleness).
        try:
            as_of_str = cached.get("as_of", "")
            if as_of_str and as_of_str != "unknown":
                data_age_days = (
                    datetime.datetime.now()
                    - datetime.datetime.strptime(as_of_str, "%Y-%m-%d")
                ).total_seconds() / 86400
                if data_age_days < 1:
                    print(f"  📦 Using cached macro indicators (as of {as_of_str})")
                    return cached
                # Data is 1+ days old — fall through to refetch
        except Exception:
            pass

    print("  🌍 Fetching macro indicators from FRED...")

    FRED_URL = "https://www.ivo-welch.info/cgi-bin/fredwrap?symbol={}"
    # Each value is a list of fallback series IDs — try left-to-right, use first that returns data.
    # FEDFUNDS occasionally drops out of the gateway; DFF (daily fed funds) is a near-equivalent fallback.
    SERIES = {
        "dgs10":    ["DGS10"],                # 10Y Treasury yield (%) — daily
        "dgs2":     ["GS2"],                  # 2Y Treasury yield (%) — monthly
        "t10y2y":   ["T10Y2Y"],               # Yield curve spread 10Y-2Y (%) — daily
        "vix":      ["VIXCLS"],               # VIX fear index — daily
        "fedfunds": ["FEDFUNDS", "DFF"],      # Fed Funds rate — monthly preferred, daily fallback
        "cpi":      ["CPIAUCSL"],             # CPI all urban consumers (for YoY calc)
        "unrate":   ["UNRATENSA"],            # Unemployment rate NSA (%)
    }

    def _fetch_one_series(series_id):
        """Fetch one FRED series; return list of (date_str, value) or None on failure."""
        try:
            r = requests.get(FRED_URL.format(series_id), timeout=12)
            if r.status_code != 200:
                return None
            lines = [ln.strip() for ln in r.text.strip().splitlines()
                     if ln.strip() and not ln.startswith("#")]
            data_rows = []
            for ln in lines:
                parts = ln.split(",")
                if len(parts) >= 2:
                    try:
                        val = float(parts[1])
                        data_rows.append((parts[0].strip(), val))
                    except ValueError:
                        continue
            return data_rows or None
        except Exception:
            return None

    raw = {}
    for key, fallback_ids in SERIES.items():
        success_id = None
        for series_id in fallback_ids:
            data_rows = _fetch_one_series(series_id)
            if data_rows:
                raw[key] = data_rows
                success_id = series_id
                if series_id != fallback_ids[0]:
                    print(f"    ℹ️ FRED {fallback_ids[0]} unavailable — using fallback {series_id}")
                break
        if not success_id:
            # All fallbacks failed — try last-known cached value if present
            cached_last = (_cache.get(cache_key) or {}).get(f"_last_{key}")
            if cached_last:
                raw[key] = cached_last
                print(f"    ℹ️ FRED {fallback_ids[0]}: all sources unavailable, using last-known cached value")
            else:
                print(f"    ⚠️ FRED {fallback_ids[0]}: unavailable and no cached fallback")

    if not raw:
        print("  ⚠️ All FRED fetches failed — macro indicators unavailable")
        return {}

    def _latest(k):
        rows = raw.get(k, [])
        # Skip NaN / missing values (FRED uses "." for missing)
        for date_str, val in reversed(rows):
            if val is not None and not (isinstance(val, float) and math.isnan(val)):
                return date_str, val
        return None, None

    # Extract latest values
    _dgs10_date, dgs10    = _latest("dgs10")
    _dgs2_date,  dgs2     = _latest("dgs2")
    _t10y2y_date, t10y2y = _latest("t10y2y")
    _vix_date,   vix      = _latest("vix")
    _ff_date,    fedfunds = _latest("fedfunds")
    _ur_date,    unrate   = _latest("unrate")

    # CPI YoY: compare latest to same month 12 months prior
    cpi_yoy = None
    cpi_as_of = None
    cpi_rows = raw.get("cpi", [])
    if len(cpi_rows) >= 13:
        # Latest value
        latest_date, latest_cpi = _latest("cpi")
        # Find value ~12 months prior (same month last year)
        if latest_cpi and latest_date:
            cpi_as_of = latest_date
            try:
                latest_yr = int(latest_date[:4])
                latest_mo = int(latest_date[4:6]) if len(latest_date) >= 6 else None
                prior_val = None
                for date_str, val in cpi_rows:
                    try:
                        yr = int(date_str[:4])
                        mo = int(date_str[4:6]) if len(date_str) >= 6 else None
                        if yr == latest_yr - 1 and mo == latest_mo:
                            prior_val = val
                            break
                    except (ValueError, IndexError):
                        continue
                if prior_val and prior_val > 0:
                    cpi_yoy = round((latest_cpi / prior_val - 1) * 100, 2)
            except (ValueError, IndexError):
                pass

    # Yield curve: prefer FRED T10Y2Y, fall back to computed
    if t10y2y is None and dgs10 is not None and dgs2 is not None:
        t10y2y = round(dgs10 - dgs2, 2)

    # Determine as_of date (most recent across all fetched series)
    as_of_candidates = [d for d in [_dgs10_date, _vix_date, _ff_date, _ur_date] if d]
    as_of = max(as_of_candidates) if as_of_candidates else "unknown"
    # Format nicely: YYYYMMDD → YYYY-MM-DD
    try:
        if len(as_of) == 8:
            as_of = f"{as_of[:4]}-{as_of[4:6]}-{as_of[6:8]}"
    except Exception:
        pass

    # ── Pre-compute signals for AI and color-coding ──────────────────────────
    def _curve_signal(v):
        if v is None: return "UNKNOWN"
        if v < -0.10: return "INVERTED"
        if v < 0.50:  return "FLAT"
        if v < 1.50:  return "NORMAL"
        return "STEEP"

    def _vix_signal(v):
        if v is None: return "UNKNOWN"
        if v > 40: return "PANIC"
        if v > 25: return "FEAR"
        if v > 18: return "CAUTION"
        return "CALM"

    def _rate_signal(v):
        if v is None: return "UNKNOWN"
        if v > 5.0: return "HIGH"
        if v > 4.0: return "ELEVATED"
        if v > 2.0: return "NORMAL"
        return "LOW"

    def _inflation_signal(v):
        if v is None: return "UNKNOWN"
        if v > 5.0: return "HOT"
        if v > 3.0: return "ABOVE_TARGET"
        if v > 2.0: return "NEAR_TARGET"
        return "LOW"

    def _labor_signal(v):
        if v is None: return "UNKNOWN"
        if v < 4.0: return "TIGHT"
        if v < 5.0: return "HEALTHY"
        if v < 6.5: return "SOFTENING"
        return "WEAK"

    macro = {
        "dgs10":            round(dgs10, 2)    if dgs10    is not None else None,
        "dgs2":             round(dgs2, 2)     if dgs2     is not None else None,
        "yield_curve":      round(t10y2y, 2)   if t10y2y   is not None else None,
        "vix":              round(vix, 1)      if vix      is not None else None,
        "fedfunds":         round(fedfunds, 2) if fedfunds is not None else None,
        "cpi_yoy":          cpi_yoy,
        "unrate":           round(unrate, 1)   if unrate   is not None else None,
        "curve_signal":     _curve_signal(t10y2y),
        "vix_signal":       _vix_signal(vix),
        "rate_signal":      _rate_signal(dgs10),
        "inflation_signal": _inflation_signal(cpi_yoy),
        "labor_signal":     _labor_signal(unrate),
        "as_of":            as_of,
    }

    # ── Historical series for HTML sparklines (last 500 pts per series) ─────
    def _clean_hist(k, n=500):
        rows = raw.get(k, [])
        return [(d, v) for d, v in rows
                if v is not None and not (isinstance(v, float) and math.isnan(v))][-n:]

    # Rolling CPI YoY series for the chart (monthly data → YoY %)
    _cpi_raw = [(d, v) for d, v in raw.get("cpi", [])
                if v is not None and not (isinstance(v, float) and math.isnan(v))]
    _cpi_yoy_hist = []
    for _i in range(12, len(_cpi_raw)):
        _cd, _cv = _cpi_raw[_i]
        _, _pv   = _cpi_raw[_i - 12]
        if _cv and _pv and _pv > 0:
            _cpi_yoy_hist.append((_cd, round((_cv / _pv - 1) * 100, 2)))

    macro["_hist_dgs10"]    = _clean_hist("dgs10")
    macro["_hist_dgs2"]     = _clean_hist("dgs2")
    macro["_hist_t10y2y"]   = _clean_hist("t10y2y")
    macro["_hist_vix"]      = _clean_hist("vix")
    macro["_hist_fedfunds"] = _clean_hist("fedfunds")
    macro["_hist_cpi_yoy"]  = _cpi_yoy_hist[-60:]   # last ~5 years of monthly data
    macro["_hist_unrate"]   = _clean_hist("unrate")

    # Persist raw last-known series for next-run fallback (C1: FEDFUNDS resilience)
    for _k, _rows in raw.items():
        if _rows:
            macro[f"_last_{_k}"] = _rows[-30:]   # last 30 datapoints is plenty for fallback

    _cache[cache_key] = macro
    fetched = [k for k in ["dgs10","dgs2","yield_curve","vix","fedfunds","cpi_yoy","unrate"]
               if macro.get(k) is not None]
    print(f"  ✅ Macro indicators fetched: {', '.join(fetched)} (as of {as_of})")
    return macro


def fetch_market_intelligence() -> dict:
    """Fetch live market intelligence for agents that benefit from real-world context.

    Three slices:
    - consumer_trends  → for Social Arbitrage (what's trending in consumer / apps)
    - tech_trends      → for Disruptive Innovation (AI / biotech / robotics news)
    - insider_activity → for Insider Track (recent cluster buying summaries)

    Cached for 4 hours. Falls back to empty strings on any error.
    """
    cache_key = "market_intelligence"
    cached = _cache.get(cache_key)
    if cached and isinstance(cached, dict) and cached.get("fetched_at"):
        try:
            age_h = ((datetime.datetime.now()
                      - datetime.datetime.fromisoformat(cached["fetched_at"])
                      ).total_seconds() / 3600)
            if age_h < 4:
                return cached
        except Exception:
            pass

    result = {
        "fetched_at": datetime.datetime.now().isoformat(),
        "consumer_trends": "",
        "tech_trends": "",
        "insider_activity": "",
        "catalyst_news": "",
    }

    # ── FMP Stock News (free on most plans) ──────────────────────────────────
    try:
        news_raw = []
        if FMP_KEY:
            r = requests.get(
                f"{FMP_BASE}/v3/stock_news",
                params={"limit": 50, "apikey": FMP_KEY},
                timeout=10,
            )
            if r.status_code == 200:
                news_raw = r.json() or []

        consumer_kw  = {"consumer", "retail", "restaurant", "app", "brand", "viral",
                        "launch", "trend", "e-commerce", "subscription", "social"}
        tech_kw      = {"ai", "artificial intelligence", "machine learning", "robotics",
                        "biotech", "genomics", "semiconductor", "cloud", "saas",
                        "quantum", "autonomous", "electric", "ev", "innovation"}
        insider_kw   = {"insider", "executive", "ceo buys", "director buys",
                        "bought shares", "form 4"}
        catalyst_kw  = {"merger", "acquisition", "acqui-hire", "spinoff", "spin-off",
                        "strategic review", "strategic alternatives", "going private",
                        "takeover", "tender offer", "proxy fight", "activist",
                        "fda", "ema", "approval", "approved", "rejection", "rejected",
                        "settlement", "doj", "sec investigation", "regulatory",
                        "restructuring", "bankruptcy", "debt exchange",
                        "buyback", "share repurchase", "special dividend",
                        "earnings beat", "earnings miss", "guidance raised",
                        "guidance cut", "restatement", "earnings surprise"}

        consumer_lines, tech_lines, insider_lines, catalyst_lines = [], [], [], []
        for item in news_raw:
            title  = (item.get("title") or "").strip()
            ticker = (item.get("symbol") or "").strip()
            date   = (item.get("publishedDate") or "")[:10]
            if not title:
                continue
            tl  = title.lower()
            tag = f"[{date}] {ticker + ': ' if ticker else ''}{title}"
            if any(w in tl for w in consumer_kw):
                consumer_lines.append(tag)
            if any(w in tl for w in tech_kw):
                tech_lines.append(tag)
            if any(w in tl for w in insider_kw):
                insider_lines.append(tag)
            if any(w in tl for w in catalyst_kw):
                catalyst_lines.append(tag)

        if consumer_lines:
            result["consumer_trends"] = (
                "RECENT CONSUMER & APP TREND NEWS (FMP, last 48h):\n"
                + "\n".join(consumer_lines[:10])
            )
        if tech_lines:
            result["tech_trends"] = (
                "RECENT TECH / INNOVATION NEWS (FMP, last 48h):\n"
                + "\n".join(tech_lines[:10])
            )
        if insider_lines:
            result["insider_activity"] = (
                "RECENT INSIDER / EXECUTIVE BUYING NEWS (FMP, last 48h):\n"
                + "\n".join(insider_lines[:8])
            )
        if catalyst_lines:
            result["catalyst_news"] = (
                "RECENT M&A / REGULATORY / CATALYST NEWS (FMP, last 48h):\n"
                + "\n".join(catalyst_lines[:12])
            )
    except Exception as e:
        pass  # fail silently — agents still work without this context

    _cache[cache_key] = result
    return result


def fetch_special_sit_news(tickers: list) -> str:
    """Fetch company-specific news for the top Special Situation candidates.

    Calls FMP /v3/stock_news with a ticker filter so the SpecSit agent has
    real headline context (M&A, FDA, restructuring, etc.) rather than having
    to infer catalysts purely from screener numbers.

    Cached 4h per unique ticker set. Falls back to empty string on any error.
    Up to 15 tickers, 3 headlines per company max.
    """
    if not tickers or not FMP_KEY:
        return ""

    tickers = [t for t in tickers if t][:15]
    cache_key = f"ss_news_{'_'.join(sorted(tickers))}"
    cached = _cache.get(cache_key)
    if cached and isinstance(cached, dict) and cached.get("fetched_at"):
        try:
            age_h = (
                (datetime.datetime.now()
                 - datetime.datetime.fromisoformat(cached["fetched_at"])
                 ).total_seconds() / 3600
            )
            if age_h < 4:
                return cached.get("block", "")
        except Exception:
            pass

    try:
        r = requests.get(
            f"{FMP_BASE}/v3/stock_news",
            params={"tickers": ",".join(tickers), "limit": 60, "apikey": FMP_KEY},
            timeout=12,
        )
        if r.status_code != 200:
            return ""

        news_raw = r.json() or []

        # Group by ticker — keep 3 most recent headlines per company
        by_ticker: dict[str, list] = {}
        for item in news_raw:
            t = (item.get("symbol") or "").strip().upper()
            if t not in by_ticker:
                by_ticker[t] = []
            if len(by_ticker[t]) < 3:
                by_ticker[t].append(item)

        lines = []
        for t in tickers:
            items = by_ticker.get(t.upper(), [])
            for item in items:
                date  = (item.get("publishedDate") or "")[:10]
                title = (item.get("title") or "").strip()
                if title:
                    lines.append(f"  {t} [{date}]: {title}")

        if not lines:
            _cache[cache_key] = {"fetched_at": datetime.datetime.now().isoformat(), "block": ""}
            return ""

        block = (
            "COMPANY-SPECIFIC NEWS FOR TOP SPECIAL-SITUATION CANDIDATES "
            "(FMP, last 7 days — up to 3 headlines each):\n"
            + "\n".join(lines)
        )
        _cache[cache_key] = {"fetched_at": datetime.datetime.now().isoformat(), "block": block}
        return block

    except Exception:
        return ""


def fmp_get(endpoint: str, params: dict = None) -> dict | list | None:
    """Make an FMP API call. Returns parsed JSON or None on failure."""
    global _fmp_call_count
    if not FMP_KEY:
        return None
    url = f"{FMP_BASE}/{endpoint}"
    p = dict(params or {})
    p["apikey"] = FMP_KEY
    try:
        r = requests.get(url, params=p, timeout=15)
        _fmp_call_count += 1
        if r.status_code == 200:
            data = r.json()
            # FMP returns {"Error Message": "..."} on some errors
            if isinstance(data, dict) and "Error Message" in data:
                print(f"  ⚠️ FMP error on {endpoint}: {data['Error Message'][:80]}")
                return None
            return data
        elif r.status_code == 403:
            print(f"  ⚠️ FMP 403 (endpoint may require subscription): {endpoint}")
            return None
        elif r.status_code == 429:
            print(f"  ⚠️ FMP rate limited — waiting 5s...")
            time.sleep(5)
            r = requests.get(url, params=p, timeout=15)
            _fmp_call_count += 1
            if r.status_code == 200:
                return r.json()
        else:
            print(f"  ⚠️ FMP {r.status_code} on {endpoint}")
    except Exception as e:
        print(f"  ⚠️ FMP request error: {e}")
    return None


def fmp_get_batch(tickers: list, endpoint_template: str, batch_size: int = 5,
                  delay: float = 0.15) -> dict:
    """Fetch data for multiple tickers, returns {ticker: data_dict}.
    endpoint_template should have {} for ticker, e.g. 'v3/rating/{}'
    """
    results = {}
    for i in range(0, len(tickers), batch_size):
        batch = tickers[i:i + batch_size]
        for t in batch:
            ep = endpoint_template.format(t)
            data = fmp_get(ep)
            if data and isinstance(data, list) and data:
                results[t] = data[0]
            elif data and isinstance(data, dict):
                results[t] = data
            time.sleep(delay)
        # Progress
        done = min(i + batch_size, len(tickers))
        if done % 50 == 0 or done == len(tickers):
            print(f"    [{done}/{len(tickers)}] fetched...")
    return results


# ─────────────────────────────────────────────
# CACHE SYSTEM
# ─────────────────────────────────────────────

_cache = {}


def load_cache() -> dict:
    """Load FMP screener cache from disk."""
    global _cache
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, "r") as f:
                data = json.load(f)
            ts = data.get("_timestamp", "")
            if ts:
                age = (datetime.datetime.now() - datetime.datetime.fromisoformat(ts)).total_seconds() / 86400
                if age < CACHE_DAYS:
                    _cache = data
                    n = len(data.get("universe", {}))
                    print(f"  📦 FMP cache loaded ({n} stocks, age: {age:.1f}d)")
                    return _cache
                else:
                    print(f"  🔄 FMP cache expired ({age:.0f}d > {CACHE_DAYS}d TTL)")
    except Exception as e:
        print(f"  ⚠️ Cache load error: {e}")
    _cache = {}
    return _cache


def save_cache():
    """Save FMP screener cache to disk."""
    _cache["_timestamp"] = datetime.datetime.now().isoformat()
    try:
        with open(CACHE_FILE, "w") as f:
            json.dump(_cache, f, default=str)
        print(f"  💾 FMP cache saved ({len(_cache.get('universe', {}))} stocks)")
    except Exception as e:
        print(f"  ⚠️ Cache save error: {e}")


# ─────────────────────────────────────────────
# PHASE 1: UNIVERSE DISCOVERY
# ─────────────────────────────────────────────

def fetch_us_universe() -> dict:
    """Fetch all US stocks from NASDAQ + NYSE + AMEX via FMP.
    Strategy 1: stock-screener endpoint (may be premium)
    Strategy 2: stock/list endpoint (free tier — gets all symbols)
    Returns {ticker: {symbol, name, sector, industry, mktCap, price, ...}}
    """
    # Always fetch fresh universe so prices are current (only 3 API calls)
    print("\n  🌐 Fetching US stock universe from FMP (fresh prices)...")
    universe = {}

    # Strategy 1: stock-screener (lowercase exchange names per FMP docs)
    for exchange in ["nasdaq", "nyse", "amex"]:
        data = fmp_get("company-screener", {
            "exchange": exchange,
            "marketCapMoreThan": 50_000_000,
            "isActivelyTrading": "true",
            "limit": 5000,
        })
        if data and isinstance(data, list) and len(data) > 10:
            for stock in data:
                t = stock.get("symbol", "")
                if not t or len(t) > 6: continue
                universe[t] = {
                    "symbol": t,
                    "name": (stock.get("companyName") or "")[:30],
                    "sector": stock.get("sector") or "Unknown",
                    "industry": stock.get("industry") or "",
                    "mktCap": stock.get("marketCap") or 0,
                    "price": stock.get("price") or 0,
                    "beta": stock.get("beta"),
                    "volume": stock.get("volume"),
                    "exchange": exchange.upper(),
                    # 52-week range — included in FMP screener response
                    "yearHigh": stock.get("yearHigh") or stock.get("52WeekHigh"),
                    "yearLow":  stock.get("yearLow")  or stock.get("52WeekLow"),
                }
            print(f"    ✅ {exchange.upper()}: {len(data)} stocks")
        else:
            print(f"    ⚠️ {exchange.upper()} screener: {len(data) if data else 'no'} results")
        time.sleep(0.3)

    # Strategy 2 fallback: if screener failed, use symbol list + profiles
    if len(universe) < 100:
        print("  🔄 Screener returned too few results — using symbol list fallback...")
        # Get all tradeable symbols
        all_symbols = fmp_get("stock-list")
        if all_symbols and isinstance(all_symbols, list):
            us_symbols = [s for s in all_symbols
                          if s.get("exchangeShortName") in ("NASDAQ", "NYSE", "AMEX")
                          and s.get("type") == "stock"
                          and s.get("symbol")
                          and len(s.get("symbol", "")) <= 5
                          and "." not in s.get("symbol", "")
                          and "-" not in s.get("symbol", "")]
            print(f"    📋 Found {len(us_symbols)} US stock symbols")

            # Batch fetch profiles (30 at a time)
            for i in range(0, len(us_symbols), 30):
                batch = us_symbols[i:i + 30]
                batch_str = ",".join(s["symbol"] for s in batch)
                profiles = fmp_get(f"profile", {"symbol": batch_str})
                if profiles and isinstance(profiles, list):
                    for p in profiles:
                        t = p.get("symbol", "")
                        mktCap = p.get("mktCap") or 0
                        if not t or mktCap < 50_000_000: continue
                        if not p.get("isActivelyTrading", True): continue
                        universe[t] = {
                            "symbol": t,
                            "name": (p.get("companyName") or "")[:30],
                            "sector": p.get("sector") or "Unknown",
                            "industry": p.get("industry") or "",
                            "mktCap": mktCap,
                            "price": p.get("price") or 0,
                            "beta": p.get("beta"),
                            "volume": p.get("volAvg"),
                            "exchange": p.get("exchangeShortName") or "",
                        }
                done = min(i + 30, len(us_symbols))
                if done % 300 == 0:
                    print(f"    [{done}/{len(us_symbols)}] profiles loaded...")
                time.sleep(0.15)

    _cache["universe"] = universe
    print(f"  ✅ US Universe: {len(universe)} stocks loaded")
    return universe


# ─────────────────────────────────────────────
# PHASE 2: BULK FUNDAMENTAL DATA
# ─────────────────────────────────────────────

def fetch_bulk_ratios(universe: dict) -> dict:
    """Fetch key financial ratios for all stocks.
    Uses FMP key-metrics-ttm endpoint (1 call per stock, batched).
    Returns {ticker: {pe, pb, roe, roa, fcfYield, ...}}
    """
    cache_key = "ratios"
    if _cache.get(cache_key):
        print(f"  📦 Using cached ratios ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    print("\n  📊 Fetching financial ratios from FMP...")
    tickers = list(universe.keys())

    # Use bulk profile endpoint — gets basic ratios for many at once
    # FMP /v3/profile/AAPL,MSFT,... supports up to ~50 tickers per call
    ratios = {}
    batch_size = 30
    for i in range(0, len(tickers), batch_size):
        batch = tickers[i:i + batch_size]
        batch_str = ",".join(batch)
        data = fmp_get(f"profile", {"symbol": batch_str})
        if data and isinstance(data, list):
            for item in data:
                t = item.get("symbol", "")
                if t:
                    ratios[t] = {
                        "pe": item.get("pe"),
                        "price": item.get("price"),
                        "mktCap": item.get("mktCap"),
                        "beta": item.get("beta"),
                        "volAvg": item.get("volAvg"),
                        "lastDiv": item.get("lastDiv"),
                        "sector": item.get("sector"),
                        "industry": item.get("industry"),
                        "name": (item.get("companyName") or "")[:30],
                        "exchange": item.get("exchangeShortName"),
                        "isActivelyTrading": item.get("isActivelyTrading"),
                    }
        done = min(i + batch_size, len(tickers))
        if done % 300 == 0:
            print(f"    [{done}/{len(tickers)}] profiles fetched...")
        time.sleep(0.15)

    _cache[cache_key] = ratios
    print(f"  ✅ Profiles loaded: {len(ratios)} stocks")
    return ratios


def _parallel_fetch(tickers: list, endpoint: str, label: str,
                    cache_key: str, params_extra: dict = None) -> dict:
    """Generic parallel fetcher for per-ticker FMP endpoints.
    Uses ThreadPoolExecutor with rate-limit-aware throttling.
    """
    if _cache.get(cache_key):
        print(f"  📦 Using cached {label} ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    print(f"\n  📊 Fetching {label}...")
    results = {}
    total = len(tickers)
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)  # max 4 concurrent — stay under 300/min

    def _fetch_one(t):
        with _throttle:
            p = {"symbol": t}
            if params_extra:
                p.update(params_extra)
            data = fmp_get(endpoint, p)
            time.sleep(0.2)  # ~200 req/min with 4 workers
            if data and isinstance(data, list) and data:
                return t, data[0]
            elif data and isinstance(data, dict) and "Error" not in str(data)[:50]:
                return t, data
            return t, None

    with ThreadPoolExecutor(max_workers=12) as pool:
        futures = {pool.submit(_fetch_one, t): t for t in tickers}
        done_count = 0
        for future in as_completed(futures):
            t, data = future.result()
            if data:
                with _lock:
                    results[t] = data
            done_count += 1
            if done_count % 100 == 0:
                print(f"    [{done_count}/{total}] {label} fetched...")

    _cache[cache_key] = results
    print(f"  ✅ {label} loaded: {len(results)} stocks")
    return results


def fetch_key_metrics(tickers: list) -> dict:
    return _parallel_fetch(tickers, "key-metrics-ttm", "key metrics", "key_metrics")


def fetch_dcf_bulk(tickers: list) -> dict:
    return _parallel_fetch(tickers, "discounted-cash-flow", "DCF values", "dcf")


def fetch_growth_estimates(tickers: list) -> dict:
    """Fetch analyst estimates — store full list for CAGR calculation."""
    cache_key = "estimates"
    if _cache.get(cache_key):
        print(f"  📦 Using cached growth estimates ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    print(f"\n  📊 Fetching growth estimates...")
    results = {}
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)

    def _f(t):
        with _throttle:
            data = fmp_get("analyst-estimates", {"symbol": t, "period": "annual", "limit": 5})
            time.sleep(0.2)
            if data and isinstance(data, list) and len(data) >= 2:
                return t, data  # full list, not data[0]
            return t, None

    with ThreadPoolExecutor(max_workers=6) as pool:
        futs = {pool.submit(_f, t): t for t in tickers}
        done = 0
        for fut in as_completed(futs):
            t, data = fut.result()
            if data:
                with _lock: results[t] = data
            done += 1
            if done % 100 == 0: print(f"    [{done}/{len(tickers)}] estimates fetched...")

    _cache[cache_key] = results
    print(f"  ✅ estimates loaded: {len(results)} stocks")
    return results


def fetch_financial_scores(tickers: list) -> dict:
    return _parallel_fetch(tickers, "financial-scores", "Piotroski scores", "scores")


def fetch_ratings(tickers: list) -> dict:
    """Ratings endpoint not available on stable API — skip."""
    print("\n  ⏭️ Skipping ratings (not available on stable API)")
    return {}


def fetch_financial_growth(tickers: list) -> dict:
    """Fetch 5 years of annual growth metrics — keeps full list for trend/acceleration detection.
    Most recent year = data[0], previous year = data[1], etc.
    """
    cache_key = "growth"
    if _cache.get(cache_key):
        print(f"  📦 Using cached growth metrics ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    print(f"\n  📊 Fetching growth metrics...")
    results = {}
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)

    def _f(t):
        with _throttle:
            data = fmp_get("financial-growth", {"symbol": t, "period": "annual", "limit": "5"})
            time.sleep(0.2)
            if data and isinstance(data, list) and data:
                return t, data  # return full list for trend detection
            return t, None

    with ThreadPoolExecutor(max_workers=12) as pool:
        futs = {pool.submit(_f, t): t for t in tickers}
        done = 0
        for fut in as_completed(futs):
            t, data = fut.result()
            if data:
                with _lock: results[t] = data
            done += 1
            if done % 100 == 0: print(f"    [{done}/{len(tickers)}] growth metrics fetched...")

    _cache[cache_key] = results
    print(f"  ✅ growth metrics loaded: {len(results)} stocks")
    return results


def fetch_ratios_ttm(tickers: list) -> dict:
    """Fetch financial ratios TTM (ROE, ROA, P/B, D/E, Div Yield, Gross Margin).
    This is the dedicated ratios endpoint — more complete than key-metrics for profitability.
    """
    return _parallel_fetch(tickers, "ratios-ttm", "financial ratios TTM", "ratios_ttm")


def fetch_balance_sheet(tickers: list) -> dict:
    """Fetch latest annual balance sheet for net cash calculation.
    Net cash = cash + short-term investments - total debt
    Used to find cash-rich companies (asset plays, turnarounds).
    """
    return _parallel_fetch(tickers, "balance-sheet-statement", "balance sheets", "balance_sheet",
                           {"period": "annual", "limit": "1"})


# ─────────────────────────────────────────────
# CAPITAL ALLOCATOR DATA — CEO + 5Y financials
# ─────────────────────────────────────────────

def fetch_balance_sheet_5y(tickers: list) -> dict:
    """6Y annual balance sheets — for shares-outstanding history, debt trend, equity.
    Returns full list (newest first) per ticker."""
    cache_key = "balance_sheet_5y"
    if _cache.get(cache_key):
        print(f"  📦 Using cached 5Y balance sheets ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    print(f"\n  📊 Fetching 5Y balance sheets...")
    results = {}
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)

    def _f(t):
        with _throttle:
            data = fmp_get("balance-sheet-statement",
                           {"symbol": t, "period": "annual", "limit": "6"})
            time.sleep(0.2)
            if data and isinstance(data, list) and len(data) >= 2:
                return t, data
            return t, None

    with ThreadPoolExecutor(max_workers=12) as pool:
        futs = {pool.submit(_f, t): t for t in tickers}
        done = 0
        for fut in as_completed(futs):
            t, data = fut.result()
            if data:
                with _lock: results[t] = data
            done += 1
            if done % 200 == 0:
                print(f"    [{done}/{len(tickers)}] 5Y balance sheets fetched...")

    _cache[cache_key] = results
    print(f"  ✅ 5Y balance sheets loaded: {len(results)} stocks")
    return results


def fetch_cash_flow_5y(tickers: list) -> dict:
    """6Y annual cash-flow statements — capex, FCF, repurchases, dividends, M&A spend."""
    cache_key = "cash_flow_5y"
    if _cache.get(cache_key):
        print(f"  📦 Using cached 5Y cash flows ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    print(f"\n  📊 Fetching 5Y cash flow statements...")
    results = {}
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)

    def _f(t):
        with _throttle:
            data = fmp_get("cash-flow-statement",
                           {"symbol": t, "period": "annual", "limit": "6"})
            time.sleep(0.2)
            if data and isinstance(data, list) and len(data) >= 2:
                return t, data
            return t, None

    with ThreadPoolExecutor(max_workers=12) as pool:
        futs = {pool.submit(_f, t): t for t in tickers}
        done = 0
        for fut in as_completed(futs):
            t, data = fut.result()
            if data:
                with _lock: results[t] = data
            done += 1
            if done % 200 == 0:
                print(f"    [{done}/{len(tickers)}] 5Y cash flows fetched...")

    _cache[cache_key] = results
    print(f"  ✅ 5Y cash flows loaded: {len(results)} stocks")
    return results


def fetch_key_executives(tickers: list) -> dict:
    """Fetch key executives — used to extract CEO name + tenure (since field).
    Returns {ticker: list of officer dicts}."""
    cache_key = "key_executives"
    if _cache.get(cache_key):
        print(f"  📦 Using cached key executives ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    print(f"\n  📊 Fetching key executives (CEO tenure)...")
    results = {}
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)

    def _f(t):
        with _throttle:
            data = fmp_get("key-executives", {"symbol": t})
            time.sleep(0.2)
            if data and isinstance(data, list) and data:
                return t, data
            return t, None

    with ThreadPoolExecutor(max_workers=12) as pool:
        futs = {pool.submit(_f, t): t for t in tickers}
        done = 0
        for fut in as_completed(futs):
            t, data = fut.result()
            if data:
                with _lock: results[t] = data
            done += 1
            if done % 200 == 0:
                print(f"    [{done}/{len(tickers)}] executives fetched...")

    _cache[cache_key] = results
    print(f"  ✅ key executives loaded: {len(results)} stocks")
    return results


def _extract_ceo(executives: list) -> dict | None:
    """Find the current CEO from the executives list. Returns
    {name, since, tenure_years} or None if no CEO found / data invalid.
    FMP's titleSince is often null/0; when missing we return tenure_years=None
    so the caller can still score on available financial data.

    Title quality ranking (higher = better match for company-level CEO):
      - Co-CEO / CEO of a division / "of [X]" qualifier → low score (divisional)
      - "Chief Executive Officer" or "Chairman & CEO" without qualifier → high score
    """
    if not executives:
        return None

    def _title_quality(raw_title: str) -> int:
        t = raw_title.lower()
        # Must contain a CEO marker
        if "chief executive officer" not in t and not t.endswith("ceo") \
           and " ceo" not in t and "ceo " not in t and t != "ceo":
            return -1
        score = 10
        # Divisional / sub-entity CEO: penalise heavily
        # Catches: "CEO of Sam's Club", "CEO of Walmart US", "CEO of Commercial & IB"
        if " of " in t and ("chief executive officer of" in t or "ceo of" in t or "& ceo of" in t):
            score -= 8
        # "Executive VP" / "EVP" prefix strongly suggests a divisional role, not the company CEO
        if t.startswith("executive vp") or t.startswith("evp") or t.startswith("executive vice president"):
            score -= 6
        # Only penalise actual co-CEO (not "co-founder")
        if "co-chief executive" in t or "co- chief executive" in t \
           or t.startswith("co-ceo") or "& co-ceo" in t or ", co-ceo" in t:
            score -= 5   # "Co-CEO" / "Co- Chief Executive Officer"
        # Prefer Chairman+CEO combos (usually the #1 executive)
        if "chairman" in t:
            score += 2
        if "president" in t and "chief executive officer" in t:
            score += 1   # "President & CEO" is the classic top-dog title
        return score

    today = datetime.date.today()
    candidates = []   # (quality_score, year_or_none, name, tenure_or_none)

    for e in executives:
        raw_title = e.get("title") or ""
        q = _title_quality(raw_title)
        if q < 0:
            continue
        name = e.get("name") or ""
        if not name:
            continue
        since_raw = e.get("since") or e.get("titleSince") or ""
        try:
            year = int(str(since_raw)[:4]) if since_raw else None
        except (ValueError, TypeError):
            year = None
        if year and (year < 1950 or year > today.year):
            year = None
        tenure = None
        if year:
            tenure = round(today.year - year + (today.month - 1) / 12.0, 1)
        candidates.append((q, year, name, tenure))

    if not candidates:
        return None

    # Sort: highest quality first, then longest tenure (known tenure beats unknown)
    candidates.sort(key=lambda x: (-x[0], -(x[3] or -1)))
    best = candidates[0]
    return {"name": best[2], "since": best[1], "tenure_years": best[3]}


def compute_per_share_growth_5y(bs_5y: list, cfs_5y: list) -> dict:
    """5Y per-share CAGRs for revenue, FCF, and book value.
    bs_5y / cfs_5y are lists of annual statements ordered newest→oldest.
    Returns: {revPerShare5yCagr, fcfPerShare5yCagr, bvPerShare5yCagr}
    """
    out = {"revPerShare5yCagr": None, "fcfPerShare5yCagr": None, "bvPerShare5yCagr": None}
    if not bs_5y or len(bs_5y) < 2:
        return out

    def _cagr(new, old, years):
        if not new or not old or new <= 0 or old <= 0 or years <= 0:
            return None
        try:
            return round((new / old) ** (1.0 / years) - 1, 4)
        except Exception:
            return None

    # Pair newest with oldest available (up to 5y span)
    newest_bs = bs_5y[0]
    oldest_bs = bs_5y[-1]
    span = len(bs_5y) - 1  # years between

    # FCF 5Y CAGR (absolute — FMP balance sheet 'commonStock' is paid-in capital $, not share count)
    if cfs_5y and len(cfs_5y) >= 2:
        fcf_new = cfs_5y[0].get("freeCashFlow")
        fcf_old = cfs_5y[-1].get("freeCashFlow")
        if fcf_new and fcf_old and fcf_new > 0 and fcf_old > 0:
            out["fcfPerShare5yCagr"] = _cagr(fcf_new, fcf_old, span)

    # Book value 5Y CAGR (absolute total equity — meaningful without share count)
    eq_new = newest_bs.get("totalStockholdersEquity") or newest_bs.get("totalEquity")
    eq_old = oldest_bs.get("totalStockholdersEquity") or oldest_bs.get("totalEquity")
    if eq_new and eq_old and eq_new > 0 and eq_old > 0:
        out["bvPerShare5yCagr"] = _cagr(eq_new, eq_old, span)

    # Net income 5Y CAGR (proxy for earnings power growth; revenue not on CF/BS)
    if cfs_5y and len(cfs_5y) >= 2:
        ni_new = cfs_5y[0].get("netIncome")
        ni_old = cfs_5y[-1].get("netIncome")
        if ni_new and ni_old and ni_new > 0 and ni_old > 0:
            out["revPerShare5yCagr"] = _cagr(ni_new, ni_old, span)

    return out


def compute_ceo_allocator_score(s: dict, bs_5y: list, cfs_5y: list,
                                 executives: list) -> dict:
    """Thorndike Outsiders-style score: did this CEO actually create per-share value?

    Returns:
      {
        score: 0-100 or None,
        grade: "A+"/"A"/.../"D" or None,
        tenure_years, ceo_name, fcf_per_share_cagr,
        shares_change_pct, roic_trend, callouts: [..]
      }
    """
    out = {
        "score": None, "grade": None, "tenure_years": None, "ceo_name": None,
        "fcf_per_share_cagr": None, "shares_change_pct": None,
        "roic_trend": None, "callouts": []
    }
    ceo = _extract_ceo(executives)
    if ceo:
        out["ceo_name"] = ceo["name"]
        out["tenure_years"] = ceo["tenure_years"]

    # Gate: need CEO found + enough financial data
    # FMP often omits titleSince; when tenure_years is None we still score using
    # the full 5Y data window (the financial track record speaks for itself).
    if not ceo:
        return out
    if not bs_5y or len(bs_5y) < 3 or not cfs_5y or len(cfs_5y) < 3:
        return out
    # If tenure is known and < 3yr, skip (too early to score meaningfully)
    if ceo["tenure_years"] is not None and ceo["tenure_years"] < 3.0:
        return out

    # Cap scoring window by tenure if known; otherwise use full available data
    if ceo["tenure_years"] is not None:
        max_years = min(int(ceo["tenure_years"]), len(bs_5y) - 1, len(cfs_5y) - 1, 5)
    else:
        max_years = min(len(bs_5y) - 1, len(cfs_5y) - 1, 5)
    if max_years < 3:
        return out

    bs_window  = bs_5y[:max_years + 1]   # newest..max_years ago
    cfs_window = cfs_5y[:max_years + 1]
    span = max_years

    # ── Component 1: FCF value creation (40 pts) — absolute FCF CAGR ──
    # Note: FMP balance sheet 'commonStock' is paid-in capital $, not share count,
    # so we use absolute FCF CAGR as the primary value-creation signal.
    fcf_new = cfs_window[0].get("freeCashFlow")
    fcf_old = cfs_window[-1].get("freeCashFlow")
    fcf_cagr = None
    if fcf_new and fcf_old and span > 0:
        if fcf_new > 0 and fcf_old > 0:
            try:
                fcf_cagr = (fcf_new / fcf_old) ** (1.0 / span) - 1
            except Exception:
                fcf_cagr = None
        elif fcf_new > 0 and fcf_old <= 0:
            fcf_cagr = 0.30   # turned FCF positive — credit it
        elif fcf_new <= 0:
            fcf_cagr = -0.10  # negative end state

    out["fcf_per_share_cagr"] = round(fcf_cagr, 4) if fcf_cagr is not None else None

    if fcf_cagr is None:
        psv_pts = 0
    elif fcf_cagr > 0.15: psv_pts = 40
    elif fcf_cagr > 0.10: psv_pts = 32
    elif fcf_cagr > 0.05: psv_pts = 22
    elif fcf_cagr > 0.0:  psv_pts = 10
    else:                 psv_pts = 0

    if fcf_cagr is not None:
        out["callouts"].append(f"FCF grew {fcf_cagr*100:+.0f}%/y over {span}y")

    # ── Component 2: Buyback/dilution discipline (20 pts) ──
    # Use netCommonStockIssuance from CF: negative = net buybacks (good), positive = dilution (bad)
    # Normalise cumulative net issuance against cumulative FCF over the window.
    total_fcf    = sum(abs(r.get("freeCashFlow") or 0) for r in cfs_window if (r.get("freeCashFlow") or 0) > 0)
    total_net_si = sum(r.get("netCommonStockIssuance") or r.get("commonStockRepurchased") or 0
                       for r in cfs_window)
    sh_pct = None
    bb_pts = 5  # neutral default when data missing
    if total_fcf > 0:
        # buyback_ratio: what fraction of FCF went to net buybacks (negative = returned capital)
        buyback_ratio = -total_net_si / total_fcf   # positive = good (returned capital)
        sh_pct = -buyback_ratio   # preserve sign convention: negative = shareholder-friendly
        out["shares_change_pct"] = round(sh_pct, 4)
        if buyback_ratio >= 0.40:
            bb_pts = 20
            out["callouts"].append(f"Returned {buyback_ratio*100:.0f}% of FCF via buybacks")
        elif buyback_ratio >= 0.15:
            bb_pts = 14
        elif buyback_ratio >= 0.0:
            bb_pts = 8
        elif buyback_ratio >= -0.20:
            bb_pts = 4   # mild net dilution
        else:
            bb_pts = 0   # heavy dilution (>20% of FCF in net stock issuance)
            out["callouts"].append(f"⚠ Net stock dilution {abs(buyback_ratio)*100:.0f}% vs FCF over tenure")
    else:
        out["shares_change_pct"] = None

    # ── Component 3: ROIC trend (15 pts) — current ROIC vs early-tenure proxy ──
    # Proxy: net income / equity over the window (rough ROE/ROIC approximation since
    # we don't have annual ROIC history without another endpoint call)
    def _roi_proxy(bs_row, cfs_row):
        ni = cfs_row.get("netIncome")
        eq = bs_row.get("totalStockholdersEquity") or bs_row.get("totalEquity")
        if ni and eq and eq > 0:
            return ni / eq
        return None
    roi_new = _roi_proxy(bs_window[0], cfs_window[0])
    roi_old = _roi_proxy(bs_window[-1], cfs_window[-1])
    roic_trend = None
    if roi_new is not None and roi_old is not None:
        delta = roi_new - roi_old
        if delta >= 0.03:    roic_trend = "rising";  rt_pts = 15
        elif delta >= -0.02: roic_trend = "flat";    rt_pts = 10
        else:                roic_trend = "falling"; rt_pts = max(0, 5 + int(delta * 50))
    else:
        rt_pts = 5
    out["roic_trend"] = roic_trend
    if roic_trend == "rising":
        out["callouts"].append(f"ROI proxy rose {roi_old*100:.0f}% → {roi_new*100:.0f}%")
    elif roic_trend == "falling":
        out["callouts"].append(f"⚠ ROI proxy fell {roi_old*100:.0f}% → {roi_new*100:.0f}%")

    # ── Component 4: Debt discipline (15 pts) — Net Debt/EBITDA stability ──
    # Use total debt growth over window vs FCF as a rough proxy
    debt_new = bs_window[0].get("totalDebt") or bs_window[0].get("longTermDebt")
    debt_old = bs_window[-1].get("totalDebt") or bs_window[-1].get("longTermDebt")
    dt_pts = 8  # neutral default
    if debt_new is not None and debt_old is not None and debt_old > 0:
        debt_change = (debt_new - debt_old) / debt_old
        # If FCF grew faster than debt, that's healthy
        if fcf_cagr is not None:
            if debt_change <= 0:                  dt_pts = 15  # de-levered
            elif debt_change < (fcf_cagr * span): dt_pts = 12  # debt grew but FCF outpaced
            elif debt_change < 1.0:               dt_pts = 8
            else:                                 dt_pts = 0   # >100% debt growth
        else:
            if debt_change <= 0:    dt_pts = 12
            elif debt_change < 0.5: dt_pts = 8
            else:                   dt_pts = 3

    # ── Component 5: Reinvestment efficiency (10 pts) — capex vs revenue trend ──
    rei_pts = 5  # neutral default
    capex_new = abs(cfs_window[0].get("capitalExpenditure", 0) or 0)
    rev_new   = cfs_window[0].get("revenue") or 1
    if capex_new and rev_new and rev_new > 0:
        cx_intensity = capex_new / rev_new
        # Low capex with rising ROI = great compounder; high capex with falling ROI = bad
        if roic_trend == "rising" and cx_intensity < 0.10:    rei_pts = 10
        elif roic_trend == "rising":                          rei_pts = 8
        elif roic_trend == "flat" and cx_intensity < 0.05:    rei_pts = 7
        elif roic_trend == "falling" and cx_intensity > 0.15: rei_pts = 1
        else:                                                 rei_pts = 5

    score = psv_pts + bb_pts + rt_pts + dt_pts + rei_pts
    score = max(0, min(100, int(score)))

    if   score >= 88: grade = "A+"
    elif score >= 78: grade = "A"
    elif score >= 70: grade = "B+"
    elif score >= 60: grade = "B"
    elif score >= 50: grade = "C+"
    elif score >= 40: grade = "C"
    else:             grade = "D"

    out["score"] = score
    out["grade"] = grade
    return out


def classify_divergence(s: dict) -> str | None:
    """Analyst-vs-Insider divergence flag.

    hidden_gem      — insiders buying strongly, sell-side neutral/bearish
    conviction_stack — insiders buying strongly, sell-side bullish (both agree)
    quiet_signal    — insiders modestly buying, sell-side neutral/bearish
    None            — no divergence signal worth surfacing
    """
    rating = (s.get("recommendation") or "").strip().lower()
    n_buys = s.get("insiderBuys") or 0
    val    = s.get("insiderValue") or 0
    insider_strong = n_buys >= 3 or val >= 250_000
    insider_any    = n_buys >= 1
    bearish_or_neutral = rating in ("hold", "sell", "strong sell", "underperform", "")
    bullish            = rating in ("buy", "strong buy", "outperform")
    if insider_strong and bearish_or_neutral:
        return "hidden_gem"
    if insider_strong and bullish:
        return "conviction_stack"
    if insider_any and bearish_or_neutral and val >= 50_000:
        return "quiet_signal"
    return None


def fetch_earnings_surprises(tickers: list) -> dict:
    """Fetch last 8 quarters of earnings surprises for beat-rate calculation.
    Beat rate = % of quarters where actual EPS beat estimated EPS.
    High beat rate = management consistently under-promises and over-delivers.
    Tries two FMP endpoint names; fails fast if neither is available on the plan.
    """
    cache_key = "earnings_surprises"
    if _cache.get(cache_key):
        print(f"  📦 Using cached earnings surprises ({len(_cache[cache_key])} stocks)")
        return _cache[cache_key]

    # Probe with the first ticker to confirm the endpoint is available before
    # attempting all 4,000 stocks (avoids 4,000 noisy 404 warnings).
    _probe_ticker = tickers[0] if tickers else None
    _endpoint = None
    for _ep in ("earnings-surprises", "historical-earnings"):
        _probe = fmp_get(_ep, {"symbol": _probe_ticker, "limit": 2}) if _probe_ticker else None
        if _probe and isinstance(_probe, list):
            _endpoint = _ep
            break

    if not _endpoint:
        print("  ⏭️ Skipping earnings surprises (endpoint not available on this plan)")
        _cache[cache_key] = {}
        return {}

    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    print(f"\n  📊 Fetching earnings surprises ({_endpoint})...")
    results = {}
    _lock = threading.Lock()
    _throttle = threading.Semaphore(4)

    def _f(t):
        with _throttle:
            data = fmp_get(_endpoint, {"symbol": t, "limit": 8})
            time.sleep(0.2)
            if data and isinstance(data, list) and len(data) >= 2:
                return t, data
            return t, None

    with ThreadPoolExecutor(max_workers=12) as pool:
        futs = {pool.submit(_f, t): t for t in tickers}
        done = 0
        for fut in as_completed(futs):
            t, data = fut.result()
            if data:
                with _lock: results[t] = data
            done += 1
            if done % 100 == 0: print(f"    [{done}/{len(tickers)}] earnings surprises fetched...")

    _cache[cache_key] = results
    print(f"  ✅ earnings surprises loaded: {len(results)} stocks")
    return results


def _parse_openinsider_table(html: str) -> list:
    """Parse the tinytable from an OpenInsider HTML page into trade dicts."""
    from html.parser import HTMLParser
    import re

    class _Parser(HTMLParser):
        def __init__(self):
            super().__init__()
            self._in_t = False; self._depth = 0
            self._in_row = False; self._in_cell = False
            self.rows = []; self._row = []; self._cell = ""

        def handle_starttag(self, tag, attrs):
            amap = dict(attrs)
            if tag == "table":
                if "tinytable" in amap.get("class", "") or "tinytable" in amap.get("id", ""):
                    self._in_t = True; self._depth = 1
                elif self._in_t:
                    self._depth += 1
            if not self._in_t: return
            if tag == "tr":
                self._in_row = True; self._row = []
            elif tag in ("td", "th") and self._in_row:
                self._in_cell = True; self._cell = ""
            elif tag == "br" and self._in_cell:
                self._cell += " "

        def handle_endtag(self, tag):
            if tag == "table" and self._in_t:
                self._depth -= 1
                if self._depth <= 0: self._in_t = False
            if not self._in_t: return
            if tag == "tr" and self._in_row:
                self._in_row = False
                if self._row: self.rows.append(self._row)
                self._row = []
            elif tag in ("td", "th") and self._in_cell:
                self._in_cell = False
                self._row.append(" ".join(self._cell.split()))
                self._cell = ""

        def handle_data(self, data):
            if self._in_cell: self._cell += data

    parser = _Parser()
    parser.feed(html)
    if not parser.rows:
        return []

    # Detect header row to find column positions dynamically
    header = [c.lower() for c in parser.rows[0]]
    def _col(names):
        for i, h in enumerate(header):
            if any(n in h for n in names): return i
        return None

    # Known OpenInsider column order (0=X, 1=FilingDate, 2=TradeDate,
    # 3=Ticker, 4=Company, 5=Name, 6=Title, 7=TradeType, 8=Price,
    # 9=Qty, 10=Owned, 11=ΔOwn, 12=Value)
    has_header = len(header) >= 5 and any(h in ("ticker", "x") for h in header[:5])
    if has_header:
        idx = {
            "date":    _col(["trade date"]) or _col(["date"]) or 2,
            "ticker":  _col(["ticker"]) or 3,
            "company": _col(["company", "issuer"]) or 4,
            "name":    _col(["insider", "owner", "reporting"]) or 5,
            "title":   _col(["title", "relationship"]) or 6,
            "type":    _col(["trade type", "transaction"]) or 7,
            "price":   _col(["price"]) or 8,
            "qty":     _col(["qty", "shares"]) or 9,
            "value":   _col(["value"]) or 12,
        }
        data_rows = parser.rows[1:]
    else:
        idx = {"date": 2, "ticker": 3, "company": 4, "name": 5,
               "title": 6, "type": 7, "price": 8, "qty": 9, "value": 12}
        data_rows = parser.rows

    def _num(s):
        s = re.sub(r'[+$,\s]', '', s).replace("(", "-").replace(")", "")
        if not s or s in ("-", ""): return 0.0
        try:
            if s.upper().endswith("M"): return float(s[:-1]) * 1_000_000
            if s.upper().endswith("K"): return float(s[:-1]) * 1_000
            if s.upper().endswith("B"): return float(s[:-1]) * 1_000_000_000
            return float(s)
        except: return 0.0

    max_idx = max(idx.values())
    trades = []
    for cells in data_rows:
        if len(cells) <= max_idx: continue
        ticker = re.sub(r'[^A-Z]', '', cells[idx["ticker"]].upper())
        if not ticker or len(ticker) > 6: continue
        trade_type = cells[idx["type"]]
        if not trade_type.upper().startswith("P"): continue  # purchases only
        price = _num(cells[idx["price"]])
        qty   = abs(_num(cells[idx["qty"]]))
        value = _num(cells[idx["value"]])
        if value == 0 and price > 0 and qty > 0:
            value = price * qty
        trades.append({
            "symbol":               ticker,
            "transactionType":      "P-Purchase",
            "transactionDate":      cells[idx["date"]],
            "securitiesTransacted": qty,
            "price":                price,
            "reportingName":        cells[idx["name"]],
            "typeOfOwner":          cells[idx["title"]],
            "_value":               value,
            "_source":              "openinsider",
            "_company":             cells[idx["company"]],
        })
    return trades


def fetch_openinsider_data(days: int = 30, min_value_k: int = 50) -> list:
    """Scrape recent insider purchases from openinsider.com (free, no API key).
    Returns trades in FMP-compatible format so downstream code works unchanged.
    Caches for 1 day.
    """
    cache_key = "openinsider"
    cached = _cache.get(cache_key, [])
    if cached:
        print(f"  📦 Using cached OpenInsider data ({len(cached)} trades)")
        return cached

    print("  🌐 Fetching insider purchases from openinsider.com...")
    session = requests.Session()
    session.headers.update({
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/123.0.0.0 Safari/537.36"),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": "https://openinsider.com/",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    })

    screener_params = (
        f"s=&o=&pl=&ph=&ll=&lh=&fd={days}&fdr=&td=0&tdr=&fdlyl=&fdlyh=&daysago=&"
        f"xp=1&vl={min_value_k}&vh=&ocl=&och=&sic1=-1&sicl=100&sich=9999&"
        f"grp=0&nfl=&nfh=&nil=&nih=&nol=&noh=&v2l=&v2h=&oc2l=&oc2h=&"
        f"sortcol=0&cnt=500&Action=1"
    )
    urls = [
        f"https://openinsider.com/screener?{screener_params}",
        "https://openinsider.com/clustered-buys",
    ]
    all_trades = []

    def _fetch_html(url, session, timeout=25):
        """Try requests first, fall back to stdlib urllib (different TLS stack)."""
        try:
            r = session.get(url, timeout=timeout, allow_redirects=True)
            if r.status_code == 200 and len(r.text) > 500:
                return r.text
        except Exception:
            pass
        # urllib fallback — different TLS fingerprint, sometimes bypasses blocks
        try:
            import urllib.request, ssl
            req = urllib.request.Request(url)
            for k, v in session.headers.items():
                req.add_header(k, v)
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
                return resp.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        return None

    for url in urls:
        html = _fetch_html(url, session)
        if html:
            trades = _parse_openinsider_table(html)
            if trades:
                all_trades.extend(trades)
                print(f"    ✅ {len(trades)} trades from openinsider.com")
                time.sleep(0.5)
        else:
            print(f"  ⚠️ OpenInsider unreachable — check Windows Firewall / antivirus")

    # Deduplicate by (ticker, date, qty)
    seen, deduped = set(), []
    for tr in all_trades:
        key = (tr["symbol"], tr["transactionDate"], tr["securitiesTransacted"])
        if key not in seen:
            seen.add(key); deduped.append(tr)

    _cache[cache_key] = deduped
    print(f"  ✅ OpenInsider: {len(deduped)} insider purchases loaded")
    return deduped


def fetch_finviz_insider_data() -> list:
    """Scrape recent insider purchases from finviz.com/insidertrading.ashx.
    Used as fallback when openinsider.com is unreachable.
    """
    import re
    cache_key = "finviz_insider"
    cached = _cache.get(cache_key, [])
    if cached:
        return cached

    url = "https://finviz.com/insidertrading.ashx?or=-10&tc=1&o=-transactionDate"
    ua  = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36")
    headers = {"User-Agent": ua, "Accept-Language": "en-US,en;q=0.9",
                "Referer": "https://finviz.com/"}
    try:
        import urllib.request, ssl
        req = urllib.request.Request(url)
        for k, v in headers.items():
            req.add_header(k, v)
        ctx = ssl.create_default_context()
        ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
        with urllib.request.urlopen(req, timeout=20, context=ctx) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        print(f"  ⚠️ Finviz insider fetch error: {str(e)[:80]}")
        return []

    # Finviz insider table: Ticker | Owner | Relationship | Date | Transaction | Cost | #Shares | Value | Total
    rows = []
    def _num(s):
        s = re.sub(r'[+$,\s]', '', s)
        try: return float(s)
        except: return 0.0

    # Find all <tr> blocks inside the insider table
    tr_blocks = re.findall(r'<tr[^>]*class="[^"]*insider[^"]*"[^>]*>(.*?)</tr>|'
                           r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL | re.IGNORECASE)
    for match in tr_blocks:
        tr = match[0] or match[1]
        tds = re.findall(r'<td[^>]*>(.*?)</td>', tr, re.DOTALL | re.IGNORECASE)
        if len(tds) < 8: continue
        strip = lambda s: re.sub(r'<[^>]+>', '', s).strip()
        cells = [strip(td) for td in tds]
        # columns: 0=Ticker, 1=Owner, 2=Relationship, 3=Date, 4=Transaction, 5=Cost, 6=#Shares, 7=Value
        txn = cells[4] if len(cells) > 4 else ""
        if "buy" not in txn.lower() and "purchase" not in txn.lower(): continue
        ticker = re.sub(r'[^A-Z]', '', cells[0].upper())
        if not ticker or len(ticker) > 6: continue
        price  = _num(cells[5]) if len(cells) > 5 else 0
        qty    = _num(cells[6]) if len(cells) > 6 else 0
        value  = _num(cells[7]) if len(cells) > 7 else price * qty
        rows.append({
            "symbol":               ticker,
            "transactionType":      "P-Purchase",
            "transactionDate":      cells[3] if len(cells) > 3 else "",
            "securitiesTransacted": qty,
            "price":                price,
            "reportingName":        cells[1] if len(cells) > 1 else "",
            "typeOfOwner":          cells[2] if len(cells) > 2 else "",
            "_value":               value,
            "_source":              "finviz",
        })

    _cache[cache_key] = rows
    if rows:
        print(f"    ✅ Finviz: {len(rows)} insider purchases")
    else:
        print(f"  ⚠️ Finviz: no purchases found (may need browser session)")
    return rows


def fetch_insider_trading(tickers: list = None) -> list:
    """Fetch recent insider purchases for the given ticker list.
    Global endpoint often fails on basic FMP tiers — per-ticker is more reliable.
    """
    cache_key = "insider"
    if _cache.get(cache_key):
        print(f"  📦 Using cached insider data ({len(_cache[cache_key])} transactions)")
        return _cache[cache_key]

    print("\n  🏦 Fetching insider trading data...")
    all_trades = []

    # Try global feed first; many FMP plans return 404 for this endpoint
    _insider_available = False
    for ep in ["insider-trading", "insider-trading-rss-feed"]:
        data = fmp_get(ep, {"limit": 500, "transactionType": "P-Purchase"})
        if data and isinstance(data, list) and len(data) > 5:
            all_trades = [tr for tr in data if _is_purchase(tr.get("transactionType", ""))]
            if all_trades:
                _insider_available = True
                print(f"    ✅ Global feed: {len(all_trades)} purchases")
                break
        time.sleep(0.3)

    # Per-ticker fallback — only attempt if global feed wasn't a 404
    if not _insider_available and tickers:
        # Probe one ticker to check if endpoint is available at all
        _probe = fmp_get("insider-trading", {"symbol": tickers[0], "limit": 1})
        if _probe is not None:  # None = 404; [] = available but no data
            print(f"    🔄 Fetching per-ticker insider data for top {min(100, len(tickers))} stocks...")
            for t in tickers[:100]:
                data = fmp_get("insider-trading", {"symbol": t, "limit": 10})
                if data and isinstance(data, list):
                    all_trades.extend(tr for tr in data if _is_purchase(tr.get("transactionType", "")))
                time.sleep(0.1)
            # Deduplicate
            seen = set()
            deduped = []
            for tr in all_trades:
                key = (tr.get("symbol"), tr.get("filingDate"), tr.get("securitiesTransacted"))
                if key not in seen:
                    seen.add(key)
                    deduped.append(tr)
            all_trades = deduped
        else:
            print("    ⚠️ insider-trading endpoint not available on this FMP plan")

    if not all_trades:
        print("    🔄 Trying openinsider.com...")
        all_trades = fetch_openinsider_data()
    if not all_trades:
        print("    🔄 Trying finviz.com insider data...")
        all_trades = fetch_finviz_insider_data()

    _cache[cache_key] = all_trades
    print(f"  ✅ Insider trades loaded: {len(all_trades)} purchases")
    return all_trades


# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _is_purchase(txn_type: str) -> bool:
    """Return True if the insider transaction is a buy/purchase (not sale/award)."""
    t = txn_type.upper()
    return (t.startswith("P") or "PURCHASE" in t or "BUY" in t) and "SALE" not in t


def _is_common_stock(s: dict) -> bool:
    """Filter out ETFs, mutual funds, preferred stocks from Lynch/quality tabs.
    These pollute the results with non-actionable picks.
    """
    ticker = s.get("ticker", "")
    name = (s.get("name") or "").lower()
    # Preferred stocks: tickers like MER-PK, BAC-PL, C-PN
    if "-" in ticker:
        return False
    # Mutual fund tickers typically 5 chars ending in X (RNNEX, IFAFX, AMECX, AMPFX)
    if len(ticker) == 5 and ticker.endswith("X"):
        return False
    # Fund/ETF name keywords
    fund_kw = (
        "fund", " etf", "income fund", "bond fund", "money market",
        "index fund", "total return fund", "growth fund", "value fund",
        "equity fund", "balanced fund", "allocation fund",
        # Closed-end fund / CEF / BDC name patterns not ending in "fund":
        "municipal income", "municipal bond", "muni bond", "amt-free",
        "nuveen ", "pimco ", "eaton vance", "kayne anderson", "eagle point",
        "western asset", "blackrock income", "calamos", "gabelli",
        # Closed-end / specialty finance vehicles with no real equity earnings:
        "credit company", "credit corp", "credit opportunity",
    )
    if any(kw in name for kw in fund_kw):
        return False
    return True


def _first(*args):
    """Return first non-None value. Handles 0.0 and negative numbers correctly
    (unlike `or` which treats 0.0 and negatives as falsy and skips them).
    """
    for a in args:
        if a is not None:
            return a
    return None


# ─────────────────────────────────────────────
# DATA ASSEMBLY — Merge all FMP data into unified stock records
# ─────────────────────────────────────────────

def assemble_stock_data(universe, profiles, key_metrics, ratios_ttm, dcf_data,
                        estimates, scores, ratings, growth_data, insider_data,
                        balance_sheet=None, earnings_surp=None,
                        bs_5y=None, cfs_5y=None, executives=None) -> dict:
    """Merge all FMP data sources into a single dict per stock."""
    print("\n  🔧 Assembling unified stock data...")


    # Build insider lookup: {ticker: {count, total_value, latest_date}}
    insider_lookup = defaultdict(lambda: {"count": 0, "totalValue": 0, "latestDate": ""})
    for trade in insider_data:
        t = trade.get("symbol", "")
        val = abs(trade.get("securitiesTransacted", 0) * (trade.get("price", 0) or 0))
        insider_lookup[t]["count"] += 1
        insider_lookup[t]["totalValue"] += val
        d = trade.get("transactionDate", "")
        if d > insider_lookup[t]["latestDate"]:
            insider_lookup[t]["latestDate"] = d

    stocks = {}
    for t, u in universe.items():
        prof = profiles.get(t, {})
        km = key_metrics.get(t, {})
        rtm = ratios_ttm.get(t, {})
        dcf = dcf_data.get(t, {})
        est = estimates.get(t, {})
        sc = scores.get(t, {})
        rat = ratings.get(t, {})
        # growth_data is now a list (most recent year first); gr = most recent, gr_prev = prior year
        _gr_list = growth_data.get(t, [])
        if isinstance(_gr_list, list) and _gr_list:
            gr      = _gr_list[0]                                     # current year
            gr_prev = _gr_list[1] if len(_gr_list) > 1 else {}        # prior year
            gr_yr2  = _gr_list[2] if len(_gr_list) > 2 else {}        # year before that
            gr_yr3  = _gr_list[3] if len(_gr_list) > 3 else {}        # 3 years ago
            gr_yr4  = _gr_list[4] if len(_gr_list) > 4 else {}        # 4 years ago
        elif isinstance(_gr_list, dict):
            gr = _gr_list; gr_prev = {}; gr_yr2 = {}; gr_yr3 = {}; gr_yr4 = {}
        else:
            gr = {}; gr_prev = {}; gr_yr2 = {}; gr_yr3 = {}; gr_yr4 = {}
        bs = (balance_sheet or {}).get(t, {})
        es = (earnings_surp or {}).get(t, [])
        ins = insider_lookup.get(t, {})

        price = prof.get("price") or u.get("price") or 0
        mktCap = prof.get("mktCap") or u.get("mktCap") or u.get("marketCap") or 0
        # PE: try multiple sources (ratios_ttm uses priceToEarningsRatioTTM)
        pe_val = (prof.get("pe") or u.get("pe") or
                  rtm.get("priceToEarningsRatioTTM") or
                  km.get("peRatioTTM") or km.get("priceEarningsRatioTTM") or
                  prof.get("peRatio") or u.get("peRatio"))
        # Fallback: compute from earningsYield
        if not pe_val and km.get("earningsYieldTTM") and km["earningsYieldTTM"] > 0.001:
            pe_val = round(1 / km["earningsYieldTTM"], 1)
        # Sanity: zero, negative, or extreme PE is not useful for screening
        # Note: must use `is not None` — `if pe_val` would silently pass pe_val=0
        if pe_val is not None and (pe_val < 0.5 or pe_val > 500):
            pe_val = None
        elif pe_val is not None:
            pe_val = round(pe_val, 1)
        if not price or price <= 0:
            continue

        # PEG calculation — primary: FMP analyst EPS CAGR from estimates endpoint
        peg = None
        growth_5y = None
        rev_growth_5y = None
        if isinstance(est, list) and len(est) >= 2:
            eps_vals = [d.get("epsAvg") or d.get("estimatedEpsAvg") for d in est
                        if (d.get("epsAvg") or d.get("estimatedEpsAvg")) and (
                                    d.get("epsAvg") or d.get("estimatedEpsAvg")) > 0]
            rev_vals = [d.get("revenueAvg") or d.get("estimatedRevenueAvg") for d in est
                        if (d.get("revenueAvg") or d.get("estimatedRevenueAvg")) and (
                                    d.get("revenueAvg") or d.get("estimatedRevenueAvg")) > 0]
            if len(eps_vals) >= 2:
                n = len(eps_vals) - 1
                try:
                    growth_5y = round((eps_vals[0] / eps_vals[-1]) ** (1 / n) - 1, 4)
                except:
                    pass
            if len(rev_vals) >= 2:
                n = len(rev_vals) - 1
                try:
                    rev_growth_5y = round((rev_vals[0] / rev_vals[-1]) ** (1 / n) - 1, 4)
                except:
                    pass
        elif isinstance(est, dict):
            growth_5y = est.get("epsGrowth5y")

        if pe_val and pe_val > 5 and growth_5y and 0.03 < growth_5y < 0.60:
            peg = round(pe_val / (growth_5y * 100), 2)
            if peg < 0.2 or peg > 50:
                peg = None

        # Fallback PEG: use trailing growth when analyst estimates unavailable
        # Lynch himself used the most recent available growth rate — this is more authentic
        if peg is None and pe_val and pe_val > 3:
            eg_tr = _first(gr.get("epsgrowth"), gr.get("epsGrowth"))
            rg_tr = _first(gr.get("revenueGrowth"), gr.get("revGrowth"))
            trailing_g = eg_tr if (eg_tr and 0.03 < eg_tr < 0.60) else (
                         rg_tr if (rg_tr and 0.03 < rg_tr < 0.60) else None)
            if trailing_g:
                _peg_fb = round(pe_val / (trailing_g * 100), 2)
                if 0.2 < _peg_fb < 50:
                    peg = _peg_fb

        # ── Forward P/E and Forward PEG ────────────────────────────────────────
        # Use est[0] (next fiscal year) for a fully consistent forward valuation.
        # Avoids the trailing-P/E ÷ forward-growth mismatch that creates fake cheap signals.
        fwd_pe = None
        fwd_peg = None
        if isinstance(est, list) and est and price > 0:
            _next_eps = est[0].get("epsAvg") or est[0].get("estimatedEpsAvg")
            if _next_eps and _next_eps > 0:
                _fpe = round(price / _next_eps, 1)
                if 4 < _fpe < 150:
                    fwd_pe = _fpe
                    if growth_5y and 0.03 < growth_5y < 0.60:
                        _fpeg = round(_fpe / (growth_5y * 100), 2)
                        if 0.1 < _fpeg < 50:
                            fwd_peg = _fpeg

        # IV from DCF — cap at 8× price to filter FMP model outliers (e.g. some insurers/HMOs)
        iv = dcf.get("dcf") or dcf.get("dcfValue")
        if iv is not None and iv <= 0:
            iv = None  # negative IV = DCF artifact (e.g. negative equity base) — not meaningful
        if iv and price > 0 and iv > price * 8:
            iv = None  # discard — model artifact, not investable signal
        mos = None
        if iv and iv > 0 and price > 0:
            mos = round((iv - price) / iv, 4)

        # ── Net cash per share (from balance sheet) ────────────────────────
        # Net cash = cash + short-term investments - total debt
        # Positive net cash = company has more cash than debt (asset play signal)
        # SKIP Financial Services: their "shortTermInvestments" is insurance float /
        # loan book / deposit-backed assets — not free cash. Net cash metric is
        # meaningless for banks, insurers, and reinsurers.
        _sector_raw = (u.get("sector") or prof.get("sector") or "").lower()
        _is_financial = "financial" in _sector_raw or "insurance" in _sector_raw
        net_cash_per_share = None
        net_cash_ratio = None  # netCashPerShare / price — how much of price is pure cash
        if bs and not _is_financial:
            _cash = _first(bs.get("cashAndCashEquivalents"), bs.get("cash")) or 0
            _st_inv = bs.get("shortTermInvestments") or 0
            _total_debt = _first(bs.get("totalDebt"), bs.get("longTermDebt")) or 0
            # Derive shares from mktCap/price first (most reliable — avoids balance sheet unit ambiguity)
            # commonStock = par value dollars, NOT shares — never use it for share count
            if mktCap and price > 0:
                _shares = mktCap / price
            else:
                _shares = _first(bs.get("sharesOutstanding"), bs.get("weightedAverageShsOut"))
            _net_cash = _cash + _st_inv - _total_debt
            if _shares and _shares > 0:
                _ncps = _net_cash / _shares
                # Sanity: if net cash per share is unreasonably large (>10× price), discard
                if price > 0 and abs(_ncps) <= price * 10:
                    net_cash_per_share = round(_ncps, 2)
                    if price > 0:
                        net_cash_ratio = round(_ncps / price, 4)

        # ── Earnings beat rate (from earnings surprises) ───────────────────
        # beat_rate = fraction of last N quarters where actual EPS > estimated EPS
        beat_rate = None
        eps_beat_streak = 0  # consecutive quarters beating estimates (most recent first)
        if isinstance(es, list) and len(es) >= 2:
            beats = []
            for rec in es:
                actual = rec.get("actualEarningResult") or rec.get("actual")
                est_eps = rec.get("estimatedEarning") or rec.get("estimated")
                if actual is not None and est_eps is not None:
                    beats.append(1 if actual >= est_eps else 0)
            if beats:
                beat_rate = round(sum(beats) / len(beats), 3)
                for b in beats:  # most recent first in FMP response
                    if b == 1:
                        eps_beat_streak += 1
                    else:
                        break

        # ── 52-week positioning ────────────────────────────────────────────
        _yr_high = u.get("yearHigh")
        _yr_low  = u.get("yearLow")
        price_vs_52h = round(price / _yr_high, 4) if (_yr_high and _yr_high > 0) else None
        price_vs_52l = round(price / _yr_low,  4) if (_yr_low  and _yr_low  > 0) else None

        # ── Per-share data corruption check ────────────────────────────────
        # FMP occasionally stores total company figures as per-share values for certain tickers
        # (e.g. MCHB: bookValuePerShare=$129,595 at price=$14 — bank total assets stored as per-share)
        # When detected, nullify all ratios derived from book/equity so they don't corrupt rankings.
        _bv_ps = _first(rtm.get("bookValuePerShareTTM"))
        _per_share_corrupt = (_bv_ps is not None and price > 0 and _bv_ps > price * 100)

        # ── Per-metric bounds checks ────────────────────────────────────────
        # P/B > 200 is not screening-useful (negative book or data error); < 0 is invalid
        _pb_raw = _first(rtm.get("priceToBookRatioTTM"))
        pb_val = (_pb_raw if (_pb_raw is not None and 0 < _pb_raw <= 200) else None)

        # ROE/ROA/ROIC: values > ±500% are almost always FMP errors (tiny denominator)
        # Round to 4 decimal places (e.g. 0.1523 → 0.1523 = 15.23%) to avoid spurious precision
        _roe_raw = _first(km.get("returnOnEquityTTM"))
        roe_val = (round(_roe_raw, 4) if (_roe_raw is not None and abs(_roe_raw) <= 5.0) else None)
        # Sanity 1: negative TOTAL equity (e.g. PEGA — heavy buybacks) makes ROE denominator invalid.
        # IMPORTANT: use total bookValuePerShare (NOT tangible).
        # Companies like ADBE have negative *tangible* equity from goodwill/intangibles but positive
        # total equity, and FMP's returnOnEquityTTM correctly uses total equity — so tangible BVPS
        # is irrelevant here and would incorrectly nullify perfectly valid ROE figures.
        _bvps_total = _first(km.get("bookValuePerShareTTM"), rtm.get("bookValuePerShareTTM"))
        if roe_val is not None and _bvps_total is not None and _bvps_total <= 0:
            roe_val = None  # negative total equity — ROE denominator is meaningless
        # Sanity 2: ROE >> ROIC by a large margin = leverage-inflated ROE, not genuine returns.
        # e.g. PEGA ROE=60%, ROIC=12% → the 60% is entirely driven by financial leverage, not operations.
        # A quality screen should use ROIC; flag ROE as unreliable when divergence is extreme.
        _roic_for_roe_check = _first(km.get("returnOnInvestedCapitalTTM"))
        if (roe_val is not None and _roic_for_roe_check is not None
                and _roic_for_roe_check > 0
                and roe_val > _roic_for_roe_check * 3.0):
            roe_val = None  # leverage-inflated — ROIC is the reliable signal here

        _roa_raw = _first(km.get("returnOnAssetsTTM"))
        roa_val = (round(_roa_raw, 4) if (_roa_raw is not None and abs(_roa_raw) <= 2.0) else None)

        _roic_raw = _first(km.get("returnOnInvestedCapitalTTM"))
        roic_val = (round(_roic_raw, 4) if (_roic_raw is not None and abs(_roic_raw) <= 5.0) else None)
        # Sanity cap: ROIC > 80% is a data artifact (near-zero invested capital base / leverage)
        # Even the best asset-light compounders (Visa, Adobe, Mastercard) run 30-50% ROIC
        # Above 0.80 is almost always a denominator artifact, not a real business signal
        if roic_val is not None and roic_val > 0.80:
            roic_val = None

        # Tangible book: if > 50× price it is certainly a data error
        _tb_raw = _first(rtm.get("tangibleBookValuePerShareTTM"))
        tb_val = (_tb_raw if (_tb_raw is None or price <= 0 or abs(_tb_raw) <= price * 50) else None)

        # D/E: values above 100 are either negative-equity artifacts or data errors
        _de_raw = _first(rtm.get("debtToEquityRatioTTM"))
        de_val = (_de_raw if (_de_raw is not None and -100 < _de_raw <= 100) else None)

        # Apply per-share corruption mask: zero out book-derived metrics
        if _per_share_corrupt:
            pb_val = roe_val = roa_val = roic_val = tb_val = None

        # ── Capital Allocator: CEO score + 5Y per-share CAGRs ──────────
        _bs5  = (bs_5y or {}).get(t) if bs_5y else None
        _cfs5 = (cfs_5y or {}).get(t) if cfs_5y else None
        _exec = (executives or {}).get(t) if executives else None
        ceo_alloc_obj = None
        per_share_cagrs = {"revPerShare5yCagr": None, "fcfPerShare5yCagr": None,
                           "bvPerShare5yCagr": None}
        if _bs5 or _cfs5 or _exec:
            try:
                ceo_alloc_obj = compute_ceo_allocator_score(
                    {}, _bs5 or [], _cfs5 or [], _exec or []
                )
            except Exception:
                ceo_alloc_obj = None
            try:
                per_share_cagrs = compute_per_share_growth_5y(_bs5 or [], _cfs5 or [])
            except Exception:
                pass

        stocks[t] = {
            "ticker": t,
            "name": u.get("name", ""),
            "sector": u.get("sector", "Unknown"),
            "industry": u.get("industry", ""),
            "exchange": u.get("exchange", ""),
            "price": price,
            "mktCap": mktCap,
            "mktCapB": round(mktCap / 1e9, 2) if mktCap else 0,
            "pe": pe_val,
            "peg": peg,
            "fwdPE": fwd_pe,
            "fwdPEG": fwd_peg,
            # P/B, D/E, Div Yield, Gross Margin → ratios-ttm (confirmed field names)
            "pb": pb_val,
            "ps": _first(rtm.get("priceToSalesRatioTTM"), km.get("evToSalesTTM")),
            # ROE/ROA/ROIC → key-metrics-ttm (bounds-checked above)
            "roe": roe_val,
            "roa": roa_val,
            "roic": roic_val,
            # FCF yield: prefer 1/P_FCF (current-price-based) over pre-computed yield
            # (pre-computed uses stale market cap from last report date — distorts fast-movers
            #  and payment processors where gross TPV flows inflate raw FCF figures)
            # Negative FCF is now preserved: P/FCF < 0 → fcfYield < 0 (cash-burning).
            # Capped at [-0.50, +0.50] to filter data errors (P/FCF near zero = extreme yield).
            "fcfYield": (lambda _pfcf, _fy: (
                round(max(min(1.0 / _pfcf, 0.50), -0.50), 4) if (_pfcf and abs(_pfcf) >= 2.0)
                else (_fy if (_fy is not None and -0.50 <= _fy <= 0.50) else None)
            ))(
                _first(km.get("priceToFreeCashFlowsRatioTTM"), rtm.get("priceToFreeCashFlowsRatioTTM")),
                _first(km.get("freeCashFlowYieldTTM"))
            ),
            # Div yield: ratios-ttm confirmed; fallback to profile lastDiv/price
            # Cap at 30% — above that is almost always a data error or imminent dividend cut
            "divYield": (lambda _dy: (min(_dy, 0.30) if _dy else None))(
                _first(
                    rtm.get("dividendYieldTTM"),
                    (prof.get("lastDiv") / price) if (prof.get("lastDiv") and price > 0) else None,
                )
            ),
            "de": de_val,
            "grossMargin": _first(rtm.get("grossProfitMarginTTM")),
            "operatingMargin": _first(rtm.get("operatingProfitMarginTTM")),
            "currentRatio": _first(rtm.get("currentRatioTTM"), km.get("currentRatioTTM")),
            "bookValue": (None if _per_share_corrupt else _bv_ps),
            "tangibleBook": tb_val,
            "iv": iv,
            "mos": mos,
            "piotroski": sc.get("piotroskiScore") or sc.get("piotroski"),
            "altmanZ": sc.get("altmanZScore") or sc.get("altmanZ"),
            "ratingScore": rat.get("ratingScore") or rat.get("score"),
            "recommendation": rat.get("ratingRecommendation") or rat.get("recommendation"),
            "epsGrowth5y": growth_5y,
            "revGrowth5y": rev_growth_5y,
            "revGrowth": _first(gr.get("revenueGrowth"), gr.get("revGrowth")),
            "epsGrowth": _first(gr.get("epsgrowth"), gr.get("epsGrowth")),
            "fcfGrowth": _first(gr.get("freeCashFlowGrowth"), gr.get("fcfGrowth")),
            "netIncomeGrowth": gr.get("netIncomeGrowth"),
            "fiveYRevGrowth": gr.get("fiveYRevenueGrowthPerShare"),
            # ── YoY growth trend — prior year for acceleration/deceleration detection ──
            "revGrowthPrev": _first(gr_prev.get("revenueGrowth"), gr_prev.get("revGrowth")),
            "epsGrowthPrev": _first(gr_prev.get("epsgrowth"), gr_prev.get("epsGrowth")),
            "revGrowthYr2":  _first(gr_yr2.get("revenueGrowth"),  gr_yr2.get("revGrowth")),
            "revGrowthYr3":  _first(gr_yr3.get("revenueGrowth"),  gr_yr3.get("revGrowth")),
            "revGrowthYr4":  _first(gr_yr4.get("revenueGrowth"),  gr_yr4.get("revGrowth")),
            # Revenue consistency: fraction of available years (up to 5) with positive growth
            # 1.0 = all years positive, 0.0 = all years negative
            "revConsistency": (lambda: (
                lambda _h: round(sum(1 for v in _h if v > 0) / len(_h), 2) if _h else None)(
                [v for v in [
                    _first(gr.get("revenueGrowth"), gr.get("revGrowth")),
                    _first(gr_prev.get("revenueGrowth"), gr_prev.get("revGrowth")),
                    _first(gr_yr2.get("revenueGrowth"), gr_yr2.get("revGrowth")),
                    _first(gr_yr3.get("revenueGrowth"), gr_yr3.get("revGrowth")),
                    _first(gr_yr4.get("revenueGrowth"), gr_yr4.get("revGrowth")),
                ] if v is not None]))(),
            # Share dilution: YoY change in weighted avg shares (negative = buybacks = good)
            "sharesGrowth": (lambda _sg: round(_sg, 4) if _sg is not None and abs(_sg) < 0.5 else None)(
                _first(gr.get("weightedAverageSharesGrowth"), gr.get("weightedAverageSharesDilutedGrowth"))
            ),
            # FCF Margin = FCF Yield × P/S ≈ FCF / Revenue (how much of each revenue dollar becomes FCF)
            # Formula: FCF/Revenue = (FCF/MktCap) × (MktCap/Revenue) = fcfYield × P/S
            # Negative margins are now preserved — a -15% FCF margin means burning 15c per $1 revenue.
            "fcfMargin": (lambda _fy, _ps: round(max(min(_fy * _ps, 1.0), -1.0), 3)
                          if (_fy is not None and _ps and _ps > 0) else None)(
                (1.0 / _first(km.get("priceToFreeCashFlowsRatioTTM"), rtm.get("priceToFreeCashFlowsRatioTTM")))
                if _first(km.get("priceToFreeCashFlowsRatioTTM"), rtm.get("priceToFreeCashFlowsRatioTTM")) else
                _first(km.get("freeCashFlowYieldTTM")),
                _first(rtm.get("priceToSalesRatioTTM"), km.get("evToSalesTTM"))
            ),
            # FCF Growth Consistency: fraction of available years (up to 5) with positive FCF growth
            "fcfGrowthConsistency": (lambda: (
                lambda _h: round(sum(1 for v in _h if v > 0) / len(_h), 2) if _h else None)(
                [v for v in [
                    _first(gr.get("freeCashFlowGrowth"), gr.get("fcfGrowth")),
                    _first(gr_prev.get("freeCashFlowGrowth"), gr_prev.get("fcfGrowth")),
                    _first(gr_yr2.get("freeCashFlowGrowth"), gr_yr2.get("fcfGrowth")),
                    _first(gr_yr3.get("freeCashFlowGrowth"), gr_yr3.get("fcfGrowth")),
                    _first(gr_yr4.get("freeCashFlowGrowth"), gr_yr4.get("fcfGrowth")),
                ] if v is not None]))(),
            # FCF Conversion = FCF Yield × P/E ≈ FCF / Net Income
            # 1.0 = FCF equals earnings (high quality); <0.5 = accounting tricks suspected
            "fcfConversion": (lambda _fy, _pe: round(_fy * _pe, 2)
                              if (_fy and _fy > 0 and _pe and 0 < _pe < 200) else None)(
                (1.0 / _first(km.get("priceToFreeCashFlowsRatioTTM"), rtm.get("priceToFreeCashFlowsRatioTTM")))
                if _first(km.get("priceToFreeCashFlowsRatioTTM"), rtm.get("priceToFreeCashFlowsRatioTTM")) else
                _first(km.get("freeCashFlowYieldTTM")),
                pe_val
            ),
            # EPS Growth Optimism: forward EPS CAGR vs 2yr historical EPS growth average
            # If analysts project EPS growing much faster than it has historically, that's a red flag
            "epsGrowthOptimism": (lambda: (
                lambda _fwd, _hist: round(_fwd / _hist - 1, 2)
                if (_fwd and _hist and _hist > 0.01) else None)(
                growth_5y,  # forward 5yr EPS CAGR from analyst estimates
                (lambda _h: sum(_h) / len(_h) if _h else None)(
                    [v for v in [
                        _first(gr_prev.get("epsgrowth"), gr_prev.get("epsGrowth")),
                        _first(gr_yr2.get("epsgrowth"), gr_yr2.get("epsGrowth")),
                    ] if v is not None and v > 0]
                )
            ))(),
            # Growth Optimism = forward analyst estimate CAGR vs 3yr historical average
            # >0 means analysts are more optimistic than history; flag >0.50 as unreliable
            "growthOptimism": (lambda: (
                lambda _fwd, _hist: round(_fwd / _hist - 1, 2)
                if (_fwd and _hist and _hist > 0.01) else None)(
                rev_growth_5y,
                (lambda _h: sum(_h) / len(_h) if _h else None)(
                    [v for v in [
                        _first(gr_prev.get("revenueGrowth"), gr_prev.get("revGrowth")),
                        _first(gr_yr2.get("revenueGrowth"), gr_yr2.get("revGrowth")),
                        _first(gr_yr3.get("revenueGrowth"), gr_yr3.get("revGrowth")),
                    ] if v is not None and v > 0]
                )
            ))(),
            "insiderBuys": ins.get("count", 0),
            "insiderValue": ins.get("totalValue", 0),
            "insiderDate": ins.get("latestDate", ""),
            "beta": prof.get("beta") or u.get("beta"),
            # ── New enrichment fields ──────────────────────────────────────
            # EV/EBITDA — better than P/E for cyclicals + capital-intensive sectors
            "evEbitda": _first(km.get("evToEbitdaTTM")),
            # EV/Revenue — useful when earnings near zero (cyclical trough, turnarounds)
            "evRevenue": _first(km.get("evToSalesTTM"), km.get("evToSalesRatioTTM"),
                                rtm.get("evToSalesRatioTTM"), rtm.get("evToSalesTTM")),
            # P/FCF — cash-flow based valuation (more reliable than P/E for many sectors)
            # Prefer direct ratio; fall back to 1/fcfYield when ratio is unavailable
            "pFcf": (lambda _r, _fy: (
                _r if (_r and 0 < _r < 500)
                else (round(1.0 / _fy, 1) if (_fy and _fy > 0) else None)
            ))(
                _first(rtm.get("priceToFreeCashFlowsRatioTTM"),
                       km.get("priceToFreeCashFlowsRatioTTM")),
                _first(km.get("freeCashFlowYieldTTM"),
                       rtm.get("freeCashFlowYieldTTM")),
            ),
            # Graham Net-Net per share — Benjamin Graham's cheapest possible valuation
            # (current assets - ALL liabilities) / shares; positive = buying below net working capital
            "grahamNetNet": _first(km.get("grahamNetNetTTM")),
            # Net Debt/EBITDA — leverage quality metric (negative = net cash position)
            "netDebtEbitda": _first(km.get("netDebtToEBITDATTM")),
            # Net cash per share (computed from balance sheet above)
            "netCashPerShare": net_cash_per_share,
            "netCashRatio": net_cash_ratio,  # netCashPerShare / price
            # Earnings beat metrics
            "beatRate": beat_rate,           # 0.0–1.0, e.g. 0.875 = beat 7 of 8 quarters
            "epsBeatStreak": eps_beat_streak,  # consecutive quarterly beats (most recent)
            # 52-week positioning
            "priceVs52H": price_vs_52h,  # price / 52wk high; <0.7 = beaten down
            "priceVs52L": price_vs_52l,  # price / 52wk low; >1.5 = well off lows
            # ── A3: Average daily dollar volume (liquidity proxy) ──────────
            # Uses volume field captured at universe fetch (screener or profile.volAvg).
            # $1M/day is the minimum the user can realistically trade; small-cap
            # agents use it as a hard filter to avoid illiquid micro-caps.
            "avgDollarVol": (lambda _v, _p: round(_v * _p, 0)
                             if (_v and _p and _v > 0 and _p > 0) else None)(
                u.get("volume") or u.get("volAvg"), price
            ),
            # ── Capital Allocator metadata ───────────────────────────────
            "ceoAllocator":      ceo_alloc_obj,          # full dict or None
            "revPerShare5yCagr": per_share_cagrs.get("revPerShare5yCagr"),
            "fcfPerShare5yCagr": per_share_cagrs.get("fcfPerShare5yCagr"),
            "bvPerShare5yCagr":  per_share_cagrs.get("bvPerShare5yCagr"),
        }
        # ── A2: Explicit Lynch category — computed after all other fields ──
        # Self-contained classifier; safe to call even when fields are None.
        try:
            stocks[t]["lynchCategory"] = _classify_lynch(stocks[t])
        except Exception:
            stocks[t]["lynchCategory"] = ""

        # ── Sprint 2 A1: Familiar Brand (consumer-observable) tag ──────────
        try:
            stocks[t]["consumerObservable"] = _classify_consumer_observable(stocks[t])
        except Exception:
            stocks[t]["consumerObservable"] = False

        # ── Capital Allocator: Analyst-vs-Insider divergence flag ──────────
        try:
            stocks[t]["divergence"] = classify_divergence(stocks[t])
        except Exception:
            stocks[t]["divergence"] = None

        # ── Sprint 3 A2: Analyst count + Under-Covered tag ──────────────────
        # FMP analyst-estimates payload includes numAnalystsRevenue/numAnalystsEps per
        # forward year.  Most-recent forward year's count is our coverage proxy.
        # Under-covered = a structural Wall Street inefficiency the user can exploit.
        try:
            _ac = None
            if isinstance(est, list):
                # Walk forward to find first entry with a positive analyst count
                for _e in est:
                    _v = _e.get("numAnalystsRevenue") or _e.get("numAnalystsEps")
                    if _v and _v > 0:
                        _ac = int(_v)
                        break
            stocks[t]["analystCount"] = _ac
            # Under-covered thresholds: small/mid cap < 8 analysts, large cap < 12
            _mc_uc = stocks[t].get("mktCap") or 0
            if _ac is not None:
                if _mc_uc < 2e9:
                    stocks[t]["underCovered"] = (_ac < 8)
                else:
                    stocks[t]["underCovered"] = (_ac < 12)
            else:
                # No analyst data at all = almost certainly under-covered (often <1 analyst)
                stocks[t]["underCovered"] = True if _mc_uc and _mc_uc > 50e6 else False
        except Exception:
            stocks[t]["analystCount"] = None
            stocks[t]["underCovered"] = False

    print(f"  ✅ Assembled {len(stocks)} stocks with full data")
    return stocks


# ─────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────

def add_title(ws, title, subtitle="", row=1):
    """Add a styled title row."""
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, name="Arial", size=14, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="1A237E")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
    ws.row_dimensions[row].height = 30
    if subtitle:
        row += 1
        ws.cell(row=row, column=1, value=subtitle).font = Font(italic=True, name="Arial", size=9, color="555555")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
    return row + 1


def write_table(ws, rows, headers, start_row, header_color="1A237E", widths=None, freeze=True):
    """Write a formatted table with headers and data rows."""
    # Headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=header_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[start_row].height = 26
    hdr_row = start_row
    start_row += 1

    # Data rows
    for ri, row in enumerate(rows):
        for ci, h in enumerate(headers, 1):
            v = row.get(h)
            cell = ws.cell(row=start_row, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.font = Font(name="Arial", size=9)
            cell.fill = ALT_FILL if ri % 2 == 0 else PLAIN_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Formatting
            if h == "Price" and isinstance(v, (int, float)):
                cell.number_format = "$#,##0.00"
            elif h == "IV" and isinstance(v, (int, float)):
                cell.number_format = "$#,##0.00"
            elif h == "ROIC" and isinstance(v, (int, float)):
                cell.number_format = "0.0%"
                # Colour-code against ROIC benchmarks: <5% weak, 5-10% ok, 10-15% good, 15-20% very good, 20%+ excellent
                if v >= 0.20:
                    cell.font = Font(bold=True, name="Arial", size=9, color="1B5E20")   # excellent — bold green
                elif v >= 0.15:
                    cell.font = Font(name="Arial", size=9, color="1B5E20")               # very good — green
                elif v >= 0.10:
                    cell.font = Font(name="Arial", size=9, color="33691E")               # good — muted green
                elif v >= 0.05:
                    cell.font = Font(name="Arial", size=9, color="827717")               # ok — amber
                else:
                    cell.font = Font(name="Arial", size=9, color="B71C1C")               # weak — red
            elif h in ("MoS", "ROE", "ROA", "FCF Yield", "Div Yield", "Gross Margin",
                       "Rev Growth", "EPS Growth", "Rev Growth 5Y", "EPS Growth 5Y",
                       "Avg MoS", "Avg FCF Yield", "Avg ROE") and isinstance(v, (int, float)):
                cell.number_format = "0.0%"
                if v > 0.01:
                    cell.font = Font(name="Arial", size=9, color="1B5E20")
                elif v < -0.01:
                    cell.font = Font(name="Arial", size=9, color="B71C1C")
            elif h in ("PEG", "Fwd PEG") and isinstance(v, (int, float)):
                cell.number_format = "0.00"
                if 0 < v < 1.0:
                    cell.fill = PatternFill("solid", fgColor="1B5E20")
                    cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                elif 0 < v < 1.5:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                elif 0 < v < 2.0:
                    cell.fill = PatternFill("solid", fgColor="FFF9C4")
            elif h in ("Shares Δ",) and isinstance(v, (int, float)):
                cell.number_format = "+0.0%;-0.0%;0%"
                if v < -0.02:   # buybacks — green
                    cell.font = Font(name="Arial", size=9, color="1B5E20")
                elif v > 0.04:  # dilution — red
                    cell.font = Font(name="Arial", size=9, color="B71C1C")
            elif h in ("Rev Consist.",) and isinstance(v, (int, float)):
                cell.number_format = "0%"
                if v >= 0.80:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                elif v < 0.40:
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif h in ("FCF Margin",) and isinstance(v, (int, float)):
                cell.number_format = "0%"
                if v >= 0.20:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")  # high margin — great
                elif v >= 0.10:
                    cell.fill = PatternFill("solid", fgColor="F1F8E9")
                elif v < 0.05:
                    cell.fill = PatternFill("solid", fgColor="FFF9C4")  # thin margin — caution
            elif h in ("Oper Margin",) and isinstance(v, (int, float)):
                cell.number_format = "0%"
                if v >= 0.25:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")  # high op margin — scalable biz
                elif v >= 0.10:
                    cell.fill = PatternFill("solid", fgColor="F1F8E9")
                elif v < 0.05:
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")  # thin/negative — caution
            elif h in ("FCF Consist.",) and isinstance(v, (int, float)):
                cell.number_format = "0%"
                if v >= 0.80:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                elif v < 0.40:
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif h in ("FCF Conv.",) and isinstance(v, (int, float)):
                cell.number_format = "0.00"
                if v >= 0.80:   # FCF ≈ earnings — high quality
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                elif v < 0.50:  # FCF << earnings — accounting concern
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif h in ("Grwth Gap", "EPS Gap") and isinstance(v, (int, float)):
                cell.number_format = "+0%;-0%;0%"
                if v > 0.50:    # analysts far too optimistic — flag red
                    cell.font = Font(name="Arial", size=9, color="B71C1C")
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")
                elif v < 0:     # analysts conservative vs history — positive signal
                    cell.font = Font(name="Arial", size=9, color="1B5E20")
            elif h in ("52w Pos", "52w vs High") and isinstance(v, (int, float)):
                cell.number_format = "0%"
                if v <= 0.70:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")   # pulled back — good entry
                elif v >= 0.95:
                    cell.fill = PatternFill("solid", fgColor="FFF9C4")   # extended — near highs
            elif h == "52w vs Low" and isinstance(v, (int, float)):
                cell.number_format = "0%"
                if v <= 1.30:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")   # near lows — early recovery
                elif v >= 1.75:
                    cell.fill = PatternFill("solid", fgColor="FFF9C4")   # well off lows — less upside
            elif h == "Net D/E" and isinstance(v, (int, float)):
                cell.number_format = "0.0"
                if v <= 2.0:
                    cell.font = Font(name="Arial", size=9, color="1B5E20")  # low leverage — green
                elif v >= 5.0:
                    cell.font = Font(name="Arial", size=9, color="B71C1C")  # high leverage — red
            elif h in ("P/E", "Fwd P/E", "P/B", "P/S", "D/E") and isinstance(v, (int, float)):
                cell.number_format = "0.0"
            elif h == "Piotroski" and isinstance(v, (int, float)):
                cell.font = Font(bold=True, name="Arial", size=10)
                if v >= 8:
                    cell.fill = PatternFill("solid", fgColor="1B5E20"); cell.font = Font(bold=True, name="Arial",
                                                                                         size=10, color="FFFFFF")
                elif v >= 6:
                    cell.fill = GREEN_FILL
                elif v <= 3:
                    cell.fill = RED_FILL
            elif h == "MktCap ($B)" and isinstance(v, (int, float)):
                cell.number_format = "$#,##0.0"
            elif h == "Cap Size" and isinstance(v, str):
                if v in ("Micro", "Small"):
                    cell.fill = PatternFill("solid", fgColor="E8F5E9")
                    cell.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                elif v == "Mid":
                    cell.fill = PatternFill("solid", fgColor="FFF9C4")
                elif v in ("Large", "Mega"):
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")
            elif h == "Rating" and isinstance(v, str):
                if "Strong Buy" in str(v):
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                elif "Buy" in str(v):
                    cell.fill = GREEN_FILL
                elif "Sell" in str(v):
                    cell.fill = RED_FILL
            elif h == "🏦 Insider" and v and str(v) != "0" and str(v) != "":
                cell.fill = GREEN_FILL;
                cell.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
            elif h == "Rank" and isinstance(v, int):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if v == 1:
                    cell.fill = PatternFill("solid", fgColor="FFD700")
                    cell.font = Font(bold=True, name="Arial", size=9, color="000000")
                elif v == 2:
                    cell.fill = PatternFill("solid", fgColor="C0C0C0")
                    cell.font = Font(bold=True, name="Arial", size=9, color="000000")
                elif v == 3:
                    cell.fill = PatternFill("solid", fgColor="CD7F32")
                    cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                elif v <= 10:
                    cell.font = Font(bold=True, name="Arial", size=9)

        ws.row_dimensions[start_row].height = 16
        start_row += 1

    # Column widths
    if widths:
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    if freeze:
        ws.freeze_panes = f"A{hdr_row + 1}"
    return start_row


def format_stock_row(s: dict) -> dict:
    """Convert assembled stock data to display-ready row."""
    ins_str = ""
    if s.get("insiderBuys", 0) >= 3:
        ins_str = f"🏦 {s['insiderBuys']}x ${s['insiderValue'] / 1000:.0f}K"
    elif s.get("insiderBuys", 0) >= 1:
        ins_str = f"🏦 {s['insiderBuys']}x"

    return {
        "Ticker": s["ticker"],
        "Company": s["name"],
        "Sector": s["sector"],
        "Price": s["price"],
        "IV": s.get("iv"),
        "MoS": s.get("mos"),
        "PEG": s.get("peg"),
        "Fwd PEG": s.get("fwdPEG"),
        "Fwd P/E": s.get("fwdPE"),
        "P/E": s.get("pe"),
        "P/B": s.get("pb"),
        "EV/EBITDA": s.get("evEbitda"),
        "EV/Rev": s.get("evRevenue"),
        "P/FCF": s.get("pFcf"),
        "ROE": s.get("roe"),
        "ROIC": s.get("roic"),
        "ROA": s.get("roa"),
        "FCF Yield": s.get("fcfYield"),
        "Div Yield": s.get("divYield"),
        "D/E": s.get("de"),
        "Net Debt/EBITDA": s.get("netDebtEbitda"),
        "Curr Ratio": s.get("currentRatio"),
        "Gross Margin": s.get("grossMargin"),
        "Oper Margin": s.get("operatingMargin"),
        "Tangible Book": s.get("tangibleBook"),
        "Net Cash/Sh": s.get("netCashPerShare"),
        "Graham NN": s.get("grahamNetNet"),
        "Beat Rate": s.get("beatRate"),
        "Beat Streak": s.get("epsBeatStreak"),
        "52w Pos": s.get("priceVs52H"),   # price / 52wk high; 0.65 = 35% off high
        "52w vs High": s.get("priceVs52H"),  # alias used in Fast Growers / QC / Stalwarts columns
        "52w vs Low": s.get("priceVs52L"),   # price / 52wk low; 1.30 = 30% off low
        "Net D/E": s.get("netDebtEbitda"),   # net debt / EBITDA; displayed in Turnarounds
        "Rev Consist.": s.get("revConsistency"),
        "Shares Δ": s.get("sharesGrowth"),
        "FCF Conv.": s.get("fcfConversion"),
        "FCF Margin": s.get("fcfMargin"),
        "FCF Consist.": s.get("fcfGrowthConsistency"),
        "Grwth Gap": s.get("growthOptimism"),
        "EPS Gap": s.get("epsGrowthOptimism"),
        "Rev Growth": s.get("revGrowth"),
        "Rev Gr Prev": s.get("revGrowthPrev"),
        "EPS Growth": s.get("epsGrowth"),
        "Rev Growth 5Y": s.get("revGrowth5y"),
        "EPS Growth 5Y": s.get("epsGrowth5y"),
        "5Y Rev/Sh Gr": s.get("fiveYRevGrowth"),
        "Piotroski": s.get("piotroski"),
        "Altman Z": s.get("altmanZ"),
        "Rating": s.get("recommendation"),
        "MktCap ($B)": s.get("mktCapB"),
        "Cap Size": _cap_size_label(s.get("mktCap", 0)),
        "Lynch Cat": s.get("lynchCategory", ""),
        "🏦 Insider": ins_str,
        # Capital Allocator metadata
        "CEO Score": (lambda _c: f"{_c['grade']} ({_c['score']})" if _c and _c.get("grade")
                      else (f"👤 New ({_c['tenure_years']:.0f}y)"
                            if _c and _c.get("tenure_years") is not None and _c['tenure_years'] < 3.0
                            else ""))(s.get("ceoAllocator")),
        "FCF/Sh 5Y": s.get("fcfPerShare5yCagr"),
        "Divergence": ({"hidden_gem": "🎯 Hidden Gem",
                        "conviction_stack": "🔥 Conviction",
                        "quiet_signal": "👁️ Quiet"}.get(s.get("divergence"), "")),
    }


# ─────────────────────────────────────────────
# TAB BUILDERS
# ─────────────────────────────────────────────

def build_iv_discount(wb, stocks):
    """Tab 2: Intrinsic Value Discount — Buffett/Lynch style: good businesses below DCF value.
    Primary signal: DCF Margin of Safety.
    Gate: Piotroski ≥ 6 (value-trap filter only — not in scoring).
    Scoring: MoS + Lynch quality (rev consistency, FCF conversion, EPS growth, buybacks)
             + FCF yield + ROIC + ROE + multi-metric cheapness + D/E + beat rate.
    """
    print("\n📊 Building Tab: IV Discount Picks...")
    qualified = []
    for t, s in stocks.items():
        if not _is_common_stock(s):
            continue
        iv    = s.get("iv")
        price = s.get("price", 0)
        mos   = s.get("mos")
        pio   = s.get("piotroski")
        fcf   = s.get("fcfYield")
        roe   = s.get("roe")
        az    = s.get("altmanZ")

        # Must have a valid DCF-based margin of safety
        if not iv or not price or price <= 0 or not mos:
            continue
        # Only undervalued stocks (positive MoS = DCF > price)
        if mos < 0.05:   # at least 5% discount — noise floor
            continue
        # Piotroski quality gate — stronger than before
        if pio is not None and pio < 6:
            continue
        # Altman Z — exclude financially distressed (bankruptcy zone)
        if az is not None and az < 1.5:
            continue
        # Earnings/cash quality gate — must have at least one positive quality signal
        # Avoids pure speculative DCF plays with no real cash flows
        has_fcf    = (fcf is not None and fcf > 0)
        has_roe    = (roe is not None and roe > 0.08)
        has_roic   = (s.get("roic") is not None and s.get("roic") > 0.08)
        has_pe     = (s.get("pe") and 0 < s.get("pe") < 40)
        if not (has_fcf or has_roe or has_roic or has_pe):
            continue

        # ── Scoring: pure IV-focus, no insider influence on rank ──────────
        score = 0

        # 1. Margin of Safety — the primary signal (40 pts max at 60% MoS)
        score += min(mos * 60, 40)

        # 2. Lynch-style business quality — replaces raw Piotroski in scoring
        #    (Piotroski still guards the gate above; here we reward Lynch metrics)
        #
        # 2a. Revenue consistency — Lynch's first litmus test: "does it grow every year?"
        rc = s.get("revConsistency")
        if rc is not None:
            if rc >= 0.80:   score += 10   # nearly all years positive — very rare quality signal
            elif rc >= 0.60: score += 6
            elif rc >= 0.40: score += 2
            else:            score -= 4    # more bad years than good — value trap warning

        # 2b. FCF conversion — Lynch wanted "real" earnings; FCF ≈ net income = no accounting tricks
        fcc = s.get("fcfConversion")
        if fcc is not None:
            if fcc >= 0.80:   score += 8   # FCF tracks earnings — credible fundamentals
            elif fcc >= 0.60: score += 4
            elif fcc < 0.40:  score -= 5   # FCF << earnings: capex drain or accrual tricks

        # 2c. EPS growth 5Y — Lynch's "PEG company" requires visible earnings growth history
        eg5 = s.get("epsGrowth5y")
        if eg5 is not None:
            if eg5 > 0.15:   score += 7   # Fast Grower territory — Lynch's sweet spot
            elif eg5 > 0.08: score += 4
            elif eg5 > 0:    score += 2
            elif eg5 < -0.05: score -= 4  # shrinking EPS = business getting worse, not just cheap

        # 2d. Share buybacks (negative sharesGrowth = shares retiring = management confidence)
        sg = s.get("sharesGrowth")
        if sg is not None:
            if sg < -0.03:   score += 5   # active buybacks — Lynch loved this signal
            elif sg < 0:     score += 2
            elif sg > 0.10:  score -= 4   # heavy dilution = not a quality business

        # 3. FCF yield — cash generation confirms DCF assumptions are credible
        if fcf and fcf > 0.12:   score += 12
        elif fcf and fcf > 0.08: score += 9
        elif fcf and fcf > 0.05: score += 6
        elif fcf and fcf > 0.02: score += 3

        # 4. ROIC — Buffett's favourite: only buy businesses that earn above cost of capital
        roic = s.get("roic")
        if roic and roic > 0.25:   score += 10
        elif roic and roic > 0.15: score += 7
        elif roic and roic > 0.10: score += 4

        # 5. ROE — return on equity confirms quality
        if roe and roe > 0.25:   score += 7
        elif roe and roe > 0.15: score += 4
        elif roe and roe > 0.08: score += 2

        # 6. Valuation confirmation: cheap on multiple metrics (multi-metric value = lower risk)
        valuation_confirms = 0
        pe  = s.get("pe")
        peg = s.get("peg")
        ps  = s.get("ps")
        if pe  and 0 < pe  < 15:   valuation_confirms += 1
        if pe  and 0 < pe  < 25:   valuation_confirms += 1
        if peg and 0 < peg < 1.0:  valuation_confirms += 1
        if peg and 0 < peg < 1.5:  valuation_confirms += 1
        if ps  and 0 < ps  < 1.0:  valuation_confirms += 1
        if s.get("pb") and 0 < s.get("pb") < 1.5: valuation_confirms += 1
        # EV/EBITDA < 12 = additional multi-metric cheap confirm
        ev_eb = s.get("evEbitda")
        if ev_eb and 0 < ev_eb < 12: valuation_confirms += 1
        score += valuation_confirms * 3   # up to 21 pts from multi-metric cheapness

        # 7. Low debt — clean balance sheet supports Buffett style hold
        de = s.get("de")
        if de is not None and de < 0.3:   score += 4
        elif de is not None and de < 0.8: score += 2

        # 8. Earnings beat rate — management consistently delivers
        br = s.get("beatRate")
        if br and br >= 0.875: score += 5
        elif br and br >= 0.75: score += 3

        row = format_stock_row(s)
        row["Score"] = round(score, 1)
        qualified.append(row)

    qualified.sort(key=lambda x: -x["Score"])
    for i, row in enumerate(qualified):
        row["Rank"] = i + 1

    # Main tab
    ws = wb.create_sheet("2. IV Discount Picks")
    ws.sheet_view.showGridLines = False
    sr = add_title(ws,
                   "💎 Intrinsic Value Discount — Good Businesses Trading Below DCF Value",
                   f"Filter: MoS≥5%, Piotroski≥6 (value-trap gate), Altman Z≥1.5, positive FCF/ROE/ROIC/PE. "
                   f"Score: MoS + Lynch quality (Rev Consistency, FCF Conversion, EPS growth, buybacks) "
                   f"+ FCF yield + ROIC + ROE + multi-metric cheapness + D/E + beat rate. {datetime.date.today()}")

    headers = [
        "Rank", "Ticker", "Company", "Sector", "Price", "IV", "MoS",
        "P/E", "EV/EBITDA", "PEG", "P/B", "FCF Yield", "ROIC", "ROE", "D/E",
        "Beat Rate", "Rev Consist.", "FCF Conv.", "EPS Growth 5Y",
        "MktCap ($B)", "Score", "🏦 Insider",
    ]
    widths = [5, 8, 22, 15, 8, 8, 7, 7, 9, 6, 6, 8, 7, 7, 7, 8, 9, 8, 9, 10, 6, 14]
    write_table(ws, qualified[:TOP_N], headers, sr, header_color="1A237E", widths=widths)
    print(f"  ✅ IV Discount tab done — {min(len(qualified), TOP_N)} stocks (from {len(qualified)} qualifying)")

    # By sector (exclude Rank from sector view for cleaner layout)
    sec_headers = headers[1:]  # drop Rank
    sec_widths = widths[1:]
    build_by_sector(wb, qualified, "2b. IV by Sector",
                    "💎 IV Discount by Sector", "1A237E", sec_headers, sec_widths)

    return qualified


def build_lynch_tab(wb, stocks, category_name, tab_number, filter_fn, sort_key,
                    color, description, custom_headers=None, custom_widths=None):
    """Generic Lynch category tab builder. Pass custom_headers/custom_widths to override defaults."""
    print(f"\n📊 Building Tab {tab_number}: {category_name}...")
    qualified = []
    for t, s in stocks.items():
        if filter_fn(s):
            row = format_stock_row(s)
            row["Score"] = round(sort_key(s), 1)
            qualified.append(row)

    qualified.sort(key=lambda x: -x["Score"])
    for i, row in enumerate(qualified):
        row["Rank"] = i + 1

    ws = wb.create_sheet(f"{tab_number}. {category_name}")
    ws.sheet_view.showGridLines = False
    sr = add_title(ws, f"📚 Lynch: {category_name}", description + f" — {datetime.date.today()}")

    headers = custom_headers or [
        "Rank", "Ticker", "Company", "Sector", "Price", "PEG", "P/E", "P/B", "IV", "MoS",
        "ROE", "FCF Yield", "Rev Growth", "EPS Growth 5Y", "Piotroski",
        "Div Yield", "MktCap ($B)", "Score", "🏦 Insider",
    ]
    widths = custom_widths or [5, 8, 22, 15, 8, 6, 7, 6, 8, 7, 7, 8, 8, 8, 7, 7, 10, 6, 14]
    write_table(ws, qualified[:TOP_N], headers, sr, header_color=color, widths=widths)

    print(f"  ✅ {category_name} tab done — {min(len(qualified), TOP_N)} (from {len(qualified)})")
    return qualified


def build_by_sector(wb, all_rows, sheet_name, title, color, headers, widths):
    """Build a per-sector breakdown sheet."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    sectors = defaultdict(list)
    for r in all_rows:
        sectors[r.get("Sector", "Unknown")].append(r)

    sr = 1
    ws.cell(row=sr, column=1, value=title).font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    ws.cell(row=sr, column=1).fill = PatternFill("solid", fgColor=color)
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
    sr += 2

    for sector in sorted(sectors.keys(), key=lambda s: -len(sectors[s])):
        rows = sectors[sector][:SECTOR_N]
        if not rows: continue
        sc = ws.cell(row=sr, column=1, value=f"📊 {sector} — {len(sectors[sector])} qualifying (top {len(rows)})")
        sc.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        sc.fill = PatternFill("solid", fgColor="37474F")
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
        sr += 1
        sr = write_table(ws, rows, headers, sr, header_color="455A64", widths=widths, freeze=False)
        sr += 1


def _repair_truncated_json(text: str) -> dict:
    """Try to salvage a JSON response that was cut off mid-stream.
    Extracts whatever complete top-level fields and pick objects exist.
    """
    import re
    result = {}

    # Extract simple string fields (synopsis, sector_rotation, macro_context, disclaimer)
    for field in ("synopsis", "sector_rotation", "macro_context", "disclaimer"):
        m = re.search(rf'"{field}"\s*:\s*"((?:[^"\\]|\\.)*)"', text)
        if m:
            result[field] = m.group(1)

    # Extract attention array items
    attn_m = re.search(r'"attention"\s*:\s*\[(.*?)(?:\]|$)', text, re.DOTALL)
    if attn_m:
        items = re.findall(r'"((?:[^"\\]|\\.)*)"', attn_m.group(1))
        if items:
            result["attention"] = items

    # Extract complete pick objects — each pick ends with a closing }
    picks = []
    pick_blocks = re.findall(r'\{[^{}]*"ticker"[^{}]*\}', text, re.DOTALL)
    for block in pick_blocks:
        try:
            p = json.loads(block)
            if p.get("ticker"):
                picks.append(p)
        except Exception:
            pass
    if picks:
        result["picks"] = picks

    return result if result.get("picks") else {}


def call_claude_analysis(picks_data: dict, stocks: dict, macro: dict = None,
                         market_intel: dict = None,
                         agent_perf: dict = None) -> dict:  # B1/B4: per-agent attribution
    """Multi-agent AI stock analysis: 5 specialist agents (parallel) + 1 judge.
    - Quality Growth: sustained compounders with ROIC leadership
    - Special Situation: event-driven, misunderstood, inflection-point plays
    - Capital Appreciation: near-term re-rating candidates with improving momentum
    - Emerging Growth: smaller fast-growers with large TAM and rising ROIC
    - Judge: synthesises all 4 reports into final 3-10 picks
    Returns {} if ANTHROPIC_KEY not set or all calls fail.
    """
    if not ANTHROPIC_KEY:
        print("  ⏭️ No ANTHROPIC_KEY — skipping AI overview")
        return {}

    print("\n  🤖 Running multi-agent AI analysis (11 specialists + judge)...")

    # ── Step 1: Cross-strategy meta-ranking ────────────────────────────────
    # A stock appearing in multiple strategies is cross-validated — stronger signal.
    # We pool the top 15 from each strategy, score by breadth + depth, deduplicate.
    meta = {}  # ticker -> {strategies, best_rank, max_score, row}
    strategy_short = {
        "IV Discount (Buffett/DCF)":        "IV Discount",
        "Quality Compounders (Buffett)":     "Quality Compounder",
        "Stalwarts (Lynch)":                 "Stalwart",
        "Fast Growers (Lynch)":              "Fast Grower",
        "Turnarounds (Lynch)":               "Turnaround",
        "Slow Growers / Income (Lynch)":     "Slow Grower",
        "Cyclicals (Lynch)":                 "Cyclical",
        "Asset Plays (Lynch)":               "Asset Play",
        "Lynch 10-Baggers":                  "10-Bagger",
    }
    for tab_name, rows in picks_data.items():
        short = strategy_short.get(tab_name, tab_name)
        for rank_i, r in enumerate(rows[:15]):
            ticker = r.get("Ticker", "?")
            score  = r.get("Score", 0) or 0
            if ticker not in meta:
                meta[ticker] = {
                    "strategies": [], "best_rank": rank_i + 1,
                    "max_score": score, "row": r,
                }
            meta[ticker]["strategies"].append(short)
            if score > meta[ticker]["max_score"]:
                meta[ticker]["max_score"] = score
            if rank_i + 1 < meta[ticker]["best_rank"]:
                meta[ticker]["best_rank"] = rank_i + 1
                meta[ticker]["row"] = r   # keep the row from its best ranking

    # Meta-score: multi-strategy breadth is the strongest signal
    for t, m in meta.items():
        n   = len(m["strategies"])
        top = max(0, 20 - m["best_rank"])          # rank 1 = 19 pts, rank 13 = 7 pts, rank 20 = 0 pts
        m["meta_score"] = n * 25 + m["max_score"] * 0.4 + top

    # Top 35 unique stocks by meta-score — used by judge and as fallback
    top_stocks = sorted(meta.values(), key=lambda x: -x["meta_score"])[:35]
    # Full unique universe — ALL stocks that appeared in any strategy tab, used for per-agent re-ranking.
    # No cap: a stock that scored #1 in Turnarounds but nowhere else is still a valid deep-value candidate.
    meta_all   = sorted(meta.values(), key=lambda x: -x["meta_score"])

    # ── Separate small-cap candidate pool for size-constrained agents ──────
    # Agents like Mayer100x, TenBagger, GoldmanSC, WallStBlind have a hard <$2B cap.
    # Passing them the general candidates_block (dominated by mid/large-caps) causes
    # the AI to pick familiar names like QLYS ($3B) or DOCS ($4.6B) in violation of their mandate.
    # Fix: build a dedicated sc pool from the full meta dict, then expand with 10-Bagger tab entries.
    _SC_CAP_LIMIT = 2.0  # $2B
    _sc_pool_all = sorted(
        [m for m in meta.values()
         if (stocks.get(m["row"].get("Ticker",""), {}).get("mktCapB") or
             m["row"].get("MktCap ($B)") or 999) < _SC_CAP_LIMIT],
        key=lambda x: -x["meta_score"]
    )
    # If fewer than 12 small-caps in meta, pull additional from the Lynch 10-Baggers rows in picks_data
    if len(_sc_pool_all) < 12:
        _existing_sc_tickers = {m["row"].get("Ticker") for m in _sc_pool_all}
        for _tab_name, _tab_rows in picks_data.items():
            for _tr in (_tab_rows or []):
                _tt = _tr.get("Ticker","")
                if not _tt or _tt in _existing_sc_tickers: continue
                _mc = (stocks.get(_tt,{}).get("mktCapB") or _tr.get("MktCap ($B)") or 999)
                if _mc < _SC_CAP_LIMIT:
                    _sc_pool_all.append({"row": _tr, "strategies": [_tab_name],
                                         "meta_score": 10, "max_score": _tr.get("Score",0),
                                         "best_rank": _tr.get("Rank",50)})
                    _existing_sc_tickers.add(_tt)
                    if len(_sc_pool_all) >= 25: break
            if len(_sc_pool_all) >= 25: break
    sc_top_stocks = _sc_pool_all[:25]  # up to 25 small-cap candidates

    # ── Step 2: Rich per-stock formatter ───────────────────────────────────
    def fmt_stock(m):
        r  = m["row"]
        t  = r.get("Ticker", "?")
        s  = stocks.get(t, {})   # raw stock data for extra fields not in row
        strats = " + ".join(m["strategies"])
        rank_tag = f"★×{len(m['strategies'])}" if len(m["strategies"]) > 1 else f"#{m['best_rank']}"

        # Core valuation
        parts = [f"{t} [{rank_tag}] ({r.get('Company','')[:20]}) | {strats}"]
        for k, lbl in [("P/E","PE"), ("PEG","PEG"), ("EV/EBITDA","EV/EB"),
                        ("P/B","PB")]:
            v = r.get(k) or s.get(k.lower().replace("/","").replace(" ",""))
            if v is not None:
                parts.append(f"{lbl}={v:.1f}" if isinstance(v, float) else f"{lbl}={v}")
        # MoS is a decimal (0.81 = 81%) — display as percentage for AI readability
        _mos_v = r.get("MoS")
        if _mos_v is not None and isinstance(_mos_v, float):
            parts.append(f"MoS={_mos_v*100:.0f}%")
        elif _mos_v is not None:
            parts.append(f"MoS={_mos_v}")

        # Quality — ROE/ROIC/FCF Yield are decimals (0.15 = 15%) — multiply by 100 for AI readability
        for k, lbl in [("ROE","ROE"), ("ROIC","ROIC"), ("FCF Yield","FCF")]:
            v = r.get(k)
            if v is not None:
                parts.append(f"{lbl}={v*100:.0f}%" if isinstance(v, float) else f"{lbl}={v}")
        for k, lbl in [("Piotroski","Pio"), ("Altman Z","AltZ")]:
            v = r.get(k)
            if v is not None:
                parts.append(f"{lbl}={v:.2f}" if isinstance(v, float) else f"{lbl}={v}")

        # Growth
        for k, lbl in [("Rev Growth","RG"), ("Rev Gr Prev","RGprev"),
                        ("EPS Growth 5Y","EG5y"), ("Rev Growth 5Y","RG5y")]:
            v = r.get(k)
            if v is not None:
                parts.append(f"{lbl}={v:+.0%}")

        # New enrichment
        ncr = s.get("netCashRatio")
        if ncr is not None:
            parts.append(f"NetCash={ncr:+.0%}")
        pvs52h = s.get("priceVs52H")
        if pvs52h is not None:
            parts.append(f"52wPos={pvs52h:.0%}")
        de = r.get("D/E") or s.get("de")
        if de is not None:
            parts.append(f"DE={de:.1f}")
        # Filter 1: ROIC quality badge — moat confirmation
        _roic_ai = s.get("roic")
        if _roic_ai and _roic_ai > 0.15:
            parts.append(f"✅ROIC>{_roic_ai*100:.0f}%")
        # Change 4: Flag when analysts project >150% EPS growth in one year (fwdPE < 40% of trailing)
        _fpe_ai = s.get("fwdPE")
        _pe_ai  = s.get("pe")
        if _fpe_ai and _pe_ai and _pe_ai > 0 and _fpe_ai < _pe_ai * 0.40:
            parts.append(f"⚠FwdPElook(fwd={_fpe_ai:.1f}vs{_pe_ai:.1f}trail)")
        # Filter 3: FCF quality checks
        _fcc_ai = s.get("fcfConversion")
        if _fcc_ai is not None:
            if _fcc_ai < 0.50:
                parts.append(f"⚠FCFConv={_fcc_ai:.1f}(low)")
            elif _fcc_ai >= 0.80:
                parts.append(f"FCFConv={_fcc_ai:.1f}(✓)")
        _fcm_ai = s.get("fcfMargin")
        if _fcm_ai is not None:
            parts.append(f"FCFMargin={_fcm_ai:.0%}")
        _fgc_ai = s.get("fcfGrowthConsistency")
        if _fgc_ai is not None and _fgc_ai < 0.50:
            parts.append(f"⚠FCFConsist={_fgc_ai:.0%}")
        # Filter 4: Growth optimism flags — forward vs historical (revenue + EPS)
        _go_ai  = s.get("growthOptimism")
        _ego_ai = s.get("epsGrowthOptimism")
        if _go_ai is not None and _go_ai > 0.50:
            parts.append(f"⚠RevGap=+{_go_ai:.0%}")
        if _ego_ai is not None and _ego_ai > 0.50:
            parts.append(f"⚠EpsGap=+{_ego_ai:.0%}")
        if (_go_ai is not None and _go_ai > 0.50) and (_ego_ai is not None and _ego_ai > 0.50):
            parts.append("🚩BOTH Rev+EPS estimates far above history")

        # Change 3: Beat rate — management execution quality
        _br_ai = s.get("beatRate")
        if _br_ai is not None:
            if _br_ai >= 0.75:
                parts.append(f"✅BeatRate={_br_ai:.0%}")
            elif _br_ai < 0.50:
                parts.append(f"⚠BeatRate={_br_ai:.0%}(low)")

        # Change 3: Gross margin — pricing power / business model quality
        _gm_ai = s.get("grossMargin")
        if _gm_ai is not None:
            if _gm_ai >= 0.60:
                parts.append(f"✅GrMgn={_gm_ai:.0%}")
            elif _gm_ai < 0.25:
                parts.append(f"⚠GrMgn={_gm_ai:.0%}(low)")
            else:
                parts.append(f"GrMgn={_gm_ai:.0%}")

        # Change 5: Operating margin — asset-light vs capital-heavy distinction
        _om_ai = s.get("operatingMargin")
        if _om_ai is not None:
            if _om_ai >= 0.25:
                parts.append(f"✅OprMgn={_om_ai:.0%}")
            elif _om_ai < 0.05:
                parts.append(f"⚠OprMgn={_om_ai:.0%}(thin)")
            else:
                parts.append(f"OprMgn={_om_ai:.0%}")

        # ── A1: Surface previously-hidden fields for the AI agents ─────────
        # Revenue consistency — fraction of last 5 years with positive growth
        _rc_ai = s.get("revConsistency")
        if _rc_ai is not None:
            if _rc_ai >= 0.80:   parts.append(f"✅RevConsist={_rc_ai:.0%}")
            elif _rc_ai < 0.40:  parts.append(f"⚠RevConsist={_rc_ai:.0%}(erratic)")
            else:                parts.append(f"RevConsist={_rc_ai:.0%}")
        # Share dilution — positive = issuing stock (bad), negative = buybacks (good)
        _shg_ai = s.get("sharesGrowth")
        if _shg_ai is not None:
            if _shg_ai > 0.15:   parts.append(f"🚩Dil=+{_shg_ai:.0%}(heavy)")
            elif _shg_ai > 0.05: parts.append(f"⚠Dil=+{_shg_ai:.0%}")
            elif _shg_ai < -0.02: parts.append(f"✅Buyback={_shg_ai:+.0%}")
        # Current ratio — short-term liquidity; <1.0 = possible cash crunch
        _cr_ai = s.get("currentRatio")
        if _cr_ai is not None:
            if _cr_ai < 1.0:     parts.append(f"⚠CR={_cr_ai:.1f}(low)")
            elif _cr_ai > 2.5:   parts.append(f"CR={_cr_ai:.1f}(strong)")
            else:                parts.append(f"CR={_cr_ai:.1f}")
        # Net Debt / EBITDA — leverage quality; >4 = over-leveraged
        _nde_ai = s.get("netDebtEbitda")
        if _nde_ai is not None:
            if _nde_ai < 0:      parts.append(f"✅NDE={_nde_ai:.1f}(net cash)")
            elif _nde_ai > 4.0:  parts.append(f"🚩NDE={_nde_ai:.1f}(over-levered)")
            elif _nde_ai > 2.5:  parts.append(f"⚠NDE={_nde_ai:.1f}")
            else:                parts.append(f"NDE={_nde_ai:.1f}")
        # Dividend yield — income component
        _dy_ai = s.get("divYield")
        if _dy_ai is not None and _dy_ai > 0:
            if _dy_ai > 0.06:    parts.append(f"⚠DY={_dy_ai:.1%}(high—check coverage)")
            elif _dy_ai > 0.025: parts.append(f"DY={_dy_ai:.1%}")
        # Graham Net-Net — cheapest-possible valuation per share (>0 = below net working capital)
        _gnn_ai = s.get("grahamNetNet")
        _price_ai = s.get("price")
        if _gnn_ai is not None and _price_ai and _gnn_ai > 0 and _price_ai > 0:
            _gnn_ratio = _gnn_ai / _price_ai
            if _gnn_ratio > 1.0: parts.append(f"💎GrahamNN={_gnn_ai:.1f}(>price—deep value)")
            elif _gnn_ratio > 0.66: parts.append(f"✅GrahamNN={_gnn_ai:.1f}")
        # Piotroski weakness flag — business quality check
        _pio_ai = s.get("piotroski")
        if _pio_ai is not None and _pio_ai < 4:
            parts.append(f"🚩PioWeak={_pio_ai}(fundamentals deteriorating)")

        # A3: Average daily dollar volume — liquidity check for small-caps
        _adv_ai = s.get("avgDollarVol")
        if _adv_ai is not None and _adv_ai > 0:
            if _adv_ai < 500_000:
                parts.append(f"🚩ADV=${_adv_ai/1e3:.0f}K(illiquid)")
            elif _adv_ai < 1_000_000:
                parts.append(f"⚠ADV=${_adv_ai/1e3:.0f}K(thin)")
            elif _adv_ai >= 10_000_000:
                parts.append(f"ADV=${_adv_ai/1e6:.0f}M")
            else:
                parts.append(f"ADV=${_adv_ai/1e6:.1f}M")

        # Size + sector
        mc = s.get("mktCapB") or r.get("MktCap ($B)")
        if mc:
            parts.append(f"Mcap=${mc:.1f}B")
        sector = s.get("sector") or r.get("Sector", "")
        if sector:
            parts.append(f"[{sector}]")
        # A2: Lynch category — helps agents match framework to business type
        _lc_ai = s.get("lynchCategory")
        if _lc_ai:
            parts.append(f"[Lynch={_lc_ai}]")
        # Sprint 2 A1: Familiar Brand — consumer-observable industries
        # Signal to agents that the user can directly evaluate the product/service
        if s.get("consumerObservable"):
            parts.append("🛒FamiliarBrand")
        # Sprint 3 A2: Under-covered (Wall Street blindspot quantified)
        _ac_ai = s.get("analystCount")
        if s.get("underCovered"):
            if _ac_ai is None:
                parts.append("🔍UnderCovered(no analysts)")
            elif _ac_ai <= 3:
                parts.append(f"🔍UnderCovered({_ac_ai}an)")
            else:
                parts.append(f"🔍UnderCovered({_ac_ai}an)")

        # ── Capital Allocator: CEO score + per-share CAGRs + divergence ──
        _ceo = s.get("ceoAllocator") or {}
        if _ceo.get("grade"):
            _t  = _ceo.get("tenure_years")
            _f  = _ceo.get("fcf_per_share_cagr")
            _sc = _ceo.get("shares_change_pct")
            _bits = [f"👑CEO={_ceo['grade']}"]
            if _t is not None: _bits.append(f"{_t:.0f}y")
            if _f is not None: _bits.append(f"FCF/sh{_f*100:+.0f}%/y")
            if _sc is not None and abs(_sc) >= 0.02:
                _bits.append(f"shares{_sc*100:+.0f}%")
            parts.append("(" + ",".join(_bits) + ")")
        elif _ceo.get("tenure_years") is not None and _ceo["tenure_years"] < 3.0:
            parts.append(f"👤NewCEO({_ceo['tenure_years']:.0f}y)")
        # Per-share growth (5Y) — explicit signal independent of CEO score
        _fcf_ps = s.get("fcfPerShare5yCagr")
        _bv_ps  = s.get("bvPerShare5yCagr")
        if _fcf_ps is not None and _ceo.get("grade") is None:
            parts.append(f"FCF/sh5y={_fcf_ps*100:+.0f}%/y")
        if _bv_ps is not None and _bv_ps > 0.10:
            parts.append(f"BV/sh5y={_bv_ps*100:+.0f}%/y")
        # Divergence flag
        _div = s.get("divergence")
        if _div == "hidden_gem":
            parts.append("🎯HiddenGem(insiders+bearishStreet)")
        elif _div == "conviction_stack":
            parts.append("🔥ConvictionStack(insiders+bullishStreet)")
        elif _div == "quiet_signal":
            parts.append("👁️QuietInsiderSignal")

        return "  " + " | ".join(parts)

    candidate_lines = [fmt_stock(m) for m in top_stocks]

    # ── Step 3: Sector context ──────────────────────────────────────────────
    sector_stats = {}
    for s in stocks.values():
        sect = s.get("sector", "Unknown")
        if sect not in sector_stats:
            sector_stats[sect] = {"pegs": [], "fcfs": [], "rgs": [], "count": 0}
        sector_stats[sect]["count"] += 1
        peg = s.get("peg")
        if peg and 0 < peg < 15: sector_stats[sect]["pegs"].append(peg)
        fcf = s.get("fcfYield")
        if fcf is not None: sector_stats[sect]["fcfs"].append(fcf)
        rg = s.get("revGrowth")
        if rg is not None: sector_stats[sect]["rgs"].append(rg)

    def _med(lst): return sorted(lst)[len(lst)//2] if lst else None
    sector_lines = []
    for sect, d in sorted(sector_stats.items(),
                          key=lambda x: _med(x[1]["pegs"]) or 99):
        if d["count"] < 5: continue
        med_peg = _med(d["pegs"])
        avg_fcf = sum(d["fcfs"]) / len(d["fcfs"]) if d["fcfs"] else None
        avg_rg  = sum(d["rgs"])  / len(d["rgs"])  if d["rgs"]  else None
        line = f"  {sect}: "
        if med_peg: line += f"PEG={med_peg:.1f} "
        if avg_fcf is not None: line += f"FCF={avg_fcf:.1%} "
        if avg_rg  is not None: line += f"RevGr={avg_rg:.1%} "
        line += f"({d['count']} stocks)"
        sector_lines.append(line)

    # ── Step 4: Shared helpers ───────────────────────────────────────────────
    def _parse_response(text):
        text = text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return _repair_truncated_json(text)

    # Model constants — specialists use Haiku (cheap, fast); judge uses Sonnet (quality)
    _SPECIALIST_MODEL = "claude-haiku-4-5"
    _JUDGE_MODEL      = "claude-sonnet-4-6"

    def _post(sys_p, usr_p, max_tok, timeout_s, model=None):
        return requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": ANTHROPIC_KEY,
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": model or _JUDGE_MODEL,
                  "max_tokens": max_tok,
                  "system": sys_p,
                  "messages": [{"role": "user", "content": usr_p}]},
            timeout=timeout_s,
        )

    candidates_block = (
        "Legend: RG=RevGrowth, RGprev=prior yr RevGrowth (trend), MoS=DCF margin of safety, "
        "NetCash=net cash/price, 52wPos=price vs 52wk high, ★×N=appears in N strategies\n"
        + chr(10).join(candidate_lines)
    )

    # Small-cap only block — pre-filtered to mktCap < $2B for size-constrained agents
    _sc_candidate_lines = [fmt_stock(m) for m in sc_top_stocks]
    sc_candidates_block = (
        "⚠️ PRE-FILTERED: ALL stocks in this list have market cap UNDER $2B. "
        "Every stock here already passes the size filter.\n"
        "Legend: RG=RevGrowth, MoS=DCF margin of safety, Mcap=market cap\n"
        + (chr(10).join(_sc_candidate_lines) if _sc_candidate_lines
           else "  (no qualifying small-cap candidates this run — return empty picks list)")
    )

    sector_block = chr(10).join(sector_lines[:10])

    # ── Strategy-specific candidate pools ────────────────────────────────────
    # Each agent gets the top-100 universe re-ranked by *their* preferred metrics.
    # This is the primary fix for agent convergence — same universe, different lens.
    def _rv(m, k):   return m["row"].get(k)
    def _sv(m, k):   return stocks.get(m["row"].get("Ticker",""), {}).get(k)

    def _agent_pool(score_fn, top_n=50, strict=False):
        """Re-rank the FULL strategy-tab universe by this agent's metric and return top_n.
        Draws from all unique stocks across all tabs (no universe cap).
        strict=True: scoring fn returns None to hard-exclude a stock (e.g. Magic Formula cutoffs).
        """
        pairs = [(score_fn(m), m) for m in meta_all]
        if strict:
            # Drop stocks where the scoring fn returned None — they don't qualify at all
            pairs = [(s, m) for s, m in pairs if s is not None]
        else:
            pairs = [((s if s is not None else -9999), m) for s, m in pairs]
        scored = sorted(pairs, key=lambda x: -x[0])
        lines = [fmt_stock(m) for _, m in scored[:top_n]]
        hdr = ("Legend: RG=RevGrowth, MoS=DCF margin of safety, 52wPos=price vs 52wk high, "
               "★×N=appears in N strategies — list pre-ranked for YOUR strategy lens\n")
        return hdr + chr(10).join(lines)

    # Quality Growth — rank by ROE + gross margin + FCF quality + Piotroski
    def _sc_quality_growth(m):
        return ((_sv(m,"roe") or 0)*25 + (_sv(m,"grossMargin") or 0)*20
                + (_sv(m,"operatingMargin") or 0)*15 + (_rv(m,"Rev Growth 5Y") or 0)*15
                + (_rv(m,"Piotroski") or 0)*2.5 + (_sv(m,"fcfConversion") or 0)*10
                + (_sv(m,"roic") or 0)*20)

    # Special Situation — rank by beaten-down + improving + catalyst signals
    def _sc_special_sit(m):
        pvs = _sv(m,"priceVs52H") or 0.8
        rg  = (_rv(m,"Rev Growth") or 0)*100
        rgp = (_rv(m,"Rev Gr Prev") or 0)*100
        de  = _sv(m,"de") or 3.0
        return ((1.0-pvs)*35 + (_rv(m,"MoS") or 0)*25
                + max(0, rg-rgp)*0.25 + max(0, 30-de*5)*0.15)

    # Capital Appreciation — rank by momentum (near 52w high = uptrend) + rev acceleration
    def _sc_cap_appreciation(m):
        pvs = (_sv(m,"priceVs52H") or 0)*100
        rg  = (_rv(m,"Rev Growth") or 0)*100
        rgp = (_rv(m,"Rev Gr Prev") or 0)*100
        fpe = _sv(m,"fwdPE") or 0
        fv  = max(0, 50-fpe) if fpe and 0 < fpe < 50 else 0
        br  = (_sv(m,"beatRate") or 0)*100
        return pvs*0.30 + rg*0.20 + max(0,rg-rgp)*0.20 + fv*0.15 + br*0.15

    # Emerging Growth — rank by highest revenue growth + small/mid-cap preference
    def _sc_emerging_growth(m):
        mc  = _sv(m,"mktCapB") or 50
        sz  = 20 if mc < 2 else 10 if mc < 5 else 5 if mc < 15 else 0
        peg = _rv(m,"PEG") or _sv(m,"peg") or 99
        pk  = max(0, 30-peg*5) if 0 < peg < 6 else 0
        return ((_rv(m,"Rev Growth") or 0)*35 + (_rv(m,"Rev Growth 5Y") or 0)*25
                + (_rv(m,"EPS Growth 5Y") or 0)*15 + sz + pk*0.25)

    # Lynch BWYK — rank by PEG (lowest = best) + consumer sector + real earnings
    def _sc_lynch(m):
        peg  = _rv(m,"PEG") or _sv(m,"peg") or 99
        ps   = max(0, 60-peg*15) if 0 < peg < 4 else 0
        sect = _sv(m,"sector") or ""
        cb   = 25 if any(w in sect.lower() for w in
                         ["consumer","retail","restaurant","food","leisure","hotel"]) else 0
        return (ps*0.45 + (_rv(m,"Rev Growth") or 0)*20
                + cb + (_rv(m,"Piotroski") or 0)*2.0
                + max(_sv(m,"fcfYield") or 0, 0)*15)

    # Social Arbitrage — rank by rev acceleration + beat rate + consumer/tech momentum
    def _sc_social_arb(m):
        rg  = (_rv(m,"Rev Growth") or 0)*100
        rgp = (_rv(m,"Rev Gr Prev") or 0)*100
        sect = _sv(m,"sector") or ""
        sb  = 20 if any(w in sect.lower() for w in
                        ["consumer","tech","software","retail","media","entertainment"]) else 0
        mc  = _sv(m,"mktCapB") or 50
        return (max(0,rg-rgp)*0.35 + (_sv(m,"beatRate") or 0)*25
                + rg*0.15 + sb + (15 if mc < 10 else 0))

    # Disruptive Innovation — rank by tech/healthcare + highest growth + scalable margins
    def _sc_cathie_wood(m):
        sect = _sv(m,"sector") or ""
        tb  = 30 if any(w in sect.lower() for w in
                        ["tech","software","healthcare","biotech","semiconductor"]) else 0
        mc  = _sv(m,"mktCapB") or 50
        return (tb + (_rv(m,"Rev Growth") or 0)*30 + (_rv(m,"Rev Growth 5Y") or 0)*20
                + (_sv(m,"grossMargin") or 0)*20 + (10 if mc < 20 else 0))

    # Magic Formula — Greenblatt's two hard cutoffs: ROIC>20% AND earnings yield>10%
    # Returns None for any stock failing either threshold — excluded from pool entirely.
    def _sc_magic_formula(m):
        pe   = _rv(m,"P/E") or _sv(m,"pe") or 0
        ey   = (100/pe) if pe and pe > 0 else 0          # earnings yield = 1/PE * 100
        roic = (_sv(m,"roic") or 0)*100
        # HARD CUTOFFS — both must pass or stock is excluded from the pool
        if roic < 20 or ey < 10:
            return None
        # Among qualifying stocks, rank by combined score (EY weighted 50%, ROIC 40%, FCF 10%)
        fcc  = (_sv(m,"fcfConversion") or 0)*100
        return ey*0.50 + roic*0.40 + fcc*0.10

    # Pabrai Asymmetric — rank by DCF margin of safety + P/B cheapness + clean balance sheet
    def _sc_pabrai(m):
        pb   = _rv(m,"P/B") or _sv(m,"pb") or 20
        pbs  = max(0, 80-pb*20) if 0 < pb < 4 else 0
        de   = _sv(m,"de") or 3
        dp   = max(0, de-1)*-10
        return ((_rv(m,"MoS") or 0)*45 + pbs*0.25
                + (_sv(m,"roic") or 0)*15 + (_rv(m,"Piotroski") or 0)*2.0 + dp)

    # Howard Marks — rank by most beaten-down yet fundamentally OK (contrarian)
    def _sc_howard_marks(m):
        pvs  = _sv(m,"priceVs52H") or 0.8
        pio  = (_rv(m,"Piotroski") or 0)
        return ((1.0-pvs)*40 + (_rv(m,"MoS") or 0)*25
                + pio*3.0 + max(_sv(m,"fcfYield") or 0, 0)*15
                + max(_rv(m,"Rev Growth") or 0, 0)*20)

    # Nick Sleep — rank by high gross margin + FCF conversion + sticky recurring revenue
    def _sc_nick_sleep(m):
        return ((_sv(m,"grossMargin") or 0)*35 + (_sv(m,"fcfConversion") or 0)*25
                + (_rv(m,"Rev Growth 5Y") or 0)*20 + (_sv(m,"operatingMargin") or 0)*10
                + (_sv(m,"roic") or 0)*10)

    # Burry Deep Value — rank by cheapest EV/EBITDA + P/B + highest MoS
    def _sc_burry(m):
        ev   = _rv(m,"EV/EBITDA") or 99
        evs  = max(0, 80-ev*4) if 0 < ev < 20 else 0
        pb   = _rv(m,"P/B") or _sv(m,"pb") or 20
        pbs  = max(0, 80-pb*20) if 0 < pb < 4 else 0
        return evs*0.35 + pbs*0.25 + (_rv(m,"MoS") or 0)*25 + (_rv(m,"Piotroski") or 0)*2.0

    # Insider Track — heavily prioritize stocks with insider buying + quality confirmation
    def _sc_insider(m):
        has_ins = 150 if m["row"].get("🏦 Insider","") else 0
        return (has_ins + (_sv(m,"roic") or 0)*25 + (_rv(m,"MoS") or 0)*20
                + max(_sv(m,"fcfYield") or 0, 0)*15 + (_rv(m,"Piotroski") or 0)*2.0)

    # Pre-build all per-agent blocks (done once, before parallel launch)
    _pool_quality_growth   = _agent_pool(_sc_quality_growth)
    _pool_special_sit      = _agent_pool(_sc_special_sit)
    _pool_cap_appreciation = _agent_pool(_sc_cap_appreciation)
    _pool_emerging_growth  = _agent_pool(_sc_emerging_growth)
    _pool_lynch            = _agent_pool(_sc_lynch)
    _pool_cathie_wood      = _agent_pool(_sc_cathie_wood)
    _pool_pabrai           = _agent_pool(_sc_pabrai)
    _pool_howard_marks     = _agent_pool(_sc_howard_marks)
    _pool_burry            = _agent_pool(_sc_burry)
    _pool_insider          = _agent_pool(_sc_insider)

    # ── Special Situation: extract top-15 tickers and fetch company-specific news ──
    _ss_top_tickers = [
        m["row"].get("Ticker", "")
        for _, m in sorted(
            [((_sc_special_sit(m) if _sc_special_sit(m) is not None else -9999), m)
             for m in meta_all],
            key=lambda x: -x[0],
        )[:15]
        if m["row"].get("Ticker", "")
    ]
    _special_sit_news = fetch_special_sit_news(_ss_top_tickers)

    # ── Market intelligence context for real-world-aware agents ──────────────
    _mi = market_intel or {}
    _consumer_intel = _mi.get("consumer_trends", "")
    _tech_intel     = _mi.get("tech_trends", "")
    _insider_intel  = _mi.get("insider_activity", "")
    _catalyst_intel = _mi.get("catalyst_news", "")

    # ── Step 5: Three specialist agents (parallel) ──────────────────────────
    SPECIALIST_JSON_SCHEMA = (
        '{"picks":['
        '{"ticker":"X","company":"Name",'
        '"business_synopsis":"2-3 sentences: what does this company do, how does it make money, who are its customers? Pure factual description — no investment thesis here.",'
        '"industry":"Specific industry or sub-sector (e.g. Cloud Security, Specialty Pharma, Industrial Automation)",'
        '"key_competitors":"Top 2-3 competitor names, comma-separated (e.g. Salesforce, HubSpot, Zoho)",'
        '"rationale":"1-2 sentences citing YOUR framework\'s specific metrics — '
        'e.g. Magic Formula: ROIC X% + EY Y%; SpecialSit: catalyst name + timeline; '
        'InsiderTrack: who bought $X on date; Pabrai: floor $X vs upside $Y = N:1. '
        'This must reflect YOUR investing lens, not generic text.",'
        '"brief_case":"one sentence — the market-misunderstanding thesis",'
        '"key_metric":"the single most compelling number","conviction":"HIGH|MEDIUM"}'
        ',...]}'
    )

    # ── B4: Agent-performance feedback block ────────────────────────────────
    # Build a one-liner per agent from B1 attribution data.  Display-only (R1 decision):
    # each specialist sees how their past picks performed so they can self-calibrate.
    _AGENT_DISPLAY = {
        "AI-QualityGrowth":   "QualityGrowth",  "AI-SpecialSit":      "SpecialSit",
        "AI-CapAppreciation": "CapAppreciation", "AI-EmergingGrowth":  "EmergingGrowth",
        "AI-TenBagger":       "TenBagger",
        "AI-LynchBWYK":       "LynchBWYK",
        "AI-CathieWood":      "CathieWood",
        "AI-Pabrai":          "Pabrai",
        "AI-HowardMarks":     "HowardMarks",
        "AI-Burry":           "Burry",           "AI-InsiderTrack":    "InsiderTrack",
    }
    _perf_lines = {}   # agent_short_name → one-liner string
    if agent_perf:
        for src, stats in agent_perf.items():
            short = _AGENT_DISPLAY.get(src)
            if not short:
                continue
            n    = stats.get("n_picks", 0)
            if n < 5:
                continue   # not enough history for meaningful feedback
            alpha  = stats.get("alpha")
            sharpe = stats.get("sharpe")
            wr     = stats.get("win_rate")
            h90    = stats.get("hit_90d")
            parts  = [f"{n} picks"]
            if alpha  is not None: parts.append(f"alpha {alpha:+.1%}")
            if sharpe is not None: parts.append(f"Sharpe {sharpe:.2f}")
            if wr     is not None: parts.append(f"win-rate {wr:.0%}")
            if h90    is not None: parts.append(f"90d hit-rate {h90:.0%}")
            note = ""
            if alpha is not None and alpha < -0.02:
                note = " ⚠ Alpha negative — tighten quality bar; fewer, higher-conviction picks."
            elif sharpe is not None and sharpe < 0:
                note = " ⚠ Risk-adjusted return negative — raise the bar on entry price and catalyst specificity."
            _perf_lines[short] = "  ".join(parts) + note

    def _perf_header(agent_name: str) -> str:
        """Return a performance context line for agent_name, or empty string."""
        line = _perf_lines.get(agent_name, "")
        if not line:
            return ""
        return f"\nYOUR RECENT TRACK RECORD (display only — context for self-calibration):\n{line}\n"

    # ── B4 (judge): build track record + specialist leaderboard block ────────
    _judge_track_block = ""
    if agent_perf:
        # 1. Judge's own historical performance
        jst = agent_perf.get("AI-Judge", {})
        j_n = jst.get("n_picks", 0)
        if j_n >= 3:
            j_alpha = jst.get("alpha")
            j_wr    = jst.get("win_rate")
            j_sharpe= jst.get("sharpe")
            j_best  = jst.get("best_ticker", "—")
            j_bret  = jst.get("best_ret")
            j_worst = jst.get("worst_ticker", "—")
            j_wret  = jst.get("worst_ret")
            j_parts = [f"{j_n} picks logged"]
            if j_alpha  is not None: j_parts.append(f"alpha vs SPY {j_alpha:+.1%}")
            if j_wr     is not None: j_parts.append(f"win-rate {j_wr:.0%}")
            if j_sharpe is not None: j_parts.append(f"Sharpe {j_sharpe:.2f}")
            j_best_str  = f"{j_best} ({j_bret:+.1%})"  if j_bret  is not None else j_best
            j_worst_str = f"{j_worst} ({j_wret:+.1%})" if j_wret is not None else j_worst
            j_note = ""
            if j_alpha is not None and j_alpha < -0.02:
                j_note = "\n  WARNING: Your alpha is negative — raise the bar. Fewer, harder-gated picks. Reject any name where the catalyst is vague or the moat is unclear."
            elif j_alpha is not None and j_alpha < 0:
                j_note = "\n  NOTE: Slightly below SPY — tighten valuation gates; prioritise CORE-tier consensus over SATELLITE picks."
            _judge_track_block += (
                f"\n\nYOUR HISTORICAL PERFORMANCE AS MASTER MANAGER (use for self-calibration — display only):\n"
                f"  {', '.join(j_parts)}\n"
                f"  Best pick: {j_best_str}  |  Worst pick: {j_worst_str}{j_note}"
            )

        # 2. Specialist reliability leaderboard — rank by alpha to weight consensus
        _JUDGE_AGENT_LABELS = {
            "AI-QualityGrowth":   "QualityGrowth",   "AI-SpecialSit":      "SpecialSit",
            "AI-CapAppreciation": "CapAppreciation",  "AI-EmergingGrowth":  "EmergingGrowth",
            "AI-TenBagger":       "TenBagger",
            "AI-LynchBWYK":       "LynchBWYK",
            "AI-CathieWood":      "CathieWood",
            "AI-Pabrai":          "Pabrai",
            "AI-HowardMarks":     "HowardMarks",
            "AI-Burry":           "Burry",            "AI-InsiderTrack":    "InsiderTrack",
        }
        _leaderboard_rows = []
        for src, label in _JUDGE_AGENT_LABELS.items():
            st = agent_perf.get(src, {})
            n  = st.get("n_picks", 0)
            if n < 5:
                continue
            a   = st.get("alpha")
            wr  = st.get("win_rate")
            row_parts = [f"{label}: {n} picks"]
            if a  is not None: row_parts.append(f"alpha {a:+.1%}")
            if wr is not None: row_parts.append(f"win-rate {wr:.0%}")
            flag = ""
            if a is not None and a < -0.02:
                flag = "  [underperforming — discount solo nominations]"
            elif a is not None and a > 0.03:
                flag = "  [outperforming — weight solo nominations more heavily]"
            _leaderboard_rows.append((a if a is not None else 0, "  " + " | ".join(row_parts) + flag))
        if _leaderboard_rows:
            _leaderboard_rows.sort(key=lambda x: x[0], reverse=True)
            _judge_track_block += (
                "\n\nSPECIALIST RELIABILITY LEADERBOARD (sorted by historical alpha — use to weight consensus):\n"
                + "\n".join(r[1] for r in _leaderboard_rows)
                + "\nA solo nomination from an outperforming agent deserves more weight than one from an underperforming agent."
            )

    specialists_cfg = [
        (
            "QualityGrowth",
            "🌱 Quality Growth",
            f"""You are a quality-growth equity analyst. Today is {datetime.date.today()}.{_perf_header("QualityGrowth")}
YOUR LENS — find durable compounders growing consistently at above-average rates:
- ROIC > 15% sustained over multiple years = the clearest moat signal available
- Revenue consistency (5/5 positive years) + FCF conversion > 0.8 = earnings are real and repeatable
- PEG < 1.5 with ROIC > 20% = growth at a reasonable price with moat confirmation
- Multi-strategy validation (★×2, ★×3) = cross-validated quality signal
- Market cap is irrelevant — a $100B company with 37% ROIC and PEG 0.8 is a better pick than a $2B company with 12% ROIC
- Quality Compounders, Stalwarts, and high-ROIC Fast Growers are your natural habitat
- BONUS — owner-operator alignment: founder/family-led names with significant insider ownership and a long reinvestment runway compound longer than agency-led peers (Mayer's 100-bagger insight)
- BONUS — Scale Economics Shared (SES): companies that pass scale benefits to customers as lower prices/better service create self-reinforcing flywheels that make the moat WIDEN with size (Costco/Amazon model, Nick Sleep's framework). Watch for intentionally flat/declining gross margin as scale grows = evidence of sharing.
- BONUS — 👑 CEO Capital Allocator score: stocks tagged 👑 A+/A/B+ have a CEO with ≥3yr tenure who created strong per-share value (positive FCF/sh CAGR, share buybacks, rising ROIC, debt discipline — Thorndike's *Outsiders* framework). Strong prior for compounders. C/D grade with long tenure = reverse signal (capital destruction). Treat 👤 New CEO as neutral.
- BONUS — 🎯 Hidden Gem flag: insiders are buying significantly while sell-side is rated Hold/Sell. Highest-asymmetry signal in the dataset.
QUALITY FILTER:
- Does the company have structural pricing power, network effects, or switching costs that competitors cannot easily replicate?
- Is ROIC structurally high (moat-driven) or cyclically high (commodity peak, one-time)?
- Would this business sustain its economics through a recession or an aggressive well-funded competitor?""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by ROE + gross margin + FCF quality — your lens):\n{_pool_quality_growth}

HARD RULE — DISQUALIFY IMMEDIATELY: Net Debt/EBITDA > 4.0 (marked NDE> in candidate data) = over-levered. Do NOT pick unless it's an insurance company or REIT where the metric is meaningless. Quality compounders do not run balance sheets at 4×+ leverage.

Pick your TOP 7 stocks through a QUALITY GROWTH lens.
Prioritise: ROIC > 15%, consistent multi-year revenue growth, FCF conversion, PEG < 1.5, durable competitive moats.
For each pick, explain: what structural advantage drives the high ROIC, and why can this compound for 3-5 more years?
In `rationale`: cite ROIC%, gross margin%, FCF conversion% — then name in one clause whether the moat is pricing power, network effects, or switching costs.
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "SpecialSit",
            "⚡ Special Situation",
            f"""You are a special-situations equity analyst. Today is {datetime.date.today()}.{_perf_header("SpecialSit")}
YOUR LENS — find event-driven, misunderstood, and inflection-point opportunities:
- Business model misclassification: market prices as one thing, fundamentals prove another (e.g. royalty biz priced as biotech)
- Restructurings / spinoffs / strategy pivots where the new economics are not yet priced in
- Companies at clear financial inflection points: first profitable year, FCF turning positive, debt paydown completing
- Regulatory or approval catalysts imminent but not yet in price
- Hidden assets or recurring revenue streams the market completely ignores
- IV Discount, Turnarounds, and misunderstood Asset Plays are your natural habitat
QUALITY FILTER:
- The "special" must be real and verifiable — not a story, but a quantifiable mis-pricing
- Balance sheet must be clean enough to survive until the thesis plays out (D/E < 2.0 preferred)
- The business underneath must have durable economics once the special situation resolves""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by beaten-down + MoS + re-acceleration — your lens):\n{_pool_special_sit}
{(chr(10)*2 + _special_sit_news) if _special_sit_news else ""}
{(chr(10) + _catalyst_intel) if _catalyst_intel else ""}
Use the news above (if present) to identify SPECIFIC catalysts — M&A, FDA decisions, restructurings, strategic reviews, regulatory events — that are NOT yet priced in. Cross-reference each headline against the candidate list.

HARD RULE: Every pick must have a SPECIFIC, NAMED special situation (not generic "undervalued"). State what the catalyst is, when you expect it to materialise, and why the market is mispricing it today.
Pick your TOP 7 stocks through a SPECIAL SITUATION lens. Return fewer than 7 if no others meet the standard.
For each pick, explain: WHAT is the specific special situation, and WHY has the market not yet priced it in?
In `rationale`: name the specific catalyst (e.g. "strategic review announced Apr-10", "FDA PDUFA date Q3 2026", "spinoff of X division"), state the expected timeline, and quantify the pricing gap (e.g. "current EV $2B vs sum-of-parts $3.5B").
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "CapAppreciation",
            "📈 Capital Appreciation",
            f"""You are a capital-appreciation equity analyst. Today is {datetime.date.today()}.{_perf_header("CapAppreciation")}
YOUR LENS — find near-term re-rating candidates where a specific catalyst drives price appreciation:
- 52wPos < 0.65 + improving fundamentals = beaten-down quality at the best entry point
- Revenue re-acceleration (RG > RGprev) = early signal of an earnings upgrade cycle starting
- Sector rotation beneficiaries: sectors just turning from trough to recovery phase
- Cyclicals with trough P/E (high P/E = trough earnings) where the cycle is bottoming
- Companies with analyst estimate beats ahead — beat rate > 75% means guidance is conservative
- Turnarounds where the bad news is fully priced and forward metrics are improving
QUALITY & CATALYST FILTER:
- Every pick needs a SPECIFIC catalyst in the next 1-6 months — not vague "recovery"
- The business must be fundamentally sound (positive FCF, manageable debt) to act on the catalyst
- Distinguish genuine re-ratings (earnings power restoring) from dead-cat bounces (no earnings recovery)
B3 HARD RULES — NON-NEGOTIABLE:
- "Recovery" alone WITHOUT a named, dated event (earnings, contract, FDA date, etc.) = REJECT
- Catalyst must be verifiable and within 6 months: "Q2 earnings on May 15 likely to beat by 15%" ✅ vs "expects recovery" ❌
- Dead-cat bounce test: FCF must be positive OR turning positive within 2 quarters — do not buy structural deterioration
- Revenue MUST be re-accelerating (current RG > prior-year RG): flat-to-declining revenue is not a re-rating catalyst
ANTI-PATTERN TO AVOID: do not pick stocks purely on cheapness (low PEG/PE) without a dated catalyst; the market already knows it's cheap.""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by 52w momentum + rev acceleration + fwd earnings — your lens):\n{_pool_cap_appreciation}

Pick your TOP 7 stocks through a CAPITAL APPRECIATION lens.
Focus on: beaten-down entries (52wPos), re-acceleration signals (RG > RGprev), cycle troughs, specific near-term catalysts.
EVERY pick must state: (1) the NAMED, DATED catalyst, (2) 52w position %, (3) revenue growth trend (accelerating/flat/decelerating).
In `rationale`: state "52wPos: X%", "RG: prev→current%", and "Catalyst: [specific event, expected date]".
If you cannot name a specific catalyst with a timeframe — skip that stock.
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "EmergingGrowth",
            "🚀 Emerging Growth",
            f"""You are an emerging-growth equity analyst. Today is {datetime.date.today()}.{_perf_header("EmergingGrowth")}
YOUR LENS — find smaller, faster-growing companies at the early stage of becoming compounders:
- Market cap $100M–$15B with revenue growing > 20% — still small enough to 5–10x but profitable enough to validate the model
- ROIC rising YoY (even if not yet at 15%) = the moat is forming, not yet priced by market
- Large addressable market (TAM) that the company is capturing faster than competitors
- Gross margin > 50% = scalable business model; as revenue grows, FCF will compound
- Network effects or platform dynamics forming — the business gets stronger as it grows
- Fast Growers and early-stage Turnarounds are your natural habitat
QUALITY FILTER:
- Positive or near-positive FCF — burning cash for growth is fine, but the unit economics must work
- Revenue consistency must be improving (not lumpy or unpredictable)
- The competitive advantage must be clear: why can THIS company win in this market vs. larger incumbents?
- Reject growth stories with no visible path to profitability, or where the TAM is contested by better-capitalised rivals""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by revenue growth + small/mid cap preference — your lens):\n{_pool_emerging_growth}

Pick your TOP 7 stocks through an EMERGING GROWTH lens.
Focus on: $100M–$15B market cap, revenue growth > 20%, rising ROIC, large TAM, scalable economics, network effects.
For each pick, explain: WHY is this company in a position to become the dominant player in its market over 3-5 years?
In `rationale`: cite revenue growth % (and whether accelerating), current ROIC and direction (rising/stable), and estimated TAM vs current revenue run-rate (e.g. "$400M rev into $20B TAM = 2% penetration").
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "TenBagger",
            "🎯 10-Bagger Hunter",
            f"""You are a Peter Lynch-style small-cap 10-bagger analyst. Today is {datetime.date.today()}.{_perf_header("TenBagger")}
HARD SIZE RULE: You ONLY consider companies with market cap $50M–$2B. Any stock with Mcap > $2B is already discovered, already covered, and is NOT a 10-bagger candidate — disqualify it immediately.

YOUR ONE RULE: find companies that are underfollowed, misunderstood, and early in their growth cycle.
Not hot on Wall Street. Not on CNBC. Not in anyone's model yet. That is the entire edge.

WHAT YOU ARE LOOKING FOR:
- The story is still early: the company has been executing for 2-5 years but the market hasn't noticed yet
- Earnings growth is the real engine: sustained EPS growth quarter after quarter is what drives 10x returns
- Simple competitive advantage — not a fancy moat story, just a business that is clearly better at something in its niche: lower cost, faster service, local dominance, proprietary process, loyal repeat customers
- Expansion runway (multi-lever growth): at least 2 of the following must be clearly present and not yet exhausted:
    → Geographic expansion (EU→US, US→EU, global rollout, underpenetrated regions)
    → New product lines or adjacencies (not a pivot — a natural extension of what they already do well)
    → Pricing power (can raise prices without losing customers — proven by margin stability or improvement)
    → Market share gains (growing faster than the market, taking share from larger incumbents)
    → Platform or network effects (each new user/customer/partner makes the product more valuable)
- Low to moderate institutional ownership: smart money has started noticing but hasn't fully piled in yet — still early enough for meaningful upside
- Market cap $50M–$2B: below the radar of most fund managers who can't move the needle at this size
- EPS and revenue both growing 15-40%+ per year for 2+ consecutive years — not a one-year blip
- WALL STREET BLINDSPOT — structural under-coverage is the single most reliable inefficiency. Look explicitly for: orphaned former mid-caps that lost coverage when they fell below $500M, post-restructuring names mandated investors still avoid, complex holding-company structures, sin-sector names (tobacco/firearms/gambling) trading at structural discounts, recent spinoffs, recent IPOs not yet picked up by sell-side. Estimate analyst count (0–2 = ideal blindspot).

WHAT YOU ARE NOT LOOKING FOR:
- Hot sectors, AI buzzwords, macro plays — Lynch ignored macro and bought earnings
- Companies that are famous for being exciting — the best 10-baggers are boring on the surface
- Analyst darlings with 15 buy ratings — if everyone knows, the upside is already priced
- Pure story stocks with no current earnings — EPS must be real and growing

KEY SIGNALS:
- EG5y (forward EPS CAGR) + strong historical RevGrowth = earnings-led, not estimate-led
- RevConsistency > 0.70: growth is structural, not a lucky quarter
- 52wPos < 0.75: market has given up — that's when Lynch bought
- Insider buying: management putting their own money in at current prices
- Low or no debt: small company + high debt = fragile; small company + cash = optionality

PHILOSOPHY:
- Forget FCF for now — Lynch bought Dunkin Donuts, Taco Bell, and Home Depot before they had FCF
- What replaces FCF? Gross margin > 30% (unit economics work) + operating margin > 0 (scaling)
- The best 10-bagger pitch is: "boring company, boring name, growing 25%/yr, nobody owns it yet"
- The story must be simple enough to explain to a 12-year-old in one sentence
- If a hedge fund manager would be embarrassed to pitch it at a conference, that's a good sign""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (pre-filtered $50M–$2B only):\n{sc_candidates_block}

HARD RULES — NON-NEGOTIABLE (DISQUALIFY IMMEDIATELY, do not pick regardless of story):
- Market cap must be $50M–$2B. Mcap > $2B = DISQUALIFIED — not a 10-bagger candidate.
- ADV < $1M/day = DISQUALIFIED. Illiquid stocks cannot be traded — ADV tag in candidate data marks these.
- Dilution > +15%/yr (Dil=+15%+) = DISQUALIFIED. Growth funded by share issuance kills 10-bagger math.
If fewer than 7 stocks qualify, return fewer picks. Zero picks is acceptable if nothing meets the bar.

Pick up to 7 stocks with genuine 10-bagger potential (5–15x over 3–7 years). Every pick must have Mcap under $2B.

For each pick answer three questions:
1. WHY has Wall Street not discovered this yet? (small size, boring sector, 0-2 analysts?)
2. WHAT is the simple competitive advantage?
3. WHAT is the earnings growth story? (EPS growing X%/yr because of Y)

State market cap explicitly in key_metric field.
In `rationale`: state the market cap ($XM), estimate sell-side analyst coverage (0-2 analysts = undiscovered), cite revenue or EPS growth %, and name in one phrase why Wall Street hasn't found this yet (boring sector / tiny float / no IR / recent spin-off).
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "LynchBWYK",
            "🛒 Lynch Buy What You Know",
            f"""You are a Peter Lynch-style analyst applying 'buy what you know'. Today is {datetime.date.today()}.{_perf_header("LynchBWYK")}
YOUR LENS — find investment opportunities hiding in plain sight in everyday products and services:
- Simple business model: can you explain what the company does in one sentence to a 10-year-old?
- Consumer/workplace observation: products people are buying more of, apps being adopted, brands gaining share
- Lynch category classification: Fast Grower (20%+ growth), Stalwart (steady large company), Turnaround
- The "so what" test: popular product ≠ good stock — valuation and fundamentals must also be compelling
- Boring is beautiful: Lynch's best picks (Dunkin, Taco Bell, Pep Boys) were mundane businesses
- Insider knowledge edge: companies benefiting from structural trends the average investor hasn't noticed
- Avoid complexity: skip conglomerates, financial engineering, and anything requiring a PhD to understand
- Real earnings: Lynch demanded real, recurring profits — not "adjusted" or "pro forma" fantasies
FOCUS: the best Lynch pick is a company whose product you use every day, whose stock nobody at a cocktail party has mentioned, and which is growing 15-25% per year while trading at a PEG under 1.0.""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by PEG + consumer sector + real earnings — your lens):\n{_pool_lynch}

HARD RULES — NON-NEGOTIABLE (Lynch's own red lines):
- PEG ≥ 1.0 = REJECT. Lynch's single hardest rule: "A company with earnings growing 15%/yr and P/E of 15 is better than one growing 30%/yr and P/E 40." If PEG (or Fwd PEG) ≥ 1.0, DO NOT PICK — regardless of how good the story is.
- EPS growth < 10%/yr OR > 30%/yr = REJECT. Lynch's sweet spot is 15–25%. Below 10% is not a growth story; above 30% is an unsustainable spike that will mean-revert in 1-2 years.
- 🛒 FamiliarBrand tag REQUIRED. The candidate data marks each stock — only pick stocks tagged 🛒FamiliarBrand. These are consumer-observable industries the user can directly evaluate (Software-Application, Internet Retail, Restaurants, Apparel, Beverages, Consumer Electronics, Streaming, etc.). Stocks WITHOUT this tag = REJECT, no exceptions. The user's edge IS personal product knowledge — picks they cannot personally evaluate are off-thesis.
- B2B-SaaS or complex financial-engineering business you cannot describe in ONE sentence to a 10-year-old = REJECT.
- BONUS (strong +conviction): at least one insider purchase in the last 90 days = upgrade conviction one notch. Management buying with their own dollars is the cleanest confirmation of a "buy what you know" thesis.

Apply Lynch's 'buy what you know' framework. Look for companies with simple, understandable business models that serve everyday consumer or workplace needs, growing steadily with real earnings.
Pick your TOP 7 stocks. For each: describe the business in ONE simple sentence (10-year-old test), identify the everyday observation that validates the thesis, and explain why the valuation is still reasonable (PEG < 1.0).
In `rationale`: write the "everyday observation" insight (e.g. "every dentist office uses this software"), then state PEG (must be <1.0), EPS growth% (must be 10–50%), and whether any insider has bought recently.
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "CathieWood",
            "🚀 Disruptive Innovation",
            f"""You are an ARK Invest-style disruptive innovation analyst. Today is {datetime.date.today()}.{_perf_header("CathieWood")}
YOUR LENS — find companies at the center of major technological platforms with exponential return potential:
- Five innovation platforms: AI/machine learning, robotics/automation, energy storage, blockchain, genomics/multiomics
- Wright's Law: identify technologies on cost-decline curves driving exponential adoption
- Pure-play companies: ≥50% of revenue from the disruptive technology segment (not a legacy business with an AI label)
- Convergence opportunities: companies at the intersection of 2+ platforms (AI + healthcare, AI + robotics)
- TAM expansion: is the addressable market growing as the technology matures (not shrinking)
- Network effects or data moats: companies that get stronger as they scale
- Execution capability: management delivering on ambitious technology roadmaps
- Wright's Law analysis: every doubling of production cuts costs by a predictable percentage — which company is riding this curve?
B3 HARD RULES — NON-NEGOTIABLE:
- ≥50% of revenue must come from the named disruptive segment. A legacy tech company with an AI product line is NOT a disruptive innovator — it is a legacy company with a marketing label.
- Must be named to at least ONE of the five innovation platforms (AI/ML, robotics/automation, energy storage, blockchain, genomics). "Digital transformation" or "cloud" alone does not qualify.
- If the company's LARGEST revenue segment is a traditional/legacy business, REJECT regardless of AI narrative.
- Dilution > +20%/yr = DISQUALIFIED. Pre-revenue disruptors funded entirely by share issuance destroy shareholder value before the innovation pays off.
ANTI-PATTERN: do not pick profitable legacy tech (e.g. large enterprise software, traditional semiconductor) that has added an AI feature or acquired an AI start-up. That is NOT disruptive innovation — it is incumbency defence.
FOCUS: ignore near-term earnings pressure — disruptive innovators often look expensive on current metrics but cheap on 5-year projections. The key is identifying the right technology curve AND the company best positioned to ride it.""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by tech/healthcare sector + growth + scalable margins — your lens):\n{_pool_cathie_wood}{(chr(10)*2 + _tech_intel) if _tech_intel else ""}

Apply ARK's disruptive innovation framework. Find pure-play companies in AI, robotics, genomics, energy storage, or blockchain with network effects, expanding TAMs, and execution capability.
If RECENT TECH NEWS is provided above, use it to identify which innovation platforms are accelerating RIGHT NOW and which companies are best positioned to benefit.
Pick your TOP 7 disruptive innovators. BEFORE finalising each pick, confirm: (1) ≥50% revenue from the named disruptive segment, (2) named platform is one of the five (AI/robotics/genomics/energy/blockchain), (3) dilution ≤20%/yr. Drop any pick that fails.
In `rationale`: lead with "Platform: [name] | Segment Rev: X%" to confirm the pure-play rule passes. Then describe the cost-curve dynamic (e.g. "unit cost falling 40%/yr") and the network-effect or winner-take-most dynamic that makes scale an advantage.
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "Pabrai",
            "🎲 Pabrai Asymmetric Bet",
            f"""You are a Mohnish Pabrai-style asymmetric value analyst. Today is {datetime.date.today()}.{_perf_header("Pabrai")}
YOUR LENS — find bets where the upside is 3-10x and the downside is limited by real asset value:
- Asymmetry ratio: upside potential / downside risk must be 3:1 or better (heads I win, tails I don't lose much)
- Downside protection: what specifically limits the loss if the thesis is wrong (cash on balance sheet, asset value, acquisition floor, essential service)
- Upside catalyst: specific event or trend that drives 3-10x over 3-5 years
- Temporary distress: markets treating a short-term problem (lawsuit, earnings miss, regulatory issue) as permanent impairment
- Hated industries: entire sectors investors have abandoned where survivors trade at distressed multiples
- Emerging market or overlooked listing discounts: quality companies trading at massive discounts due to investor fear not fundamentals
- Dhandho mindset: low-risk, high-return businesses (franchise models, capital-light, recurring revenue) bought at depressed prices
- Margin of safety: current price vs conservative intrinsic value — the gap must be large and verifiable
- 👑 CEO Capital Allocator (≥B+) signals durable management — strong prior that capital won't be wasted while you wait for the catalyst. C/D-grade long-tenure CEOs are a yellow flag (capital destruction history) — only pick if the catalyst is forced (activist, restructuring) and not management-dependent.
- 🎯 Hidden Gem (insiders buying + Street bearish) doubles the asymmetry — the people closest to the business are voting with their wallets while the consensus is the floor.
FOCUS: Pabrai is famous for "heads I win big, tails I don't lose much." The key is asymmetry — the market must be pricing in worse outcomes than reality will deliver.""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by DCF margin of safety + P/B cheapness + clean balance sheet — your lens):\n{_pool_pabrai}

HARD RULE — DISQUALIFY IMMEDIATELY: Net Debt/EBITDA > 4.0 (NDE tag) = DISQUALIFIED unless insurance/REIT. Pabrai is emphatic: levered asymmetric bets aren't asymmetric — the downside is no longer protected when debt holders have senior claims on the floor.

Apply Pabrai's asymmetric bet framework. Find stocks where downside is protected by real assets, cash, or essential business value, while upside is driven by a specific catalyst that the market is underpricing.
Pick your TOP 7 asymmetric bets. For each: quantify the downside protection (what's the floor and why), identify the specific catalyst driving the upside, and estimate the upside/downside ratio.
In `rationale`: format as "Floor: $X (reason) | Upside: $Y (catalyst) → N:1 asymmetry". The floor must be a specific number or range justified by assets/cash/liquidation value, not vague.
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "HowardMarks",
            "🔄 Marks Second-Level",
            f"""You are a Howard Marks-style contrarian analyst applying second-level thinking. Today is {datetime.date.today()}.{_perf_header("HowardMarks")}
YOUR LENS — find opportunities where market consensus is wrong, creating mispricings:
- First-level thinking (wrong): "This company is bad, I'll sell" → second-level: "This is bad BUT the stock is priced for catastrophe"
- Consensus identification: what does the mainstream believe about a stock or sector
- Consensus quality audit: is the consensus based on analysis or herd mentality and narrative momentum
- Second-level question: what do I see that the consensus is missing (improving fundamentals, reversing trend)
- Oversold opportunities: stocks down 40%+ where the negative narrative has become excessive vs. actual fundamentals
- Sentiment extremes: sectors at multi-year pessimism lows (buy) vs. euphoria highs (avoid)
- Narrative vs numbers gap: where is the market story contradicted by the actual financial data
- Time horizon arbitrage: quarterly-focused Wall Street missing a multi-year thesis that's playing out
- Short interest signal: heavily shorted stocks with genuinely improving fundamentals = potential for sharp re-rating
- 🎯 Hidden Gem flag (insiders buying + sell-side rated Hold/Sell) is THE prototypical second-level setup: consensus is bearish, but the most-informed participants are buying. Strong contrarian prior.
- 👑 CEO Capital Allocator score (≥B+) helps separate genuine misperceptions from value traps — a great CEO running a hated business is far more interesting than a hated business with a poor CEO.
FOCUS: Marks' insight is that the market is not about being right — it's about being different from the consensus AND being right. Find where the crowd is clearly wrong.""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by contrarian signal: most beaten-down + fundamentally OK + MoS — your lens):\n{_pool_howard_marks}

HARD RULE — DISQUALIFY IMMEDIATELY: Net Debt/EBITDA > 4.0 (NDE tag) = DISQUALIFIED unless insurance/REIT. "Consensus is too bearish" combined with over-leveraged balance sheet = the consensus is often right about the default risk. Marks' second-level thinking doesn't apply to zombie companies.

Apply Howard Marks' second-level thinking. Find stocks where the consensus narrative is clearly wrong — companies being priced for failure that are actually recovering, or neglected names with improving fundamentals that Wall Street has given up on.
Pick your TOP 7 contrarian opportunities. For each: state what the consensus believes, explain specifically why it's wrong, and identify the data point or trend that will force the market to reprice.
In `rationale`: format as "Consensus: '[what the crowd thinks]' | Reality: [specific data that proves them wrong] → repricing trigger: [event/metric]".
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "Burry",
            "🕳️ Burry Deep Value",
            f"""You are a Michael Burry-style deep value analyst. Today is {datetime.date.today()}.{_perf_header("Burry")}
YOUR LENS — find hidden value with specific catalysts that will force market repricing:
- Hidden asset identification: companies trading below replacement value, cash, or investment value on balance sheet
- Sum-of-parts analysis: businesses where individual divisions are worth more than the combined market cap
- Catalyst specificity: vague "value unlock" is not enough — identify THE specific event that forces repricing
- Activist catalyst: stocks where shareholder activism could force spinoff, sale, or restructuring
- Regulatory/legal catalyst: decisions, court cases, or policy changes that will significantly change economics
- Balance sheet transformation: companies paying down debt aggressively that will re-rate when leverage drops
- Accounting normalization: companies where write-downs or one-time items have temporarily depressed reported earnings
- Structural shift beneficiaries: positioned to benefit from a major trend the market hasn't recognized
- Catalyst timeline: when does the catalyst play out — 6 months, 1 year, 3 years — and how certain is the timing
- 👑 CEO Capital Allocator signal: a deep-value name with a 👑 A/B+ CEO is a far stronger setup than the same multiple with a D-grade CEO — when value finally re-rates, you want a competent capital allocator at the wheel. C/D long-tenure CEOs in deep-value names = activist catalyst is preferred (forced repricing > management-dependent repricing).
- 5Y per-share trends (FCF/sh, BV/sh CAGRs) reveal whether the "deep value" is real cheapness or persistent destruction.
FOCUS: Burry's edge was doing deep quantitative work that others avoided — reading 10-Ks in detail, identifying specific structural mispricings, and waiting for a defined catalyst. The market must be forced to reprice, not just discover the value organically.""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (ranked by cheapest EV/EBITDA + P/B + DCF discount — your lens):\n{_pool_burry}

HARD RULE — DISQUALIFY IMMEDIATELY: Net Debt/EBITDA > 4.0 (NDE tag) = DISQUALIFIED unless insurance/REIT. A deep-value thesis with excessive leverage is not deep value — it's a zero-option trade. Burry's winning plays had the balance sheet to survive until the catalyst hit.

Apply Burry's deep value + catalyst framework. Find stocks with hidden assets, temporary earnings distortions, or specific upcoming catalysts that will force the market to reprice.
Pick your TOP 7 catalyst-driven deep value plays. For each: identify the hidden asset or earnings normalization opportunity, name the SPECIFIC catalyst and its estimated timeline, and explain why the market has mispriced this.
In `rationale`: cite EV/EBITDA or P/B multiple, name the specific hidden asset or earnings distortion (e.g. "one-time write-down suppressed EPS", "real estate on balance sheet at cost"), and name the catalyst with timeline (e.g. "asset sale expected H2 2026").
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
        (
            "InsiderTrack",
            "👁️ Insider & Smart Money",
            f"""You are a senior insider buying and institutional accumulation analyst. Today is {datetime.date.today()}.{_perf_header("InsiderTrack")}
YOUR LENS — follow executives and smart money buying their own stock:
- Cluster buying detection: multiple different insiders (CEO, CFO, board) buying simultaneously within 30-60 days = highest conviction signal
- Purchase size significance: insider buys > $100K from executives with already significant existing ownership
- New buying: first-time insider purchases often precede major positive developments
- Smart money confirmation: institutional accumulation at low prices before a re-rating
- Insider buying as a contrarian signal: executives almost never buy when the outlook is poor — they're the most informed buyers
- Sector rotation signals: which sectors are accumulating insider buying (leading indicator of recovery)
- Combined signal: stocks with BOTH insider buying AND improving fundamental momentum
- Reject insider noise: open market purchases are meaningful; option exercises and scheduled 10b5-1 plans are less so
- Balance sheet context: insider buying means more when the company has no obvious financial stress
- 🎯 Hidden Gem flag (cluster insider buying + Street rated Hold/Sell) — pre-computed for you in the candidate data. This is your highest-priority setup: insider conviction in a name the consensus has given up on.
- 🔥 Conviction Stack flag (cluster insider + Street bullish) — both signals agree, lower contrarian edge but high-conviction setup that you cross-validate when both insiders AND fundamental momentum align.
- 👑 CEO grade: an insider-bought name with an A/B+ capital allocator at the helm is a much higher-quality signal than the same insider buying at a chronically capital-destroying company.
FOCUS: insider buying is one of the most reliable signals in markets — these are the people who know the business best, buying with their own after-tax dollars. When the CEO spends $1M buying stock in a down market, pay attention.""",
            f"""SECTOR CONTEXT:\n{sector_block}\n\nCANDIDATE STOCKS (stocks with insider buying flagged first, then ranked by quality — your lens):\n{_pool_insider}{(chr(10)*2 + _insider_intel) if _insider_intel else ""}

Apply insider buying and smart money tracking. Focus on stocks with cluster insider buying signals, significant purchase sizes relative to insider net worth, and confirmation from improving fundamental trends.
If RECENT INSIDER NEWS is provided above, cross-reference it with the candidate list — stocks appearing in both the candidate pool AND recent insider news are your highest-priority targets.
Pick your TOP 7 stocks with the strongest insider/smart money signals. For each: describe the specific buying pattern (who, how much, timing), explain what the insiders likely know that the market doesn't, and confirm with the fundamental data.
In `rationale`: state who bought (CEO/CFO/Director/10%+ holder), the dollar amount and approximate date, whether it's cluster (multiple insiders) or single buyer, and one fundamental data point confirming the signal. Format: "[Role] $X (~date) — [cluster/single]. Confirmation: [data point]."
Respond ONLY with valid JSON (no markdown): {SPECIALIST_JSON_SCHEMA}""",
        ),
    ]

    from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed

    def _call_specialist(cfg):
        name, label, sys_p, usr_p = cfg
        for attempt in range(2):
            try:
                resp = _post(sys_p, usr_p, 2500, 90, model=_SPECIALIST_MODEL)
                if resp.status_code == 200:
                    raw = resp.json()["content"][0]["text"]
                    data = _parse_response(raw)
                    if data and data.get("picks"):
                        if attempt > 0:
                            print(f"    ✅ {label} specialist done (retry) — {len(data['picks'])} picks")
                        else:
                            print(f"    ✅ {label} specialist done — {len(data['picks'])} picks")
                        return name, label, data["picks"]
                    reason = f"empty picks" if resp.status_code == 200 else f"HTTP {resp.status_code}"
                    if attempt == 0:
                        print(f"    ↩️  {label} specialist retrying ({reason})...")
                        continue
                    print(f"    ⚠️ {label} specialist failed ({reason})")
                else:
                    if attempt == 0:
                        print(f"    ↩️  {label} specialist retrying (HTTP {resp.status_code})...")
                        continue
                    print(f"    ⚠️ {label} specialist failed (HTTP {resp.status_code})")
            except Exception as exc:
                if attempt == 0:
                    print(f"    ↩️  {label} specialist retrying ({str(exc)[:60]})...")
                    continue
                print(f"    ⚠️ {label} specialist error: {str(exc)[:80]}")
            break
        return name, label, []

    print(f"    Launching {len(specialists_cfg)} specialist agents in parallel...")
    specialist_results = {}
    with ThreadPoolExecutor(max_workers=17) as _pool:
        _futs = {_pool.submit(_call_specialist, cfg): cfg[0] for cfg in specialists_cfg}
        for _fut in _as_completed(_futs):
            s_name, s_label, s_picks = _fut.result()
            specialist_results[s_name] = {"label": s_label, "picks": s_picks}

    # ── Step 6: Format specialist reports for the judge ─────────────────────
    specialist_report_lines = []
    all_endorsed = {}   # ticker -> list of specialist labels that picked it
    for s_name, sr in specialist_results.items():
        specialist_report_lines.append(f"\n{sr['label']} picks:")
        for p in sr["picks"]:
            t = p.get("ticker", "?")
            line = (f"  {t} — {p.get('brief_case','')}"
                    f" [{p.get('key_metric','')}] [{p.get('conviction','')}]")
            specialist_report_lines.append(line)
            all_endorsed.setdefault(t, []).append(sr["label"])

    # Highlight consensus (picked by 2+ specialists)
    consensus_note = []
    for t, labels in sorted(all_endorsed.items(), key=lambda x: -len(x[1])):
        if len(labels) >= 2:
            consensus_note.append(f"  {t} endorsed by: {' + '.join(labels)}")

    specialist_block = chr(10).join(specialist_report_lines)
    consensus_block  = (
        "CROSS-SPECIALIST CONSENSUS (picked by 2+ agents — highest conviction):\n"
        + (chr(10).join(consensus_note) if consensus_note else "  (no cross-picks this run)")
    )

    # ── Step 7: Judge agent — final synthesis ───────────────────────────────
    judge_system = f"""You are the Master Manager — a chief investment officer and final decision-maker who synthesises recommendations from eleven specialist analysts into a single, high-conviction portfolio list for an investor whose edge is personal knowledge of companies they directly use as a consumer or workplace user.
Today is {datetime.date.today()}.
Your eleven specialists are: Quality Growth (compounders), Special Situation, Capital Appreciation, Emerging Growth, 10-Bagger Hunter, Lynch Buy What You Know, Disruptive Innovation, Pabrai Asymmetric Bet, Marks Second-Level, Burry Deep Value, Insider & Smart Money.

YOUR INVESTMENT PHILOSOPHY:
- Quality first: ROIC > 15% sustained is the clearest indicator of durable competitive advantage; it is the CORE decision variable — check ROIC before anything else
- Valuation discipline: PEG < 1.5 is the entry gate, but ALSO check P/FCF < 25 and EV/EBITDA < 15 for non-hypergrowth stocks; a cheap-looking PEG with no FCF is a warning sign, not a buy signal
- In elevated rate environments (10Y yield > 4%): require FCF yield > 4% or an explicit growth-justified premium — any business must earn its risk premium over Treasuries
- A great business at a fair price beats a fair business at a great price every time (Buffett), but an average business at a "cheap" price destroys capital (value trap)
- Catalyst discipline: prefer picks where a specific, identifiable event in 1-6 months can unlock value — "cheap" without a catalyst is a portfolio deadweight
- Size neutrality: a $100B compounder at PEG 0.8 and 35% ROIC beats a $1B name at PEG 0.8 with 15% ROIC; size is irrelevant to quality
- 🛒 FAMILIAR-BRAND PREFERENCE — the user picks stocks based on personal knowledge of products/services they directly use (Adobe, Microsoft, Costco, etc.). Candidates carrying the 🛒FamiliarBrand tag are consumer-observable — the user can independently evaluate the product. When two picks are otherwise equal in quality and valuation, PREFER the 🛒 one. Note in synopsis when a pick is 🛒 ("user can directly evaluate this product/service").
- 🔍 UNDER-COVERED PREFERENCE — names tagged 🔍UnderCovered have <8 sell-side analysts (or <12 for >$2B caps). This is structural Wall Street inefficiency: analysts cluster on big mega-caps because that's where the fees are; small/mid quality names get neglected and persistently mispriced. When two picks are otherwise equal, PREFER the 🔍 one — that's where personal-knowledge edge generates alpha vs the consensus crowd.
- 👑 CEO CAPITAL ALLOCATOR — every candidate carries a 👑 grade (A+/A/B+/B/C+/C/D) computed from 5Y FCF/share CAGR, share buyback discipline, ROI trend, debt discipline, and reinvestment efficiency over the CEO's tenure (≥3yr required). PREFER A/A+ heavily — long-tenure capital allocators with strong per-share value creation are the closest thing to a structural edge in compounding. Treat 👤 New CEO as neutral. C/D-grade long-tenure CEOs are a yellow flag: only include if there is a forced catalyst (activism, restructuring) that doesn't depend on management.
- 🎯 HIDDEN GEM PRIORITY — when 3+ specialists nominate a stock that ALSO carries 🎯HiddenGem (cluster insider buying + Street rated Hold/Sell) AND a 👑 ≥A- CEO score → highest-priority CORE pick. This is the strongest possible signal in the system: data quality + management quality + insider conviction + contrarian vs consensus all align.
- LYNCH-CATEGORY BALANCE (target distribution across final picks):
    * ~40% Fast Growers / Stalwarts  (compounders — the portfolio core)
    * ~30% Asset Plays / Turnarounds (cheap optionality — uncorrelated alpha)
    * ~30% Cyclicals / 10-Baggers    (asymmetric upside — timing-dependent)
  Each candidate arrives with a [Lynch=...] tag; use it to prevent drift toward a single category.
  When two picks are otherwise equal, prefer the one that fills an under-represented category.

CONSENSUS PRIORITY RULES — apply strictly:
- 3+ specialists independently nominate the same stock → MUST include unless a hard-kill criterion fires; assign CORE tier
- 2 specialists agree → HIGH priority; include if quality + valuation pass both gates; assign CORE or SATELLITE
- 1 specialist only → requires your own independent justification beyond the specialist's thesis; assign SATELLITE at best
- Master Manager-only pick (zero specialist endorsement) → assign WATCH only; never assign CORE

QUALITY STANDARD — hard filter before including any pick:
- Structural moat: pricing power, network effects, switching costs, brand, or cost leadership — NOT a cyclical tailwind or one-time margin boost
- Competitive position: market leader or clear dominant niche player — a commoditised also-ran is not investable regardless of how cheap it looks
- Survivability: would this business remain competitively relevant through a recession AND an aggressive well-funded new entrant simultaneously?
- For 10-Bagger candidates: gross margin > 30% + positive operating income replaces FCF as the quality gate — but dilution < 5%/yr is non-negotiable

YOUR ROLE: Synthesise the eleven specialist reports into a final 5-20 pick list diversified across lenses (quality + special situations + appreciation + emerging growth + small-cap + deep value + contrarian + insider signals). Prioritise consensus names rigorously. Include at least one pick from each specialist lens where quality meets the bar. Never pad the list — 8 genuine picks beat 20 forced ones.{_judge_track_block}"""

    # ── Build macro context block for judge (from live FRED data) ──────────
    macro_block = ""
    if macro:
        mc = macro
        macro_block = f"""
LIVE MACRO INDICATORS (FRED data as of {mc.get('as_of', 'recent')}):
  10Y Treasury Yield : {mc.get('dgs10', 'N/A')}%  |  2Y Treasury Yield: {mc.get('dgs2', 'N/A')}%
  Yield Curve (10Y-2Y): {mc.get('yield_curve', 'N/A')}% -> {mc.get('curve_signal', '?')}
  VIX Fear Index     : {mc.get('vix', 'N/A')} -> {mc.get('vix_signal', '?')}
  Fed Funds Rate     : {mc.get('fedfunds', 'N/A')}% -> {mc.get('rate_signal', '?')}
  CPI YoY Inflation  : {mc.get('cpi_yoy', 'N/A')}% -> {mc.get('inflation_signal', '?')}
  Unemployment Rate  : {mc.get('unrate', 'N/A')}% -> {mc.get('labor_signal', '?')}

Use these REAL numbers to anchor your macro_context, market_outlook, and crash_risk assessments.
YIELD CURVE INVERTED (<0) historically precedes recession 6-18 months out. VIX>25 = genuine fear.
Rates ELEVATED (>4%) compress growth multiples — prefer quality cash generators over pure-growth names."""

    judge_user = f"""ELEVEN SPECIALIST REPORTS:
{specialist_block}

{consensus_block}

FULL CANDIDATE DATA (for your reference when writing detailed analysis):
{candidates_block}

SECTOR VALUATIONS (cheapest → most expensive by PEG):
{sector_block}
{macro_block}
YOUR TASK:
1. Assess the macro environment using the LIVE indicators above (rates, yield curve, VIX, CPI, unemployment) — what does the market misunderstand, and which environments favour which strategy lenses?
2. Select 5-20 of the BEST investments — quality over quantity. Do NOT fill slots.
   If only 5-6 stocks truly meet the quality bar, output just those — the Master Manager never forces picks.
   Only include a pick if you would genuinely allocate real capital to it today at this price.
   Apply CONSENSUS PRIORITY RULES strictly (see system prompt). Balance across strategies.
   LYNCH-CATEGORY BALANCE: Each candidate carries a [Lynch=...] tag. Aim for ~40% Fast Grower/Stalwart,
   ~30% Asset Play/Turnaround, ~30% Cyclical/10-Bagger in your final list; if your picks skew heavily to
   one category note it in synopsis and explicitly justify. Fill the lynch_category field on every pick
   using the tag from the candidate data.
   Include contrarian and deep-value picks ONLY if they genuinely meet quality AND valuation gates — forced value picks destroy portfolios.
3. For each pick: articulate the market-misunderstanding thesis — what specific thing does the consensus miss? What is the verifiable catalyst?
4. Assess competitive position with specifics: who are the top 2-3 competitors, and what structural advantage makes this company hard to displace?
5. Survivability check: be explicit — what happens to revenues and FCF in a -20% GDP recession scenario?

QUALITY FILTERS — verify before including any pick:
  1. ROIC > 15% preferred (✅ROIC flag in data) — this is the primary quality gate; check it first
  2. Valuation: PEG < 1.5 AND (P/FCF < 25 OR EV/EBITDA < 15) — two valuation confirmations required for non-hypergrowth
  3. FCF conversion ≥ 0.6 (FCFConv in data) — earnings quality gate; growth without FCF conversion is accounting, not business performance
  4. Flag ⚠GrwthGap and ⚠EpsGap picks explicitly — analyst optimism significantly ahead of track record is a red flag, not a buy signal
  5. Rate adjustment: if 10Y > 4%, verify FCF yield > 4% or explicitly justify why growth premium is warranted

KILL CRITERIA — hard rejections, no exceptions:
  ❌ FCF negative (unless 10-Bagger candidate with gross margin > 30% AND operating income positive)
  ❌ ROIC < 8% — reject unless it is a genuine financial inflection-point turnaround with explicit evidence in the data
  ❌ Revenue declining majority of years (revConsistency < 0.40) — structural decline, not cyclical
  ❌ High D/E (> 2.5) + negative FCF + decelerating growth = value trap; reject
  ❌ Commodity business with zero pricing power and no cost moat — permanently uninvestable at any PEG

Respond with ONLY valid JSON (no markdown, no preamble):
{{
  "synopsis": "2-3 sentences: what does the market get wrong right now? Where is the real opportunity?",
  "sector_rotation": "1-2 sentences: which sectors are at trough/peak and why — cite specific data",
  "macro_context": "2-3 sentences: how do current rates, yield curve, VIX, and inflation create specific opportunities or risks?",
  "macro_dashboard": {{
    "rate_environment": "1 sentence: what do current Treasury yields mean for equity valuations right now",
    "recession_risk": "LOW | MODERATE | HIGH — cite specific yield curve + unemployment evidence",
    "fed_policy": "HAWKISH | NEUTRAL | DOVISH — based on current fed funds rate vs inflation"
  }},
  "market_outlook": {{
    "near_term_bias": "BULLISH | NEUTRAL | CAUTIOUS | BEARISH",
    "long_term_bias": "BULLISH | NEUTRAL | CAUTIOUS | BEARISH",
    "crash_risk": "LOW | ELEVATED | HIGH",
    "rationale": "2 sentences citing specific macro numbers (e.g. 10Y at X%, VIX at Y, curve at Z%) to justify this view"
  }},
  "attention": ["specific risk #1 with ticker impact", "specific risk #2", "specific risk #3"],
  "specialist_consensus": "{'; '.join(consensus_note[:5]) if consensus_note else 'none this run'}",
  "picks": [
    {{
      "ticker": "TICKER",
      "company": "Company Name",
      "business_synopsis": "2-3 sentences: what does this company do, how does it make money, who pays them? Pure factual description — no investment thesis here.",
      "industry": "Specific industry or sub-sector (e.g. Cloud Security, Specialty Pharma, Industrial Automation)",
      "key_competitors": "Top 2-3 competitor names, comma-separated",
      "sector": "Sector",
      "strategy": "Fast Grower | 10-Bagger | Stalwart | Turnaround | Asset Play | Cyclical | Slow Grower | IV Discount | Quality Compounder",
      "lynch_category": "FastGrower | Stalwart | SlowGrower | Cyclical | Turnaround | AssetPlay (use the [Lynch=...] tag from the candidate data; if multi-label, pick the primary one that drives the thesis)",
      "endorsed_by": "QualityGrowth + SpecialSit | EmergingGrowth only | CapAppreciation + QualityGrowth | etc.",
      "position_tier": "CORE | SATELLITE | WATCH",
      "headline": "One-liner market-misunderstanding thesis — what is the market missing in plain English?",
      "story": "2-3 sentences: WHAT DOES THE MARKET NOT UNDERSTAND? What specific mispricing exists today?",
      "industry_context": "1-2 sentences: where is this industry in its cycle, and what drives the next phase?",
      "competitive_position": "1-2 sentences: top competitors named, and what specifically makes this company hard to displace",
      "survivability": "1 sentence: explicit assessment — how do revenues and FCF hold in a downturn?",
      "catalyst": "The specific named event or metric shift in next 1-6 months that unlocks value",
      "watch": "The single biggest risk that breaks this thesis — name it specifically",
      "conviction": "HIGH | MEDIUM",
      "urgency": "ACT NOW | WITHIN WEEKS | WITHIN MONTHS | WATCH | AVOID"
    }}
  ],
  "disclaimer": "Brief disclaimer"
}}

Position tier guide: CORE=3+ specialists endorse OR 2 specialists + exceptional quality; SATELLITE=1-2 specialists + quality pass; WATCH=Master Manager view only, no specialist endorsement.
Urgency guide: ACT NOW=catalyst imminent + entry compelling today; WITHIN WEEKS=good entry window 1-4wks; WITHIN MONTHS=patient accumulation thesis; WATCH=wait for confirmation signal; AVOID=thesis broken or kill criterion fires."""

    try:
        print("    Calling judge agent for final synthesis...")
        resp = _post(judge_system, judge_user, 12000, 300)
        if resp.status_code != 200:
            print(f"  ⚠️ Judge agent error {resp.status_code}: {resp.text[:200]}")
            if resp.status_code in (503, 529, 502, 524):
                # Server overloaded — retry once with compact prompt (same path as timeout fallback)
                import time as _time_judge
                print("  ⏳ Server overloaded — retrying judge with compact prompt in 15s...")
                _time_judge.sleep(15)
                compact_user_503 = (
                    f"Specialist reports:\n{specialist_block}\n\n{consensus_block}\n\n"
                    f"Pick the best 6 investments. JSON only:\n"
                    '{"synopsis":"...","sector_rotation":"...","macro_context":"...",'
                    '"market_outlook":{"near_term_bias":"CAUTIOUS","long_term_bias":"NEUTRAL",'
                    '"crash_risk":"ELEVATED","rationale":"..."},'
                    '"attention":["risk1","risk2","risk3"],"specialist_consensus":"see above",'
                    '"picks":[{"ticker":"T","company":"C","sector":"S","strategy":"S","endorsed_by":"...",'
                    '"headline":"...","story":"...","industry_context":"...","competitive_position":"...",'
                    '"survivability":"...","catalyst":"...","watch":"...","conviction":"HIGH","urgency":"WITHIN MONTHS"}],'
                    '"disclaimer":"Not investment advice."}'
                )
                try:
                    r_retry = _post(judge_system, compact_user_503, 3500, 120)
                    if r_retry.status_code == 200:
                        result_retry = _parse_response(r_retry.json()["content"][0]["text"])
                        if result_retry:
                            print(f"  ✅ Judge retry succeeded — {len(result_retry.get('picks', []))} picks")
                            result_retry["_specialist_picks"] = specialist_results
                            return result_retry
                    print(f"  ⚠️ Judge retry status {r_retry.status_code}")
                except Exception as e_retry:
                    print(f"  ⚠️ Judge retry also failed: {str(e_retry)[:80]}")
            return {}
        result = _parse_response(resp.json()["content"][0]["text"])
        if not result:
            print("  ⚠️ Judge JSON unrecoverable")
            return {}
        n_picks = len(result.get("picks", []))
        n_consensus = len(consensus_note)
        print(f"  ✅ Multi-agent analysis complete — {n_picks} picks ({n_consensus} cross-specialist consensus)")
        result["_specialist_picks"] = specialist_results   # for performance tracking
        return result

    except Exception as e:
        err = str(e)
        if "timed out" in err.lower() or "timeout" in err.lower():
            # Fallback: judge with just specialist reports, no full candidate data
            print("  ⚠️ Judge timed out — retrying with compact input...")
            compact_user = (
                f"Specialist reports:\n{specialist_block}\n\n{consensus_block}\n\n"
                f"Pick the best 6 investments. JSON only:\n"
                '{"synopsis":"...","sector_rotation":"...","macro_context":"...",'
                '"market_outlook":{"near_term_bias":"CAUTIOUS","long_term_bias":"NEUTRAL",'
                '"crash_risk":"ELEVATED","rationale":"..."},'
                '"attention":["risk1","risk2","risk3"],"specialist_consensus":"see above",'
                '"picks":[{"ticker":"T","company":"C","sector":"S","strategy":"S","endorsed_by":"...",'
                '"headline":"...","story":"...","industry_context":"...","competitive_position":"...",'
                '"survivability":"...","catalyst":"...","watch":"...","conviction":"HIGH","urgency":"WITHIN MONTHS"}],'
                '"disclaimer":"Not investment advice."}'
            )
            try:
                r2 = _post(judge_system, compact_user, 3500, 120)
                if r2.status_code == 200:
                    result2 = _parse_response(r2.json()["content"][0]["text"])
                    if result2:
                        print(f"  ✅ Judge retry succeeded — {len(result2.get('picks', []))} picks")
                        result2["_specialist_picks"] = specialist_results
                        return result2
            except Exception as e2:
                print(f"  ⚠️ Judge retry also failed: {str(e2)[:80]}")
        else:
            print(f"  ⚠️ Judge agent failed: {err[:120]}")
        return {}


def call_mall_manager(judge_result: dict, stocks: dict,
                      macro: dict = None, market_intel: dict = None,
                      fast_growers: list = None,
                      quality_compounders: list = None) -> dict:
    """🛍️ Mall Manager — Peter Lynch "shopping mall" rationaliser.

    Takes the existing judge's picks + all 11 specialists' picks (already in
    judge_result["_specialist_picks"]) and re-screens them through a single lens:
    is there a real-world consumer/observable thesis Wall Street is missing?

    Returns a short list (5-8) of high-conviction names, each with a
    `consumer_thesis` field describing what you'd literally observe in daily life.
    Stocks with great financials but no plausible Lynch story are REJECTED.
    """
    if not ANTHROPIC_KEY:
        return {}
    if not judge_result or not judge_result.get("picks"):
        return {}

    print("\n  🛍️ Running AI Mall Manager (Lynch Shopping Mall rationaliser)...")

    _MODEL = "claude-sonnet-4-6"

    # Build candidate pool: specialist picks + Fast Growers + Quality Compounders
    # The judge's output is shown separately as context but does NOT seed the pool.
    candidate_tickers = []
    seen = set()
    # 1. Specialist nominations (primary signal)
    for spec_name, sr in judge_result.get("_specialist_picks", {}).items():
        for p in sr.get("picks", []):
            t = (p.get("ticker") or "").upper()
            if t and t not in seen:
                candidate_tickers.append(t); seen.add(t)
    # 2. Fast Growers tab (Lynch-aligned: high-growth, PEG-driven)
    for row in (fast_growers or []):
        t = (row.get("Ticker") or row.get("ticker") or "").upper()
        if t and t not in seen:
            candidate_tickers.append(t); seen.add(t)
    # 3. Quality Compounders tab (durable moats, high ROIC — some are consumer brands)
    for row in (quality_compounders or []):
        t = (row.get("Ticker") or row.get("ticker") or "").upper()
        if t and t not in seen:
            candidate_tickers.append(t); seen.add(t)

    # Cap at 150 (specialists ~65 + fast growers 50 + quality compounders 50)
    candidate_tickers = candidate_tickers[:150]
    if not candidate_tickers:
        print("  ⚠️ Mall Manager: no candidates to evaluate")
        return {}

    # Build a compact per-stock format for the prompt — focus on consumer/identity signals
    cand_lines = []
    for t in candidate_tickers:
        s = stocks.get(t, {})
        co  = (s.get("companyName") or s.get("name") or t)[:50]
        ind = (s.get("industry") or "")[:35]
        sec = (s.get("sector") or "")[:20]
        mc  = s.get("mktCap")
        mc_s = (f"${mc/1e9:.1f}B" if mc and mc >= 1e9
                else f"${mc/1e6:.0f}M" if mc else "—")
        px  = s.get("price")
        px_s = f"${px:.0f}" if px else "—"
        rg  = s.get("revGrowth") or s.get("revenueGrowthYoy")
        rg_s = f"RG {rg*100:.0f}%" if rg else ""
        roic = s.get("roic")
        roic_s = f"ROIC {roic*100:.0f}%" if roic else ""
        tags = []
        if s.get("consumerObservable"):
            tags.append("🛒FamiliarBrand")
        if s.get("underCovered"):
            ac = s.get("analystCount")
            tags.append(f"🔍UnderCovered({ac}an)" if ac else "🔍UnderCovered")
        tag_s = " ".join(tags)
        cand_lines.append(f"  {t} | {co} | {ind} ({sec}) | {mc_s} @ {px_s} | {rg_s} {roic_s} {tag_s}".rstrip())

    candidates_block = "\n".join(cand_lines)

    # Macro hint (compact — one line)
    macro_line = ""
    if macro:
        try:
            yc = macro.get("yield_curve")
            vix = macro.get("vix")
            macro_line = f"Macro: 10Y={macro.get('dgs10','?')}, VIX={vix}, yield curve={yc}"
        except Exception:
            macro_line = ""

    sys_prompt = f"""You are the 🛍️ Mall Manager — a Peter Lynch-style investor whose ONLY job is to find stocks where there is a real-world, consumer-observable edge that the data-driven Wall Street consensus is structurally missing.

You are NOT a financial analyst. You don't care about beating consensus on EPS by 2¢. You care about ONE thing: walking through a shopping mall, scrolling TikTok, opening your kid's bedroom door, glancing at your office software dock, or queuing at an airport — and noticing that people are using a product MORE than the data shows yet. That observation, before it shows up in next quarter's revenue, is the edge.

CORE RULES — apply strictly:

1. NO LYNCH STORY → NO PICK. If you cannot articulate a plausible "I would observe this in daily life" thesis in one sentence, REJECT the stock — regardless of how great the financials look. A name with 30% ROIC and great moats but no consumer-observable surface is the JUDGE'S job, not yours.

2. STRONGLY PREFER 🛒FamiliarBrand candidates. That tag means the industry is consumer-observable (apps, retail, restaurants, EVs, streaming, healthcare you can name from CVS shelves). That's your hunting ground.

3. STRONGLY PREFER 🔍UnderCovered candidates when a Lynch story exists. Wall Street has <8 sell-side analysts on these names. The combination of consumer-observable + analyst-neglected is the most fertile inefficiency you can exploit.

4. REJECT pure B2B SaaS, enterprise data infra, industrial gear, financial services back-office, or commoditised commodity producers — UNLESS there is a true consumer-facing product surface. Examples that QUALIFY: Block (Cash App you actually use), Tesla (cars you see on the road), Spotify (app on your phone). Examples that DO NOT qualify: enterprise data warehouses, niche industrial valves, reinsurance carriers, B2B payment rails to merchants you never interact with.

5. WRITE THE consumer_thesis FIELD AS YOU'D TELL A FRIEND AT DINNER. No P/E ratios. No ROIC. No PEG. Real-world product observation only. Examples of GREAT consumer_thesis lines:
   - "Every coffee shop I walk into has a Square POS now — small merchants are switching from legacy terminals because the hardware is free and the app is faster."
   - "My teenage daughter and three of her friends are all using Duolingo Max with the AI tutor — they've replaced their school's $200/yr language tutor with a $7/mo app."
   - "Three of my colleagues started taking Halozyme-formulated subcutaneous Darzalex this month — patients are choosing the 5-min injection over the 4-hour IV infusion every time."

6. PICK 5-8 NAMES. NOT MORE. The judge already gives a longer list. Your value is selectivity — saying "out of these 30 specialist nominations, only THESE have a real Lynch story". Forcing more dilutes the signal.

7. CONTRARIAN ANGLE — explicitly call out, in the `wall_street_blindspot` field, what data-driven Wall Street is MISREADING. Examples: "Sell-side models a sub-pandemic baseline because they don't see how AI tools have made it sticky", "Street treats this as a melting ice cube because the legacy product is shrinking, but the new mobile app is doubling MAUs unmodeled".

8. TIEBREAKER ONLY — 👑 CEO grade and 🎯 Hidden Gem flags appear in the candidate data. Your primary job is the consumer-observable thesis, not capital allocation. But when two candidates have equally strong Lynch stories, prefer the one with 👑 ≥B+ CEO and/or 🎯 Hidden Gem. NEVER substitute these for an actual Lynch story — a great CEO at a B2B name with no consumer surface still gets REJECTED.

YOUR OUTPUT — JSON ONLY, no other text. Schema:

{{
  "synopsis": "One paragraph: today's overall Lynch-edge thesis — what consumer themes you noticed in the candidate pool",
  "picks": [
    {{
      "ticker": "TICKER",
      "company": "Company Name",
      "sector": "Sector",
      "consumer_thesis": "ONE SENTENCE — what you would observe in real life that confirms this thesis (no financial jargon)",
      "headline": "Short punchy 1-line summary of the investment idea",
      "story": "2-3 sentences: WHY this consumer trend translates to investable upside; what's accelerating; how durable",
      "wall_street_blindspot": "ONE SENTENCE — what data-driven analysis is structurally missing or misreading",
      "catalyst": "Specific 1-6 month event/metric where the consumer trend will become visible to the Street",
      "watch": "Single biggest risk that breaks the consumer thesis (NOT a financial risk — a product/usage risk)",
      "conviction": "HIGH | MEDIUM"
    }}
  ],
  "rejected_examples": "Brief 1-2 sentence note on 2-3 specialist-nominated stocks you rejected and WHY (no consumer-observable story present)"
}}

Today is {datetime.date.today()}. {macro_line}"""

    judge_context = "\n".join(
        f"  {p.get('ticker','?')} — {p.get('headline','')[:80]}"
        for p in judge_result.get('picks', [])[:11]
    )
    user_prompt = f"""You have {len(candidate_tickers)} candidate stocks drawn from today's 11 specialist analysts, the Fast Growers tab, and the Quality Compounders tab. Your job is to identify the 5-8 with the strongest Peter Lynch "shopping mall" consumer-observable thesis.

CANDIDATE POOL (specialist nominations + Fast Growers + Quality Compounders):
{candidates_block}

FOR REFERENCE ONLY — what the data-driven Master Manager picked today (you are fully independent and can agree, disagree, or overlap):
{judge_context}

YOUR TASK:
1. Walk through the specialist candidate pool. For each one, ask: "Could I plausibly observe this trend in my daily life — in a mall, on TikTok, in my kid's room, on my phone, at an airport?"
2. If yes, write the consumer_thesis. If no, REJECT.
3. Output 5-8 picks where the Lynch story is strongest. Quality over quantity.
4. In `rejected_examples`, name 2-3 stocks from the specialist pool you rejected and explain why (e.g., "RNR — reinsurance carrier, zero consumer surface", "PLMR — specialty B2B insurance, no observable product").

JSON only. No markdown, no preamble."""

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": ANTHROPIC_KEY,
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": _MODEL, "max_tokens": 6000,
                  "system": sys_prompt,
                  "messages": [{"role": "user", "content": user_prompt}]},
            timeout=240,
        )
        if resp.status_code != 200:
            print(f"  ⚠️ Mall Manager error {resp.status_code}: {resp.text[:200]}")
            return {}
        text = resp.json()["content"][0]["text"].strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        text = text.strip().rstrip("`")
        try:
            result = json.loads(text)
        except Exception as je:
            print(f"  ⚠️ Mall Manager JSON parse failed: {str(je)[:100]}")
            return {}
        n = len(result.get("picks", []))
        print(f"  ✅ Mall Manager picks: {n}")
        return result
    except Exception as e:
        print(f"  ⚠️ Mall Manager failed: {str(e)[:120]}")
        return {}


def log_ai_picks(ai_result: dict, stocks: dict, mall_result: dict = None):
    """Auto-log AI picks to fmp_ai_picks_log.csv for performance tracking.
    Logs judge picks (source=AI-Judge) AND specialist picks (source=AI-Bull/Value/Contrarian).
    CSV fields: date, source, ticker, company, strategy, conviction, entry_price, headline
    """
    today = datetime.date.today().isoformat()
    rows = []

    def _valid_ticker(ticker_str):
        """Reject agent labels, empty strings, or anything that looks like non-ticker."""
        if not ticker_str:
            return False
        if ticker_str.startswith("AI-"):   # agent labels: AI-Judge, AI-Bull etc.
            return False
        if len(ticker_str) > 10:           # real tickers are ≤ 5 chars (BRK.B = 5 + dot)
            return False
        return True

    # A8: Tickers already in today's strategy picks log (to flag strategy-echo overlaps)
    _strategy_today = set()
    try:
        if os.path.exists(PICKS_LOG):
            with open(PICKS_LOG, "r", encoding="utf-8") as _psf:
                for _pr in csv.DictReader(_psf):
                    if _pr.get("date") == today:
                        _strategy_today.add(_pr.get("ticker", "").upper())
    except Exception:
        pass

    # ── Specialist picks FIRST (so judge can detect echoes) ───────────────
    # A8: Build specialist-ticker set — judge picks matching become "AI-Judge-Echo"
    specialist_tickers_today = set()
    specialist_rows = []
    for spec_name, sr in ai_result.get("_specialist_picks", {}).items():
        source = f"AI-{spec_name}"   # "AI-Bull", "AI-Value", "AI-Contrarian"
        for p in sr.get("picks", []):
            t = p.get("ticker", "").upper()
            s = stocks.get(t, {})
            price = s.get("price")
            if _valid_ticker(t) and price:
                _hl = p.get("key_metric", "")[:80]
                if t in _strategy_today:
                    # Strategy-echo flag: the same ticker was logged as a strategy pick today
                    _hl = (f"[strat-echo] {_hl}")[:80]
                specialist_rows.append({
                    "date": today, "source": source,
                    "ticker": t,
                    "company": s.get("name", t)[:30],
                    "strategy": p.get("brief_case", "")[:50],
                    "conviction": p.get("conviction", ""),
                    "entry_price": round(price, 2),
                    "headline": _hl,
                    "prompt_version": PROMPT_VERSION,   # B6
                    "strategy_echo": "1" if t in _strategy_today else "",  # A8-fix
                    "synopsis": p.get("business_synopsis", "")[:300],
                    "industry": p.get("industry", "")[:60],
                    "key_competitors": p.get("key_competitors", "")[:120],
                })
                specialist_tickers_today.add(t)

    # ── Judge picks — top 5 only (judge already ranks by conviction/urgency) ──
    # A8: If a specialist already picked the same ticker today, tag the judge row
    # as "AI-Judge-Echo" — it's a confirming endorsement, not an independent pick,
    # and should NOT get its own performance credit in agent attribution.
    judge_rows = []
    for p in ai_result.get("picks", [])[:5]:
        t = p.get("ticker", "").upper()
        s = stocks.get(t, {})
        price = s.get("price")
        if _valid_ticker(t) and price:
            _source = "AI-Judge-Echo" if t in specialist_tickers_today else "AI-Judge"
            _hl = p.get("headline", "")[:80]
            if t in _strategy_today and "strat-echo" not in _hl:
                _hl = (f"[strat-echo] {_hl}")[:80]
            judge_rows.append({
                "date": today, "source": _source,
                "ticker": t,
                "company": p.get("company", s.get("name", ""))[:30],
                "strategy": p.get("strategy", ""),
                "conviction": p.get("conviction", ""),
                "entry_price": round(price, 2),
                "headline": _hl,
                "prompt_version": PROMPT_VERSION,   # B6
                "strategy_echo": "1" if t in _strategy_today else "",  # A8-fix
                "synopsis": p.get("business_synopsis", "")[:300],
                "industry": p.get("industry", "")[:60],
                "key_competitors": p.get("key_competitors", "")[:120],
            })

    # ── Mall Manager picks (Lynch consumer-observable lens) ───────────────
    mall_rows = []
    for p in (mall_result or {}).get("picks", []) if mall_result else []:
        t = p.get("ticker", "").upper()
        s = stocks.get(t, {})
        price = s.get("price")
        if _valid_ticker(t) and price:
            _hl = (p.get("consumer_thesis") or p.get("headline", ""))[:80]
            if t in _strategy_today and "strat-echo" not in _hl:
                _hl = (f"[strat-echo] {_hl}")[:80]
            mall_rows.append({
                "date": today, "source": "AI-MallManager",
                "ticker": t,
                "company": p.get("company", s.get("name", ""))[:30],
                "strategy": "Lynch Mall",
                "conviction": p.get("conviction", ""),
                "entry_price": round(price, 2),
                "headline": _hl,
                "prompt_version": PROMPT_VERSION,
                "strategy_echo": "1" if t in _strategy_today else "",
                "synopsis": (p.get("story") or p.get("consumer_thesis", ""))[:300],
                "industry": s.get("industry", "")[:60],
                "key_competitors": "",
            })

    # Judge first (preserved ordering), then mall manager, then specialists
    rows = judge_rows + mall_rows + specialist_rows

    if not rows:
        return

    # B6/A8-fix: prompt_version + strategy_echo columns — backfill empty for old rows
    NEW_FIELDS = ["date", "source", "ticker", "company", "strategy",
                  "conviction", "entry_price", "headline", "prompt_version",
                  "strategy_echo",   # A8-fix: "1" if ticker was also a strategy pick today
                  "synopsis", "industry", "key_competitors"]

    # Auto-migrate: old schemas missing any of the new columns
    file_exists = os.path.exists(AI_PICKS_LOG)
    if file_exists:
        with open(AI_PICKS_LOG, "r", encoding="utf-8") as _f:
            _reader = csv.DictReader(_f)
            _old_headers = _reader.fieldnames or []
            needs_migration = any(col not in _old_headers for col in NEW_FIELDS)
            if needs_migration:
                _old_rows = list(_reader)
                _migrated = [{**{k: "" for k in NEW_FIELDS},
                              **{k: r.get(k, "") for k in _old_headers if k in NEW_FIELDS}}
                             for r in _old_rows]
                with open(AI_PICKS_LOG, "w", newline="", encoding="utf-8") as _fw:
                    _w = csv.DictWriter(_fw, fieldnames=NEW_FIELDS)
                    _w.writeheader()
                    _w.writerows(_migrated)

    # De-duplicate: skip source+ticker combos already logged today
    file_exists = os.path.exists(AI_PICKS_LOG)
    existing_today = set()
    if file_exists:
        with open(AI_PICKS_LOG, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("date") == today:
                    existing_today.add((row.get("source", "AI-Judge"), row.get("ticker", "")))
    new_rows = [r for r in rows if (r["source"], r["ticker"]) not in existing_today]
    if not new_rows:
        print(f"  ℹ️ AI picks already logged for today")
        return
    with open(AI_PICKS_LOG, "a", newline="", encoding="utf-8") as f:
        fieldnames = NEW_FIELDS   # B6: keeps schema in sync with NEW_FIELDS constant above
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        if not file_exists:
            writer.writeheader()
        writer.writerows(new_rows)
    n_judge = sum(1 for r in new_rows if r["source"] == "AI-Judge")
    n_echo  = sum(1 for r in new_rows if r["source"] == "AI-Judge-Echo")
    n_mall  = sum(1 for r in new_rows if r["source"] == "AI-MallManager")
    n_spec  = sum(1 for r in new_rows if r["source"] not in ("AI-Judge", "AI-Judge-Echo", "AI-MallManager"))
    echo_note = f" ({n_echo} echoes)" if n_echo else ""
    mall_note = f" + {n_mall} mall" if n_mall else ""
    print(f"  📝 AI picks logged: {n_judge} judge{echo_note}{mall_note} + {n_spec} specialist → {AI_PICKS_LOG}")


def build_agent_reports_tab(wb, ai_result: dict, stocks: dict):
    """Build a tab showing each specialist agent's individual picks and reasoning."""
    print("\n📊 Building Tab: Agent Reports...")
    ws = wb.create_sheet("1d. Agent Reports")
    ws.sheet_view.showGridLines = False
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    AGENT_DESCRIPTIONS = {
        "QualityGrowth":  ("🌱 Quality Growth",
                          "Finds durable compounders with ROIC >15%, consistent multi-year revenue growth, and structural competitive moats. Prioritises FCF conversion and PEG <1.5 as quality gates."),
        "SpecialSit":     ("⚡ Special Situation",
                          "Identifies event-driven, misunderstood, and inflection-point opportunities. Focuses on business model misclassification, regulatory catalysts, and hidden assets not yet priced in."),
        "CapAppreciation":("📈 Capital Appreciation",
                          "Finds near-term re-rating candidates with specific catalysts in 1-6 months. Targets beaten-down quality (52wPos <65%), revenue re-acceleration signals, and cycle trough entries."),
        "EmergingGrowth": ("🚀 Emerging Growth",
                          "Identifies smaller fast-growing companies ($100M-$15B) at the early stage of becoming compounders. Targets 20%+ revenue growth, rising ROIC, and large underserved TAMs."),
        "TenBagger":      ("🎯 10-Bagger Hunter",
                          "Peter Lynch-style small-cap hunter. Looks for underfollowed companies ($50M-$2B) with 15-40% EPS+revenue growth, simple competitive advantages, at least 2 expansion levers not yet exhausted, and structural Wall Street under-coverage (orphans, post-restructuring, complex/sin-sector names)."),
        "LynchBWYK":      ("🛒 Lynch Buy What You Know",
                          "Lynch's 'buy what you know' approach — simple, understandable businesses serving everyday consumer or workplace needs, growing 15-25%/yr with real earnings at PEG <1.0."),
        "CathieWood":     ("🚀 Disruptive Innovation",
                          "ARK Invest-style innovation screen. Identifies pure-play companies in AI, robotics, genomics, energy storage, or blockchain riding Wright's Law cost curves with network effects and expanding TAMs."),
        "Pabrai":         ("🎲 Pabrai Asymmetric Bet",
                          "Mohnish Pabrai's 'heads I win, tails I don't lose much' framework. Finds bets with 3:1+ upside/downside asymmetry where downside is protected by real assets or essential business value."),
        "HowardMarks":    ("🔄 Marks Second-Level",
                          "Howard Marks' second-level thinking — contrarian analysis where market consensus is factually wrong. Targets oversold stocks where the negative narrative exceeds actual fundamental deterioration."),
        "Burry":          ("🕳️ Burry Deep Value",
                          "Michael Burry-style catalyst-driven deep value. Identifies hidden assets, temporary earnings distortions, and specific upcoming events that will FORCE the market to reprice."),
        "InsiderTrack":   ("👁️ Insider & Smart Money",
                          "Tracks cluster insider buying (multiple executives buying simultaneously) and significant open-market purchases. Insiders are the most informed buyers — their conviction buying with personal capital is a high-signal indicator."),
    }

    AGENT_COLORS = {
        "QualityGrowth": "2E7D32", "SpecialSit": "6A1B9A", "CapAppreciation": "1565C0",
        "EmergingGrowth": "E65100", "TenBagger": "BF360D",
        "LynchBWYK": "880E4F",
        "CathieWood": "0D47A1", "Pabrai": "33691E",
        "HowardMarks": "827717", "Burry": "3E2723",
        "InsiderTrack": "263238",
    }

    specialist_picks = ai_result.get("_specialist_picks", {})
    if not specialist_picks:
        ws.cell(row=1, column=1, value="No agent reports available for this run.").font = Font(bold=True)
        print("  ⚠️ No specialist data for Agent Reports tab")
        return

    # Title row
    title_cell = ws.cell(row=1, column=1, value=f"🤖 Agent Reports — All Specialist Picks  ({datetime.date.today()})")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    title_cell.fill = PatternFill("solid", fgColor="263238")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 24

    subtitle = ws.cell(row=2, column=1, value="Each agent applies a distinct investment philosophy. Picks are unfiltered specialist output — the Judge synthesises these into the final AI Top Picks.")
    subtitle.font = Font(italic=True, size=9, color="546E7A", name="Arial")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 14

    ws.column_dimensions["A"].width = 7   # Rank
    ws.column_dimensions["B"].width = 8   # Ticker
    ws.column_dimensions["C"].width = 26  # Company
    ws.column_dimensions["D"].width = 10  # Conviction
    ws.column_dimensions["E"].width = 22  # Key Metric
    ws.column_dimensions["F"].width = 60  # Brief Case
    ws.column_dimensions["G"].width = 10  # Price
    ws.column_dimensions["H"].width = 10  # MktCap

    sr = 4  # current row

    # Order: existing 5 first, then 12 new
    agent_order = ["QualityGrowth", "SpecialSit", "CapAppreciation", "EmergingGrowth", "TenBagger",
                   "LynchBWYK", "CathieWood",
                   "Pabrai", "HowardMarks", "Burry", "InsiderTrack"]

    for agent_key in agent_order:
        if agent_key not in specialist_picks:
            continue
        sr_data = specialist_picks[agent_key]
        picks = sr_data.get("picks", [])
        label_str = sr_data.get("label", agent_key)
        desc_info = AGENT_DESCRIPTIONS.get(agent_key, (label_str, ""))
        _, desc_text = desc_info
        color_hex = AGENT_COLORS.get(agent_key, "455A64")

        # Agent header
        hdr = ws.cell(row=sr, column=1, value=label_str)
        hdr.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
        hdr.fill = PatternFill("solid", fgColor=color_hex)
        hdr.alignment = Alignment(vertical="center", horizontal="left")
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=8)
        ws.row_dimensions[sr].height = 20
        sr += 1

        # Philosophy description
        desc_cell = ws.cell(row=sr, column=1, value=desc_text)
        desc_cell.font = Font(italic=True, size=9, color="37474F", name="Arial")
        desc_cell.fill = PatternFill("solid", fgColor="ECEFF1")
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=8)
        ws.row_dimensions[sr].height = 30
        sr += 1

        # Column headers
        col_headers = ["#", "Ticker", "Company", "Conviction", "Key Metric", "Brief Case / Thesis", "Price", "MktCap"]
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=sr, column=ci, value=h)
            c.font = Font(bold=True, size=9, color="FFFFFF", name="Arial")
            c.fill = PatternFill("solid", fgColor="78909C")
            c.alignment = Alignment(horizontal="center" if ci in (1, 4, 7, 8) else "left", vertical="center")
        ws.row_dimensions[sr].height = 16
        sr += 1

        # Picks
        for pi, pick in enumerate(picks, 1):
            tk = pick.get("ticker", "")
            s_data = stocks.get(tk, {})
            price = s_data.get("price")
            mktcap = s_data.get("mktCapB")
            row_fill = "FAFAFA" if pi % 2 == 0 else "FFFFFF"
            row_data = [
                pi,
                tk,
                pick.get("company", s_data.get("name", tk))[:28],
                pick.get("conviction", ""),
                pick.get("key_metric", "")[:24],
                pick.get("brief_case", "")[:200],
                round(price, 2) if price else "",
                f"${mktcap:.1f}B" if mktcap else "",
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=sr, column=ci, value=val)
                c.font = Font(size=9, name="Arial",
                              bold=(ci == 2),
                              color=("1565C0" if ci == 2 else "212121"))
                c.fill = PatternFill("solid", fgColor=row_fill)
                c.alignment = Alignment(
                    horizontal="center" if ci in (1, 4, 7, 8) else "left",
                    vertical="top", wrap_text=(ci == 6)
                )
            ws.row_dimensions[sr].height = 14 if len(pick.get("brief_case", "")) < 100 else 28
            sr += 1

        sr += 1  # blank row between agents

    print(f"  ✅ Agent Reports tab done — {len([k for k in agent_order if k in specialist_picks])} agents")


def build_ai_picks_tab(wb, ai_result: dict, stocks: dict, ws=None):
    """Tab 1b: AI Top Picks — Claude's curated picks with collapsible full stories."""
    print("\n📊 Building Tab: AI Top Picks...")
    if ws is None:
        ws = wb.create_sheet("1b. AI Top Picks")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    today = datetime.date.today()

    if not ai_result:
        ws.cell(row=1, column=1,
                value="Set ANTHROPIC_API_KEY environment variable to enable AI analysis.").font = \
            Font(italic=True, name="Arial", size=11, color="888888")
        return

    picks = ai_result.get("picks", [])
    TC = 16

    # Column widths: #, Ticker, Company, Sector, Strategy, Conv, Agents, Urgency, Price, PEG, P/E, MoS, ROE, FCF, Pio, Headline
    for col, w in zip("ABCDEFGHIJKLMNOP",
                      [5, 9, 22, 14, 14, 7, 9, 13, 8, 6, 6, 7, 7, 7, 5, 50]):
        ws.column_dimensions[col].width = w

    def _agents_short(endorsed_by_str):
        """Convert agent name string to short display label."""
        s = (endorsed_by_str or "").lower()
        parts = []
        if "qualitygrowth" in s or "quality growth" in s: parts.append("QGrwth")
        if "specialsit"    in s or "special sit"    in s: parts.append("SpSit")
        if "capappreciation" in s or "capital appreciation" in s: parts.append("CapAp")
        if "emerginggrowth" in s or "emerging growth" in s: parts.append("EmGrwth")
        # legacy fallback
        if not parts:
            if "bull"    in s: parts.append("Bull")
            if "value"   in s: parts.append("Val")
            if "contra"  in s: parts.append("Cont")
        return "+".join(parts) if parts else "Master Manager"

    # Consensus fill: 3 agents=dark green, 2=amber, 1=blue, judge only=grey
    _AGENTS_FILL = {3: "1B5E20", 2: "E65100", 1: "1565C0", 0: "546E7A"}

    HIGH_FILL = PatternFill("solid", fgColor="1B5E20")
    MED_FILL  = PatternFill("solid", fgColor="1565C0")

    def _hdr(row, value, fill_hex, font_size=9, height=18):
        c = ws.cell(row=row, column=1, value=value)
        c.font = Font(bold=True, name="Arial", size=font_size, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TC)
        ws.row_dimensions[row].height = height
        return row + 1

    def _col_hdrs(row, headers, fill_hex):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=fill_hex)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 26
        return row + 1

    r = 1
    r = _hdr(r, f"  🤖  AI Top Picks  —  Claude Sonnet Analysis    {today}",
             "0D1B2A", font_size=13, height=30)

    # ── Synopsis + Sector rotation side by side ────────────────────
    synopsis = ai_result.get("synopsis", "")
    sr_text  = ai_result.get("sector_rotation", "")
    SL = 9
    if synopsis or sr_text:
        r = _hdr(r, "  MARKET SYNOPSIS  &  SECTOR ROTATION", "1A237E", font_size=8, height=16)
        if synopsis:
            c = ws.cell(row=r, column=1, value=synopsis)
            c.font = Font(name="Arial", size=9, italic=True, color="1A237E")
            c.fill = PatternFill("solid", fgColor="E8EAF6")
            c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SL)
        if sr_text:
            c2 = ws.cell(row=r, column=SL+1, value=f"SECTOR ROTATION\n{sr_text}")
            c2.font = Font(name="Arial", size=8, italic=True, color="1B5E20")
            c2.fill = PatternFill("solid", fgColor="E8F5E9")
            c2.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.merge_cells(start_row=r, start_column=SL+1, end_row=r, end_column=TC)
        ws.row_dimensions[r].height = 42
        r += 1

    # ── Macro context ──────────────────────────────────────────────
    macro_ctx = ai_result.get("macro_context", "")
    if macro_ctx:
        r = _hdr(r, "  🌍  GEOPOLITICAL & MACRO CONTEXT", "4A148C", font_size=8, height=16)
        mc = ws.cell(row=r, column=1, value=macro_ctx)
        mc.font = Font(name="Arial", size=9, italic=True, color="4A148C")
        mc.fill = PatternFill("solid", fgColor="F3E5F5")
        mc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
        ws.row_dimensions[r].height = 36
        r += 1

    # ── Market outlook / crash risk banner ────────────────────────
    mo = ai_result.get("market_outlook", {})
    if mo:
        nt  = (mo.get("near_term_bias") or "NEUTRAL").upper()
        lt  = (mo.get("long_term_bias") or "NEUTRAL").upper()
        cr  = (mo.get("crash_risk") or "LOW").upper()
        mo_rationale = mo.get("rationale", "")
        # Color scheme: BEARISH/HIGH = red; CAUTIOUS/ELEVATED = amber; BULLISH = green; NEUTRAL = teal
        _BIAS_FILL  = {"BULLISH": "1B5E20", "NEUTRAL": "006064",
                       "CAUTIOUS": "E65100", "BEARISH": "B71C1C"}
        _CRASH_FILL = {"LOW": "1B5E20", "ELEVATED": "E65100", "HIGH": "B71C1C"}
        nt_fill  = _BIAS_FILL.get(nt,  "006064")
        lt_fill  = _BIAS_FILL.get(lt,  "006064")
        cr_fill  = _CRASH_FILL.get(cr, "E65100")
        # Header
        r = _hdr(r, "  📊  MARKET OUTLOOK  &  CRASH RISK ASSESSMENT", "263238",
                 font_size=8, height=16)
        # Three badge cells + rationale
        BADGE_W = 4  # columns per badge (~4 cols each = 12 total, rationale in last 3)
        badge_data = [
            (f"NEAR-TERM\n{nt}",  nt_fill,  1,             BADGE_W),
            (f"LONG-TERM\n{lt}",  lt_fill,  BADGE_W+1,     BADGE_W),
            (f"CRASH RISK\n{cr}", cr_fill,  BADGE_W*2+1,   BADGE_W),
        ]
        for label, fill_hex, start_col, span in badge_data:
            bc = ws.cell(row=r, column=start_col, value=label)
            bc.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            bc.fill = PatternFill("solid", fgColor=fill_hex)
            bc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.merge_cells(start_row=r, start_column=start_col,
                           end_row=r, end_column=start_col + span - 1)
        if mo_rationale:
            rc = ws.cell(row=r, column=BADGE_W*3+1, value=mo_rationale)
            rc.font = Font(name="Arial", size=8, italic=True, color="37474F")
            rc.fill = PatternFill("solid", fgColor="ECEFF1")
            rc.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
            ws.merge_cells(start_row=r, start_column=BADGE_W*3+1,
                           end_row=r, end_column=TC)
        ws.row_dimensions[r].height = 30
        r += 1

    # ── Key themes (single compact row) ───────────────────────────
    themes = ai_result.get("attention", [])
    if themes:
        r = _hdr(r, "  ⚠  KEY THEMES & RISKS", "B71C1C", font_size=8, height=16)
        bullets = "     ".join(f"▶ {t}" for t in themes)
        bc = ws.cell(row=r, column=1, value=bullets)
        bc.font = Font(name="Arial", size=8)
        bc.fill = PatternFill("solid", fgColor="FFEBEE")
        bc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1  # spacer

    # ── Picks table ────────────────────────────────────────────────
    r = _hdr(r,
             f"  TOP {len(picks)} AI PICKS  —  click ▶ row numbers to expand full story",
             "0D47A1", font_size=9, height=20)

    col_hdrs = ["#", "Ticker", "Company", "Sector", "Strategy", "Conv.",
                "Agents", "Urgency", "Price", "PEG", "P/E", "MoS%", "ROE%", "FCF%", "Pio",
                "Headline  (expand row ▶ for full story)"]
    r = _col_hdrs(r, col_hdrs, "1A237E")

    fmts = [None, None, None, None, None, None, None, None,
            "$#,##0.00", "0.00", "0.0", "0%", "0%", "0%", None, None]

    # Urgency colour map
    URGENCY_STYLES = {
        "ACT NOW":       ("B71C1C", "FFFFFF"),  # deep red
        "WITHIN WEEKS":  ("E65100", "FFFFFF"),  # deep orange
        "WITHIN MONTHS": ("1565C0", "FFFFFF"),  # blue
        "WATCH":         ("37474F", "FFFFFF"),  # dark grey
        "AVOID":         ("757575", "FFFFFF"),  # grey
    }

    # Fetch live prices for AI picks via v3 quote (stable API doesn't reliably serve live quotes)
    live_prices = {}
    _pick_tickers = [p.get("ticker","").upper() for p in picks if p.get("ticker")]
    for _t in _pick_tickers:
        _p = fetch_live_price(_t)
        if _p:
            live_prices[_t] = _p
        time.sleep(0.15)

    for i, p in enumerate(picks, 1):
        t        = p.get("ticker", "").upper()
        s        = stocks.get(t, {})
        conv     = (p.get("conviction") or "MEDIUM").upper()
        urgency  = (p.get("urgency") or "WATCH").upper()
        agents   = _agents_short(p.get("endorsed_by", ""))
        n_agents = agents.count("+") + (0 if agents == "Master Manager" else 1)
        rf       = ALT_FILL if i % 2 == 0 else PLAIN_FILL
        # Use live price if available, fall back to cached
        live_price = live_prices.get(t) or s.get("price")

        # ── Summary row (always visible) ──────────────────────────
        sum_vals = [
            i, t,
            p.get("company", s.get("name",""))[:22],
            p.get("sector", s.get("sector",""))[:14],
            p.get("strategy","")[:14],
            conv,
            agents,    # col 7 — new Agents column
            urgency,
            live_price, s.get("peg"), s.get("pe"),
            s.get("mos"), s.get("roe"), s.get("fcfYield"),
            s.get("piotroski"),
            p.get("headline",""),
        ]
        for ci, (v, fmt) in enumerate(zip(sum_vals, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.fill = rf
            cell.border = THIN_BORDER
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=(ci == 16))
            if fmt and isinstance(v, (int, float)):
                cell.number_format = fmt
            if ci == 2:
                cell.font = Font(bold=True, name="Arial", size=10, color="0D47A1")
            elif ci == 6:  # conviction badge
                cell.fill = HIGH_FILL if conv == "HIGH" else MED_FILL
                cell.font = Font(bold=True, name="Arial", size=8, color="FFFFFF")
            elif ci == 7:  # agents badge — color by consensus count
                _af = _AGENTS_FILL.get(n_agents, "546E7A")
                cell.fill = PatternFill("solid", fgColor=_af)
                cell.font = Font(bold=True, name="Arial", size=8, color="FFFFFF")
            elif ci == 8:  # urgency badge
                ufg, utf = URGENCY_STYLES.get(urgency, ("757575", "FFFFFF"))
                cell.fill = PatternFill("solid", fgColor=ufg)
                cell.font = Font(bold=True, name="Arial", size=8, color=utf)
            elif ci == 10 and isinstance(v, (int, float)):   # PEG
                if 0 < v < 1.0:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                    cell.font = Font(bold=True, name="Arial", size=9)
                elif 0 < v < 1.5:
                    cell.fill = PatternFill("solid", fgColor="FFF9C4")
            elif ci in (12, 13, 14) and isinstance(v, (int, float)):  # MoS, ROE, FCF
                cell.font = Font(name="Arial", size=9,
                                 color="1B5E20" if v > 0.01 else ("B71C1C" if v < -0.01 else "000000"))
            elif ci == 15 and isinstance(v, (int, float)):  # Piotroski
                cell.font = Font(bold=True, name="Arial", size=9)
                if v >= 8:
                    cell.fill = PatternFill("solid", fgColor="1B5E20")
                    cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                elif v >= 7:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
            elif ci == 16:
                cell.alignment = Alignment(wrap_text=True, vertical="center",
                                           horizontal="left")
        ws.row_dimensions[r].height = 16
        r += 1

        # ── Detail row (collapsible, hidden by default) ────────────
        story       = p.get("story","")
        ind_ctx     = p.get("industry_context","")
        comp_pos    = p.get("competitive_position","")
        survivabil  = p.get("survivability","")
        catalyst    = p.get("catalyst","")
        watch       = p.get("watch","")
        detail   = (f"  📖 MARKET EDGE: {story}"
                    + (f"\n\n  🏭 INDUSTRY: {ind_ctx}" if ind_ctx else "")
                    + (f"\n\n  🏆 COMPETITIVE POSITION: {comp_pos}" if comp_pos else "")
                    + (f"\n\n  🛡 SURVIVABILITY: {survivabil}" if survivabil else "")
                    + f"\n\n  ⚡ CATALYST: {catalyst}\n\n  ⚠ WATCH: {watch}")
        dc = ws.cell(row=r, column=1, value=detail)
        dc.font = Font(name="Arial", size=9, italic=True, color="37474F")
        dc.fill = PatternFill("solid", fgColor="F5F5F5")
        dc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
        ws.row_dimensions[r].height = 105
        ws.row_dimensions[r].outline_level = 1
        ws.row_dimensions[r].hidden = True
        r += 1

    disc = ai_result.get("disclaimer", "")
    if disc:
        dc2 = ws.cell(row=r, column=1, value=disc)
        dc2.font = Font(name="Arial", size=7, italic=True, color="AAAAAA")
        dc2.fill = PatternFill("solid", fgColor="FAFAFA")
        dc2.alignment = Alignment(indent=1, vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
        ws.row_dimensions[r].height = 13
        r += 1

    # ── Performance history (from log) ────────────────────────────
    r += 1
    if os.path.exists(AI_PICKS_LOG):
        r = _hdr(r, "  📈  AI PICKS PERFORMANCE HISTORY", "263238", font_size=9, height=20)
        logged = []
        with open(AI_PICKS_LOG, "r", encoding="utf-8") as f:
            for row_data in csv.DictReader(f):
                logged.append(row_data)

        if logged:
            ph2 = ["Date", "Ticker", "Company", "Strategy", "Conv.",
                   "Entry $", "Current $", "Return", "Days", "Status"]
            r = _col_hdrs(r, ph2, "263238")

            for ri, row_data in enumerate(logged):
                t2    = row_data.get("ticker", "")
                try:
                    entry = float(row_data.get("entry_price") or 0)
                except (ValueError, TypeError):
                    entry = 0
                curr  = stocks.get(t2, {}).get("price", 0) or 0
                ret = (curr - entry) / entry if entry > 0 and curr > 0 else None
                try:
                    days = (datetime.date.today() - datetime.date.fromisoformat(row_data["date"])).days
                except Exception:
                    days = None
                if ret is None:   status = "N/A"
                elif ret > 0.20:  status = "WINNER"
                elif ret > 0.05:  status = "UP"
                elif ret > -0.05: status = "FLAT"
                elif ret > -0.15: status = "DOWN"
                else:             status = "LOSS"

                _src = row_data.get("source", "AI-Judge")
                row_vals = [row_data.get("date",""), t2,
                            row_data.get("company","")[:20],
                            _src[:12],
                            row_data.get("conviction","")[:4],
                            entry, curr if curr else None, ret, days, status]
                alt = ri % 2 == 0
                for ci, v in enumerate(row_vals, 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.border = THIN_BORDER
                    cell.font = Font(name="Arial", size=9)
                    cell.fill = ALT_FILL if alt else PLAIN_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if ci in (6, 7) and isinstance(v, (int, float)):
                        cell.number_format = "$#,##0.00"
                    elif ci == 8 and isinstance(v, float):
                        cell.number_format = "0.0%"
                        cell.font = Font(name="Arial", size=9,
                                         color="1B5E20" if v > 0 else "B71C1C")
                    elif ci == 10:
                        if status == "WINNER":
                            cell.fill = PatternFill("solid", fgColor="1B5E20")
                            cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                        elif status == "UP":
                            cell.fill = PatternFill("solid", fgColor="C8E6C9")
                        elif status in ("DOWN", "LOSS"):
                            cell.fill = PatternFill("solid", fgColor="FFCDD2")
                ws.row_dimensions[r].height = 16
                r += 1

    ws.freeze_panes = "A3"
    print(f"  ✅ AI Top Picks tab done — {len(picks)} picks")


def build_overview_tab(ws, stocks, iv_rows, stalwarts, fast_growers, turnarounds,
                       slow_growers, cyclicals, asset_plays, quality_compounders,
                       fmp_call_count, ai=None, portfolio=None,
                       portfolio_nav=None, portfolio_ret=None, portfolio_spy_ret=None,
                       macro=None, ten_baggers=None):
    """Build the Overview tab: AI analysis at top, strategy tables below."""
    ws.sheet_view.showGridLines = False
    today = datetime.date.today()
    TC = 15  # total columns A–O

    # Enable row outline controls (for collapsible story rows)
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Column widths: #, Ticker, Company, Sector, Strategy, Conv, Urgency, Price, PEG, P/E, MoS, ROE, Pio, _, Headline
    for col, w in zip("ABCDEFGHIJKLMNO",
                      [5, 9, 22, 14, 14, 7, 13, 8, 6, 6, 7, 7, 5, 5, 45]):
        ws.column_dimensions[col].width = w

    # ── Helpers ─────────────────────────────────────────────────────
    def _hdr(row, value, fill_hex, font_size=9, height=18):
        """Full-width dark header row (matches write_table header style)."""
        c = ws.cell(row=row, column=1, value=value)
        c.font = Font(bold=True, name="Arial", size=font_size, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TC)
        ws.row_dimensions[row].height = height
        return row + 1

    def _text_row(row, value, fill_hex, font_color="222222", font_size=9,
                  italic=False, height=36, cols=TC):
        """Full-width wrapped text content row."""
        c = ws.cell(row=row, column=1, value=value)
        c.font = Font(bold=False, italic=italic, name="Arial",
                      size=font_size, color=font_color)
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        ws.row_dimensions[row].height = height
        return row + 1

    def _col_hdrs(row, headers, fill_hex):
        """Write a column-header row matching write_table style."""
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=fill_hex)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 26
        return row + 1

    r = 1

    # ── Title bar ──────────────────────────────────────────────────
    r = _hdr(r, f"  FMP Stock Screener  —  Professional Fundamentals Edition    {today}",
             "0D1B2A", font_size=13, height=30)
    r = _text_row(r,
                  f"Universe: {len(stocks):,} US stocks  |  FMP API calls: {fmp_call_count}"
                  f"  |  Data: Financial Modeling Prep  |  AI: Anthropic Claude Sonnet",
                  "0D1B2A", font_color="AAAAAA", font_size=8, height=14)

    # ── MACRO DASHBOARD ────────────────────────────────────────────
    # Always shown (doesn't depend on AI running) — displays real FRED data
    if macro:
        mc = macro
        r = _hdr(r, f"  🌍  MACRO DASHBOARD  —  Live FRED Data  (as of {mc.get('as_of', '?')})",
                 "1B2838", font_size=9, height=18)

        # Color helpers for metric tiles
        def _tile_fill(signal):
            """Return fill hex based on severity signal (GREEN/YELLOW/RED)."""
            sig = signal.upper() if signal else ""
            if sig in ("CALM", "NORMAL", "STEEP", "NEAR_TARGET", "LOW",
                       "HEALTHY", "TIGHT"):
                return "1B5E20"   # green
            if sig in ("CAUTION", "FLAT", "ABOVE_TARGET", "ELEVATED",
                       "SOFTENING"):
                return "E65100"   # orange/yellow
            if sig in ("PANIC", "FEAR", "INVERTED", "HOT", "HIGH",
                       "WEAK"):
                return "B71C1C"   # red
            return "455A64"       # grey for UNKNOWN / N/A

        # ── Per-metric signal helpers (each tile uses its own value's thresholds) ──
        def _dgs2_signal(v):
            # 2Y yield signal: same thresholds as 10Y (both drive real rates)
            if v is None: return "UNKNOWN"
            if v > 5.0: return "HIGH"
            if v > 4.0: return "ELEVATED"
            if v > 2.0: return "NORMAL"
            return "LOW"

        def _fedfunds_signal(v):
            # Fed Funds: HAWKISH when above 4.5%, ELEVATED 3.5-4.5%, NEUTRAL below
            if v is None: return "UNKNOWN"
            if v > 4.5: return "HAWKISH"
            if v > 3.5: return "ELEVATED"
            if v > 2.0: return "NORMAL"
            return "DOVISH"

        # ── Row A: 7 metric tiles ──────────────────────────────────
        tiles = [
            # (label, value_str, signal_str, col_start, col_span)
            ("10Y YIELD",
             f"{mc['dgs10']}%" if mc.get('dgs10') is not None else "N/A",
             mc.get('rate_signal', 'UNKNOWN'), 1, 2),
            ("2Y YIELD",
             f"{mc['dgs2']}%" if mc.get('dgs2') is not None else "N/A",
             _dgs2_signal(mc.get('dgs2')), 3, 2),
            ("YIELD CURVE",
             (f"+{mc['yield_curve']}%" if mc.get('yield_curve', 0) >= 0
              else f"{mc['yield_curve']}%") if mc.get('yield_curve') is not None else "N/A",
             mc.get('curve_signal', 'UNKNOWN'), 5, 3),
            ("VIX",
             str(mc['vix']) if mc.get('vix') is not None else "N/A",
             mc.get('vix_signal', 'UNKNOWN'), 8, 2),
            ("FED FUNDS",
             f"{mc['fedfunds']}%" if mc.get('fedfunds') is not None else "N/A",
             _fedfunds_signal(mc.get('fedfunds')), 10, 2),
            ("CPI YoY",
             f"{mc['cpi_yoy']}%" if mc.get('cpi_yoy') is not None else "N/A",
             mc.get('inflation_signal', 'UNKNOWN'), 12, 2),
            ("UNEMPLOYMT",
             f"{mc['unrate']}%" if mc.get('unrate') is not None else "N/A",
             mc.get('labor_signal', 'UNKNOWN'), 14, 2),
        ]

        for label, val_str, signal, col_start, col_span in tiles:
            fill_hex = _tile_fill(signal)
            # Label cell (top mini-row in merged area — we use two lines in one cell)
            tc = ws.cell(row=r, column=col_start,
                         value=f"{label}\n{val_str}")
            tc.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            tc.fill = PatternFill("solid", fgColor=fill_hex)
            tc.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
            if col_span > 1:
                ws.merge_cells(start_row=r, start_column=col_start,
                               end_row=r, end_column=col_start + col_span - 1)
        ws.row_dimensions[r].height = 28
        r += 1

        # ── Row B: signal tags row ─────────────────────────────────
        for label, val_str, signal, col_start, col_span in tiles:
            fill_hex = _tile_fill(signal)
            # Slightly lighter shade: append "99" alpha-equivalent via a fixed light shade
            light_fills = {
                "1B5E20": "C8E6C9",   # green light
                "E65100": "FFE0B2",   # orange light
                "B71C1C": "FFCDD2",   # red light
                "455A64": "ECEFF1",   # grey light
            }
            light = light_fills.get(fill_hex, "ECEFF1")
            sc = ws.cell(row=r, column=col_start, value=signal)
            sc.font = Font(bold=True, name="Arial", size=8,
                           color=fill_hex)
            sc.fill = PatternFill("solid", fgColor=light)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            if col_span > 1:
                ws.merge_cells(start_row=r, start_column=col_start,
                               end_row=r, end_column=col_start + col_span - 1)
        ws.row_dimensions[r].height = 14
        r += 1

        # ── Row C: interpretation line ─────────────────────────────
        interp_parts = []
        vc = mc.get("yield_curve")
        if vc is not None:
            cs = mc.get("curve_signal", "")
            if cs == "INVERTED":
                interp_parts.append(f"Yield curve INVERTED ({vc:+.2f}%) - historically precedes recession 6-18mo")
            elif cs == "FLAT":
                interp_parts.append(f"Yield curve FLAT ({vc:+.2f}%) - slowdown risk, watch carefully")
            else:
                interp_parts.append(f"Yield curve {cs.lower()} ({vc:+.2f}%) - no imminent recession signal")
        vv = mc.get("vix")
        if vv is not None:
            vs = mc.get("vix_signal", "")
            interp_parts.append(f"VIX {vv} ({vs.lower()} fear)")
        ci = mc.get("cpi_yoy")
        if ci is not None:
            inf_s = mc.get("inflation_signal", "")
            interp_parts.append(f"CPI {ci}% YoY ({inf_s.lower().replace('_', ' ')})")
        ff = mc.get("fedfunds")
        if ff is not None:
            rs = mc.get("rate_signal", "")
            interp_parts.append(f"Fed Funds {ff}% ({rs.lower()})")

        interp_text = "  |  ".join(interp_parts) if interp_parts else "Macro data unavailable"
        r = _text_row(r, interp_text, "1B2838", font_color="90A4AE",
                      font_size=8, height=14)

    # ── AI SECTION ─────────────────────────────────────────────────
    if ai:
        r = _hdr(r, "  🤖  AI MARKET OVERVIEW  —  Claude Sonnet Analysis",
                 "1A237E", font_size=10, height=20)

        # Synopsis + sector rotation side by side
        synopsis = ai.get("synopsis", "")
        sr_text  = ai.get("sector_rotation", "")
        SL = 9  # left columns for synopsis
        if synopsis or sr_text:
            if synopsis:
                c = ws.cell(row=r, column=1, value=synopsis)
                c.font = Font(name="Arial", size=9, italic=True, color="1A237E")
                c.fill = PatternFill("solid", fgColor="E8EAF6")
                c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SL)
            if sr_text:
                c2 = ws.cell(row=r, column=SL+1,
                             value=f"SECTOR ROTATION\n{sr_text}")
                c2.font = Font(name="Arial", size=8, italic=True, color="1B5E20")
                c2.fill = PatternFill("solid", fgColor="E8F5E9")
                c2.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                ws.merge_cells(start_row=r, start_column=SL+1, end_row=r, end_column=TC)
            ws.row_dimensions[r].height = 42
            r += 1

        # Macro context
        macro_ctx = ai.get("macro_context", "")
        if macro_ctx:
            r = _hdr(r, "  🌍  GEOPOLITICAL & MACRO CONTEXT", "4A148C", font_size=8, height=16)
            r = _text_row(r, macro_ctx, "F3E5F5", font_color="4A148C",
                          italic=True, height=36)

        # Market outlook / crash risk banner
        mo = ai.get("market_outlook", {})
        if mo:
            nt  = (mo.get("near_term_bias") or "NEUTRAL").upper()
            lt  = (mo.get("long_term_bias") or "NEUTRAL").upper()
            cr  = (mo.get("crash_risk") or "LOW").upper()
            mo_rationale = mo.get("rationale", "")
            _BIAS_FILL  = {"BULLISH": "1B5E20", "NEUTRAL": "006064",
                           "CAUTIOUS": "E65100", "BEARISH": "B71C1C"}
            _CRASH_FILL = {"LOW": "1B5E20", "ELEVATED": "E65100", "HIGH": "B71C1C"}
            nt_fill  = _BIAS_FILL.get(nt,  "006064")
            lt_fill  = _BIAS_FILL.get(lt,  "006064")
            cr_fill  = _CRASH_FILL.get(cr, "E65100")
            r = _hdr(r, "  📊  MARKET OUTLOOK  &  CRASH RISK ASSESSMENT", "263238",
                     font_size=8, height=16)
            BADGE_W = 4
            badge_data = [
                (f"NEAR-TERM\n{nt}",  nt_fill,  1,           BADGE_W),
                (f"LONG-TERM\n{lt}",  lt_fill,  BADGE_W+1,   BADGE_W),
                (f"CRASH RISK\n{cr}", cr_fill,  BADGE_W*2+1, BADGE_W),
            ]
            for label, fill_hex, start_col, span in badge_data:
                bc = ws.cell(row=r, column=start_col, value=label)
                bc.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
                bc.fill = PatternFill("solid", fgColor=fill_hex)
                bc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.merge_cells(start_row=r, start_column=start_col,
                               end_row=r, end_column=start_col + span - 1)
            if mo_rationale:
                rc = ws.cell(row=r, column=BADGE_W*3+1, value=mo_rationale)
                rc.font = Font(name="Arial", size=8, italic=True, color="37474F")
                rc.fill = PatternFill("solid", fgColor="ECEFF1")
                rc.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
                ws.merge_cells(start_row=r, start_column=BADGE_W*3+1,
                               end_row=r, end_column=TC)
            ws.row_dimensions[r].height = 30
            r += 1

        # AI macro_dashboard interpretation (rate env, recession risk, fed policy)
        md = ai.get("macro_dashboard", {})
        if md:
            _RECV_FILL = {"LOW": "1B5E20", "MODERATE": "E65100", "HIGH": "B71C1C"}
            _FED_FILL  = {"DOVISH": "1B5E20", "NEUTRAL": "006064", "HAWKISH": "B71C1C"}
            recv_risk = (md.get("recession_risk") or "").split()[0].upper()  # first word only
            fed_pol   = (md.get("fed_policy") or "NEUTRAL").upper()
            r_fill = _RECV_FILL.get(recv_risk, "455A64")
            f_fill = _FED_FILL.get(fed_pol, "006064")
            # Badges for recession risk + fed policy
            MBADGE_W = 3
            mb_data = [
                (f"RECESSION RISK\n{recv_risk}", r_fill, 1,            MBADGE_W),
                (f"FED POLICY\n{fed_pol}",       f_fill, MBADGE_W+1,   MBADGE_W),
            ]
            r = _hdr(r, "  🏦  AI MACRO INTERPRETATION", "37474F", font_size=8, height=16)
            for label, fill_hex, start_col, span in mb_data:
                bc = ws.cell(row=r, column=start_col, value=label)
                bc.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                bc.fill = PatternFill("solid", fgColor=fill_hex)
                bc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.merge_cells(start_row=r, start_column=start_col,
                               end_row=r, end_column=start_col + span - 1)
            rate_env = md.get("rate_environment", "")
            if rate_env:
                re_c = ws.cell(row=r, column=MBADGE_W*2+1, value=rate_env)
                re_c.font = Font(name="Arial", size=8, italic=True, color="37474F")
                re_c.fill = PatternFill("solid", fgColor="ECEFF1")
                re_c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
                ws.merge_cells(start_row=r, start_column=MBADGE_W*2+1,
                               end_row=r, end_column=TC)
            ws.row_dimensions[r].height = 28
            r += 1

        # Key themes (compact horizontal list in one row)
        themes = ai.get("attention", [])
        if themes:
            r = _hdr(r, "  ⚠  KEY THEMES & RISKS", "B71C1C", font_size=8, height=16)
            bullets = "     ".join(f"▶ {t}" for t in themes)
            r = _text_row(r, bullets, "FFEBEE", font_size=8, height=20)

        # AI Picks table
        picks = ai.get("picks", [])
        if picks:
            r += 1
            r = _hdr(r,
                     f"  TOP {len(picks)} AI PICKS  —  click ▶ row numbers to expand full story",
                     "0D47A1", font_size=9, height=20)

            pick_hdrs = ["#", "Ticker", "Company", "Sector", "Strategy",
                         "Conv.", "Urgency", "Price", "PEG", "P/E", "MoS%", "ROE%", "Pio",
                         "Headline  (expand row to see full thesis)"]
            r = _col_hdrs(r, pick_hdrs, "1A237E")

            HIGH_F = PatternFill("solid", fgColor="1B5E20")
            MED_F  = PatternFill("solid", fgColor="1565C0")
            fmts   = [None, None, None, None, None, None, None,
                      "$#,##0.00", "0.00", "0.0", "0%", "0%", None, None]

            URGENCY_STYLES = {
                "ACT NOW":       ("B71C1C", "FFFFFF"),
                "WITHIN WEEKS":  ("E65100", "FFFFFF"),
                "WITHIN MONTHS": ("1565C0", "FFFFFF"),
                "WATCH":         ("37474F", "FFFFFF"),
                "AVOID":         ("757575", "FFFFFF"),
            }

            # Fetch live prices for overview picks (same as AI tab)
            _ov_live = {}
            for _p in picks:
                _t2 = _p.get("ticker", "").upper()
                if _t2:
                    _lp = fetch_live_price(_t2)
                    if _lp:
                        _ov_live[_t2] = _lp
                    time.sleep(0.15)

            for i, p in enumerate(picks, 1):
                t      = p.get("ticker", "").upper()
                s      = stocks.get(t, {})
                conv   = (p.get("conviction") or "MEDIUM").upper()
                urgency = (p.get("urgency") or "WATCH").upper()
                rf     = ALT_FILL if i % 2 == 0 else PLAIN_FILL

                # ── Summary row (always visible) ──────────────────
                _ov_price = _ov_live.get(t) or s.get("price")
                sum_vals = [
                    i, t,
                    p.get("company", s.get("name",""))[:22],
                    p.get("sector", s.get("sector",""))[:14],
                    p.get("strategy","")[:14],
                    conv, urgency,
                    _ov_price, s.get("peg"), s.get("pe"),
                    s.get("mos"), s.get("roe"), s.get("piotroski"),
                    p.get("headline",""),
                ]
                for ci, (v, fmt) in enumerate(zip(sum_vals, fmts), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.fill = rf
                    cell.border = THIN_BORDER
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center",
                                               wrap_text=(ci == 14))
                    if fmt and isinstance(v, (int, float)):
                        cell.number_format = fmt
                    if ci == 2:
                        cell.font = Font(bold=True, name="Arial", size=9, color="0D47A1")
                    elif ci == 6:  # conviction
                        cell.fill = HIGH_F if conv == "HIGH" else MED_F
                        cell.font = Font(bold=True, name="Arial", size=8, color="FFFFFF")
                    elif ci == 7:  # urgency
                        ufg, utf = URGENCY_STYLES.get(urgency, ("757575", "FFFFFF"))
                        cell.fill = PatternFill("solid", fgColor=ufg)
                        cell.font = Font(bold=True, name="Arial", size=8, color=utf)
                    elif ci == 10 and isinstance(v, (int, float)):
                        cell.number_format = "0.0"
                    elif ci == 11 and isinstance(v, float):
                        cell.number_format = "0%"
                        if v > 0.2: cell.fill = PatternFill("solid", fgColor="C8E6C9")
                    elif ci == 13 and isinstance(v, (int, float)):
                        if v >= 8:
                            cell.fill = PatternFill("solid", fgColor="1B5E20")
                            cell.font = Font(bold=True, name="Arial", size=8, color="FFFFFF")
                        elif v >= 7:
                            cell.fill = PatternFill("solid", fgColor="C8E6C9")
                    elif ci == 14:
                        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                                   horizontal="left")
                ws.row_dimensions[r].height = 16
                r += 1

                # ── Detail row (collapsible — outline level 1) ────
                story    = p.get("story","")
                ind_ctx  = p.get("industry_context","")
                catalyst = p.get("catalyst","")
                watch    = p.get("watch","")
                detail   = (f"  {t}  |  📖 MARKET EDGE: {story}"
                            + (f"\n\n  🏭 INDUSTRY: {ind_ctx}" if ind_ctx else "")
                            + f"\n\n  ⚡ CATALYST: {catalyst}\n\n  ⚠ WATCH: {watch}")
                dc = ws.cell(row=r, column=1, value=detail)
                dc.font = Font(name="Arial", size=8, italic=True, color="37474F")
                dc.fill = PatternFill("solid", fgColor="F5F5F5")
                dc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
                ws.row_dimensions[r].height = 75
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = True   # collapsed by default
                r += 1

            disc = ai.get("disclaimer", "")
            if disc:
                r = _text_row(r, disc, "FAFAFA", font_color="AAAAAA",
                              font_size=7, height=13)
    else:
        r = _hdr(r,
                 "  Set ANTHROPIC_API_KEY in .env to enable AI market analysis",
                 "37474F", font_size=9, height=20)

    # ── PORTFOLIO MANAGER SUMMARY ──────────────────────────────────
    if portfolio and portfolio.get("holdings") is not None:
        r += 1
        r = _hdr(r, "  💼  AI PORTFOLIO MANAGER  —  Paper Portfolio Summary  (see tab 1c for full detail)",
                 "1B2631", font_size=9, height=20)

        holdings_ov = portfolio.get("holdings", [])
        cash_ov     = portfolio.get("cash", PORTFOLIO_INITIAL_CASH)
        initial_ov  = portfolio.get("initial_cash", PORTFOLIO_INITIAL_CASH)
        nav_ov      = portfolio_nav or (cash_ov + sum(
            h["shares"] * (stocks.get(h["ticker"], {}).get("price") or h["entry_price"])
            for h in holdings_ov))
        ret_ov      = portfolio_ret if portfolio_ret is not None else ((nav_ov - initial_ov) / initial_ov)
        spy_ov      = portfolio_spy_ret
        alpha_ov    = (ret_ov - spy_ov) if spy_ov is not None else None

        # Stats row
        stat_labels = ["NAV", "Total Return", "SPY (same period)", "Alpha",
                       "Cash", f"Positions ({len(holdings_ov)}/{PORTFOLIO_MAX_POSITIONS})", "Started"]
        stat_values = [
            f"${nav_ov:,.0f}",
            f"{ret_ov:+.1%}",
            f"{spy_ov:+.1%}" if spy_ov is not None else "n/a",
            f"{alpha_ov:+.1%}" if alpha_ov is not None else "n/a",
            f"${cash_ov:,.0f}",
            ", ".join(h["ticker"] for h in holdings_ov[:6]) + ("…" if len(holdings_ov) > 6 else ""),
            portfolio.get("started", "—"),
        ]
        for ci, (lbl, val) in enumerate(zip(stat_labels, stat_values), 1):
            lc = ws.cell(row=r, column=ci*2-1, value=lbl)
            lc.font = Font(bold=True, name="Arial", size=8, color="AAAAAA")
            lc.fill = PatternFill("solid", fgColor="1B2631")
            lc.alignment = Alignment(horizontal="right")
            vc = ws.cell(row=r, column=ci*2, value=val)
            is_pos = isinstance(val, str) and val.startswith("+")
            is_neg = isinstance(val, str) and val.startswith("-")
            vc.font = Font(bold=True, name="Arial", size=9,
                           color=("00C853" if is_pos else ("FF1744" if is_neg else "FFFFFF")))
            vc.fill = PatternFill("solid", fgColor="1B2631")
        ws.row_dimensions[r].height = 16
        r += 1

        # Portfolio thesis
        thesis = portfolio.get("_last_thesis", "")
        if thesis:
            tc2 = ws.cell(row=r, column=1, value=f"PM Thesis: {thesis}")
            tc2.font = Font(name="Arial", size=8, italic=True, color="455A64")
            tc2.fill = PatternFill("solid", fgColor="ECEFF1")
            tc2.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
            ws.row_dimensions[r].height = 28
            r += 1

    # ── STRATEGY SUMMARY TABLE ─────────────────────────────────────
    r += 1
    r = _hdr(r, "  STRATEGY PICKS SUMMARY", "263238", font_size=10, height=22)

    tab_configs = [
        ("💎 IV Discount",      iv_rows,           "1A237E"),
        ("🏆 Quality Comp.",    quality_compounders,"7B1FA2"),
        ("🏛 Stalwarts",        stalwarts,          "4A148C"),
        ("🚀 Fast Growers",     fast_growers,       "1B5E20"),
        ("🔧 Turnarounds",      turnarounds,        "B71C1C"),
        ("💰 Slow Growers",     slow_growers,       "455A64"),
        ("🔄 Cyclicals",        cyclicals,          "E65100"),
        ("🏗 Asset Plays",      asset_plays,        "0D47A1"),
        ("🎯 10-Baggers",       ten_baggers or [],  "BF360D"),
    ]
    sum_hdrs = ["Strategy", "# Qualifying", "Top 3 Tickers", "Top Pick", "PEG", "MoS", "Piotroski"]
    r = _col_hdrs(r, sum_hdrs, "37474F")

    for label, tab_rows, color in tab_configs:
        top3   = ", ".join(p.get("Ticker","") for p in tab_rows[:3])
        top1   = tab_rows[0] if tab_rows else {}
        vals   = [label, len(tab_rows), top3,
                  top1.get("Company","")[:22],
                  top1.get("PEG"), top1.get("MoS"), top1.get("Piotroski")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.fill = ALT_FILL if r % 2 == 0 else PLAIN_FILL
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="left" if ci in (1,3,4) else "center",
                                       vertical="center")
            if ci == 1:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            elif ci == 5 and isinstance(v, float):
                cell.number_format = "0.00"
                if 0 < v < 1.0: cell.fill = PatternFill("solid", fgColor="C8E6C9")
                elif 0 < v < 1.5: cell.fill = PatternFill("solid", fgColor="FFF9C4")
            elif ci == 6 and isinstance(v, float):
                cell.number_format = "0%"
                if v > 0.2: cell.fill = PatternFill("solid", fgColor="C8E6C9")
            elif ci == 7 and isinstance(v, (int, float)):
                if v >= 8:
                    cell.fill = PatternFill("solid", fgColor="1B5E20")
                    cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                elif v >= 7:
                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
        ws.row_dimensions[r].height = 16
        r += 1

    # ── MINI PICK TABLES PER STRATEGY ─────────────────────────────
    r += 1
    r = _hdr(r, "  STRATEGY DETAIL  —  Top 8 picks per strategy",
             "263238", font_size=10, height=22)

    MINI_HDRS = ["#", "Ticker", "Company", "Sector", "Price",
                 "PEG", "P/E", "MoS%", "ROE%", "FCF%", "Pio", "Score"]
    MINI_FMTS = [None,None,None,None,"$#,##0.00","0.00","0.0","0%","0%","0%",None,"0.0"]

    for label, tab_rows, color in tab_configs:
        r = _hdr(r, f"  {label}", color, font_size=9, height=18)
        r = _col_hdrs(r, MINI_HDRS, color)

        for rank, pick in enumerate(tab_rows[:8], 1):
            rf2 = ALT_FILL if rank % 2 == 0 else PLAIN_FILL
            vals2 = [f"#{rank}", pick.get("Ticker",""), (pick.get("Company") or "")[:22],
                     (pick.get("Sector") or "")[:14], pick.get("Price"),
                     pick.get("PEG"), pick.get("P/E"), pick.get("MoS"),
                     pick.get("ROE"), pick.get("FCF Yield"),
                     pick.get("Piotroski"), pick.get("Score")]
            for ci, (v, fmt) in enumerate(zip(vals2, MINI_FMTS), 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.font = Font(name="Arial", size=9)
                c.fill = rf2
                c.alignment = Alignment(
                    horizontal="left" if ci in (2,3,4) else "center",
                    vertical="center")
                c.border = THIN_BORDER
                if fmt and isinstance(v, (int, float)): c.number_format = fmt
                if ci == 8 and isinstance(v, float):
                    if v > 0.3: c.fill = PatternFill("solid", fgColor="C8E6C9")
                    elif v > 0.1: c.fill = PatternFill("solid", fgColor="E8F5E9")
                if ci == 11 and isinstance(v, (int, float)):
                    if v >= 8:
                        c.fill = PatternFill("solid", fgColor="1B5E20")
                        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                    elif v >= 7: c.fill = PatternFill("solid", fgColor="C8E6C9")
                if ci == 6 and isinstance(v, float) and v > 0:
                    if v < 1.0: c.fill = PatternFill("solid", fgColor="C8E6C9")
                    elif v < 1.5: c.fill = PatternFill("solid", fgColor="FFF9C4")
            ws.row_dimensions[r].height = 16
            r += 1
        r += 1  # gap between strategies

    ws.freeze_panes = "A3"


def log_picks(picks_by_strategy: dict, prices: dict):
    """Auto-log today's top picks to CSV for performance tracking. Skips duplicates."""
    today = datetime.date.today().isoformat()
    file_exists = os.path.exists(PICKS_LOG)

    # Don't duplicate — collect already-logged ticker+strategy combos for today
    existing_today = set()
    if file_exists:
        with open(PICKS_LOG, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("date") == today:
                    existing_today.add((row.get("ticker"), row.get("strategy")))

    rows = []
    for strategy, pick_rows in picks_by_strategy.items():
        for r in pick_rows[:5]:
            t = r.get("Ticker", "")
            price = prices.get(t, r.get("Price"))
            if t and price and (t, strategy) not in existing_today:
                rows.append({"date": today, "strategy": strategy,
                             "ticker": t, "entry_price": round(price, 2),
                             "company": r.get("Company", "")})
    if not rows:
        if existing_today:
            print(f"  ℹ️ Strategy picks already logged for today")
        return
    with open(PICKS_LOG, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["date", "strategy", "ticker", "entry_price", "company"])
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)
    print(f"  📝 Strategy picks logged: {len(rows)} entries → {PICKS_LOG}")


def load_portfolio() -> dict:
    """Load the paper portfolio from JSON file. Creates a fresh one if it doesn't exist."""
    if os.path.exists(PORTFOLIO_FILE):
        try:
            with open(PORTFOLIO_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "started": datetime.date.today().isoformat(),
        "initial_cash": PORTFOLIO_INITIAL_CASH,
        "cash": PORTFOLIO_INITIAL_CASH,
        "holdings": [],
        "transactions": [],
    }


def save_portfolio(p: dict):
    """Persist portfolio to JSON."""
    with open(PORTFOLIO_FILE, "w", encoding="utf-8") as f:
        json.dump(p, f, indent=2)


def run_portfolio_manager(portfolio: dict, candidates_block: str,
                          sector_block: str, stocks: dict) -> dict:
    """Call Claude as a patient long-term portfolio manager.
    Returns a decisions dict with 'review' (hold/sell) + 'buys' lists, or {} on failure.
    """
    if not ANTHROPIC_KEY:
        return {}

    today = datetime.date.today()

    # Build current-holdings context with live performance
    holdings_lines = []
    for h in portfolio.get("holdings", []):
        t = h["ticker"]
        s = stocks.get(t, {})
        current_price = s.get("price") or h["entry_price"]
        ret = (current_price - h["entry_price"]) / h["entry_price"] if h["entry_price"] > 0 else 0
        try:
            days = (today - datetime.date.fromisoformat(h["entry_date"])).days
        except Exception:
            days = 0
        roic_v = s.get("roic")
        fcf_v  = s.get("fcfYield")
        rg_v   = s.get("revGrowth")
        peg_v  = s.get("fwdPEG") or s.get("peg")
        line = (f"  {t} ({h.get('company','')[:20]}) | Held {days}d | "
                f"Entry ${h['entry_price']:.2f} → Now ${current_price:.2f} ({ret:+.1%}) | "
                f"ROIC={roic_v*100:.0f}% " if roic_v else f"  {t} | Held {days}d | Entry ${h['entry_price']:.2f} → Now ${current_price:.2f} ({ret:+.1%}) | ")
        if roic_v:  line += f"ROIC={roic_v*100:.0f}% "
        if fcf_v:   line += f"FCF={fcf_v:.1%} "
        if rg_v:    line += f"RG={rg_v:+.0%} "
        if peg_v:   line += f"PEG={peg_v:.2f} "
        line += f"| Thesis: {h.get('rationale','')[:80]}"
        holdings_lines.append(line)

    n_holdings = len(portfolio.get("holdings", []))
    open_slots  = PORTFOLIO_MAX_POSITIONS - n_holdings
    portfolio_value = portfolio.get("cash", PORTFOLIO_INITIAL_CASH)
    for h in portfolio.get("holdings", []):
        t = h["ticker"]
        s = stocks.get(t, {})
        current_price = s.get("price") or h["entry_price"]
        portfolio_value += h["shares"] * current_price

    holdings_block = "\n".join(holdings_lines) if holdings_lines else "  (no holdings yet — deploy capital)"

    # Sector concentration summary for PM awareness
    _sector_counts: dict = {}
    for h in portfolio.get("holdings", []):
        _sec = stocks.get(h["ticker"], {}).get("sector") or h.get("sector") or "Unknown"
        _sector_counts[_sec] = _sector_counts.get(_sec, 0) + 1
    _conc_lines = [f"  {sec}: {cnt} position{'s' if cnt > 1 else ''}{' ⚠️ AT CAP' if cnt >= 4 else ''}"
                   for sec, cnt in sorted(_sector_counts.items(), key=lambda x: -x[1])]
    sector_conc_block = "\n".join(_conc_lines) if _conc_lines else "  (no holdings)"

    # B7: Lynch-category distribution for PM awareness
    _lynch_pm_counts: dict = {}
    for h in portfolio.get("holdings", []):
        _raw_lc = stocks.get(h["ticker"], {}).get("lynchCategory") or h.get("lynch_category") or ""
        for _cat in str(_raw_lc).split("+"):
            _cat = _cat.strip()
            if _cat:
                _lynch_pm_counts[_cat] = _lynch_pm_counts.get(_cat, 0) + 1
    _LYNCH_CAPS = {"FastGrower": 4, "Stalwart": 5, "SlowGrower": 6, "Cyclical": 2, "Turnaround": 3, "AssetPlay": 4}
    _lynch_pm_lines = []
    for _cat, _cnt in sorted(_lynch_pm_counts.items(), key=lambda x: -x[1]):
        _cap = _LYNCH_CAPS.get(_cat, 5)
        _warn = f" ⚠️ AT LYNCH CAP (max {_cap})" if _cnt >= _cap else ""
        _lynch_pm_lines.append(f"  {_cat}: {_cnt}{_warn}")
    lynch_dist_block = "\n".join(_lynch_pm_lines) if _lynch_pm_lines else "  (no Lynch categories assigned)"

    # Recent transactions (last 10) for context
    recent_txns = portfolio.get("transactions", [])[-10:]
    txn_lines = [f"  {tx['date']} {tx['action']} {tx['ticker']} @ ${tx['price']:.2f} — {tx.get('rationale','')[:60]}"
                 for tx in recent_txns]
    txn_block = "\n".join(txn_lines) if txn_lines else "  (no transactions yet)"

    sys_prompt = f"""You are a patient, long-horizon portfolio manager running a paper portfolio of US equities.
Today is {today}. Portfolio started: {portfolio.get('started', today.isoformat())}.

PORTFOLIO RULES:
- Maximum {PORTFOLIO_MAX_POSITIONS} positions, roughly equal-weight (~${PORTFOLIO_TARGET_POSITION:,.0f} per position)
- Time horizon: 1-3 years. You are NOT a trader.
- Hold bias: do NOT sell unless the thesis is clearly broken or a dramatically superior opportunity exists
- A position held <60 days needs strong justification to sell — short-term volatility is NOT a reason
- Every SELL must state the exact reason the original thesis is broken
- Every BUY must have: ROIC > 12%, positive FCF, PEG < 2.5, and a clear moat or catalyst
- Diversify across sectors and strategies — hard cap of 4 positions per sector (Technology included)
- If existing holdings already have 4 in one sector, DO NOT buy another in that sector regardless of quality
- Lynch category caps (B7): ≤4 FastGrowers, ≤2 Cyclicals, ≤3 Turnarounds, ≤5 Stalwarts, ≤6 SlowGrowers, ≤4 AssetPlays at any time. At-cap categories are marked ⚠️ — do NOT add another without first selling one of the same category.
- Cash not deployed is drag — fill open slots if quality candidates exist

CURRENT PORTFOLIO STATE:
Cash available: ${portfolio.get('cash', PORTFOLIO_INITIAL_CASH):,.2f}
Holdings ({n_holdings}/{PORTFOLIO_MAX_POSITIONS}):
{holdings_block}

SECTOR CONCENTRATION (current holdings):
{sector_conc_block}

LYNCH CATEGORY DISTRIBUTION (current holdings):
{lynch_dist_block}

RECENT TRANSACTIONS:
{txn_block}"""

    usr_prompt = f"""SECTOR CONTEXT:
{sector_block}

CANDIDATE STOCKS (quality-ranked pool):
{candidates_block}

YOUR TASK:
1. Review each current holding: HOLD or SELL (with explicit thesis-broken reasoning)
2. Propose BUYs for open slots (currently {open_slots} open) — only if a genuinely compelling candidate exists
3. If a current holding has badly deteriorated AND a clearly superior replacement exists, you may propose a SWAP (SELL + BUY)

Respond ONLY with valid JSON (no markdown):
{{
  "portfolio_thesis": "2-3 sentences: your overall portfolio positioning and market view",
  "review": [
    {{
      "ticker": "QLYS",
      "decision": "HOLD",
      "rationale": "Thesis intact: ROIC 28%, revenue re-accelerating, PEG 1.6 still fair for quality"
    }},
    {{
      "ticker": "XYZ",
      "decision": "SELL",
      "rationale": "Thesis broken: revenue declined 3 consecutive quarters — no longer a quality compounder"
    }}
  ],
  "buys": [
    {{
      "ticker": "HIG",
      "company": "The Hartford Financial",
      "rationale": "Best-in-class insurer at trough valuation; rate environment is a tailwind; ROIC 28%",
      "conviction": "HIGH",
      "sell_trigger": "Combined ratio rises above 100% for 2 consecutive quarters"
    }}
  ]
}}"""

    try:
        import requests as _req
        headers = {
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        body = {
            "model": "claude-haiku-4-5",
            "max_tokens": 2500,
            "system": sys_prompt,
            "messages": [{"role": "user", "content": usr_prompt}],
        }
        resp = _req.post("https://api.anthropic.com/v1/messages",
                         headers=headers, json=body, timeout=120)
        if resp.status_code == 200:
            raw = resp.json()["content"][0]["text"]
            # Strip markdown fences if present
            raw = raw.strip()
            if raw.startswith("```"):
                raw = "\n".join(raw.split("\n")[1:])
            if raw.endswith("```"):
                raw = "\n".join(raw.split("\n")[:-1])
            # Extract just the JSON object — discard any trailing text Haiku
            # appends after the closing brace ("Extra data" parse error)
            start = raw.find("{")
            end   = raw.rfind("}") + 1
            if start != -1 and end > start:
                raw = raw[start:end]
            result = json.loads(raw)
            print(f"  ✅ Portfolio manager done — "
                  f"{len(result.get('review',[]))} holdings reviewed, "
                  f"{len(result.get('buys',[]))} buys proposed")
            return result
        else:
            print(f"  ⚠️ Portfolio manager HTTP {resp.status_code}: {resp.text[:200]}")
    except Exception as e:
        print(f"  ⚠️ Portfolio manager error: {str(e)[:80]}")
    return {}


def check_exit_rules(holding: dict, stock: dict) -> "str | None":
    """C2: Three automatic hard-exit rules evaluated BEFORE the PM runs.
    Returns a rationale string if the position should be force-sold, else None.

    Per-Lynch-category stop thresholds (plan R8):
      AssetPlay / Turnaround: -40%
      FastGrower:             -25%
      Stalwart / default:     -20%

    Rules:
    1. Relative weakness: down > threshold AND sector performance > -15%
       (indicates stock-specific failure, not macro)
    2. Thesis stale: held > 365d AND ROIC < 8% AND FCF yield < 0
    3. Revenue deceleration: current rev growth < prior-year rev growth - 20pp
    """
    t       = holding.get("ticker", "")
    entry   = float(holding.get("entry_price") or 0)
    curr    = stock.get("price") or entry
    ret     = ((curr - entry) / entry) if entry > 0 else 0

    try:
        days = (datetime.date.today() -
                datetime.date.fromisoformat(holding.get("entry_date", "2000-01-01"))).days
    except Exception:
        days = 0

    lynch_cat = stock.get("lynchCategory") or holding.get("lynch_category") or ""

    # ── Per-category stop threshold ──────────────────────────────────────
    if "AssetPlay" in lynch_cat or "Turnaround" in lynch_cat:
        stop_threshold = -0.40
    elif "FastGrower" in lynch_cat:
        stop_threshold = -0.25
    else:
        stop_threshold = -0.20   # Stalwart, SlowGrower, Cyclical, unknown

    # ── Rule 1: Relative weakness ────────────────────────────────────────
    # 52-week return used as proxy for absolute stock performance
    year_high = stock.get("yearHigh") or 0
    year_low  = stock.get("yearLow")  or 0
    if ret <= stop_threshold and ret < -0.05:
        # Sector context: if sector up or flat, the weakness is stock-specific
        # We use sector P/E trend as a rough proxy — but we don't have easy sector
        # return data here. Use a simplified rule: if down > threshold AND position
        # held > 60 days (not just a temporary dip in a new position)
        if days >= 60:
            return (f"C2-Rule1: Relative weakness — down {ret:+.1%} "
                    f"(threshold {stop_threshold:.0%} for {lynch_cat or 'default'} category), "
                    f"held {days}d")

    # ── Rule 2: Thesis stale ─────────────────────────────────────────────
    roic    = stock.get("roic")    # decimal, e.g. 0.08 = 8%
    fcf_y   = stock.get("fcfYield")  # decimal
    if days > 365:
        if roic is not None and roic < 0.08:
            if fcf_y is not None and fcf_y < 0:
                return (f"C2-Rule2: Thesis stale — held {days}d, "
                        f"ROIC={roic*100:.1f}% (<8%), FCF yield={fcf_y:.1%} (negative)")

    # ── Rule 3: Revenue deceleration ─────────────────────────────────────
    rev_g_curr  = stock.get("revGrowth")     # TTM or most-recent annual
    rev_g_prior = stock.get("revGrowthPrev") # prior-year rev growth (assembled at line ~2082)
    if (rev_g_curr is not None and rev_g_prior is not None and
            (rev_g_curr - rev_g_prior) < -0.20):   # deceleration ≥ 20pp
        return (f"C2-Rule3: Revenue deceleration — "
                f"current RG={rev_g_curr:+.1%} vs prior={rev_g_prior:+.1%} "
                f"(Δ {rev_g_curr - rev_g_prior:+.1%})")

    return None


def compute_position_size(ticker: str, price: float, portfolio: dict) -> float:
    """C1: Vol-scaled position sizing.
    Target: risk 1% of portfolio value per position using 60-day realised vol.
    Result is clamped between 0.5× and 2× the equal-weight baseline.
    Falls back to PORTFOLIO_TARGET_POSITION on any data failure.
    """
    if price <= 0:
        return PORTFOLIO_TARGET_POSITION

    # Portfolio current total value
    portfolio_value = portfolio.get("cash", PORTFOLIO_INITIAL_CASH)
    for h in portfolio.get("holdings", []):
        portfolio_value += h.get("shares", 0) * h.get("entry_price", 0)

    equal_weight = PORTFOLIO_TARGET_POSITION
    risk_budget   = portfolio_value * 0.01  # 1% portfolio risk target per position

    try:
        # Populate cache via fetch_price_on_date (side-effect: fills _hist_price_cache)
        if ticker not in _hist_price_cache:
            recent = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
            fetch_price_on_date(ticker, recent)   # triggers cache fetch

        prices_dict = _hist_price_cache.get(ticker, {})
        if len(prices_dict) < 10:
            return equal_weight

        _sorted_prices = [v for _, v in sorted(prices_dict.items())][-63:]  # up to 63 trading days
        if len(_sorted_prices) < 5:
            return equal_weight

        # Daily log returns
        log_rets = []
        for i in range(1, len(_sorted_prices)):
            p0, p1 = _sorted_prices[i-1], _sorted_prices[i]
            if p0 > 0 and p1 > 0:
                import math
                log_rets.append(math.log(p1 / p0))

        if len(log_rets) < 5:
            return equal_weight

        avg_lr = sum(log_rets) / len(log_rets)
        daily_vol = (sum((r - avg_lr)**2 for r in log_rets) / (len(log_rets) - 1)) ** 0.5
        if daily_vol <= 0:
            return equal_weight

        # Annualised vol → daily dollar risk per share
        # Position size = risk_budget / (price * daily_vol * √1)
        # (1-day 1σ move in dollar terms per share = price * daily_vol)
        vol_sized = risk_budget / (price * daily_vol)

        # Clamp to [0.5×, 2×] equal-weight
        lo = equal_weight * 0.5
        hi = equal_weight * 2.0
        return max(lo, min(hi, vol_sized))

    except Exception:
        return equal_weight



def apply_portfolio_decisions(portfolio: dict, decisions: dict, stocks: dict) -> dict:
    """Apply PM decisions to portfolio: execute sells then buys. Returns updated portfolio."""
    if not decisions:
        return portfolio
    today = datetime.date.today().isoformat()
    holdings_by_ticker = {h["ticker"]: h for h in portfolio.get("holdings", [])}

    # ── Backfill lynch_category / sector for legacy holdings (Fix-9b) ──
    for h in portfolio.get("holdings", []):
        t = h["ticker"]
        s = stocks.get(t, {})
        if not h.get("lynch_category") and s.get("lynchCategory"):
            h["lynch_category"] = s["lynchCategory"]
        if not h.get("sector") and s.get("sector"):
            h["sector"] = s["sector"]

    # ── C2: Hard exit rules (fire BEFORE PM review) ───────────────────
    forced_sells = {}
    for h in list(portfolio.get("holdings", [])):
        t = h["ticker"]
        s = stocks.get(t, {})
        reason = check_exit_rules(h, s)
        if reason:
            forced_sells[t] = reason
            print(f"    🚨 FORCE-SELL {t}: {reason[:80]}")

    # Merge forced sells into decisions["review"] (add or upgrade to SELL)
    existing_review_tickers = {r.get("ticker","").upper() for r in decisions.get("review", [])}
    for t_fs, reason_fs in forced_sells.items():
        if t_fs not in existing_review_tickers:
            decisions.setdefault("review", []).append({
                "ticker":   t_fs,
                "decision": "SELL",
                "rationale": reason_fs,
            })
        else:
            # Override existing HOLD/SELL with force-sell reasoning
            for r in decisions["review"]:
                if r.get("ticker","").upper() == t_fs:
                    r["decision"]  = "SELL"
                    r["rationale"] = reason_fs
                    break

    # ── Process SELLs ────────────────────────────────────────────────
    for rev in decisions.get("review", []):
        if rev.get("decision") != "SELL":
            continue
        t = rev.get("ticker", "").upper()
        if t not in holdings_by_ticker:
            continue
        h = holdings_by_ticker[t]
        s = stocks.get(t, {})
        sell_price = s.get("price") or h["entry_price"]
        gross_proceeds = h["shares"] * sell_price
        # C3: apply slippage (you get slightly less on a sell) and commission
        tc_sell = round(gross_proceeds * PORTFOLIO_SLIPPAGE_RATE + PORTFOLIO_COMMISSION, 4)
        proceeds = gross_proceeds - tc_sell
        ret = (sell_price - h["entry_price"]) / h["entry_price"] if h["entry_price"] > 0 else 0
        portfolio["cash"] = portfolio.get("cash", 0) + proceeds
        portfolio["transactions"].append({
            "date": today,
            "action": "SELL",
            "ticker": t,
            "company": h.get("company", ""),
            "shares": h["shares"],
            "price": round(sell_price, 2),
            "value": round(proceeds, 2),
            "return_pct": round(ret * 100, 2),
            "transaction_cost": round(tc_sell, 2),  # C3
            "rationale": rev.get("rationale", ""),
        })
        del holdings_by_ticker[t]
        print(f"    💰 SELL {t} @ ${sell_price:.2f} ({ret:+.1%}) [tc=${tc_sell:.2f}] — {rev.get('rationale','')[:60]}")

    # ── Process BUYs ─────────────────────────────────────────────────
    # C1-Fix: pre-populate _hist_price_cache for all buy candidates in one pass
    # so compute_position_size doesn't fire individual API calls inside the loop
    _buy_tickers = [b.get("ticker","").upper() for b in decisions.get("buys", [])
                    if b.get("ticker","") and b.get("ticker","").upper() not in _hist_price_cache]
    if _buy_tickers:
        _recent = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
        for _bt in _buy_tickers:
            fetch_price_on_date(_bt, _recent)   # fills _hist_price_cache[_bt] once

    for buy in decisions.get("buys", []):
        t = buy.get("ticker", "").upper()
        if not t or t in holdings_by_ticker:
            continue  # already held
        if len(holdings_by_ticker) >= PORTFOLIO_MAX_POSITIONS:
            break
        s = stocks.get(t, {})
        price = s.get("price")
        if not price or price <= 0:
            continue
        # C1: vol-scaled position size, clamped to [0.5×, 2×] equal-weight
        invest = min(compute_position_size(t, price, portfolio), portfolio.get("cash", 0))
        if invest < price:
            continue  # can't afford even 1 share
        shares = max(1, int(invest / price))
        cost = shares * price
        # C3: apply slippage (you pay slightly more on a buy) and commission
        tc_buy = round(cost * PORTFOLIO_SLIPPAGE_RATE + PORTFOLIO_COMMISSION, 4)
        total_cost = cost + tc_buy
        if total_cost > portfolio.get("cash", 0):
            continue  # insufficient cash including transaction cost
        portfolio["cash"] = portfolio.get("cash", 0) - total_cost
        holdings_by_ticker[t] = {
            "ticker": t,
            "company": buy.get("company", s.get("name", t))[:30],
            "shares": shares,
            "entry_price": round(price, 2),
            "entry_date": today,
            "rationale": buy.get("rationale", "")[:200],
            "conviction": buy.get("conviction", "MEDIUM"),
            "sell_trigger": buy.get("sell_trigger", "")[:150],
            "lynch_category": s.get("lynchCategory", ""),   # Fix-9: persist for C2 round-trip
            "sector": s.get("sector", ""),                   # Fix-9: persist sector for B7 round-trip
        }
        portfolio["transactions"].append({
            "date": today,
            "action": "BUY",
            "ticker": t,
            "company": buy.get("company", s.get("name", t))[:30],
            "shares": shares,
            "price": round(price, 2),
            "value": round(cost, 2),
            "return_pct": 0.0,
            "transaction_cost": round(tc_buy, 2),  # C3
            "rationale": buy.get("rationale", "")[:200],
        })
        print(f"    🛒 BUY  {t} {shares}sh @ ${price:.2f} (${cost:,.0f}) [tc=${tc_buy:.2f}] — {buy.get('rationale','')[:60]}")

    portfolio["holdings"] = list(holdings_by_ticker.values())
    portfolio["last_updated"] = today
    return portfolio


def build_portfolio_tab(wb, portfolio: dict, stocks: dict, spy_prices: dict = None,
                        spy_today: float = None, ws=None):
    """Tab 1c: AI Portfolio Manager — persistent paper portfolio with full trade log."""
    print("\n📊 Building Tab: AI Portfolio...")
    if ws is None:
        ws = wb.create_sheet("1c. AI Portfolio")
    ws.sheet_view.showGridLines = False

    today = datetime.date.today()
    holdings = portfolio.get("holdings", [])
    transactions = portfolio.get("transactions", [])
    cash = portfolio.get("cash", PORTFOLIO_INITIAL_CASH)
    initial = portfolio.get("initial_cash", PORTFOLIO_INITIAL_CASH)

    # ── Compute portfolio value ──────────────────────────────────────
    total_mkt = 0.0
    for h in holdings:
        t = h["ticker"]
        s = stocks.get(t, {})
        cp = s.get("price") or h["entry_price"]
        total_mkt += h["shares"] * cp
    nav = cash + total_mkt
    total_ret = (nav - initial) / initial if initial > 0 else 0

    # SPY return since portfolio started
    spy_ret_overall = None
    started = portfolio.get("started")
    if started and spy_prices and spy_today:
        for delta in range(8):
            d = (datetime.date.fromisoformat(started) - datetime.timedelta(days=delta)).isoformat()
            if d in spy_prices and spy_prices[d] > 0:
                spy_ret_overall = (spy_today - spy_prices[d]) / spy_prices[d]
                break

    # ── Column widths ────────────────────────────────────────────────
    col_widths = [8, 9, 25, 14, 8, 9, 9, 9, 9, 9, 8, 8, 50, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    TC = len(col_widths)

    def _hdr(row, txt, color, size=9, height=18):
        c = ws.cell(row=row, column=1, value=txt)
        c.font = Font(bold=True, name="Arial", size=size, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TC)
        ws.row_dimensions[row].height = height
        return row + 1

    r = 1
    r = _hdr(r, f"  💼  AI Portfolio Manager  —  Paper Portfolio  |  {today}", "0D1B2A", size=12, height=28)

    # ── Stats bar ────────────────────────────────────────────────────
    stat_items = [
        ("NAV", f"${nav:,.0f}"),
        ("Cash", f"${cash:,.0f}"),
        ("Invested", f"${total_mkt:,.0f}"),
        ("Total Return", f"{total_ret:+.1%}"),
        ("SPY (same period)", f"{spy_ret_overall:+.1%}" if spy_ret_overall is not None else "n/a"),
        ("Alpha", f"{(total_ret - spy_ret_overall):+.1%}" if spy_ret_overall is not None else "n/a"),
        ("Positions", f"{len(holdings)}/{PORTFOLIO_MAX_POSITIONS}"),
        ("Started", portfolio.get("started", "—")),
    ]
    for ci, (lbl, val) in enumerate(stat_items, 1):
        lc = ws.cell(row=r, column=ci*2-1, value=lbl)
        lc.font = Font(bold=True, name="Arial", size=8, color="AAAAAA")
        lc.fill = PatternFill("solid", fgColor="0D1B2A")
        lc.alignment = Alignment(horizontal="right")
        vc = ws.cell(row=r, column=ci*2, value=val)
        vc.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        vc.fill = PatternFill("solid", fgColor="0D1B2A")
    ws.row_dimensions[r].height = 16
    r += 1

    # ── Holdings table ───────────────────────────────────────────────
    r = _hdr(r, "  📊  Current Holdings", "1A237E")
    hdrs = ["Entry Date", "Ticker", "Company", "Sector", "Shares",
            "Entry $", "Current $", "Return", "SPY Ret", "Rel Ret",
            "Days", "Conviction", "Rationale", "Sell Trigger"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A237E")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[r].height = 26
    r += 1

    for ri, h in enumerate(holdings):
        t = h["ticker"]
        s = stocks.get(t, {})
        cp = s.get("price") or h["entry_price"]
        ep = h["entry_price"]
        ret_h = (cp - ep) / ep if ep > 0 else None
        try:
            days_h = (today - datetime.date.fromisoformat(h["entry_date"])).days
        except Exception:
            days_h = None
        # SPY for same period
        spy_h_entry = None
        if spy_prices and h.get("entry_date"):
            for delta in range(8):
                d = (datetime.date.fromisoformat(h["entry_date"]) - datetime.timedelta(days=delta)).isoformat()
                if d in spy_prices and spy_prices[d] > 0:
                    spy_h_entry = spy_prices[d]
                    break
        spy_h_ret = ((spy_today - spy_h_entry) / spy_h_entry
                     if spy_h_entry and spy_today else None)
        rel_h = (ret_h - spy_h_ret) if ret_h is not None and spy_h_ret is not None else None

        fill = ALT_FILL if ri % 2 == 0 else PLAIN_FILL
        row_vals = [h.get("entry_date",""), t, h.get("company","")[:24],
                    s.get("sector","")[:13], h["shares"],
                    ep, round(cp, 2), ret_h, spy_h_ret, rel_h,
                    days_h, h.get("conviction",""),
                    h.get("rationale","")[:120], h.get("sell_trigger","")[:80]]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(name="Arial", size=9)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left" if ci >= 3 else "center",
                                    vertical="center", wrap_text=(ci >= 13))
            if hdrs[ci-1] in ("Entry $", "Current $"):
                c.number_format = "$#,##0.00"
            elif hdrs[ci-1] in ("Return", "SPY Ret"):
                c.number_format = "0.0%"
                if isinstance(v, float):
                    c.font = Font(name="Arial", size=9,
                                  color="1B5E20" if v > 0 else "B71C1C")
            elif hdrs[ci-1] == "Rel Ret":
                c.number_format = "0.0%"
                if isinstance(v, float):
                    c.fill = PatternFill("solid", fgColor="C8E6C9" if v > 0 else "FFCDD2")
                    c.font = Font(bold=True, name="Arial", size=9,
                                  color="1B5E20" if v > 0 else "B71C1C")
        ws.row_dimensions[r].height = 15
        r += 1

    if not holdings:
        ws.cell(row=r, column=1,
                value="No holdings yet — portfolio will deploy capital on next run."
                ).font = Font(italic=True, name="Arial", size=9, color="888888")
        r += 1

    # ── Transaction log ──────────────────────────────────────────────
    r += 1
    r = _hdr(r, "  📋  Transaction Log  (all buys & sells)", "263238")
    txn_hdrs = ["Date", "Action", "Ticker", "Company", "Shares", "Price", "Value", "Return", "Rationale"]
    txn_widths_map = [10, 8, 8, 22, 7, 9, 10, 8, 60]
    for ci, (h, w) in enumerate(zip(txn_hdrs, txn_widths_map), 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="263238")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 20
    r += 1

    for ri, tx in enumerate(reversed(transactions)):  # newest first
        fill = ALT_FILL if ri % 2 == 0 else PLAIN_FILL
        is_buy = tx.get("action") == "BUY"
        action_fill = PatternFill("solid", fgColor="C8E6C9" if is_buy else "FFCDD2")
        row_vals = [tx.get("date",""), tx.get("action",""), tx.get("ticker",""),
                    tx.get("company","")[:22], tx.get("shares"),
                    tx.get("price"), tx.get("value"),
                    (tx.get("return_pct", 0) / 100) if tx.get("return_pct") is not None else None,
                    tx.get("rationale","")[:120]]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(name="Arial", size=9)
            c.fill = action_fill if ci == 2 else fill
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left" if ci >= 4 else "center",
                                    vertical="center", wrap_text=(ci == 9))
            if txn_hdrs[ci-1] in ("Price", "Value"):
                c.number_format = "$#,##0.00"
            elif txn_hdrs[ci-1] == "Return" and isinstance(v, float):
                c.number_format = "0.0%"
                c.font = Font(name="Arial", size=9,
                              color="1B5E20" if (v or 0) >= 0 else "B71C1C")
        ws.row_dimensions[r].height = 15
        r += 1

    print(f"  ✅ Portfolio tab done — {len(holdings)} holdings, {len(transactions)} transactions")
    return nav, total_ret, spy_ret_overall


def compute_agent_performance(spy_prices: dict = None, spy_today: float = None) -> dict:
    """B1: Read fmp_ai_picks_log.csv and compute per-agent attribution metrics.

    Returns a dict keyed by source string (e.g. 'AI-QualityGrowth', 'AI-Judge'):
        n_picks        : int   — total logged picks
        avg_ret        : float — mean current return (entry→today)
        avg_spy        : float — mean SPY return over same holding periods
        alpha          : float — avg_ret - avg_spy
        sharpe         : float — annualised Sharpe on alpha series (None if <3 picks)
        win_rate       : float — fraction with ret > 0
        hit_30d        : float — fraction where Ret30d > 0  (picks ≥30 days old only)
        hit_90d        : float — fraction where Ret90d > 0  (picks ≥90 days old only)
        hit_180d       : float — fraction where Ret180d > 0 (picks ≥180 days old only)
        best_ticker    : str
        best_ret       : float
        worst_ticker   : str
        worst_ret      : float
        med_hold_days  : float — median holding period in days
        prompt_versions: set   — all prompt_version values seen (for filtering)
    """
    if not os.path.exists(AI_PICKS_LOG):
        return {}

    import statistics as _stats

    # --- helpers -----------------------------------------------------------
    def _spy_ret_for_date(entry_date_str: str) -> float | None:
        """SPY return from entry_date to today."""
        if not spy_prices or not spy_today:
            return None
        try:
            target = datetime.date.fromisoformat(entry_date_str)
        except ValueError:
            return None
        for delta in range(8):
            d = (target - datetime.timedelta(days=delta)).isoformat()
            if d in spy_prices and spy_prices[d] > 0:
                return (spy_today - spy_prices[d]) / spy_prices[d]
        return None

    # --- read log ----------------------------------------------------------
    rows = []
    try:
        with open(AI_PICKS_LOG, "r", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                rows.append(r)
    except Exception:
        return {}

    # --- gather current prices for unique tickers -------------------------
    unique_t = list({r["ticker"] for r in rows if r.get("ticker")})
    live_px: dict = {}
    for t in unique_t:
        p = fetch_live_price(t)
        if p:
            live_px[t] = p
        time.sleep(0.05)

    # --- aggregate by agent -----------------------------------------------
    _per: dict = {}   # source → lists of per-pick stats

    for r in rows:
        src = r.get("source", "unknown")
        if not src:
            continue
        t = r.get("ticker", "")
        try:
            entry = float(r.get("entry_price") or 0)
        except (ValueError, TypeError):
            entry = 0
        curr = live_px.get(t, 0)
        ret = ((curr - entry) / entry) if entry > 0 and curr > 0 else None

        date_str = r.get("date", "")
        try:
            days = (datetime.date.today() - datetime.date.fromisoformat(date_str)).days
        except Exception:
            days = None

        spy_r = _spy_ret_for_date(date_str)
        alpha_r = (ret - spy_r) if ret is not None and spy_r is not None else ret

        # B2 checkpoint rets (lazy fetch via fetch_price_on_date cache)
        ret_30d  = _checkpoint_ret_b1(t, entry, date_str, 30,  days)
        ret_90d  = _checkpoint_ret_b1(t, entry, date_str, 90,  days)
        ret_180d = _checkpoint_ret_b1(t, entry, date_str, 180, days)

        if src not in _per:
            _per[src] = {"rets": [], "alphas": [], "spy_rets": [], "days_list": [],
                         "wins": 0, "n": 0,
                         "hits_30": [], "hits_90": [], "hits_180": [],
                         "best_ret": None, "best_t": "—",
                         "worst_ret": None, "worst_t": "—",
                         "pv_set": set()}
        ag = _per[src]
        ag["n"] += 1
        ag["pv_set"].add(r.get("prompt_version", ""))

        if ret is not None:
            ag["rets"].append(ret)
            if ret > 0:
                ag["wins"] += 1
            if ag["best_ret"] is None or ret > ag["best_ret"]:
                ag["best_ret"] = ret; ag["best_t"] = t
            if ag["worst_ret"] is None or ret < ag["worst_ret"]:
                ag["worst_ret"] = ret; ag["worst_t"] = t
        if alpha_r is not None:
            ag["alphas"].append(alpha_r)
        if spy_r is not None:
            ag["spy_rets"].append(spy_r)
        if days is not None:
            ag["days_list"].append(days)
        if ret_30d is not None:
            ag["hits_30"].append(1 if ret_30d > 0 else 0)
        if ret_90d is not None:
            ag["hits_90"].append(1 if ret_90d > 0 else 0)
        if ret_180d is not None:
            ag["hits_180"].append(1 if ret_180d > 0 else 0)

    # --- compute summary stats per agent ----------------------------------
    out = {}
    for src, ag in _per.items():
        rets    = ag["rets"]
        alphas  = ag["alphas"]
        dl      = ag["days_list"]
        n       = ag["n"]

        avg_ret  = sum(rets)   / len(rets)   if rets   else None
        avg_spy  = sum(ag["spy_rets"]) / len(ag["spy_rets"]) if ag["spy_rets"] else None
        alpha    = sum(alphas) / len(alphas) if alphas else None
        win_rate = ag["wins"] / len(rets)    if rets   else None

        # Annualised Sharpe on alpha series
        sharpe = None
        if len(alphas) >= 3:
            try:
                _std = _stats.stdev(alphas)
                if _std > 0:
                    avg_days = sum(dl) / len(dl) if dl else 30
                    sharpe = (alpha / _std) * ((252 / max(avg_days, 1)) ** 0.5)
            except Exception:
                pass

        out[src] = {
            "n_picks":       n,
            "avg_ret":       avg_ret,
            "avg_spy":       avg_spy,
            "alpha":         alpha,
            "sharpe":        sharpe,
            "win_rate":      win_rate,
            "hit_30d":       (sum(ag["hits_30"])  / len(ag["hits_30"]))  if ag["hits_30"]  else None,
            "hit_90d":       (sum(ag["hits_90"])  / len(ag["hits_90"]))  if ag["hits_90"]  else None,
            "hit_180d":      (sum(ag["hits_180"]) / len(ag["hits_180"])) if ag["hits_180"] else None,
            "best_ticker":   ag["best_t"],
            "best_ret":      ag["best_ret"],
            "worst_ticker":  ag["worst_t"],
            "worst_ret":     ag["worst_ret"],
            "med_hold_days": (_stats.median(dl) if dl else None),
            "prompt_versions": ag["pv_set"],
        }
    return out


def _checkpoint_ret_b1(ticker: str, entry: float, entry_date_str: str,
                        horizon_days: int, hold_days) -> float | None:
    """Thin wrapper used by both build_picks_tracking._build_row and compute_agent_performance.

    C5: If the ticker has been fetched but the cache is empty (no price history),
    the stock is presumed delisted. Returns _DELISTED_RETURN (-100%) and adds
    ticker to _delisted_tickers for visual flagging downstream.
    Returns None when the horizon hasn't elapsed yet.
    """
    global _delisted_tickers
    if entry <= 0 or hold_days is None or hold_days < horizon_days:
        return None
    try:
        entry_dt = datetime.date.fromisoformat(entry_date_str)
    except ValueError:
        return None
    target_date = (entry_dt + datetime.timedelta(days=horizon_days)).isoformat()
    price_at_horizon = fetch_price_on_date(ticker, target_date)
    if price_at_horizon and price_at_horizon > 0:
        return (price_at_horizon - entry) / entry
    # C5: If cache was populated but empty → presumed delisted
    if ticker in _hist_price_cache and not _hist_price_cache[ticker]:
        _delisted_tickers.add(ticker)
        return _DELISTED_RETURN   # conservative -100%
    return None


def build_picks_tracking(wb, stocks):
    """Tab 11: Picks Tracking — measure performance of logged picks.
    Reads fmp_picks_log.csv, fetches current prices, shows P&L.
    """
    print("\n📊 Building Tab: Picks Tracking...")
    ws = wb.create_sheet("11. Picks Tracking")
    ws.sheet_view.showGridLines = False

    sr = add_title(ws, "📈 Picks Tracking — Strategy Performance Log",
                   f"Auto-logged every run. Top 5 per strategy tracked at entry price. {datetime.date.today()}")

    if not os.path.exists(PICKS_LOG):
        ws.cell(row=sr, column=1,
                value="No picks logged yet — will appear automatically after first run.").font = \
            Font(italic=True, name="Arial", size=10, color="888888")
        print("  ℹ️ No picks log found")
        return

    # ── Read strategy picks (fmp_picks_log.csv) ────────────────────────────
    logged = []
    with open(PICKS_LOG, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            row["_kind"] = "strategy"
            row["_display_strategy"] = row.get("strategy", "")[:22]
            logged.append(row)

    # ── Read AI agent picks (fmp_ai_picks_log.csv) — if exists ─────────────
    ai_logged = []
    _agent_icons = {
        "AI-QualityGrowth":    "🌱 Qual.Growth",
        "AI-SpecialSit":       "⚡ Special Sit",
        "AI-CapAppreciation":  "📈 Cap.Apprecn",
        "AI-EmergingGrowth":   "🚀 Emerg.Growth",
        "AI-TenBagger":        "🎯 10-Bagger",
        "AI-LynchBWYK":        "🛒 Lynch BWYK",
        "AI-CathieWood":       "🚀 Disruptive",
        "AI-Pabrai":           "🎲 Pabrai",
        "AI-HowardMarks":      "🔄 Contrarian",
        "AI-Burry":            "🕳️ Deep Value",
        "AI-InsiderTrack":     "👁️ Insider",
        "AI-Judge":            "⚖️ Master Manager",
        "AI-MallManager":      "🛍️ Mall Manager",
        # legacy labels kept for old log entries
        "AI-GoldmanSC":        "🏦 Goldman SC (retired)",
        "AI-SocialArb":        "📱 Social Arb (retired)",
        "AI-Mayer100x":        "💯 100-Bagger (retired)",
        "AI-MagicFormula":     "🔢 Magic Formula (retired)",
        "AI-NickSleep":        "🌀 Scale Econ (retired)",
        "AI-WallStBlind":      "🔍 WallStBlind (retired)",
        "AI-Bull":             "🐂 Bull (legacy)",
        "AI-Value":            "🛡️ Value (legacy)",
        "AI-Contrarian":       "🔄 Contrarian (legacy)",
    }
    if os.path.exists(AI_PICKS_LOG):
        with open(AI_PICKS_LOG, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                src = row.get("source", "AI-Judge")
                row["_kind"] = "ai"
                row["_display_strategy"] = _agent_icons.get(src, src)
                row["ticker"] = row.get("ticker", "")
                row["entry_price"] = row.get("entry_price", "")
                ai_logged.append(row)

    all_logged = logged + ai_logged
    if not all_logged:
        return

    # ── Fetch LIVE prices for all unique tickers ────────────────────────────
    unique_tickers = list({r["ticker"] for r in all_logged if r.get("ticker")})
    current_prices = {}
    print(f"    Fetching live prices for {len(unique_tickers)} tracked tickers...")
    for _t in unique_tickers:
        _p = fetch_live_price(_t)
        if _p:
            current_prices[_t] = _p
        time.sleep(0.15)
    # Fallback: cached price
    for t in unique_tickers:
        if t not in current_prices:
            s = stocks.get(t, {})
            if s.get("price"):
                current_prices[t] = s["price"]

    # ── Fetch SPY benchmark history (single batch call) ────────────────────
    _all_dates = []
    for r in all_logged:
        try:
            _all_dates.append(datetime.date.fromisoformat(r["date"]))
        except Exception:
            pass
    spy_prices = {}
    spy_today  = None
    if _all_dates:
        _earliest = min(_all_dates).isoformat()
        print(f"    Fetching SPY history from {_earliest} for benchmark...")
        spy_prices = fetch_spy_history(_earliest)
        spy_today  = fetch_live_price("SPY")

    def _spy_entry_price(entry_date_str: str) -> float | None:
        """Find SPY close on or before entry date (handles weekends/holidays)."""
        if not spy_prices:
            return None
        try:
            target = datetime.date.fromisoformat(entry_date_str)
        except Exception:
            return None
        # Scan backwards up to 7 calendar days to find nearest prior trading day
        for delta in range(8):
            d = (target - datetime.timedelta(days=delta)).isoformat()
            if d in spy_prices and spy_prices[d] > 0:
                return spy_prices[d]
        return None

    # B2: delegate to module-level helper (shared with compute_agent_performance)

    # ── Build display rows ──────────────────────────────────────────────────
    def _build_row(r):
        t = r.get("ticker", "")
        try:
            entry = float(r.get("entry_price") or 0)
        except (ValueError, TypeError):
            entry = 0
        current = current_prices.get(t, 0)
        ret = ((current - entry) / entry) if entry > 0 and current > 0 else None
        try:
            days = (datetime.date.today() - datetime.date.fromisoformat(r["date"])).days
        except Exception:
            days = None
        # SPY benchmark return for the same holding period
        spy_entry = _spy_entry_price(r.get("date", ""))
        spy_ret = ((spy_today - spy_entry) / spy_entry
                   if spy_entry and spy_today and spy_entry > 0 else None)
        rel_ret = ((ret - spy_ret) if ret is not None and spy_ret is not None else None)

        # B2: Fixed-horizon checkpoint returns (only populated when old enough)
        entry_date_str = r.get("date", "")
        ret_30d  = _checkpoint_ret_b1(t, entry, entry_date_str, 30,  days)
        ret_90d  = _checkpoint_ret_b1(t, entry, entry_date_str, 90,  days)
        ret_180d = _checkpoint_ret_b1(t, entry, entry_date_str, 180, days)

        return {
            "Date Logged": entry_date_str,
            "Agent / Strategy": r.get("_display_strategy", "")[:22],
            "Ticker": t,
            "Company": r.get("company", "")[:25],
            "Entry $": round(entry, 2) if entry else None,
            "Current $": round(current, 2) if current else None,
            "Return": ret,
            "SPY Ret": spy_ret,
            "Rel Ret": rel_ret,
            "Days": days,
            "Ret 30d": ret_30d,    # B2: price 30 days after entry / entry - 1
            "Ret 90d": ret_90d,    # B2: price 90 days after entry / entry - 1
            "Ret 180d": ret_180d,  # B2: price 180 days after entry / entry - 1
            "Delisted": "⚠ DELISTED" if t in _delisted_tickers else "",  # C5
            "_kind": r.get("_kind", "strategy"),
            "_pv":    r.get("prompt_version", ""),  # B6: for backtest filtering
        }

    strat_rows = [_build_row(r) for r in logged]
    ai_rows    = [_build_row(r) for r in ai_logged]
    strat_rows.sort(key=lambda x: (x.get("Agent / Strategy", ""), x.get("Date Logged", "")))
    ai_rows.sort(key=lambda x: (x.get("Agent / Strategy", ""), x.get("Date Logged", "")))

    # B2: Added Ret 30d / Ret 90d / Ret 180d columns; C5: Delisted flag
    headers = ["Date Logged", "Agent / Strategy", "Ticker", "Company",
               "Entry $", "Current $", "Return", "SPY Ret", "Rel Ret", "Days",
               "Ret 30d", "Ret 90d", "Ret 180d", "Delisted"]
    widths  = [12, 22, 8, 25, 9, 9, 9, 9, 9, 6, 9, 9, 9, 10]

    def _write_section(title_val, title_color, row_list, start_row):
        r = start_row
        # Section title
        tc = ws.cell(row=r, column=1, value=title_val)
        tc.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        tc.fill = PatternFill("solid", fgColor=title_color)
        ws.row_dimensions[r].height = 18
        r += 1
        # Header
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1A237E")
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        r += 1
        # Rows
        for ri, row in enumerate(row_list):
            fill = ALT_FILL if ri % 2 == 0 else PLAIN_FILL
            for ci, h in enumerate(headers, 1):
                v = row.get(h)
                c = ws.cell(row=r, column=ci, value=v)
                c.font = Font(name="Arial", size=9)
                c.border = THIN_BORDER
                c.fill = fill
                c.alignment = Alignment(
                    horizontal="left" if ci in (2, 4) else "center",
                    vertical="center")
                if h in ("Entry $", "Current $") and isinstance(v, (int, float)):
                    c.number_format = "$#,##0.00"
                elif h == "Return" and isinstance(v, (int, float)):
                    c.number_format = "0.0%"
                    if v > 0.10:
                        c.fill = PatternFill("solid", fgColor="C8E6C9")
                        c.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                    elif v > 0:
                        c.fill = PatternFill("solid", fgColor="E8F5E9")
                    elif v < -0.10:
                        c.fill = PatternFill("solid", fgColor="FFCDD2")
                        c.font = Font(name="Arial", size=9, color="B71C1C")
                    elif v < 0:
                        c.fill = PatternFill("solid", fgColor="FFEBEE")
                elif h == "SPY Ret" and isinstance(v, (int, float)):
                    c.number_format = "0.0%"
                elif h == "Rel Ret" and isinstance(v, (int, float)):
                    c.number_format = "0.0%"
                    if v > 0:
                        c.fill = PatternFill("solid", fgColor="C8E6C9")
                        c.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                    elif v < 0:
                        c.fill = PatternFill("solid", fgColor="FFCDD2")
                        c.font = Font(name="Arial", size=9, color="B71C1C")
            ws.row_dimensions[r].height = 15
            r += 1
        return r

    def _write_summary(title_val, title_color, row_list, start_row):
        r = start_row + 1
        tc = ws.cell(row=r, column=1, value=title_val)
        tc.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        tc.fill = PatternFill("solid", fgColor=title_color)
        ws.row_dimensions[r].height = 18
        r += 1
        # Column headers for summary — A7 adds Sharpe + Best + Worst
        for ci, lbl in enumerate(["Agent / Strategy", "Picks", "Avg Return", "SPY Avg",
                                  "Avg Alpha", "Win Rate", "Sharpe", "Best", "Worst"], 1):
            hc = ws.cell(row=r, column=ci, value=lbl)
            hc.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor="455A64")
            hc.alignment = Alignment(horizontal="center")
        r += 1
        by_agent = {}
        for row in row_list:
            ag = row.get("Agent / Strategy", "Unknown")
            if ag not in by_agent:
                by_agent[ag] = {"returns": [], "spy_rets": [], "rel_rets": [],
                                "days": [], "best_t": None, "best_v": None,
                                "worst_t": None, "worst_v": None}
            ret = row.get("Return")
            spy_r = row.get("SPY Ret")
            rel_r = row.get("Rel Ret")
            dys  = row.get("Days")
            tkr  = row.get("Ticker")
            if ret is not None:
                by_agent[ag]["returns"].append(ret)
                # Track best and worst raw return for per-agent spotlight
                if (by_agent[ag]["best_v"] is None) or (ret > by_agent[ag]["best_v"]):
                    by_agent[ag]["best_v"] = ret
                    by_agent[ag]["best_t"] = tkr
                if (by_agent[ag]["worst_v"] is None) or (ret < by_agent[ag]["worst_v"]):
                    by_agent[ag]["worst_v"] = ret
                    by_agent[ag]["worst_t"] = tkr
            if spy_r is not None:
                by_agent[ag]["spy_rets"].append(spy_r)
            if rel_r is not None:
                by_agent[ag]["rel_rets"].append(rel_r)
            if dys is not None and dys > 0:
                by_agent[ag]["days"].append(dys)
        for ag, data in sorted(by_agent.items()):
            returns  = data["returns"]
            spy_rets = data["spy_rets"]
            rel_rets = data["rel_rets"]
            days     = data["days"]
            avg     = sum(returns)  / len(returns)  if returns  else None
            avg_spy = sum(spy_rets) / len(spy_rets) if spy_rets else None
            avg_rel = sum(rel_rets) / len(rel_rets) if rel_rets else None
            win_rate = (sum(1 for v in rel_rets if v > 0) / len(rel_rets)
                        if rel_rets else None)
            # Annualised Sharpe on alpha (rel_rets): need ≥3 picks and non-zero std
            sharpe = None
            if len(rel_rets) >= 3 and avg_rel is not None:
                _n = len(rel_rets)
                _var = sum((v - avg_rel) ** 2 for v in rel_rets) / (_n - 1)
                _std = _var ** 0.5
                if _std > 0:
                    # Convert per-pick alpha into annualised Sharpe using median hold-days
                    _avg_days = (sum(days) / len(days)) if days else 30
                    # periods per year; clamp to at least 1 to avoid div-by-zero spike
                    _periods = max(1.0, 252.0 / max(_avg_days, 1.0))
                    sharpe = (avg_rel / _std) * (_periods ** 0.5)
            ws.cell(row=r, column=1, value=ag).font = Font(bold=True, name="Arial", size=9)
            ws.cell(row=r, column=2, value=len(returns))
            avg_cell = ws.cell(row=r, column=3, value=avg if avg is not None else "no data")
            if isinstance(avg, float):
                avg_cell.number_format = "0.0%"
                avg_cell.font = Font(bold=True, name="Arial", size=9,
                                     color="1B5E20" if avg > 0 else "B71C1C")
            spy_cell = ws.cell(row=r, column=4, value=avg_spy)
            if isinstance(avg_spy, float):
                spy_cell.number_format = "0.0%"
                spy_cell.font = Font(name="Arial", size=9)
            rel_cell = ws.cell(row=r, column=5, value=avg_rel)
            if isinstance(avg_rel, float):
                rel_cell.number_format = "0.0%"
                rel_cell.font = Font(bold=True, name="Arial", size=9,
                                     color="1B5E20" if avg_rel > 0 else "B71C1C")
            wr_cell = ws.cell(row=r, column=6, value=win_rate)
            if isinstance(win_rate, float):
                wr_cell.number_format = "0%"
                wr_cell.font = Font(bold=True, name="Arial", size=9,
                                    color="1B5E20" if win_rate >= 0.50 else "B71C1C")
            # Sharpe
            sh_cell = ws.cell(row=r, column=7, value=sharpe)
            if isinstance(sharpe, float):
                sh_cell.number_format = "0.00"
                sh_cell.font = Font(bold=True, name="Arial", size=9,
                                    color="1B5E20" if sharpe > 0 else "B71C1C")
            else:
                sh_cell.value = "—"
                sh_cell.font = Font(name="Arial", size=9, color="888888")
            # Best / worst picks
            _bt = data["best_t"]; _bv = data["best_v"]
            _wt = data["worst_t"]; _wv = data["worst_v"]
            if _bt and _bv is not None:
                _b_cell = ws.cell(row=r, column=8,
                                  value=f"{_bt} {_bv*100:+.0f}%")
                _b_cell.font = Font(name="Arial", size=9, color="1B5E20")
            if _wt and _wv is not None:
                _w_cell = ws.cell(row=r, column=9,
                                  value=f"{_wt} {_wv*100:+.0f}%")
                _w_cell.font = Font(name="Arial", size=9, color="B71C1C")
            r += 1
        return r

    # ── Write strategy picks section ────────────────────────────────────────
    sr = _write_section("📊  Strategy Picks", "1A237E", strat_rows, sr)
    sr = _write_summary("Strategy Performance Summary", "263238", strat_rows, sr)

    # ── Write AI agent picks section ────────────────────────────────────────
    sr += 1
    if ai_rows:
        sr = _write_section("🤖  AI Agent Picks  (Bull · Value · Contrarian · Judge)",
                            "4A148C", ai_rows, sr)
        sr = _write_summary("AI Agent Performance Summary", "37474F", ai_rows, sr)
    else:
        ws.cell(row=sr, column=1,
                value="AI agent picks will appear here after the first multi-agent run."
                ).font = Font(italic=True, name="Arial", size=9, color="888888")

    print(f"  ✅ Picks Tracking done — {len(logged)} strategy picks, "
          f"{len(ai_logged)} AI agent picks, {len(unique_tickers)} unique tickers")


def build_quality_compounders(wb, stocks):
    """Tab 3b: Quality Compounders — Buffett's 'wonderful businesses at fair prices'.
    ROE > 15%, ROIC > 12%, consistent growth, low debt, Piotroski >= 7.
    These are businesses with durable competitive advantages (moats).
    """
    print("\n📊 Building Tab: Quality Compounders...")
    qualified = []
    for t, s in stocks.items():
        if not _is_common_stock(s): continue
        roe = s.get("roe")
        roic = s.get("roic")
        pio = s.get("piotroski")
        fcf = s.get("fcfYield")
        de = s.get("de")
        rg = s.get("revGrowth")
        pe = s.get("pe")

        # Quality gates: ROIC > 15% (Filter 1), ROE > 15%, Piotroski >= 7, FCF positive (Filter 3)
        # ROIC > ROE: ROIC measures true capital efficiency regardless of leverage
        # Real Estate excluded: land appreciation inflates ROIC; lumpy/asset-heavy; not a moat business
        # Basic Materials excluded: commodity-cycle peaks inflate ROIC/FCF for miners; not durable moats
        _sector_qc = (s.get("sector") or "")
        if _sector_qc == "Real Estate": continue
        if _sector_qc == "Basic Materials": continue
        if not roic or roic < 0.15: continue   # raised from 12% — moat signal
        if not roe or roe < 0.15: continue
        if not pio or pio < 7: continue
        if not fcf or fcf <= 0: continue        # Filter 3: FCF must be positive
        if not s.get("mktCap", 0) > 1e9: continue
        if de is not None and de > 3.0: continue  # not excessively leveraged

        # Kill: RevConsistency floor — compounders must show steady revenue
        rc_kill = s.get("revConsistency")
        if rc_kill is not None and rc_kill < 0.60:
            continue

        # Kill: Declining EPS projection — not a compounder if earnings expected to shrink
        # Override: ROIC > 25% is exception (genuine moat may face temporary headwind)
        eg5_kill = s.get("epsGrowth5y")
        if eg5_kill is not None and eg5_kill < -0.03 and not (roic and roic > 0.25):
            continue

        # Score: ROIC weighted MORE than ROE — avoids rewarding leverage-inflated returns
        sc = 0
        # ROIC is the primary capital efficiency signal (Filter 1)
        if roic > 0.30:      sc += 18
        elif roic > 0.25:    sc += 14
        elif roic > 0.20:    sc += 10
        elif roic > 0.15:    sc += 5

        # ROE is secondary — can be inflated by debt, so lower weight than ROIC
        if roe > 0.30:       sc += 8
        elif roe > 0.25:     sc += 6
        elif roe > 0.20:     sc += 4
        elif roe > 0.15:     sc += 2

        if pio >= 9:         sc += 10
        elif pio >= 8:       sc += 7
        elif pio >= 7:       sc += 4

        if fcf > 0.08:       sc += 8
        elif fcf > 0.05:     sc += 5
        elif fcf > 0.02:     sc += 2

        best_peg_qc = s.get("fwdPEG") or s.get("peg")
        _fin_qc       = "financial" in (s.get("sector") or "").lower()
        _basic_mat_qc = "basic materials" in (s.get("sector") or "").lower()
        _energy_qc    = (_sector_qc) == "Energy"
        peg_sc_qc = 0
        if best_peg_qc and 0 < best_peg_qc < 1.0:   peg_sc_qc = 10
        elif best_peg_qc and 0 < best_peg_qc < 1.5: peg_sc_qc = 6
        elif best_peg_qc and 0 < best_peg_qc < 2.5: peg_sc_qc = 3
        # Quality bonus: ROIC+FCF validates the PEG is moat-backed
        if peg_sc_qc and roic and roic > 0.15 and fcf and fcf > 0.04:
            peg_sc_qc = min(10, round(peg_sc_qc * 1.25))
        # Deceleration penalty
        _rg_qc = rg or 0; _rg_p_qc = s.get("revGrowthPrev")
        if peg_sc_qc and _rg_p_qc is not None and (_rg_qc - _rg_p_qc) < -0.10:
            peg_sc_qc = round(peg_sc_qc * 0.60)
        if peg_sc_qc and _fin_qc and best_peg_qc and best_peg_qc < 1.5:
            peg_sc_qc = round(peg_sc_qc * 0.50)
        if peg_sc_qc and _basic_mat_qc and best_peg_qc and best_peg_qc < 1.5:
            peg_sc_qc = round(peg_sc_qc * 0.60)  # commodity miners: spot-price cycle inflates apparent cheapness
        if peg_sc_qc and _energy_qc and best_peg_qc and best_peg_qc < 1.5:
            peg_sc_qc = round(peg_sc_qc * 0.60)  # E&P: commodity-driven earnings distort PEG reliability
        sc += peg_sc_qc

        if rg and rg > 0.10: sc += 5
        elif rg and rg > 0.05: sc += 2

        if s.get("mos") and s.get("mos") > 0.2: sc += 5

        # Revenue consistency bonus
        rc_qc = s.get("revConsistency")
        if rc_qc is not None:
            if rc_qc >= 0.80:   sc += 6
            elif rc_qc >= 0.60: sc += 2
            elif rc_qc < 0.40:  sc -= 4
        # Share buybacks as capital allocation quality signal
        sg_qc = s.get("sharesGrowth")
        if sg_qc is not None:
            if sg_qc < -0.03:   sc += 5
            elif sg_qc < 0:     sc += 2
            elif sg_qc > 0.05:  sc -= 4
        # Filter 3: FCF conversion — FCF ≈ Net Income = no accounting tricks
        fcc_qc = s.get("fcfConversion")
        if fcc_qc is not None:
            if fcc_qc >= 0.80:  sc += 8   # FCF quality confirmed — earnings are real cash
            elif fcc_qc >= 0.60: sc += 4
            elif fcc_qc < 0.40:  sc -= 6  # FCF << earnings: accruals, capex drain, or tricks
        # FCF Margin: high FCF margin = scalable, asset-light business
        fcm_qc = s.get("fcfMargin")
        if fcm_qc is not None:
            if fcm_qc >= 0.20:   sc += 8
            elif fcm_qc >= 0.12: sc += 5
            elif fcm_qc >= 0.06: sc += 2
            elif fcm_qc < 0.02:  sc -= 4
        # FCF growth consistency
        fgc_qc = s.get("fcfGrowthConsistency")
        if fgc_qc is not None:
            if fgc_qc >= 0.80:   sc += 5
            elif fgc_qc >= 0.60: sc += 2
            elif fgc_qc < 0.40:  sc -= 4
        # Operating margin — scalable, asset-light business model signal
        om_qc = s.get("operatingMargin")
        if om_qc is not None:
            if om_qc >= 0.25:   sc += 5
            elif om_qc >= 0.15: sc += 2
            elif om_qc < 0.05:  sc -= 4
        # Filter 4: Growth quality — penalise over-optimistic analyst estimates
        go_qc = s.get("growthOptimism")
        if go_qc is not None and go_qc > 0.50:
            sc -= 5   # analysts projecting >50% more growth than history suggests

        row = format_stock_row(s)
        row["Score"] = round(sc, 1)
        qualified.append(row)

    qualified.sort(key=lambda x: -x["Score"])
    for i, row in enumerate(qualified):
        row["Rank"] = i + 1

    ws = wb.create_sheet("3b. Quality Compounders")
    ws.sheet_view.showGridLines = False
    sr = add_title(ws,
                   "🏆 Quality Compounders — Buffett's Wonderful Businesses",
                   f"ROE >15% + ROIC >12% + Piotroski ≥7 + Positive FCF. "
                   f"Durable moat businesses at reasonable prices. {datetime.date.today()}")

    headers = ["Rank", "Ticker", "Company", "Sector", "Price", "Fwd PEG", "PEG", "Fwd P/E", "P/E",
               "IV", "MoS", "ROIC", "ROE", "FCF Yield", "FCF Margin", "Oper Margin", "FCF Conv.", "FCF Consist.",
               "Rev Growth", "Grwth Gap", "Rev Consist.", "Shares Δ", "EPS Growth 5Y",
               "Piotroski", "52w vs High", "Div Yield",
               "CEO Score", "FCF/Sh 5Y", "Divergence",
               "MktCap ($B)", "Score", "🏦 Insider"]
    widths = [5, 8, 22, 15, 8, 7, 6, 7, 7, 8, 7, 7, 7, 8, 9, 8, 8, 9, 8, 8, 9, 8, 8, 7, 9, 7, 12, 9, 14, 10, 6, 14]
    write_table(ws, qualified[:TOP_N], headers, sr, header_color="B71C1C", widths=widths)
    print(f"  ✅ Quality Compounders tab done — {min(len(qualified), TOP_N)} (from {len(qualified)})")
    return qualified


def build_hold_forever_tab(wb, stocks):
    """Sprint 3 A4: 💎 Hold Forever — the user's natural long-term shortlist.

    Stricter than Quality Compounders. Only names that pass:
    - ROIC > 15% (current TTM proxy for sustained quality)
    - Revenue consistency >= 80% (4/5 years positive growth)
    - FCF growth consistency >= 60% (FCF positive most years)
    - 5Y revenue CAGR 5-25% (steady but not unsustainable)
    - Gross margin > 40% OR Operating margin > 15% (moat proxy)
    - Net Debt/EBITDA < 3 (or net cash)
    - Net buybacks (sharesGrowth <= 0.02 — no significant dilution)
    - Piotroski >= 6 (financial health)
    - mktCap > $1B (need scale to compound for years)

    Limit 25 names. The user-philosophy shortlist for personal-knowledge evaluation.
    """
    print("\n📊 Building Tab: 💎 Hold Forever...")
    qualified = []
    for t, s in stocks.items():
        if not _is_common_stock(s):
            continue
        roic = s.get("roic")
        rc   = s.get("revConsistency")
        fgc  = s.get("fcfGrowthConsistency")
        rg5  = s.get("revGrowth5y") or s.get("fiveYRevGrowth")
        gm   = s.get("grossMargin")
        om   = s.get("operatingMargin")
        nde  = s.get("netDebtEbitda")
        sg   = s.get("sharesGrowth")
        pio  = s.get("piotroski")
        mc   = s.get("mktCap") or 0
        sec  = s.get("sector") or ""

        # Hard gates
        if mc < 1e9: continue
        if not roic or roic < 0.15: continue
        if rc is None or rc < 0.80: continue
        if fgc is not None and fgc < 0.60: continue   # may be missing for new IPOs — allow None
        if rg5 is None or rg5 < 0.05 or rg5 > 0.25: continue
        # Moat proxy: high gross margin OR high operating margin (business has pricing power)
        moat = (gm and gm > 0.40) or (om and om > 0.15)
        if not moat: continue
        if nde is not None and nde > 3.0: continue
        if sg is not None and sg > 0.02: continue   # net buybacks (or flat) — no significant dilution
        if pio is not None and pio < 6: continue
        # Skip cyclical commodity-driven sectors — moat is structural, not cyclical
        if sec in ("Basic Materials", "Energy"): continue

        # Score: how strongly does this fit the "compound for a decade" pattern?
        score = 0
        score += min(40, roic * 100)            # ROIC contributes up to 40
        score += rc * 20                        # consistency up to 20
        if fgc is not None: score += fgc * 15
        score += min(15, rg5 * 50)              # 5Y CAGR up to 15
        if gm: score += min(10, gm * 12)        # gross margin up to 10
        if sg is not None and sg < 0: score += min(8, abs(sg) * 100)   # buyback bonus
        if nde is not None and nde < 0: score += 5   # net cash bonus
        # Familiar Brand bonus — user can directly evaluate
        if s.get("consumerObservable"): score += 8
        # Under-covered bonus — Wall St hasn't piled in
        if s.get("underCovered"): score += 5

        row = format_stock_row(s)
        row["Score"] = round(score, 1)
        # Add Hold Forever-specific columns the user wants to see
        row["Rev Consist."] = rc
        row["FCF Consist."] = fgc
        row["Familiar"]     = "🛒" if s.get("consumerObservable") else ""
        row["Under-Cov"]    = "🔍" if s.get("underCovered") else ""
        qualified.append(row)

    qualified.sort(key=lambda x: -x["Score"])
    for i, row in enumerate(qualified):
        row["Rank"] = i + 1

    ws = wb.create_sheet("3c. Hold Forever")
    ws.sheet_view.showGridLines = False
    sr = add_title(ws,
                   "💎 Hold Forever — Buy-and-Forget Quality",
                   f"ROIC >15% + 80%+ rev consistency + steady 5–25% CAGR + moat (gross margin >40% or op margin >15%) "
                   f"+ no dilution + Piotroski ≥6. Top 25 names — your natural long-term shortlist. {datetime.date.today()}")

    headers = ["Rank", "Ticker", "Company", "Sector", "Familiar", "Under-Cov", "Price", "ROIC", "ROE",
               "FCF Yield", "Gross Mgn", "Oper Mgn", "Rev Consist.", "FCF Consist.",
               "Rev Growth 5Y", "Shares Δ", "Net Debt/EBITDA", "Piotroski",
               "CEO Score", "FCF/Sh 5Y", "Divergence",
               "MktCap ($B)", "Score"]
    widths = [5, 8, 24, 14, 8, 9, 8, 7, 7, 9, 9, 9, 11, 11, 12, 8, 14, 9, 12, 9, 14, 11, 7]
    write_table(ws, qualified[:25], headers, sr, header_color="6A1B9A", widths=widths)
    print(f"  ✅ Hold Forever tab done — {min(len(qualified), 25)} (from {len(qualified)} qualifying)")
    return qualified[:25]


def fetch_etf_returns(etf_tickers: list) -> dict:
    """Fetch 1Y of daily closes for each ETF (including SPY) and compute
    1M/3M/6M/1Y returns. Returns {ticker: {"1M": float, "3M": float, "6M": float, "1Y": float}}.
    Uses a single historical-price-eod/light call per ticker.
    """
    global _fmp_call_count
    if not FMP_KEY:
        return {}
    today = datetime.date.today()
    from_date = (today - datetime.timedelta(days=380)).isoformat()
    results = {}
    all_tickers = list(set(etf_tickers + ["SPY"]))
    for ticker in all_tickers:
        try:
            r = requests.get(f"{FMP_BASE}/historical-price-eod/light",
                             params={"symbol": ticker, "from": from_date,
                                     "to": today.isoformat(), "apikey": FMP_KEY},
                             timeout=20)
            _fmp_call_count += 1
            if r.status_code != 200:
                continue
            data = r.json()
            if not isinstance(data, list) or not data:
                continue
            # Build date→price dict; sort by date descending
            # FMP stable endpoint returns "price" field (not "close")
            price_map = {rec["date"]: float(rec.get("price") or rec.get("close") or 0)
                         for rec in data if rec.get("date") and (rec.get("price") or rec.get("close"))}
            if not price_map:
                continue
            sorted_dates = sorted(price_map.keys(), reverse=True)
            latest_price = price_map[sorted_dates[0]]

            def _price_on_or_before(target_date):
                for delta in range(10):
                    d = (target_date - datetime.timedelta(days=delta)).isoformat()
                    if d in price_map and price_map[d] > 0:
                        return price_map[d]
                return None

            p1w  = _price_on_or_before(today - datetime.timedelta(days=7))
            p1m  = _price_on_or_before(today - datetime.timedelta(days=21))
            p3m  = _price_on_or_before(today - datetime.timedelta(days=63))
            p6m  = _price_on_or_before(today - datetime.timedelta(days=126))
            p1y  = _price_on_or_before(today - datetime.timedelta(days=252))

            # ── Technical indicators — computed from same price_map, zero extra API calls ──
            # 52-week high/low (use up to 252 trading days = ~365 calendar days)
            prices_252 = [price_map[d] for d in sorted_dates[:252] if price_map.get(d, 0) > 0]
            hi52  = max(prices_252) if prices_252 else None
            lo52  = min(prices_252) if prices_252 else None
            vs52h = round(latest_price / hi52, 4) if hi52 and hi52 > 0 else None
            vs52l = round(latest_price / lo52, 4) if lo52 and lo52 > 0 else None

            # Simple moving averages
            prices_chrono = [price_map[d] for d in sorted(sorted_dates) if price_map.get(d, 0) > 0]
            ma50  = round(sum(prices_chrono[-50:])  / min(50,  len(prices_chrono)), 2) if len(prices_chrono) >= 10 else None
            ma200 = round(sum(prices_chrono[-200:]) / min(200, len(prices_chrono)), 2) if len(prices_chrono) >= 10 else None
            vs_ma50  = round(latest_price / ma50,  4) if ma50  and ma50  > 0 else None
            vs_ma200 = round(latest_price / ma200, 4) if ma200 and ma200 > 0 else None

            # 14-period Wilder RSI from last 15 daily price changes
            _rsi_prices = prices_chrono[-16:] if len(prices_chrono) >= 16 else []
            rsi14 = None
            if len(_rsi_prices) >= 15:
                _changes = [_rsi_prices[i] - _rsi_prices[i-1] for i in range(1, len(_rsi_prices))]
                _gains = [c for c in _changes if c > 0]; _losses = [abs(c) for c in _changes if c < 0]
                _avg_gain = sum(_gains) / 14 if _gains else 0
                _avg_loss = sum(_losses) / 14 if _losses else 0
                if _avg_loss == 0:
                    rsi14 = 100.0
                else:
                    _rs = _avg_gain / _avg_loss
                    rsi14 = round(100 - (100 / (1 + _rs)), 1)

            # Momentum direction: is 1M return stronger than recent 3M average rate?
            _r1m = (latest_price - p1m) / p1m if p1m else None
            _r3m = (latest_price - p3m) / p3m if p3m else None
            if _r1m is not None and _r3m is not None:
                momentum_dir = "▲ Accel" if _r1m > (_r3m / 3) else "▼ Decel"
            else:
                momentum_dir = None

            results[ticker] = {
                "current": latest_price,
                "1W":  (latest_price - p1w)  / p1w  if p1w  else None,
                "1M":  _r1m,
                "3M":  _r3m,
                "6M":  (latest_price - p6m)  / p6m  if p6m  else None,
                "1Y":  (latest_price - p1y)  / p1y  if p1y  else None,
                # 52-week range
                "52H": hi52, "52L": lo52,
                "vs52H": vs52h, "vs52L": vs52l,
                # Moving averages
                "ma50": ma50, "ma200": ma200,
                "vs_ma50": vs_ma50, "vs_ma200": vs_ma200,
                # Oscillator
                "rsi14": rsi14,
                # Trend direction
                "momentum_dir": momentum_dir,
            }
            time.sleep(0.12)
        except Exception:
            continue

    # Compute vs-SPY alpha for each ticker
    spy = results.get("SPY", {})
    for ticker, ret in results.items():
        if ticker == "SPY":
            continue
        for period in ("1M", "3M", "6M", "1Y"):
            etf_r = ret.get(period)
            spy_r = spy.get(period)
            ret[f"{period}_alpha"] = (etf_r - spy_r) if (etf_r is not None and spy_r is not None) else None
    return results


def build_sector_valuations(wb, stocks):
    """Tab 9: Sector Rotation Dashboard — valuation, momentum, quality, signal."""
    print("\n📊 Building Tab: Sector Valuations...")
    ws = wb.create_sheet("9. Sector Rotation")
    ws.sheet_view.showGridLines = False

    # ── Fetch live sector ETF performance vs SPY ───────────────────────────
    # Reverse map: sector name → ETF ticker
    _sector_to_etf = {v: k for k, v in SECTOR_ETFS.items()}
    print("    Fetching sector ETF performance vs SPY...")
    _etf_returns = fetch_etf_returns(list(SECTOR_ETFS.keys()))
    spy_1m  = (_etf_returns.get("SPY") or {}).get("1M")
    spy_3m  = (_etf_returns.get("SPY") or {}).get("3M")
    spy_6m  = (_etf_returns.get("SPY") or {}).get("6M")
    spy_1y  = (_etf_returns.get("SPY") or {}).get("1Y")
    print(f"    SPY: 1M={spy_1m:.1%} 3M={spy_3m:.1%} 6M={spy_6m:.1%} 1Y={spy_1y:.1%}"
          if all(x is not None for x in [spy_1m, spy_3m, spy_6m, spy_1y]) else
          "    SPY data partially unavailable")

    # ── Economic cycle context (static, well-established rotation pattern) ──
    CYCLE_PHASE = {
        "Technology":             "Mid/Late",
        "Financial Services":     "Early",
        "Financials":             "Early",
        "Consumer Cyclical":      "Early/Mid",
        "Consumer Discretionary": "Early/Mid",
        "Industrials":            "Mid",
        "Basic Materials":        "Mid/Late",
        "Materials":              "Mid/Late",
        "Energy":                 "Late",
        "Healthcare":             "Late/Recession",
        "Consumer Defensive":     "Recession",
        "Consumer Staples":       "Recession",
        "Utilities":              "Recession",
        "Real Estate":            "Early/Mid",
        "Communication Services": "Mid",
    }

    # ── Aggregate per sector ──
    sector_agg = defaultdict(lambda: {
        "pegs": [], "mos": [], "fcf": [], "roe": [], "rg": [], "epsg": [],
        "pe": [], "pb": [], "de": [], "pio": [], "count": 0, "stocks": []
    })
    for t, s in stocks.items():
        if not _is_common_stock(s): continue
        sect = s.get("sector") or "Unknown"
        agg = sector_agg[sect]
        agg["count"] += 1
        agg["stocks"].append(s)
        if s.get("peg") and 0 < s.get("peg") < 30:   agg["pegs"].append(s.get("peg"))
        if s.get("mos") is not None:               agg["mos"].append(s.get("mos"))
        if s.get("fcfYield") is not None:          agg["fcf"].append(s.get("fcfYield"))
        if s.get("roe") is not None:               agg["roe"].append(s.get("roe"))
        if s.get("revGrowth") is not None:         agg["rg"].append(s.get("revGrowth"))
        if s.get("epsGrowth5y") is not None:       agg["epsg"].append(s.get("epsGrowth5y"))
        if s.get("pe") and 0 < s.get("pe") < 200:     agg["pe"].append(s.get("pe"))
        if s.get("pb") and 0 < s.get("pb") < 50:      agg["pb"].append(s.get("pb"))
        if s.get("de") is not None:                agg["de"].append(s.get("de"))
        if s.get("piotroski") is not None:         agg["pio"].append(s.get("piotroski"))

    def _med(lst):
        if not lst: return None
        s = sorted(lst)
        n = len(s)
        return round(s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2, 4)

    def _avg(lst):
        return round(sum(lst) / len(lst), 4) if lst else None

    def _pct_above(lst, threshold):
        if not lst: return None
        return round(sum(1 for x in lst if x > threshold) / len(lst), 3)

    rows = []
    for sector, agg in sorted(sector_agg.items(), key=lambda x: -x[1]["count"]):
        if agg["count"] < 5: continue

        med_peg   = _med(agg["pegs"])
        med_pe    = _med(agg["pe"])
        med_pb    = _med(agg["pb"])
        avg_mos   = _avg(agg["mos"])
        avg_fcf   = _avg(agg["fcf"])
        avg_roe   = _avg(agg["roe"])
        avg_rg    = _avg(agg["rg"])
        avg_epsg  = _avg(agg["epsg"])
        avg_pio   = _avg(agg["pio"])
        pct_under = _pct_above(agg["mos"], 0)   # % stocks with positive MoS
        pct_qual  = _pct_above(agg["pio"], 6.5) # % stocks with Piotroski ≥ 7
        n_qual    = sum(1 for x in agg["pio"] if x >= 7)

        # ── Rotation Score (0-100): value + quality + momentum ──
        score = 0
        # Valuation (40 pts): low PEG + high % undervalued
        if med_peg:
            if med_peg < 1.0:    score += 20
            elif med_peg < 1.5:  score += 15
            elif med_peg < 2.0:  score += 10
            elif med_peg < 3.0:  score += 5
        if pct_under is not None:
            score += round(pct_under * 20)  # up to 20pts for 100% undervalued
        # Quality (30 pts): FCF + ROE + Piotroski breadth
        if avg_fcf and avg_fcf > 0.05:   score += 10
        elif avg_fcf and avg_fcf > 0.02: score += 5
        if avg_roe and avg_roe > 0.15:   score += 10
        elif avg_roe and avg_roe > 0.08: score += 5
        if pct_qual and pct_qual > 0.3:  score += 10
        elif pct_qual and pct_qual > 0.15: score += 5
        # Momentum (30 pts): revenue + EPS growth
        if avg_rg and avg_rg > 0.10:    score += 15
        elif avg_rg and avg_rg > 0.05:  score += 8
        elif avg_rg and avg_rg > 0:     score += 3
        if avg_epsg and avg_epsg > 0.10: score += 15
        elif avg_epsg and avg_epsg > 0.05: score += 8
        elif avg_epsg and avg_epsg > 0:  score += 3

        # ETF performance for this sector
        _etf_tkr  = _sector_to_etf.get(sector)
        _etf_data = _etf_returns.get(_etf_tkr, {}) if _etf_tkr else {}
        rows.append({
            "Sector":        sector,
            "Signal":        None,   # assigned by relative ranking below
            "Rot. Score":    score,
            "# Stocks":      agg["count"],
            "# Quality":     n_qual,
            "% Underval.":   pct_under,
            "Med PEG":       med_peg,
            "Med P/E":       med_pe,
            "Med P/B":       med_pb,
            "ETF":           _etf_tkr or "—",
            "1M vs SPY":     _etf_data.get("1M_alpha"),
            "3M vs SPY":     _etf_data.get("3M_alpha"),
            "6M vs SPY":     _etf_data.get("6M_alpha"),
            "1Y vs SPY":     _etf_data.get("1Y_alpha"),
            "Avg FCF Yield": avg_fcf,
            "Avg ROE":       avg_roe,
            "Avg Rev Grw":   avg_rg,
            "Avg EPS Grw":   avg_epsg,
            "Cycle Phase":   CYCLE_PHASE.get(sector, "—"),
            "_stocks":       agg["stocks"],  # for top picks mini-table
        })

    rows.sort(key=lambda x: -x["Rot. Score"])

    # ── Relative signal assignment ─────────────────────────────────────────
    # Absolute thresholds inflate BUY in any bull market.
    # Sector rotation is about WHERE to put capital, not whether to invest at all.
    # Rule: top 3 → BUY, bottom 3 → AVOID, rest → HOLD.
    # Tie-break: minimum absolute score of 40 required for BUY (prevents BUY in a crash).
    n = len(rows)
    for i, row in enumerate(rows):
        sc = row["Rot. Score"]
        if i < 3 and sc >= 40:
            row["Signal"] = "BUY"
        elif i >= n - 3:
            row["Signal"] = "AVOID"
        else:
            row["Signal"] = "HOLD"

    # ── Main rotation table ──
    sr = add_title(ws,
                   "🔄 Sector Rotation Dashboard",
                   f"Rotation Score = Valuation (40) + Quality (30) + Momentum (30). "
                   f"Medians from {len(stocks)} US common stocks. {datetime.date.today()}")

    main_headers = ["Sector", "Signal", "Rot. Score", "# Stocks", "# Quality",
                    "% Underval.", "Med PEG", "Med P/E", "Med P/B",
                    "ETF", "1M vs SPY", "3M vs SPY", "6M vs SPY", "1Y vs SPY",
                    "Avg FCF Yield", "Avg ROE",
                    "Avg Rev Grw", "Avg EPS Grw", "Cycle Phase"]
    main_widths  = [22, 8, 9, 8, 8, 9, 8, 8, 8, 6, 9, 9, 9, 9, 10, 8, 10, 10, 12]

    # Write header row
    for ci, h in enumerate(main_headers, 1):
        c = ws.cell(row=sr, column=ci, value=h)
        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="263238")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[sr].height = 26
    sr += 1

    BUY_FILL   = PatternFill("solid", fgColor="1B5E20")
    HOLD_FILL  = PatternFill("solid", fgColor="F57F17")
    AVOID_FILL = PatternFill("solid", fgColor="B71C1C")
    BUY_ROW    = PatternFill("solid", fgColor="E8F5E9")
    HOLD_ROW   = PatternFill("solid", fgColor="FFF8E1")
    AVOID_ROW  = PatternFill("solid", fgColor="FFEBEE")

    for ri, row in enumerate(rows):
        sig = row["Signal"]
        row_fill = BUY_ROW if sig == "BUY" else (HOLD_ROW if sig == "HOLD" else AVOID_ROW)
        for ci, h in enumerate(main_headers, 1):
            v = row.get(h)
            cell = ws.cell(row=sr, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.font = Font(name="Arial", size=9)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if h == "Signal":
                if sig == "BUY":
                    cell.fill = BUY_FILL
                    cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                elif sig == "HOLD":
                    cell.fill = HOLD_FILL
                    cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                else:
                    cell.fill = AVOID_FILL
                    cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            elif h == "Rot. Score":
                cell.font = Font(bold=True, name="Arial", size=9)
                if row["Rot. Score"] >= 55:
                    cell.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                elif row["Rot. Score"] < 35:
                    cell.font = Font(bold=True, name="Arial", size=9, color="B71C1C")
            elif h in ("% Underval.", "Avg FCF Yield", "Avg ROE",
                       "Avg Rev Grw", "Avg EPS Grw") and isinstance(v, float):
                cell.number_format = "0%"
                if v and v > 0.01:
                    cell.font = Font(name="Arial", size=9, color="1B5E20")
                elif v and v < -0.01:
                    cell.font = Font(name="Arial", size=9, color="B71C1C")
            elif h in ("1M vs SPY", "3M vs SPY", "6M vs SPY", "1Y vs SPY"):
                if isinstance(v, float):
                    cell.number_format = "+0.0%;-0.0%;0%"
                    if v >= 0.03:
                        cell.fill = PatternFill("solid", fgColor="C8E6C9")
                        cell.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                    elif v > 0:
                        cell.fill = PatternFill("solid", fgColor="E8F5E9")
                        cell.font = Font(name="Arial", size=9, color="2E7D32")
                    elif v <= -0.03:
                        cell.fill = PatternFill("solid", fgColor="FFCDD2")
                        cell.font = Font(bold=True, name="Arial", size=9, color="B71C1C")
                    elif v < 0:
                        cell.fill = PatternFill("solid", fgColor="FFEBEE")
                        cell.font = Font(name="Arial", size=9, color="C62828")
                else:
                    cell.value = "n/a"
                    cell.font = Font(name="Arial", size=8, color="AAAAAA")
            elif h in ("Med PEG", "Med P/E", "Med P/B") and isinstance(v, float):
                cell.number_format = "0.1"
            elif h == "Sector":
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True, name="Arial", size=9)
            elif h == "Cycle Phase":
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[sr].height = 16
        sr += 1

    for ci, w in enumerate(main_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"

    sr += 2  # gap

    # ── Cycle Reference Guide ──
    guide_title = ws.cell(row=sr, column=1, value="📚 Economic Cycle — Sector Rotation Guide")
    guide_title.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    guide_title.fill = PatternFill("solid", fgColor="37474F")
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(main_headers))
    sr += 1

    cycle_data = [
        ("🌱 Early Cycle (Recovery)",     "Financials, Consumer Discretionary, Real Estate, Industrials",
         "Credit loosens, consumer spending rebounds, capex picks up"),
        ("☀️  Mid Cycle (Expansion)",      "Technology, Industrials, Materials, Communication Services",
         "Earnings growth broad, rates stable, business investment peaks"),
        ("🌇 Late Cycle (Slowdown)",       "Energy, Materials, Healthcare, Consumer Staples",
         "Inflation rises, margins compress, defensives outperform"),
        ("🌧️  Recession",                  "Consumer Staples, Healthcare, Utilities, Bonds",
         "Revenue falls, dividends valued, capital preservation focus"),
    ]
    cycle_headers = ["Phase", "Favored Sectors", "Key Dynamic"]
    cycle_widths  = [22, 45, 42]
    for ci, h in enumerate(cycle_headers, 1):
        c = ws.cell(row=sr, column=ci, value=h)
        c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="546E7A")
        c.border = THIN_BORDER
    sr += 1
    for phase, favored, dynamic in cycle_data:
        ws.cell(row=sr, column=1, value=phase).font = Font(bold=True, name="Arial", size=9)
        ws.cell(row=sr, column=2, value=favored).font = Font(name="Arial", size=9)
        ws.cell(row=sr, column=3, value=dynamic).font = Font(name="Arial", size=9, color="546E7A")
        for ci in range(1, 4):
            ws.cell(row=sr, column=ci).border = THIN_BORDER
            ws.cell(row=sr, column=ci).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[sr].height = 18
        sr += 1
    for ci, w in enumerate(cycle_widths, 1):
        # don't override already-set widths for col 1; use max
        existing = ws.column_dimensions[get_column_letter(ci)].width or 0
        ws.column_dimensions[get_column_letter(ci)].width = max(existing, w)

    sr += 2

    # ── Top 3 picks per sector (sorted by Score) ──
    picks_title = ws.cell(row=sr, column=1, value="🏆 Top Picks per Sector (by Score)")
    picks_title.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    picks_title.fill = PatternFill("solid", fgColor="1A237E")
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(main_headers))
    sr += 1

    pick_headers = ["Ticker", "Company", "Price", "PEG", "P/E", "MoS", "ROE", "FCF Yield",
                    "Piotroski", "Rev Growth", "Score"]
    pick_widths_local = [8, 24, 8, 7, 7, 7, 7, 9, 7, 9, 7]

    for row in rows:
        sector_name = row["Sector"]
        sig = row["Signal"]
        sc_fill = BUY_FILL if sig == "BUY" else (HOLD_FILL if sig == "HOLD" else AVOID_FILL)

        # Section header
        sh = ws.cell(row=sr, column=1,
                     value=f"  {sector_name}  [{sig}] Score: {row['Rot. Score']}")
        sh.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        sh.fill = sc_fill
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(main_headers))
        sr += 1

        # Score the stocks in this sector
        sector_stocks = row["_stocks"]
        scored = []
        for s in sector_stocks:
            if not _is_common_stock(s): continue
            sc2 = 0
            peg = s.get("peg")
            mos = s.get("mos")
            pio = s.get("piotroski")
            if peg and 0 < peg < 1.0:   sc2 += 15
            elif peg and 0 < peg < 1.5: sc2 += 10
            elif peg and 0 < peg < 2.0: sc2 += 5
            if mos and mos > 0.2:  sc2 += 10
            elif mos and mos > 0:  sc2 += 5
            if pio and pio >= 8:   sc2 += 10
            elif pio and pio >= 6: sc2 += 5
            if s.get("roe") and s.get("roe") > 0.15: sc2 += 5
            if s.get("fcfYield") and s.get("fcfYield") > 0.05: sc2 += 5
            if s.get("revGrowth") and s.get("revGrowth") > 0.1: sc2 += 5
            scored.append((sc2, s))
        scored.sort(key=lambda x: -x[0])
        top3 = scored[:3]

        if top3:
            # mini header
            for ci, h in enumerate(pick_headers, 1):
                c = ws.cell(row=sr, column=ci, value=h)
                c.font = Font(bold=True, name="Arial", size=8, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="455A64")
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center")
            sr += 1
            for pick_sc, s in top3:
                vals = [s.get("ticker"), s.get("name", "")[:28],
                        s.get("price"), s.get("peg"), s.get("pe"),
                        s.get("mos"), s.get("roe"), s.get("fcfYield"),
                        s.get("piotroski"), s.get("revGrowth"), round(pick_sc, 1)]
                for ci, (h, v) in enumerate(zip(pick_headers, vals), 1):
                    cell = ws.cell(row=sr, column=ci, value=v)
                    cell.border = THIN_BORDER
                    cell.font = Font(name="Arial", size=8)
                    cell.fill = ALT_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if h == "Price" and isinstance(v, float):
                        cell.number_format = "$#,##0.00"
                    elif h in ("MoS", "ROE", "FCF Yield", "Rev Growth") and isinstance(v, float):
                        cell.number_format = "0%"
                    elif h in ("PEG", "P/E") and isinstance(v, float):
                        cell.number_format = "0.1"
                ws.row_dimensions[sr].height = 14
                sr += 1
        sr += 1  # gap between sectors

    print(f"  ✅ Sector Valuations done — {len(rows)} sectors")
    return rows, _etf_returns  # etf_returns reused by build_sector_etf_rotation (no re-fetch)


def build_sector_etf_rotation(wb, stocks, etf_returns: dict, sector_fund_scores: dict = None):
    """Tab 10: ETF Sector Rotation — technical + fundamental signals for ETF rotation trading.
    Uses same ETF price history already fetched by build_sector_valuations (zero new API calls).
    Adds: 52w positioning, MA50/200, RSI-14, momentum direction, combined rotation signal.
    """
    print("\n📊 Building Tab: ETF Sector Rotation...")
    ws = wb.create_sheet("10. ETF Rotation")
    ws.sheet_view.showGridLines = False

    CYCLE_PHASE = {
        "Technology": "Mid/Late", "Financial Services": "Early",
        "Consumer Cyclical": "Early/Mid", "Industrials": "Mid",
        "Basic Materials": "Mid/Late", "Energy": "Late",
        "Healthcare": "Late/Recession", "Consumer Defensive": "Recession",
        "Utilities": "Recession", "Real Estate": "Early/Mid",
        "Communication Services": "Mid",
    }
    # Reverse map: sector → ETF
    _sec2etf = {v: k for k, v in SECTOR_ETFS.items()}

    # ── Aggregate fundamental quality per sector from stocks dict ─────────────
    _fund_scores = sector_fund_scores or {}
    if not _fund_scores:
        _sec_agg = {}
        for t, s in stocks.items():
            if not _is_common_stock(s): continue
            sec = s.get("sector", "Unknown")
            if sec not in _sec_agg:
                _sec_agg[sec] = {"pegs": [], "mos": [], "fcf": [], "roe": [], "rg": []}
            if s.get("peg") and 0 < s["peg"] < 20:  _sec_agg[sec]["pegs"].append(s["peg"])
            if s.get("mos") is not None:             _sec_agg[sec]["mos"].append(s["mos"])
            if s.get("fcfYield") is not None:        _sec_agg[sec]["fcf"].append(s["fcfYield"])
            if s.get("roe") is not None:             _sec_agg[sec]["roe"].append(s["roe"])
            if s.get("revGrowth") is not None:       _sec_agg[sec]["rg"].append(s["revGrowth"])
        def _med(lst): return sorted(lst)[len(lst)//2] if lst else None
        def _avg(lst): return sum(lst)/len(lst) if lst else None
        for sec, agg in _sec_agg.items():
            sc = 0
            peg = _med(agg["pegs"])
            if peg:
                if peg < 1.0: sc += 20
                elif peg < 1.5: sc += 15
                elif peg < 2.0: sc += 10
                elif peg < 3.0: sc += 5
            pct_val = sum(1 for v in agg["mos"] if v and v > 0) / len(agg["mos"]) if agg["mos"] else 0
            sc += min(20, int(pct_val * 20))
            fcf_avg = _avg(agg["fcf"])
            if fcf_avg:
                if fcf_avg > 0.05: sc += 10
                elif fcf_avg > 0.02: sc += 5
            roe_avg = _avg(agg["roe"])
            if roe_avg:
                if roe_avg > 0.15: sc += 10
                elif roe_avg > 0.08: sc += 5
            rg_avg = _avg(agg["rg"])
            if rg_avg:
                if rg_avg > 0.10: sc += 15
                elif rg_avg > 0.05: sc += 8
                elif rg_avg > 0: sc += 3
            _fund_scores[sec] = sc

    # ── Compute ETF rotation score for each sector ─────────────────────────────
    etf_rows = []
    spy_data = etf_returns.get("SPY", {})

    for etf_ticker, sector_name in SECTOR_ETFS.items():
        ed = etf_returns.get(etf_ticker, {})
        if not ed: continue

        sc = 0
        # Technical momentum (40 pts)
        alpha_3m = ed.get("3M_alpha")
        if alpha_3m is not None:
            if alpha_3m > 0.05:   sc += 15
            elif alpha_3m > 0:    sc += 8
            elif alpha_3m < -0.10: sc -= 12
            elif alpha_3m < -0.05: sc -= 5
        vs_ma50  = ed.get("vs_ma50")
        vs_ma200 = ed.get("vs_ma200")
        if vs_ma50 is not None:
            if vs_ma50 > 1.0: sc += 10
            else:             sc -= 8
        if vs_ma200 is not None:
            if vs_ma200 > 1.0: sc += 10
            else:              sc -= 5
        if ed.get("momentum_dir") == "▲ Accel": sc += 5

        # Mean-reversion opportunity (30 pts)
        vs52h = ed.get("vs52H")
        rsi   = ed.get("rsi14")
        if vs52h is not None:
            if vs52h <= 0.65:   sc += 25
            elif vs52h <= 0.75: sc += 18
            elif vs52h <= 0.85: sc += 10
            elif vs52h >= 0.98: sc -= 12
        if rsi is not None:
            if rsi <= 30:  sc += 10
            elif rsi >= 70: sc -= 10

        # Fundamental quality (30 pts)
        fund_sc = _fund_scores.get(sector_name, 0)
        sc += round(fund_sc / 100 * 30)

        # Signal
        if sc >= 65:   sig = "🟢 ROTATE IN"
        elif sc >= 48: sig = "🟡 HOLD"
        elif sc >= 35: sig = "🟠 TAKE PROFITS"
        else:          sig = "🔴 AVOID"

        etf_rows.append({
            "sector": sector_name, "etf": etf_ticker, "signal": sig, "score": sc,
            "price":    ed.get("current"),
            "vs52H":    vs52h,
            "vs52L":    ed.get("vs52L"),
            "1W":       ed.get("1W"),
            "1M":       ed.get("1M"),
            "3M":       ed.get("3M"),
            "6M":       ed.get("6M"),
            "1Y":       ed.get("1Y"),
            "1M_alpha": ed.get("1M_alpha"),
            "3M_alpha": alpha_3m,
            "6M_alpha": ed.get("6M_alpha"),
            "1Y_alpha": ed.get("1Y_alpha"),
            "vs_ma50":  vs_ma50,
            "vs_ma200": vs_ma200,
            "rsi14":    rsi,
            "trend":    ed.get("momentum_dir"),
            "cycle":    CYCLE_PHASE.get(sector_name, "—"),
            "fund_sc":  fund_sc,
        })

    etf_rows.sort(key=lambda x: -x["score"])

    # ── SECTION A: Title ──────────────────────────────────────────────────────
    sr = add_title(ws,
                   "🔄 ETF Sector Rotation — Technical + Fundamental Signals",
                   f"52w positioning · MA50/200 · RSI-14 · Alpha vs SPY · Zero extra API calls — {datetime.date.today()}")

    # SPY benchmark row
    spy_row_r = sr
    sr += 1
    _spy_lbl_cell = ws.cell(row=spy_row_r, column=1, value="📊 SPY Benchmark")
    _spy_lbl_cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
    _spy_lbl_cell.fill = PatternFill("solid", fgColor="263238")
    _spy_data_map = {
        "1W": spy_data.get("1W"), "1M": spy_data.get("1M"), "3M": spy_data.get("3M"),
        "6M": spy_data.get("6M"), "1Y": spy_data.get("1Y"),
        "vs52H": spy_data.get("vs52H"), "vs52L": spy_data.get("vs52L"),
        "RSI": spy_data.get("rsi14"),
        "vs MA50": spy_data.get("vs_ma50"), "vs MA200": spy_data.get("vs_ma200"),
    }
    _spy_col = 2
    for lbl, val in _spy_data_map.items():
        hc = ws.cell(row=spy_row_r - 1 if spy_row_r > 1 else spy_row_r, column=_spy_col)
        dc = ws.cell(row=spy_row_r, column=_spy_col)
        dc.value = val
        dc.font = Font(bold=True, name="Arial", size=9)
        dc.border = THIN_BORDER
        dc.alignment = Alignment(horizontal="center")
        if isinstance(val, float):
            if lbl in ("1W","1M","3M","6M","1Y","vs MA50","vs MA200"):
                dc.number_format = "+0.0%;-0.0%;0%"
                if val > 0: dc.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                elif val < 0: dc.font = Font(bold=True, name="Arial", size=9, color="B71C1C")
            elif lbl in ("vs52H","vs52L"):
                dc.number_format = "0%"
            elif lbl == "RSI":
                dc.number_format = "0.0"
        _spy_col += 1
    # Market regime note
    spy_vs200 = spy_data.get("vs_ma200")
    spy_rsi   = spy_data.get("rsi14")
    if spy_vs200 is not None:
        if spy_vs200 >= 1.05:   regime = "BULL MARKET — SPY above MA200 (+5%)"
        elif spy_vs200 >= 1.0:  regime = "RECOVERY — SPY above MA200 (marginal)"
        elif spy_vs200 >= 0.93: regime = "CORRECTION — SPY below MA200"
        else:                   regime = "BEAR MARKET — SPY deeply below MA200"
        rc = ws.cell(row=spy_row_r, column=_spy_col, value=regime)
        rc.font = Font(bold=True, name="Arial", size=9,
                       color="1B5E20" if spy_vs200 >= 1.0 else "B71C1C")
        rc.alignment = Alignment(horizontal="left")

    # ── SECTION B: Main ETF Dashboard table ───────────────────────────────────
    sr += 1  # blank row
    hdr_row = sr
    _hdr_cols = [
        "Sector", "ETF", "Signal", "ETF Score",
        "Price", "52w vs High", "52w vs Low",
        "1W", "1M", "3M", "6M", "1Y",
        "1M α", "3M α", "6M α", "1Y α",
        "vs MA50", "vs MA200", "RSI-14", "Trend", "Cycle",
    ]
    _hdr_widths = [22, 6, 14, 9, 8, 11, 10, 7, 8, 8, 8, 8, 8, 8, 8, 8, 9, 10, 7, 10, 14]

    for ci, h in enumerate(_hdr_cols, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font  = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="263238")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 28
    for ci, w in enumerate(_hdr_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"
    sr += 1

    # Row fill by signal
    _SIG_FILL = {
        "🟢 ROTATE IN":    ("E8F5E9", "1B5E20"),   # light green / dark green text
        "🟡 HOLD":         ("FFFDE7", "F57F17"),   # light yellow / amber text
        "🟠 TAKE PROFITS": ("FFF3E0", "E65100"),   # light orange / orange text
        "🔴 AVOID":        ("FFEBEE", "B71C1C"),   # light red / red text
    }

    for row_data in etf_rows:
        sig     = row_data["signal"]
        fill_clr, txt_clr = _SIG_FILL.get(sig, ("FFFFFF", "000000"))
        row_fill = PatternFill("solid", fgColor=fill_clr)

        for ci, h in enumerate(_hdr_cols, 1):
            key_map = {
                "Sector": "sector", "ETF": "etf", "Signal": "signal", "ETF Score": "score",
                "Price": "price", "52w vs High": "vs52H", "52w vs Low": "vs52L",
                "1W": "1W", "1M": "1M", "3M": "3M", "6M": "6M", "1Y": "1Y",
                "1M α": "1M_alpha", "3M α": "3M_alpha", "6M α": "6M_alpha", "1Y α": "1Y_alpha",
                "vs MA50": "vs_ma50", "vs MA200": "vs_ma200",
                "RSI-14": "rsi14", "Trend": "trend", "Cycle": "cycle",
            }
            v = row_data.get(key_map.get(h, ""))
            cell = ws.cell(row=sr, column=ci, value=v)
            cell.fill   = row_fill
            cell.border = THIN_BORDER
            cell.font   = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                horizontal="left" if h in ("Sector", "Signal", "Trend", "Cycle") else "center",
                vertical="center")

            # Column-specific formatting
            if h == "Signal":
                cell.font = Font(bold=True, name="Arial", size=9, color=txt_clr)
            elif h == "ETF Score":
                cell.font = Font(bold=True, name="Arial", size=9,
                                 color="1B5E20" if (v or 0) >= 55 else ("B71C1C" if (v or 0) < 35 else "000000"))
            elif h == "Price" and isinstance(v, float):
                cell.number_format = "$#,##0.00"
            elif h in ("1W","1M","3M","6M","1Y","1M α","3M α","6M α","1Y α") and isinstance(v, float):
                cell.number_format = "+0.0%;-0.0%;0%"
                color = "1B5E20" if v > 0.005 else ("B71C1C" if v < -0.005 else "555555")
                cell.font = Font(name="Arial", size=9, color=color)
            elif h in ("52w vs High","52w vs Low","vs MA50","vs MA200") and isinstance(v, float):
                cell.number_format = "0%"
                # 52w vs High: low = green (beaten down = opportunity)
                if h == "52w vs High":
                    if v <= 0.70:   cell.fill = PatternFill("solid", fgColor="C8E6C9"); cell.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                    elif v >= 0.95: cell.fill = PatternFill("solid", fgColor="FFCDD2"); cell.font = Font(name="Arial", size=9, color="B71C1C")
                elif h == "52w vs Low":
                    if v <= 1.15:   cell.fill = PatternFill("solid", fgColor="C8E6C9")
                    elif v >= 1.75: cell.fill = PatternFill("solid", fgColor="FFF9C4")
                elif h in ("vs MA50","vs MA200"):
                    if v >= 1.02:   cell.font = Font(name="Arial", size=9, color="1B5E20")
                    elif v <= 0.97: cell.font = Font(name="Arial", size=9, color="B71C1C")
            elif h == "RSI-14" and isinstance(v, float):
                cell.number_format = "0.0"
                if v <= 30:   cell.fill = PatternFill("solid", fgColor="C8E6C9"); cell.font = Font(bold=True, name="Arial", size=9, color="1B5E20")
                elif v >= 70: cell.fill = PatternFill("solid", fgColor="FFCDD2"); cell.font = Font(bold=True, name="Arial", size=9, color="B71C1C")
            elif h == "Trend" and v:
                cell.font = Font(name="Arial", size=9, color="1B5E20" if "Accel" in str(v) else "B71C1C")

        ws.row_dimensions[sr].height = 16
        sr += 1

    # ── SECTION C: Rotation Opportunity Matrix ─────────────────────────────────
    sr += 1
    mat_hdr = ws.cell(row=sr, column=1, value="📊 Rotation Opportunity Matrix — Best Risk/Reward by Quadrant")
    mat_hdr.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    mat_hdr.fill = PatternFill("solid", fgColor="37474F")
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=8)
    ws.row_dimensions[sr].height = 18
    sr += 1

    # Sub-header
    _mat_hdrs = ["Sector", "ETF", "Opportunity Type", "ETF Score", "52w vs High", "Fund. Score", "RSI", "Entry Rationale"]
    _mat_wids = [22, 6, 18, 9, 11, 10, 7, 45]
    for ci, (h, w) in enumerate(zip(_mat_hdrs, _mat_wids), 1):
        c = ws.cell(row=sr, column=ci, value=h)
        c.font  = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="455A64")
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    sr += 1

    # Classify each sector ETF into opportunity quadrants
    _opp_rows = []
    for row_data in etf_rows:
        vs52h   = row_data.get("vs52H")
        fund_sc = row_data.get("fund_sc", 0)
        a3m     = row_data.get("3M_alpha")
        vs200   = row_data.get("vs_ma200")
        rsi     = row_data.get("rsi14")
        sc      = row_data["score"]

        if vs52h is not None and vs52h <= 0.75 and fund_sc >= 40:
            opp_type = "🏆 Turnaround"
            rationale = f"ETF {round(vs52h*100)}% of 52w high — quality sector beaten down, fundamentals intact"
        elif a3m is not None and a3m > 0.03 and vs200 is not None and vs200 >= 1.0:
            top3_alpha = sorted(etf_rows, key=lambda x: -(x.get("3M_alpha") or -99))[:3]
            if row_data in top3_alpha:
                opp_type = "📈 Momentum"
                rationale = f"Top-3 3M alpha vs SPY ({round((a3m or 0)*100,1)}%), above MA200 — ride the trend"
            else:
                opp_type = "🟡 Hold Trend"
                rationale = "Above MA200, positive momentum but not top-ranked"
        elif vs52h is not None and vs52h <= 0.75 and fund_sc < 35:
            opp_type = "⚠️ Value Trap Risk"
            rationale = "Cheap ETF but fundamentals weak — caution, could be structural decline"
        elif vs52h is not None and vs52h >= 0.95 and fund_sc >= 45:
            opp_type = "💰 Extended Quality"
            rationale = f"Strong fundamentals but ETF near 52w high — consider trimming or wait for pullback"
        elif sc >= 48:
            opp_type = "🟡 Hold"
            rationale = "Mixed signals — no strong edge either direction"
        else:
            opp_type = "🔴 Avoid / Rotate Out"
            rationale = "Weak technicals + weak fundamentals — better opportunities elsewhere"
        _opp_rows.append((row_data, opp_type, rationale))

    # Sort: Turnaround first, then Momentum, then Hold, then Avoid
    _opp_priority = {"🏆 Turnaround": 0, "📈 Momentum": 1, "🟡 Hold Trend": 2,
                     "💰 Extended Quality": 3, "🟡 Hold": 4, "⚠️ Value Trap Risk": 5, "🔴 Avoid / Rotate Out": 6}
    _opp_rows.sort(key=lambda x: _opp_priority.get(x[1], 9))

    for row_data, opp_type, rationale in _opp_rows:
        _opp_fill_map = {
            "🏆 Turnaround":        "C8E6C9", "📈 Momentum": "E8F5E9",
            "🟡 Hold Trend":        "FFFDE7", "🟡 Hold": "FFFDE7",
            "💰 Extended Quality":  "FFF3E0", "⚠️ Value Trap Risk": "FFF9C4",
            "🔴 Avoid / Rotate Out":"FFEBEE",
        }
        _fill = PatternFill("solid", fgColor=_opp_fill_map.get(opp_type, "FFFFFF"))
        vals = [
            row_data["sector"], row_data["etf"], opp_type, row_data["score"],
            row_data.get("vs52H"), row_data.get("fund_sc"),
            row_data.get("rsi14"), rationale,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=sr, column=ci, value=v)
            c.fill   = _fill
            c.border = THIN_BORDER
            c.font   = Font(name="Arial", size=9)
            c.alignment = Alignment(
                horizontal="left" if ci in (1, 3, 8) else "center",
                vertical="center", wrap_text=(ci == 8))
            if ci == 5 and isinstance(v, float): c.number_format = "0%"
            elif ci == 7 and isinstance(v, float): c.number_format = "0.0"
            elif ci == 3:
                c.font = Font(bold=True, name="Arial", size=9)
        ws.row_dimensions[sr].height = 16
        sr += 1

    # ── Footer: legend ──────────────────────────────────────────────────────────
    sr += 1
    _legend = [
        "📖  How to use:  ROTATE IN = ETF beaten down + fundamentals support recovery. "
        "TAKE PROFITS = near 52w high, trim position.  AVOID = weak technicals + fundamentals.  "
        "RSI ≤30 = oversold (buy signal).  RSI ≥70 = overbought (sell signal).  "
        "vs MA200 <1.0 = long-term downtrend (caution).  "
        "52w vs High ≤70% = deep discount — highest mean-reversion potential."
    ]
    lc = ws.cell(row=sr, column=1, value=_legend[0])
    lc.font = Font(italic=True, name="Arial", size=8, color="546E7A")
    lc.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(_hdr_cols))
    ws.row_dimensions[sr].height = 36

    print(f"  ✅ ETF Rotation tab done — {len(etf_rows)} sector ETFs")
    return etf_rows


def build_insider_tab(wb, stocks, insider_data):
    """Tab 10: Insider Buying — cluster buys + individual large purchases."""
    print("\n📊 Building Tab: Insider Buying...")
    ws = wb.create_sheet("10. Insider Buying")
    ws.sheet_view.showGridLines = False

    if not insider_data:
        sr = add_title(ws, "🏦 Insider Buying — Recent Open Market Purchases",
                       "No insider data available.")
        ws.cell(row=sr, column=1,
                value="No insider trades loaded. Check FMP plan or openinsider.com connectivity.").font = \
            Font(italic=True, name="Arial", size=10, color="888888")
        print("  ✅ Insider Buying tab done — 0 trades")
        return

    _src_map = {"openinsider": "openinsider.com", "finviz": "finviz.com"}
    source = _src_map.get(insider_data[0].get("_source", ""), "FMP")

    def _trade_value(tr):
        v = tr.get("_value", 0)
        if not v: v = abs((tr.get("securitiesTransacted") or 0) * (tr.get("price") or 0))
        return v

    # ── Title mapping for importance scoring ──
    TITLE_SCORE = {
        "ceo": 10, "chief executive": 10, "president": 8, "coo": 7,
        "cfo": 7, "chief financial": 7, "chairman": 8, "cto": 6,
        "director": 5, "vp": 4, "vice president": 4, "officer": 3,
    }
    def _title_score(title):
        t = title.lower()
        for k, v in TITLE_SCORE.items():
            if k in t: return v
        return 2

    # ── Group by ticker ──
    ticker_trades = defaultdict(list)
    for trade in insider_data:
        t = trade.get("symbol", "")
        ticker_trades[t].append(trade)

    # ── Score each company ──
    company_rows = []
    for t, trades in ticker_trades.items():
        s = stocks.get(t, {})
        total_val  = sum(_trade_value(tr) for tr in trades)
        latest     = max((tr.get("transactionDate") or "") for tr in trades)
        n_insiders = len({tr.get("reportingName", "") for tr in trades})
        n_buys     = len(trades)
        max_title  = max((_title_score(tr.get("typeOfOwner") or tr.get("reportingCik","")) for tr in trades), default=2)

        # Conviction score
        score = 0
        score += min(n_buys, 5) * 8           # up to 40 pts for 5+ buys
        score += min(n_insiders - 1, 4) * 10   # cluster bonus: +10 per additional insider
        score += max_title * 3                  # senior insider bonus
        if total_val > 1_000_000: score += 20
        elif total_val > 500_000: score += 12
        elif total_val > 100_000: score += 5
        if s.get("piotroski") and s.get("piotroski") >= 7: score += 5
        if s.get("mos") and s.get("mos") > 0.1: score += 5

        # Cluster label
        if n_insiders >= 3:   cluster = "CLUSTER"
        elif n_insiders == 2: cluster = "DOUBLE"
        else:                 cluster = "SINGLE"

        names = "; ".join(
            f"{tr.get('reportingName','?')[:18]} ({(tr.get('typeOfOwner') or '')[:8]})"
            for tr in sorted(trades, key=lambda x: -_trade_value(x))[:3]
        )

        row = format_stock_row(s) if s else {"Ticker": t, "Company": t}
        row["# Buys"]   = n_buys
        row["# Insiders"] = n_insiders
        row["Cluster"]  = cluster
        row["Total $M"] = round(total_val / 1_000_000, 3)
        row["Latest"]   = latest
        row["Who Bought"] = names
        row["Score"]    = score
        company_rows.append(row)

    company_rows.sort(key=lambda x: -x["Score"])

    sr = add_title(ws,
                   "🏦 Insider Buying — Recent Open Market Purchases",
                   f"Source: {source}. {len(insider_data)} total purchases across "
                   f"{len(company_rows)} companies. "
                   f"Cluster = 3+ insiders buying same stock. {datetime.date.today()}")

    # ── Cluster buys section ──
    clusters = [r for r in company_rows if r["Cluster"] in ("CLUSTER", "DOUBLE")]
    singles  = [r for r in company_rows if r["Cluster"] == "SINGLE"]

    headers = ["Score", "Cluster", "Ticker", "Company", "Sector", "Price",
               "PEG", "P/E", "MoS", "Piotroski", "# Buys", "# Insiders",
               "Total $M", "Latest", "Who Bought"]
    widths = [7, 8, 8, 22, 15, 8, 6, 7, 7, 7, 6, 8, 9, 10, 45]

    if clusters:
        ch = ws.cell(row=sr, column=1,
                     value=f"🔥 CLUSTER BUYS — {len(clusters)} companies with 2+ insiders buying")
        ch.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        ch.fill = PatternFill("solid", fgColor="B71C1C")
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
        ws.row_dimensions[sr].height = 18
        sr += 1
        sr = write_table(ws, clusters[:30], headers, sr, header_color="C62828", widths=widths)
        sr += 2

    sh = ws.cell(row=sr, column=1,
                 value=f"📋 ALL INSIDER PURCHASES — top {min(len(singles), TOP_N)} by conviction score")
    sh.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    sh.fill = PatternFill("solid", fgColor="1B5E20")
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(headers))
    ws.row_dimensions[sr].height = 18
    sr += 1
    write_table(ws, singles[:TOP_N], headers, sr, header_color="2E7D32", widths=widths)

    # Extra formatting: color Cluster column
    # (write_table already handled; add cluster color override here)
    for row in ws.iter_rows(min_row=sr, max_row=ws.max_row):
        for cell in row:
            if cell.value == "CLUSTER":
                cell.fill = PatternFill("solid", fgColor="B71C1C")
                cell.font = Font(bold=True, name="Arial", size=8, color="FFFFFF")
            elif cell.value == "DOUBLE":
                cell.fill = PatternFill("solid", fgColor="E57373")
                cell.font = Font(bold=True, name="Arial", size=8, color="FFFFFF")

    print(f"  ✅ Insider Buying tab done — {len(clusters)} cluster + {len(singles)} single buys "
          f"({len(insider_data)} trades total)")


# ─────────────────────────────────────────────
# HTML DASHBOARD
# ─────────────────────────────────────────────

def build_html_report(stocks, iv_rows, stalwarts, fast_growers, turnarounds,
                      slow_growers, cyclicals, asset_plays, quality_compounders,
                      sector_rows=None, etf_rows=None, ai=None, macro=None,
                      portfolio=None, fmp_call_count=0, ten_baggers=None,
                      agent_perf=None,     # B1: per-agent attribution dict
                      sparklines=None,
                      hold_forever=None,
                      mall=None) -> str:  # 🛍️ Mall Manager picks (Lynch consumer-observable)
    """Generate a self-contained mobile-responsive HTML dashboard.
    Reads the same data structures that feed the Excel — no extra computation.
    Returns the full HTML string.
    """
    today = datetime.date.today().strftime("%Y-%m-%d")
    now   = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── CSS ───────────────────────────────────────────────────────────────
    css = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
       background: #0f1117; color: #e0e0e0; font-size: 14px; }
a { color: #90caf9; }
h1 { font-size: 1.1rem; font-weight: 700; color: #fff; }
h2 { font-size: .85rem; font-weight: 700; color: #90caf9;
     text-transform: uppercase; letter-spacing: .05em; }
.header { background: #1a237e; padding: 12px 16px;
          display: flex; justify-content: space-between; align-items: center; }
.header small { color: #9fa8da; font-size: .75rem; }
.nav { background: #1e1e2e; display: flex; flex-wrap: wrap; gap: 2px;
       padding: 6px 8px; position: sticky; top: 0; z-index: 99;
       border-bottom: 1px solid #333; }
.nav button { background: #2a2a3e; color: #aaa; border: none; border-radius: 4px;
              padding: 5px 10px; font-size: .75rem; cursor: pointer; }
.nav button.active, .nav button:hover { background: #1a237e; color: #fff; }
section { padding: 12px 16px; display: none; }
section.active { display: block; }
.section-title { margin: 0 0 10px; padding: 6px 10px; background: #1a237e;
                 border-radius: 4px; font-size: .8rem; font-weight: 700;
                 color: #fff; text-transform: uppercase; }
.tbl-wrap { overflow-x: auto; -webkit-overflow-scrolling: touch; }
table { border-collapse: collapse; width: 100%; font-size: .78rem;
        white-space: nowrap; }
th { background: #1a237e; color: #fff; padding: 6px 8px;
     text-align: left; cursor: pointer; user-select: none; }
th:hover { background: #283593; }
td { padding: 5px 8px; border-bottom: 1px solid #2a2a2a; }
tr:hover td { background: #1a1a2e !important; }
tr.alt td { background: #161622; }
.g  { background: #1b3a1e !important; color: #a5d6a7; }
.a  { background: #3e2a00 !important; color: #ffe082; }
.r  { background: #3e1111 !important; color: #ef9a9a; }
.badge { display: inline-block; border-radius: 3px; padding: 2px 6px;
         font-size: .7rem; font-weight: 700; }
.badge-high   { background: #1b5e20; color: #a5d6a7; }
.badge-med    { background: #0d47a1; color: #90caf9; }
.badge-actnow { background: #b71c1c; color: #fff; }
.badge-weeks  { background: #e65100; color: #fff; }
.badge-months { background: #1565c0; color: #fff; }
.badge-micro  { background: #4e342e; color: #bcaaa4; }
.badge-small  { background: #1b5e20; color: #c8e6c9; }
.badge-mid    { background: #0d47a1; color: #bbdefb; }
.badge-large  { background: #4a148c; color: #e1bee7; }
.badge-mega   { background: #e65100; color: #fff;    }
.badge-watch  { background: #37474f; color: #cfd8dc; }
.badge-bull   { background: #1b5e20; color: #a5d6a7; }
.badge-bear   { background: #b71c1c; color: #ef9a9a; }
.badge-neut   { background: #006064; color: #b2ebf2; }
.badge-caut   { background: #e65100; color: #ffe0b2; }
.badge-buy    { background: #1b5e20; color: #a5d6a7; }
.badge-hold   { background: #006064; color: #b2ebf2; }
.badge-avoid  { background: #b71c1c; color: #ef9a9a; }
.badge-elev   { background: #e65100; color: #ffe0b2; }
/* A9: Lynch category badges — colour-coded per Peter Lynch category */
.badge-lynch          { background: #283593; color: #c5cae9; font-weight: 600; }
.badge-lynch-fast     { background: #1b5e20; color: #a5d6a7; font-weight: 600; } /* Fast Grower — green */
.badge-lynch-stalwart { background: #0d47a1; color: #90caf9; font-weight: 600; } /* Stalwart — blue   */
.badge-lynch-slow     { background: #37474f; color: #cfd8dc; font-weight: 600; } /* Slow Grower — grey */
.badge-lynch-cyclical { background: #e65100; color: #ffe0b2; font-weight: 600; } /* Cyclical — orange */
.badge-lynch-turn     { background: #6a1b9a; color: #e1bee7; font-weight: 600; } /* Turnaround — purple */
.badge-lynch-asset    { background: #4e342e; color: #d7ccc8; font-weight: 600; } /* Asset Play — brown  */
.macro-grid { display: grid;
              grid-template-columns: repeat(auto-fill, minmax(120px, 1fr));
              gap: 8px; margin: 10px 0; }
.macro-tile { background: #1e1e2e; border-radius: 6px; padding: 10px;
              text-align: center; }
.macro-tile .lbl { font-size: .65rem; color: #9e9e9e; text-transform: uppercase; }
.macro-tile .val { font-size: 1.1rem; font-weight: 700; color: #fff;
                   margin: 4px 0 2px; }
.macro-tile .sig { font-size: .65rem; font-weight: 700; border-radius: 3px;
                   padding: 1px 5px; display: inline-block; }
.mt-clickable { cursor: pointer; transition: background .15s, outline .15s; }
.mt-clickable:hover { background: #252540 !important; }
.mt-active { background: #1a2744 !important; outline: 1px solid #3d5afe; }
.macro-detail { background: #131325; border-radius: 6px; padding: 14px 16px;
                margin: 2px 0 14px; border-left: 3px solid #3d5afe;
                animation: mdFadeIn .18s ease; }
@keyframes mdFadeIn { from { opacity:0; transform:translateY(-5px); } to { opacity:1; transform:none; } }
.sig-g { background: #1b3a1e; color: #a5d6a7; }
.sig-a { background: #3e2a00; color: #ffe082; }
.sig-r { background: #3e1111; color: #ef9a9a; }
.pick-card { background: #1e1e2e; border-radius: 6px; padding: 12px;
             margin-bottom: 10px; border-left: 3px solid #1a237e; }
.pick-card .pick-hdr { display: flex; align-items: center; gap: 8px;
                       flex-wrap: wrap; margin-bottom: 6px; }
.pick-ticker { font-size: 1.1rem; font-weight: 700; color: #fff; }
.pick-co     { color: #9e9e9e; font-size: .82rem; }
.pick-hl     { color: #e0e0e0; margin: 6px 0 4px; font-style: italic; }
.pick-story  { color: #bdbdbd; font-size: .8rem; line-height: 1.5; }
.pick-meta   { display: flex; gap: 12px; flex-wrap: wrap; margin-top: 8px;
               font-size: .75rem; }
.pick-meta span { color: #9e9e9e; }
.pick-meta b   { color: #e0e0e0; }
.interp { background: #1a1a2e; padding: 8px 12px; border-radius: 4px;
          font-size: .78rem; color: #9e9e9e; margin-top: 8px; }
.synopsis { background: #1a237e22; border-left: 3px solid #1a237e;
            padding: 10px 14px; border-radius: 0 4px 4px 0;
            margin-bottom: 12px; font-style: italic; color: #c5cae9;
            font-size: .85rem; line-height: 1.6; }
.outlook-row { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 12px; }
.outlook-tile { border-radius: 6px; padding: 10px 14px; text-align: center;
                min-width: 90px; }
.outlook-tile .o-lbl { font-size: .65rem; text-transform: uppercase; opacity: .8; }
.outlook-tile .o-val { font-size: .95rem; font-weight: 700; margin-top: 3px; }
.agent-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(150px, 1fr));
              gap: 8px; margin: 10px 0 16px; }
.agent-card { background: #1e1e2e; border-radius: 6px; padding: 10px 12px;
              border-left: 3px solid #1a237e; }
.agent-card .ag-name { font-size: .78rem; font-weight: 700; color: #fff; margin-bottom: 6px; }
.agent-card .ag-stat { font-size: .72rem; color: #9e9e9e; line-height: 1.7; }
.agent-card .ag-stat b { color: #e0e0e0; }
.agent-card.win { border-left-color: #1b5e20; }
.agent-card.loss { border-left-color: #b71c1c; }
.agent-section { margin-bottom: 18px; }
.agent-hdr { border-radius: 5px 5px 0 0; padding: 9px 14px; }
.agent-hdr .ag-title { font-size: .9rem; font-weight: 700; color: #fff; }
.agent-hdr .ag-desc { font-size: .75rem; color: rgba(255,255,255,.75); margin-top: 3px; line-height: 1.5; }
.agent-picks-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(280px,1fr));
                    gap: 8px; padding: 10px; background: #16161f; border-radius: 0 0 5px 5px; }
/* Specialist cards now share the mm-card base style; only left border differs per agent */
.agent-pick-card { background: #1e1e2e; border-radius: 4px; padding: 8px 10px;
                   border-left: 2px solid #1a237e; }
.agent-pick-card .ap-ticker { font-weight: 700; color: #fff; font-size: .85rem; }
.agent-pick-card .ap-co { color: #9e9e9e; font-size: .72rem; white-space: nowrap;
                           overflow: hidden; text-overflow: ellipsis; }
.agent-pick-card .ap-rationale { color: #c5cae9; font-size: .72rem; margin-top: 5px;
                                  line-height: 1.45; font-style: italic; }
.agent-pick-card .ap-thesis { color: #9e9e9e; font-size: .70rem; margin-top: 4px; line-height: 1.4; }
.agent-pick-card .ap-metric { font-size: .68rem; color: #78909c; margin-top: 4px; }
.sparkline-wrap { background:#12121e; border-radius:4px; padding:6px 8px; margin-bottom:8px;
                  display:flex; flex-direction:column; gap:4px; }
.sparkline-label { font-size:.60rem; color:#546e7a; display:flex; justify-content:space-between; }
.xcard { cursor: pointer; transition: border-color .15s; }
.xcard:hover { filter: brightness(1.07); }
.xbody { margin-top: 8px; border-top: 1px solid #2a2a3e; padding-top: 8px; }
.xarrow { float: right; font-size: .75rem; color: #546e7a; margin-left: 6px; }
.mm-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(280px,1fr));
           gap: 8px; margin-bottom: 14px; }
.mm-card { background: #1e1e2e; border-radius: 6px; padding: 10px 12px;
           border-left: 3px solid #ffd54f; cursor: pointer; }
.mm-card:hover { filter: brightness(1.08); }
.mm-card .mm-rank { font-size: .72rem; color: #78909c; }
.mm-card .mm-ticker { font-size: 1rem; font-weight: 700; color: #fff; margin-right: 6px; }
.mm-card .mm-co { font-size: .75rem; color: #9e9e9e; }
.mm-card .mm-hl { font-size: .78rem; font-style: italic; color: #c5cae9;
                  margin: 6px 0 4px; line-height: 1.4; }
.mm-card .mm-story { font-size: .75rem; color: #bdbdbd; line-height: 1.5; }
.mm-card .mm-meta { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 6px;
                    font-size: .72rem; }
.mm-card .mm-meta span { color: #9e9e9e; }
.mm-card .mm-meta b { color: #e0e0e0; }
.mm-card .mm-attr { margin-top: 6px; line-height: 1.9; }
"""

    # ── JS (tab switch + table sort) ──────────────────────────────────────
    js = """
function showTab(id, btn) {
  document.querySelectorAll('section').forEach(s => s.classList.remove('active'));
  document.querySelectorAll('.nav button').forEach(b => b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  if (btn) btn.classList.add('active');
}
function sortTable(th) {
  var table = th.closest('table');
  var tbody = table.querySelector('tbody');
  var rows  = Array.from(tbody.querySelectorAll('tr'));
  var col   = Array.from(th.parentNode.children).indexOf(th);
  var asc   = th.dataset.asc !== 'true';
  var dateRe = /^\\d{4}-\\d{2}-\\d{2}$/;
  rows.sort(function(a, b) {
    var av = a.children[col] ? a.children[col].innerText.replace(/[%$,]/g,'').trim() : '';
    var bv = b.children[col] ? b.children[col].innerText.replace(/[%$,]/g,'').trim() : '';
    if (dateRe.test(av) && dateRe.test(bv)) return asc ? av.localeCompare(bv) : bv.localeCompare(av);
    var an = parseFloat(av), bn = parseFloat(bv);
    if (!isNaN(an) && !isNaN(bn)) return asc ? an - bn : bn - an;
    return asc ? av.localeCompare(bv) : bv.localeCompare(av);
  });
  rows.forEach(r => tbody.appendChild(r));
  th.dataset.asc = asc ? 'true' : 'false';
}
document.querySelectorAll('th').forEach(th => th.addEventListener('click', function(){sortTable(this);}));
function toggleExpand(card) {
  var body = card.querySelector('.xbody');
  var arrow = card.querySelector('.xarrow');
  if (!body) return;
  var open = body.style.display !== 'none';
  body.style.display = open ? 'none' : 'block';
  if (arrow) arrow.textContent = open ? '▼' : '▲';
}
function showMacroDetail(el, id) {
  var wasActive = el.classList.contains('mt-active');
  document.querySelectorAll('.macro-detail').forEach(function(d){d.style.display='none';});
  document.querySelectorAll('.macro-tile').forEach(function(t){t.classList.remove('mt-active');});
  if (!wasActive) {
    el.classList.add('mt-active');
    var p = document.getElementById('md-'+id);
    if (p) { p.style.display='block'; }
  }
}
"""

    # ── Helpers ───────────────────────────────────────────────────────────
    def _pct(v, decimals=1):
        if v is None: return "—"
        return f"{v*100:.{decimals}f}%"

    def _num(v, decimals=1):
        if v is None: return "—"
        return f"{v:.{decimals}f}"

    def _money(v):
        if v is None: return "—"
        return f"${v:.2f}"

    def _cell(val, raw, cls=""):
        """Format a table cell with optional color class."""
        display = "—" if val == "—" or val is None else val
        if cls:
            return f'<td class="{cls}">{display}</td>'
        return f"<td>{display}</td>"

    def _peg_cls(v):
        if v is None: return ""
        if v < 1.0:  return "g"
        if v < 1.5:  return "a"
        if v > 2.5:  return "r"
        return ""

    def _roic_cls(v):
        if v is None: return ""
        if v > 0.20: return "g"
        if v > 0.10: return "a"
        return "r"

    def _roe_cls(v):
        if v is None: return ""
        if v > 0.15: return "g"
        if v > 0.08: return "a"
        return "r"

    def _mos_cls(v):
        if v is None: return ""
        if v > 0.30: return "g"
        if v > 0.10: return "a"
        if v < 0:    return "r"
        return ""

    def _sig_cls(sig):
        s = (sig or "").upper()
        if s in ("CALM", "NORMAL", "STEEP", "NEAR_TARGET", "LOW",
                 "HEALTHY", "TIGHT", "DOVISH"):
            return "sig-g"
        if s in ("CAUTION", "FLAT", "ABOVE_TARGET", "ELEVATED",
                 "SOFTENING", "NEUTRAL"):
            return "sig-a"
        if s in ("PANIC", "FEAR", "INVERTED", "HOT", "HIGH",
                 "WEAK", "HAWKISH"):
            return "sig-r"
        return "sig-a"

    def _bias_badge(v):
        v = (v or "NEUTRAL").upper()
        cls = {"BULLISH": "badge-bull", "BEARISH": "badge-bear",
               "CAUTIOUS": "badge-caut"}.get(v, "badge-neut")
        return f'<span class="badge {cls}">{v}</span>'

    def _crash_badge(v):
        v = (v or "ELEVATED").upper()
        cls = {"LOW": "badge-buy", "HIGH": "badge-bear"}.get(v, "badge-elev")
        return f'<span class="badge {cls}">{v}</span>'

    def _signal_badge(v):
        v = (v or "HOLD").upper()
        cls = {"BUY": "badge-buy", "AVOID": "badge-avoid"}.get(v, "badge-hold")
        return f'<span class="badge {cls}">{v}</span>'

    def _urgency_badge(u):
        u = (u or "WATCH").upper()
        cls = {"ACT NOW": "badge-actnow", "WITHIN WEEKS": "badge-weeks",
               "WITHIN MONTHS": "badge-months"}.get(u, "badge-watch")
        return f'<span class="badge {cls}">{u}</span>'

    def _conv_badge(c):
        c = (c or "MEDIUM").upper()
        return f'<span class="badge {"badge-high" if c=="HIGH" else "badge-med"}">{c}</span>'

    def _cap_badge(mktcap):
        """Return a coloured cap-size badge for Master Manager pick cards."""
        mc = mktcap or 0
        if   mc >= 200e9: lbl, cls = "Mega",  "badge-mega"
        elif mc >=  10e9: lbl, cls = "Large", "badge-large"
        elif mc >=   2e9: lbl, cls = "Mid",   "badge-mid"
        elif mc >= 300e6: lbl, cls = "Small", "badge-small"
        elif mc >       0: lbl, cls = "Micro", "badge-micro"
        else: return ""
        return f'<span class="badge {cls}">{lbl} Cap</span>'

    # A9: Lynch-category badge. Takes either the raw lynchCategory string (may
    # contain multiple labels joined with "+") or None. Returns empty string when
    # unknown so call sites degrade gracefully.
    _LYNCH_CLS = {
        "FastGrower":  "badge-lynch-fast",
        "Stalwart":    "badge-lynch-stalwart",
        "SlowGrower":  "badge-lynch-slow",
        "Cyclical":    "badge-lynch-cyclical",
        "Turnaround":  "badge-lynch-turn",
        "AssetPlay":   "badge-lynch-asset",
    }
    _LYNCH_LABEL = {
        "FastGrower":  "Fast Grower",
        "Stalwart":    "Stalwart",
        "SlowGrower":  "Slow Grower",
        "Cyclical":    "Cyclical",
        "Turnaround":  "Turnaround",
        "AssetPlay":   "Asset Play",
    }
    def _lynch_badge(raw):
        """Render one or more Lynch-category badges (multi-label split on '+')."""
        if not raw:
            return ""
        cats = [c.strip() for c in str(raw).split("+") if c.strip()]
        out = []
        for c in cats:
            cls = _LYNCH_CLS.get(c, "badge-lynch")
            lbl = _LYNCH_LABEL.get(c, c)
            out.append(f'<span class="badge {cls}" style="font-size:.63rem">{lbl}</span>')
        return " ".join(out)

    # Sprint 2 A1: 🛒 Familiar Brand badge — surfaces consumer-observable names
    # to help the user spot picks they can evaluate from personal experience.
    def _familiar_badge(stock):
        if stock and stock.get("consumerObservable"):
            return ('<span class="badge" style="font-size:.63rem;background:#4a148c;'
                    'color:#fff;padding:2px 6px;border-radius:3px" '
                    'title="Familiar brand — consumer-observable industry, you can evaluate this from personal experience">'
                    '🛒 Familiar</span>')
        return ""

    # Sprint 3 A2: 🔍 Under-Covered badge — Wall Street has not noticed yet
    def _undercovered_badge(stock):
        if stock and stock.get("underCovered"):
            ac = stock.get("analystCount")
            label = f"🔍 {ac} analyst{'s' if ac != 1 else ''}" if ac else "🔍 Uncovered"
            return ('<span class="badge" style="font-size:.63rem;background:#0d47a1;'
                    'color:#fff;padding:2px 6px;border-radius:3px" '
                    'title="Wall Street has not noticed — low analyst coverage, structural inefficiency to exploit">'
                    f'{label}</span>')
        return ""

    # 👑 CEO Capital Allocator badge (Thorndike Outsiders-style score)
    def _ceo_badge(stock):
        if not stock:
            return ""
        ceo = stock.get("ceoAllocator") or {}
        grade = ceo.get("grade")
        tenure = ceo.get("tenure_years")
        # New CEO (<3yr): show neutral chip
        if not grade and tenure is not None and tenure < 3.0:
            name = (ceo.get("ceo_name") or "")[:32]
            tip = f"New CEO ({tenure:.1f}y) — {name}. Too early to score capital allocation."
            return ('<span class="badge" style="font-size:.63rem;background:#37474f;'
                    f'color:#cfd8dc;padding:2px 6px;border-radius:3px" title="{tip}">'
                    f'👤 New CEO ({tenure:.0f}y)</span>')
        if not grade:
            return ""
        # Color by grade
        if grade in ("A+", "A"):
            bg, fg = "#b8860b", "#fff8e1"     # gold
        elif grade in ("B+", "B"):
            bg, fg = "#1565c0", "#e3f2fd"     # blue
        elif grade in ("C+", "C"):
            bg, fg = "#546e7a", "#eceff1"     # grey
        else:  # D
            bg, fg = "#5d4037", "#efebe9"     # dim brown
        name = (ceo.get("ceo_name") or "")[:30]
        fcf  = ceo.get("fcf_per_share_cagr")
        sh   = ceo.get("shares_change_pct")
        bits = []
        if fcf is not None: bits.append(f"FCF/sh {fcf*100:+.0f}%/y")
        if sh  is not None and abs(sh) >= 0.02: bits.append(f"shares {sh*100:+.0f}%")
        if ceo.get("roic_trend"): bits.append(f"ROI {ceo['roic_trend']}")
        callouts = " · ".join(bits)
        tenure_s = f"{tenure:.0f}y" if tenure else "?"
        tip = f"{name} ({tenure_s}). Score: {ceo.get('score','?')}/100. {callouts}"
        return ('<span class="badge" style="font-size:.63rem;'
                f'background:{bg};color:{fg};padding:2px 6px;border-radius:3px;'
                f'font-weight:700" title="{tip}">'
                f'👑 {grade} ({tenure_s})</span>')

    # 🎯 Divergence badge (Hidden Gem / Conviction Stack / Quiet Signal)
    def _divergence_badge(stock):
        if not stock:
            return ""
        d = stock.get("divergence")
        if d == "hidden_gem":
            return ('<span class="badge" style="font-size:.63rem;background:#bf360c;'
                    'color:#fff3e0;padding:2px 6px;border-radius:3px;font-weight:700" '
                    'title="Insiders are buying significantly while sell-side is rated Hold/Sell — highest-asymmetry signal">'
                    '🎯 Hidden Gem</span>')
        if d == "conviction_stack":
            return ('<span class="badge" style="font-size:.63rem;background:#c62828;'
                    'color:#ffebee;padding:2px 6px;border-radius:3px;font-weight:700" '
                    'title="Insiders + sell-side both bullish — conviction stack, both signals agree">'
                    '🔥 Conviction</span>')
        if d == "quiet_signal":
            return ('<span class="badge" style="font-size:.63rem;background:#455a64;'
                    'color:#cfd8dc;padding:2px 6px;border-radius:3px" '
                    'title="Modest insider buying with neutral/bearish sell-side rating">'
                    '👁️ Quiet</span>')
        return ""

    # Capital Allocation mini-block for expanded cards
    def _capalloc_block(stock):
        if not stock:
            return ""
        ceo = stock.get("ceoAllocator") or {}
        rev = stock.get("revPerShare5yCagr")
        fcf = stock.get("fcfPerShare5yCagr")
        bv  = stock.get("bvPerShare5yCagr")
        if not ceo.get("grade") and rev is None and fcf is None and bv is None:
            return ""
        rows = []
        if ceo.get("grade"):
            _ten = ceo.get("tenure_years")
            _ten_s = f"{_ten:.0f}y" if _ten else "tenure?"
            rows.append(f'<div><b style="color:#ffc107">👑 CEO {ceo["grade"]} ({ceo.get("score","?")}/100)</b> '
                        f'— {ceo.get("ceo_name","")} · {_ten_s}</div>')
        per_bits = []
        if rev is not None: per_bits.append(f"NI: {rev*100:+.0f}%/y")
        if fcf is not None: per_bits.append(f"FCF: {fcf*100:+.0f}%/y")
        if bv  is not None: per_bits.append(f"Equity: {bv*100:+.0f}%/y")
        if per_bits:
            rows.append(f'<div style="color:#b0bec5;font-size:.66rem;margin-top:2px">'
                        f'5Y growth: {" · ".join(per_bits)}</div>')
        for c in (ceo.get("callouts") or [])[:3]:
            color = "#ef9a9a" if c.startswith("⚠") else "#a5d6a7"
            rows.append(f'<div style="color:{color};font-size:.65rem">{c}</div>')
        return ('<div style="background:#1a1a2488;border-left:2px solid #ffc10755;'
                'border-radius:3px;padding:5px 8px;margin-top:6px">'
                '<div style="font-size:.60rem;font-weight:700;color:#ffc107;'
                'text-transform:uppercase;letter-spacing:.05em;margin-bottom:3px">'
                '👑 Capital Allocation</div>'
                + "".join(rows) + '</div>')

    _sparklines_data = sparklines or {}

    def _sparkline_svg(ticker, w=200, h=44):
        """Inline SVG sparkline from 5Y sampled prices. Returns '' if no data."""
        prices = _sparklines_data.get(ticker)
        if not prices or len(prices) < 4:
            return ""
        mn, mx = min(prices), max(prices)
        rng = mx - mn
        if rng == 0:
            return ""
        is_up  = prices[-1] >= prices[0]
        stroke = "#66bb6a" if is_up else "#ef5350"
        # safe ID — no dots/dashes that break SVG id references
        gid    = "sg" + ticker.replace(".", "X").replace("-", "X").replace("/", "X")
        pad    = 3
        n      = len(prices)
        def _xy(i, p):
            x = pad + i / (n - 1) * (w - 2 * pad)
            y = pad + (1 - (p - mn) / rng) * (h - 2 * pad)
            return x, y
        pts  = [_xy(i, p) for i, p in enumerate(prices)]
        poly = " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
        fill = poly + f" {pts[-1][0]:.1f},{h} {pts[0][0]:.1f},{h}"
        lx, ly = pts[-1]
        return (
            f'<svg viewBox="0 0 {w} {h}" width="{w}" height="{h}" '
            f'style="display:block;overflow:visible;border-radius:3px">'
            f'<defs><linearGradient id="{gid}" x1="0" y1="0" x2="0" y2="1">'
            f'<stop offset="0%" stop-color="{stroke}" stop-opacity="0.22"/>'
            f'<stop offset="100%" stop-color="{stroke}" stop-opacity="0"/>'
            f'</linearGradient></defs>'
            f'<polygon points="{fill}" fill="url(#{gid})"/>'
            f'<polyline points="{poly}" fill="none" stroke="{stroke}" '
            f'stroke-width="1.5" stroke-linejoin="round" stroke-linecap="round"/>'
            f'<circle cx="{lx:.1f}" cy="{ly:.1f}" r="2.8" fill="{stroke}"/>'
            f'</svg>'
        )

    def _trend_badge(ticker):
        """Small badge showing 3-month vs prior-3-month price trend."""
        prices = _sparklines_data.get(ticker)
        if not prices or len(prices) < 12:
            return ""
        recent = sum(prices[-4:]) / 4
        prior  = sum(prices[-12:-8]) / 4
        if prior <= 0:
            return ""
        chg = (recent - prior) / prior
        if chg > 0.05:
            return ('<span style="font-size:.60rem;background:#1b5e20;color:#a5d6a7;'
                    'border-radius:3px;padding:1px 5px;white-space:nowrap">▲ Up</span>')
        if chg < -0.05:
            return ('<span style="font-size:.60rem;background:#3e1111;color:#ef9a9a;'
                    'border-radius:3px;padding:1px 5px;white-space:nowrap">▼ Down</span>')
        return ('<span style="font-size:.60rem;background:#1a1a2e;color:#78909c;'
                'border:1px solid #2a2a3e;border-radius:3px;padding:1px 5px;'
                'white-space:nowrap">→ Flat</span>')

    def _strategy_table(rows, cols, title_id, title_label, description=""):
        """Render a strategy tab as a sortable table."""
        if not rows:
            return f'<section id="{title_id}"><p style="color:#666">No data</p></section>'
        desc_html = (f'<p style="background:#1a1a2e;padding:8px 12px;border-radius:4px;'
                     f'font-size:.75rem;color:#9e9e9e;margin-bottom:10px;line-height:1.6">'
                     f'<b style="color:#90caf9">Filter: </b>{description}</p>') if description else ""
        hdr_html = "".join(f"<th>{c}</th>" for c in cols)
        body_rows = []
        for i, r in enumerate(rows):
            alt = ' class="alt"' if i % 2 == 0 else ''
            cells = []
            for c in cols:
                v = r.get(c)
                # Apply color logic per column
                if c in ("PEG", "Fwd PEG"):
                    cells.append(_cell(_num(v), v, _peg_cls(v)))
                elif c == "ROIC":
                    cells.append(_cell(_pct(v), v, _roic_cls(v)))
                elif c in ("ROE", "ROE%"):
                    cells.append(_cell(_pct(v), v, _roe_cls(v)))
                elif c in ("MoS", "MoS%"):
                    cells.append(_cell(_pct(v), v, _mos_cls(v)))
                elif c == "Price":
                    cells.append(_cell(_money(v), v))
                elif c in ("FCF Yield", "Rev Growth", "EPS Growth", "Rev Growth 5Y",
                           "EPS Growth 5Y", "Div Yield", "Gross Margin", "Oper Margin"):
                    cells.append(_cell(_pct(v), v))
                elif c == "Conviction":
                    cells.append(f"<td>{_conv_badge(v)}</td>")
                elif c == "Urgency":
                    cells.append(f"<td>{_urgency_badge(v)}</td>")
                elif c in ("P/E", "P/B", "P/FCF", "EV/EBITDA", "Fwd P/E", "Score", "Rank",
                           "Piotroski", "MktCap ($B)", "Beta", "Net Debt/EBITDA"):
                    cells.append(_cell(_num(v, 1) if isinstance(v, float) else (str(v) if v is not None else "—"), v))
                elif c == "FCF/Sh 5Y":
                    cells.append(_cell(_pct(v) if v is not None else "—", v))
                elif c == "CEO Score":
                    # Color-coded by grade
                    grade_color = ""
                    if v and isinstance(v, str):
                        if v.startswith(("A+", "A ")): grade_color = "color:#ffc107;font-weight:700"
                        elif v.startswith("B"):        grade_color = "color:#90caf9"
                        elif v.startswith("C"):        grade_color = "color:#9e9e9e"
                        elif v.startswith("D"):        grade_color = "color:#ef9a9a"
                        elif v.startswith("👤"):        grade_color = "color:#78909c"
                    cells.append(f'<td style="{grade_color}">{v if v else "—"}</td>')
                elif c == "Divergence":
                    cells.append(f"<td>{v if v else '—'}</td>")
                else:
                    cells.append(f"<td>{v if v is not None else '—'}</td>")
            body_rows.append(f"<tr{alt}>{''.join(cells)}</tr>")
        return f"""
<section id="{title_id}">
  <div class="section-title">{title_label}</div>
  {desc_html}
  <div class="tbl-wrap">
    <table>
      <thead><tr>{hdr_html}</tr></thead>
      <tbody>{"".join(body_rows)}</tbody>
    </table>
  </div>
</section>"""

    # ── AGENT ANALYSIS SECTION ───────────────────────────────────────────
    def _agent_section():
        """HTML tab: each specialist agent's philosophy + individual picks."""
        _sp = (ai or {}).get("_specialist_picks", {})
        if not _sp:
            return '<section id="agents"><p style="color:#666">No agent data available</p></section>'

        _ADESC = {
            "QualityGrowth":  ("🌱 Quality Growth",  "2E7D32",
                "Finds durable compounders with ROIC >15%, consistent multi-year revenue growth, "
                "and structural competitive moats. Prioritises FCF conversion and PEG <1.5 as quality gates."),
            "SpecialSit":     ("⚡ Special Situation", "6A1B9A",
                "Identifies event-driven, misunderstood, and inflection-point opportunities. Focuses on "
                "business model misclassification, regulatory catalysts, and hidden assets not yet priced in."),
            "CapAppreciation":("📈 Capital Appreciation", "1565C0",
                "Finds near-term re-rating candidates with specific catalysts in 1–6 months. Targets "
                "beaten-down quality (52wPos <65%), revenue re-acceleration signals, and cycle trough entries."),
            "EmergingGrowth": ("🚀 Emerging Growth",  "E65100",
                "Identifies smaller fast-growing companies ($100M–$15B) at the early stage of becoming "
                "compounders. Targets 20%+ revenue growth, rising ROIC, and large underserved TAMs."),
            "TenBagger":      ("🎯 10-Bagger Hunter", "BF360D",
                "Peter Lynch-style small-cap hunter. Looks for underfollowed companies ($50M–$2B) with "
                "15–40% EPS+revenue growth, simple competitive advantages, at least 2 expansion levers, and "
                "structural Wall Street under-coverage (orphans, post-restructuring, complex/sin-sector names)."),
            "LynchBWYK":      ("🛒 Lynch Buy What You Know","880E4F",
                "Lynch's 'buy what you know' approach — simple, understandable businesses serving everyday "
                "consumer or workplace needs, growing 15–25%/yr with real earnings at PEG <1.0."),
            "CathieWood":     ("🚀 Disruptive Innovation","0D47A1",
                "ARK Invest-style innovation screen. Identifies pure-play companies in AI, robotics, "
                "genomics, energy storage, or blockchain riding Wright's Law cost curves with network effects."),
            "Pabrai":         ("🎲 Pabrai Asymmetric Bet","33691E",
                "Mohnish Pabrai's 'heads I win, tails I don't lose much' framework. Finds bets with "
                "3:1+ upside/downside asymmetry where downside is protected by real assets or essential value."),
            "HowardMarks":    ("🔄 Marks Second-Level","827717",
                "Howard Marks' second-level thinking — contrarian analysis where market consensus is factually "
                "wrong. Targets oversold stocks where the negative narrative exceeds actual deterioration."),
            "Burry":          ("🕳️ Burry Deep Value",  "3E2723",
                "Michael Burry-style catalyst-driven deep value. Identifies hidden assets, temporary earnings "
                "distortions, and specific upcoming events that will FORCE the market to reprice."),
            "InsiderTrack":   ("👁️ Insider & Smart Money","263238",
                "Tracks cluster insider buying (multiple executives buying simultaneously) and significant "
                "open-market purchases. Insiders are the most informed buyers with personal-capital conviction."),
        }
        _AGENT_ORDER = ["QualityGrowth", "SpecialSit", "CapAppreciation", "EmergingGrowth", "TenBagger",
                        "LynchBWYK", "CathieWood",
                        "Pabrai", "HowardMarks", "Burry", "InsiderTrack"]

        sections_html = []
        for ak in _AGENT_ORDER:
            if ak not in _sp:
                continue
            sr_data = _sp[ak]
            agent_picks = sr_data.get("picks", [])
            label, color_hex, desc = _ADESC.get(ak, (ak, "37474F", ""))

            pick_cards_html = []
            for pp in agent_picks:
                tk = pp.get("ticker","")
                s2 = stocks.get(tk, {})
                mc_b = s2.get("mktCapB")
                prc  = s2.get("price")
                conv2 = (pp.get("conviction") or "MEDIUM").upper()
                conv_color = "#1b5e20" if conv2 == "HIGH" else "#0d47a1"
                rationale2 = pp.get("rationale", "") or pp.get("brief_case", "")
                # A9: Lynch category badge from assembled stock dict
                _lb = _lynch_badge(s2.get("lynchCategory"))
                pick_cards_html.append(f"""
<div class="agent-pick-card" style="border-left-color:#{color_hex}">
  <div style="display:flex;justify-content:space-between;align-items:center">
    <span class="ap-ticker">{tk}</span>
    <span style="font-size:.65rem;background:{conv_color};color:#fff;border-radius:3px;padding:1px 5px">{conv2}</span>
  </div>
  <div class="ap-co">{pp.get("company",s2.get("name",tk))[:32]}</div>
  {f'<div style="margin-top:4px">{_lb}</div>' if _lb else ''}
  <div class="ap-rationale">{rationale2[:200]}</div>
  <div class="ap-metric">{pp.get("key_metric","")[:60]}{"  ·  $"+f"{prc:.0f}" if prc else ""}{"  ·  $"+f"{mc_b:.1f}B" if mc_b else ""}</div>
</div>""")

            sections_html.append(f"""
<div class="agent-section">
  <div class="agent-hdr" style="background:#{color_hex}">
    <div class="ag-title">{label}</div>
    <div class="ag-desc">{desc}</div>
  </div>
  <div class="agent-picks-grid">{"".join(pick_cards_html)}</div>
</div>""")

        n_agents = len(sections_html)
        return f"""
<section id="agents">
  <div class="section-title">🔬 Agent Analysis — {n_agents} Specialist Reports</div>
  <p style="background:#1a1a2e;padding:8px 12px;border-radius:4px;font-size:.75rem;
     color:#9e9e9e;margin-bottom:14px;line-height:1.6">
    <b style="color:#90caf9">{n_agents} specialist agents</b> each apply a distinct investment philosophy
    to the same universe of {len(stocks):,} stocks. Their unfiltered picks are synthesised by the
    <b style="color:#90caf9">Master Manager</b> into the final <b style="color:#90caf9">AI Top Picks</b>.
  </p>
  {"".join(sections_html)}
</section>"""

    # ── NAV TABS ──────────────────────────────────────────────────────────
    tabs = [
        ("ai",       "🤖 AI Analysis"),
        ("macro",    "🌍 Macro"),
        ("hold",     "💎 Hold Forever"),
        ("qual",     "🏆 Quality Comp."),
        ("fastg",    "🚀 Fast Growers"),
        ("tenb",     "🎯 10-Baggers"),
        ("turn",     "🔁 Turnarounds"),
        ("asset",    "🏗 Asset Plays"),
        ("cycl",     "🔄 Cyclicals"),
        ("slowg",    "🐢 Slow Growers"),
        ("sector",   "🗺 Sectors"),
        ("perf",     "📈 Performance"),
        # Legacy tabs (still rendered to allow deep-links / Excel cross-reference but not in primary nav)
        # ("iv",       "📊 IV Discount"),    # B4: merged into Quality Compounders
        # ("stalwart", "🏛 Stalwarts"),       # B4: merged into Quality Compounders
    ]
    nav_html = "\n".join(
        f'<button onclick="showTab(\'{tid}\', this)" class="{"active" if i==0 else ""}">{label}</button>'
        for i, (tid, label) in enumerate(tabs)
    )

    # ── MACRO SECTION ─────────────────────────────────────────────────────
    def _macro_section():
        if not macro:
            return '<section id="macro"><p style="color:#666">Macro data unavailable</p></section>'
        mc = macro

        def _dgs2_sig(v):
            if v is None: return "UNKNOWN"
            return "HIGH" if v>5 else "ELEVATED" if v>4 else "NORMAL" if v>2 else "LOW"
        def _ff_sig(v):
            if v is None: return "UNKNOWN"
            return "HAWKISH" if v>4.5 else "ELEVATED" if v>3.5 else "NORMAL" if v>2 else "DOVISH"

        # ── SVG sparkline helpers ──────────────────────────────────────────
        def _sparkline_svg(hist_key, color="#90caf9", w=230, h=72, ref_val=None):
            """Large chart for the detail panel."""
            pts = [(d, v) for d, v in mc.get(hist_key, [])
                   if v is not None and not (isinstance(v, float) and math.isnan(v))]
            if len(pts) < 3:
                return (f'<svg width="{w}" height="{h}">'
                        f'<text x="10" y="22" fill="#555" font-size="11">No history cached — '
                        f'will appear after next data refresh</text></svg>')
            vals = [v for _, v in pts]
            mn, mx = min(vals), max(vals)
            rng = max(mx - mn, 0.001)
            pad = 10
            coords = []
            for i, (_, v) in enumerate(pts):
                x = pad + (i / (len(pts) - 1)) * (w - 2 * pad)
                y = h - pad - ((v - mn) / rng) * (h - 2 * pad)
                coords.append((x, y))
            poly  = " ".join(f"{x:.1f},{y:.1f}" for x, y in coords)
            area  = f"{coords[0][0]:.1f},{h} {poly} {coords[-1][0]:.1f},{h}"
            ex, ey = coords[-1]
            ref_svg = ""
            if ref_val is not None and mn <= ref_val <= mx:
                ry = h - pad - ((ref_val - mn) / rng) * (h - 2 * pad)
                ref_svg = (
                    f'<line x1="{pad}" y1="{ry:.1f}" x2="{w-pad}" y2="{ry:.1f}" '
                    f'stroke="#444" stroke-width="1" stroke-dasharray="4,3"/>'
                    f'<text x="{w-pad+3}" y="{ry+3:.1f}" font-size="8" fill="#666">{ref_val}</text>'
                )
            lbl_svg = (
                f'<text x="{pad}" y="{h-1}" font-size="8" fill="#555">{mn:.2f}</text>'
                f'<text x="{w-pad}" y="{h-1}" font-size="8" fill="#555" text-anchor="end">{mx:.2f}</text>'
            )
            return (
                f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}" style="display:block;max-width:100%">'
                f'<polygon points="{area}" fill="{color}" fill-opacity="0.10"/>'
                f'<polyline points="{poly}" fill="none" stroke="{color}" stroke-width="1.8" '
                f'stroke-linecap="round" stroke-linejoin="round"/>'
                f'{ref_svg}'
                f'<circle cx="{ex:.1f}" cy="{ey:.1f}" r="3" fill="{color}"/>'
                f'{lbl_svg}'
                f'</svg>'
            )

        def _mini_spark(hist_key, color="#90caf9"):
            """Tiny 78×20 sparkline inside the collapsed tile."""
            pts = [(d, v) for d, v in mc.get(hist_key, [])
                   if v is not None and not (isinstance(v, float) and math.isnan(v))]
            pts = pts[-90:]
            if len(pts) < 3:
                return ""
            vals = [v for _, v in pts]
            mn, mx = min(vals), max(vals)
            rng = max(mx - mn, 0.001)
            tw, th, pad = 78, 20, 2
            coords = []
            for i, (_, v) in enumerate(pts):
                x = pad + (i / (len(pts) - 1)) * (tw - 2 * pad)
                y = th - pad - ((v - mn) / rng) * (th - 2 * pad)
                coords.append((x, y))
            poly = " ".join(f"{x:.1f},{y:.1f}" for x, y in coords)
            area = f"{coords[0][0]:.1f},{th} {poly} {coords[-1][0]:.1f},{th}"
            ex, ey = coords[-1]
            return (
                f'<svg width="{tw}" height="{th}" viewBox="0 0 {tw} {th}" '
                f'style="display:block;margin:5px auto 0">'
                f'<polygon points="{area}" fill="{color}" fill-opacity="0.12"/>'
                f'<polyline points="{poly}" fill="none" stroke="{color}" stroke-width="1.3"/>'
                f'<circle cx="{ex:.1f}" cy="{ey:.1f}" r="2.2" fill="{color}"/>'
                f'</svg>'
            )

        def _hist_compare(hist_key, cur_val):
            """Table showing 3M / 6M / 1Y ago values vs current."""
            pts = [(d, v) for d, v in mc.get(hist_key, [])
                   if v is not None and not (isinstance(v, float) and math.isnan(v))]
            if not pts or cur_val is None:
                return ""
            as_of_str = mc.get("as_of", "")
            try:
                ref_dt = (datetime.datetime.strptime(as_of_str, "%Y-%m-%d")
                          if as_of_str else datetime.datetime.now())
            except Exception:
                ref_dt = datetime.datetime.now()

            def _parse_d(s):
                for fmt in ("%Y-%m-%d", "%Y%m%d"):
                    try: return datetime.datetime.strptime(str(s), fmt)
                    except: pass
                return None

            def _closest(target):
                best_dt, best_v = None, None
                for d, v in pts:
                    dt = _parse_d(d)
                    if dt is None: continue
                    if best_dt is None or abs((dt-target).days) < abs((best_dt-target).days):
                        best_dt, best_v = dt, v
                return best_v

            ago3  = _closest(ref_dt - datetime.timedelta(days=91))
            ago6  = _closest(ref_dt - datetime.timedelta(days=182))
            ago12 = _closest(ref_dt - datetime.timedelta(days=365))

            def _row(lbl, old_v):
                if old_v is None:
                    return f'<tr><td style="color:#666;padding-right:14px">{lbl}</td><td>—</td><td>—</td></tr>'
                chg = cur_val - old_v
                cls = "color:#a5d6a7" if chg < 0 else "color:#ef9a9a" if chg > 0 else "color:#9e9e9e"
                return (f'<tr>'
                        f'<td style="color:#9e9e9e;padding-right:14px">{lbl}</td>'
                        f'<td style="padding-right:10px">{old_v:.2f}</td>'
                        f'<td style="{cls}">{chg:+.2f}</td>'
                        f'</tr>')

            return (
                '<table style="font-size:.71rem;border-collapse:collapse;line-height:1.9">'
                '<thead><tr>'
                '<th style="color:#666;font-weight:400;text-align:left;padding-right:14px">Period</th>'
                '<th style="color:#666;font-weight:400;padding-right:10px">Value</th>'
                '<th style="color:#666;font-weight:400">Δ vs now</th>'
                '</tr></thead><tbody>'
                + _row("3M ago", ago3)
                + _row("6M ago", ago6)
                + _row("1Y ago", ago12)
                + '</tbody></table>'
            )

        # ── Signal → colour map ────────────────────────────────────────────
        _SIG_COLOR = {
            "HIGH": "#ef5350", "ELEVATED": "#ff9800", "NORMAL": "#66bb6a", "LOW": "#42a5f5",
            "HAWKISH": "#ef5350", "DOVISH": "#66bb6a",
            "HOT": "#ef5350", "ABOVE_TARGET": "#ff9800", "NEAR_TARGET": "#ffee58",
            "INVERTED": "#ef5350", "FLAT": "#ff9800", "STEEP": "#42a5f5",
            "PANIC": "#ef5350", "FEAR": "#ff9800", "CAUTION": "#ffee58", "CALM": "#66bb6a",
            "TIGHT": "#66bb6a", "HEALTHY": "#66bb6a", "SOFTENING": "#ff9800", "WEAK": "#ef5350",
            "UNKNOWN": "#9e9e9e",
        }

        # ── Investor context per indicator × signal ────────────────────────
        _CTX = {
            "dgs10": {
                "HIGH":     "High 10Y yields compress equity multiples via DCF — every 1% rise cuts fair value ~15% for growth stocks. Watch for credit market stress and spread widening.",
                "ELEVATED": "Above-neutral 10Y creates headwinds for high-P/E growth stocks. Bond competition for yield-seeking capital is tangible at these levels.",
                "NORMAL":   "Goldilocks zone. 10Y in 2–4% range is historically compatible with solid equity returns and reasonable multiples across sectors.",
                "LOW":      "Ultra-low yields push investors into equities for returns. Supports very high multiples but signals weak nominal growth expectations.",
            },
            "dgs2": {
                "HIGH":     "High short rates signal Fed has not finished tightening. Restrictive for leveraged companies and credit. Watch for HY spread widening.",
                "ELEVATED": "2Y yield elevated — markets pricing restrictive Fed. Rate cut expectations drive the narrative more than the absolute level.",
                "NORMAL":   "2Y yield near neutral — Fed in pause or mild easing mode. Supports credit conditions and risk-asset re-rating.",
                "LOW":      "Very low 2Y = market pricing aggressive cuts, often signals recession concern. Historically coincides with equity volatility.",
            },
            "yield_curve": {
                "INVERTED":  "⚠️ Inverted yield curve is the #1 historical recession predictor (6–18mo lead time). Not a timing signal — a risk flag. Raise quality in portfolio.",
                "FLAT":      "Flat curve signals slowing growth or Fed near peak rates. Historically precedes slowdown. Mid-cycle defensive tilt warranted.",
                "NORMAL":    "Normal upward-sloping curve = healthy growth expectations. Supports bank profitability, lending activity, and credit creation.",
                "STEEP":     "Steep curve signals strong growth expectations or early recovery from inversion. Historically bullish for cyclicals and financials.",
            },
            "vix": {
                "PANIC":     "VIX>40: Extreme fear. Historically one of the best medium-term entry signals — 12M forward S&P returns average +30% from these levels. Short-term pain persists.",
                "FEAR":      "VIX 25–40: Elevated fear. Markets pricing significant tail risk. 12M forward returns historically above average — phased buying is opportunistic.",
                "CAUTION":   "VIX 18–25: Moderate uncertainty. Market is alert but not panicked. Avoid aggressive new sizing; watch for directional break.",
                "CALM":      "VIX<18: Low fear / complacency. Poor timing for new aggressive longs when VIX dips below 15 — historically precedes a spike.",
            },
            "fedfunds": {
                "HAWKISH":   "Fed funds still elevated / hawkish. Restrictive conditions drag on housing, CapEx, and leveraged companies. Rate-cut timing is the key re-rating catalyst.",
                "ELEVATED":  "Rate cycle near peak. Equities typically rally 6–12mo after the last hike as forward rate expectations improve. First-cut timing closely watched.",
                "NORMAL":    "Neutral Fed funds (2–3.5%). Neither stimulative nor restrictive. Standard backdrop — company fundamentals drive returns.",
                "DOVISH":    "Stimulative rates. Risk assets historically outperform in first 12mo of easing cycle. Small caps and long-duration growth stocks benefit most.",
            },
            "cpi_yoy": {
                "HOT":       "CPI well above 2% target. Fed has limited room to cut — inflation persistence is bearish for rate-sensitive assets. Pricing power stocks outperform.",
                "ABOVE_TARGET": "CPI above target but decelerating. Fed in hold mode. Watch for disinflation confirmation to become bullish for growth stocks.",
                "NEAR_TARGET":  "CPI approaching Fed's 2% target. Rate cuts become actionable. Historically bullish for long-duration assets and growth names.",
                "LOW":       "Below-target inflation. Fed free to cut aggressively. Very supportive for bonds and high-multiple growth stocks.",
            },
            "unrate": {
                "TIGHT":     "Very tight labor market (<4%). Consumer spending strong; wage inflation risk keeps Fed cautious. Good backdrop for consumer discretionary.",
                "HEALTHY":   "Healthy unemployment (4–5%). GDP growing, consumer confident. Standard pro-growth backdrop for equities across sectors.",
                "SOFTENING": "Labor market softening — cracks in employment signal late-cycle dynamics. Monitor for acceleration; reduce cyclical overweight.",
                "WEAK":      "Rising unemployment signals recessionary pressure. Consumer spending declines. Shift to defensive sectors: healthcare, utilities, staples.",
            },
        }

        # ── Indicator definitions ──────────────────────────────────────────
        _vc_str = ((f"+{mc['yield_curve']}%" if mc.get('yield_curve', 0) >= 0
                    else f"{mc.get('yield_curve', '—')}%")
                   if mc.get('yield_curve') is not None else "—")
        indicators = [
            ("dgs10",       "10Y YIELD",   f"{mc.get('dgs10','—')}%",   mc.get('rate_signal','?'),      "_hist_dgs10",    mc.get('dgs10'),       None),
            ("dgs2",        "2Y YIELD",    f"{mc.get('dgs2','—')}%",    _dgs2_sig(mc.get('dgs2')),      "_hist_dgs2",     mc.get('dgs2'),        None),
            ("yield_curve", "YIELD CURVE", _vc_str,                     mc.get('curve_signal','?'),     "_hist_t10y2y",   mc.get('yield_curve'), 0),
            ("vix",         "VIX",         str(mc.get('vix','—')),      mc.get('vix_signal','?'),       "_hist_vix",      mc.get('vix'),         20),
            ("fedfunds",    "FED FUNDS",   f"{mc.get('fedfunds','—')}%", _ff_sig(mc.get('fedfunds')),   "_hist_fedfunds", mc.get('fedfunds'),    2.5),
            ("cpi_yoy",     "CPI YoY",     f"{mc.get('cpi_yoy','—')}%", mc.get('inflation_signal','?'), "_hist_cpi_yoy",  mc.get('cpi_yoy'),     2),
            ("unrate",      "UNEMPLOYMT",  f"{mc.get('unrate','—')}%",  mc.get('labor_signal','?'),     "_hist_unrate",   mc.get('unrate'),      4),
        ]

        # ── Build tiles + detail panels ────────────────────────────────────
        tiles_html   = ""
        detail_panels = ""
        for key, lbl, val_str, sig, hist_key, cur_val, ref_val in indicators:
            sc    = _sig_cls(sig)
            color = _SIG_COLOR.get(sig, "#90caf9")
            mini  = _mini_spark(hist_key, color)
            spark = _sparkline_svg(hist_key, color, ref_val=ref_val)
            htbl  = _hist_compare(hist_key, cur_val)
            ctx   = _CTX.get(key, {}).get(sig, "")

            tiles_html += (
                f'<div class="macro-tile mt-clickable" onclick="showMacroDetail(this,\'{key}\')">'
                f'<div class="lbl">{lbl}</div>'
                f'<div class="val">{val_str}</div>'
                f'<span class="sig {sc}">{sig}</span>'
                f'{mini}'
                f'</div>'
            )
            ctx_html = (f'<p style="margin-top:12px;font-size:.78rem;color:#cfd8dc;line-height:1.65;'
                        f'border-left:3px solid {color};padding-left:10px">{ctx}</p>') if ctx else ""
            detail_panels += (
                f'<div class="macro-detail" id="md-{key}" style="display:none">'
                f'<div style="font-size:.68rem;color:#9e9e9e;margin-bottom:10px;'
                f'text-transform:uppercase;letter-spacing:.06em">📊 {lbl} — historical chart &amp; context</div>'
                f'<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start">'
                f'<div style="flex:2;min-width:180px">{spark}</div>'
                f'<div style="flex:1;min-width:140px">{htbl}</div>'
                f'</div>'
                f'{ctx_html}'
                f'</div>'
            )

        # ── Stock Style Regime signal ──────────────────────────────────────
        _rate_sig  = mc.get("rate_signal", "")
        _vix_sig   = mc.get("vix_signal", "")
        _curve_sig = mc.get("curve_signal", "")
        _cpi_sig   = mc.get("inflation_signal", "")

        # Score each dimension: positive = growth-friendly, negative = value/defensive
        _style_score = 0
        if _rate_sig in ("LOW",):              _style_score += 2
        elif _rate_sig == "NORMAL":            _style_score += 1
        elif _rate_sig == "ELEVATED":          _style_score -= 1
        elif _rate_sig == "HIGH":              _style_score -= 2
        if _cpi_sig in ("LOW", "NEAR_TARGET"): _style_score += 1
        elif _cpi_sig == "ABOVE_TARGET":       _style_score -= 1
        elif _cpi_sig == "HOT":                _style_score -= 2
        if _curve_sig == "INVERTED":           _style_score -= 2
        elif _curve_sig == "FLAT":             _style_score -= 1
        elif _curve_sig in ("NORMAL", "STEEP"): _style_score += 1
        if _vix_sig in ("PANIC", "FEAR"):      _style_score -= 2
        elif _vix_sig == "CAUTION":            _style_score -= 1
        elif _vix_sig == "CALM":               _style_score += 1

        # Map score to regime
        if _style_score >= 3:
            _regime_label = "FAVOR GROWTH"
            _regime_color = "#42a5f5"
            _regime_bg    = "#0d47a122"
            _regime_icon  = "🚀"
            _regime_desc  = "Low yields + benign inflation = growth stocks outperform. High-P/E multiples supported; favor tech, biotech, emerging growth."
        elif _style_score >= 1:
            _regime_label = "BALANCED"
            _regime_color = "#66bb6a"
            _regime_bg    = "#1b5e2022"
            _regime_icon  = "⚖️"
            _regime_desc  = "Mixed signals. Both growth and value can work — quality and earnings consistency matter most. Avoid extreme duration bets."
        elif _style_score >= -1:
            _regime_label = "FAVOR VALUE / CASH FLOW"
            _regime_color = "#ff9800"
            _regime_bg    = "#e65100" + "22"
            _regime_icon  = "💰"
            _regime_desc  = "Elevated yields raise the hurdle rate for speculative growth. Cash-flow-generative companies and value names have a structural tailwind."
        elif _style_score >= -3:
            _regime_label = "DEFENSIVE — VOLATILITY RISK"
            _regime_color = "#ef5350"
            _regime_bg    = "#b71c1c22"
            _regime_icon  = "🛡️"
            _regime_desc  = "High rates + elevated VIX or curve stress. Favor quality compounders, dividend payers, and cash-generative defensives. Reduce speculative exposure."
        else:
            _regime_label = "RISK-OFF"
            _regime_color = "#ef5350"
            _regime_bg    = "#b71c1c33"
            _regime_icon  = "⚠️"
            _regime_desc  = "Multiple red flags: inverted curve, elevated inflation, and market stress. Historically strong signal to raise quality and reduce cyclical / high-leverage exposure."

        _dgs10_val = mc.get("dgs10")

        # Sector guidance per regime
        _SECTOR_GUIDANCE = {
            "FAVOR GROWTH": {
                "favor":  ["Technology", "Biotech / Healthcare Innovation", "Consumer Discretionary", "Emerging Growth"],
                "avoid":  ["Utilities", "Telecom (dividend plays)", "Traditional Banks"],
                "note":   "Low rates compress the discount rate — long-duration earnings are worth more today.",
            },
            "BALANCED": {
                "favor":  ["Quality Compounders (any sector)", "Financials", "Industrials"],
                "avoid":  ["Highly leveraged names", "Speculative unprofitable growth"],
                "note":   "Earnings quality and FCF consistency beat style tilts in neutral regimes.",
            },
            "FAVOR VALUE / CASH FLOW": {
                "favor":  ["Banks & Insurers", "Energy", "Defensive dividend companies", "Cash-generating value stocks"],
                "avoid":  ["High-P/E unprofitable growth", "Long-duration tech", "Speculative biotech"],
                "note":   "Elevated yields mean bond alternatives compete with equities — cash flow today beats growth tomorrow.",
            },
            "DEFENSIVE — VOLATILITY RISK": {
                "favor":  ["Healthcare", "Consumer Staples", "Utilities", "Banks (net interest margin)"],
                "avoid":  ["High-beta cyclicals", "Leveraged companies", "Unprofitable growth"],
                "note":   "Stress environments reward quality balance sheets and predictable earnings.",
            },
            "RISK-OFF": {
                "favor":  ["Cash / short-duration bonds", "Defensive staples", "Gold / commodity producers"],
                "avoid":  ["Cyclicals", "Leveraged buyouts / high-debt", "Speculative growth"],
                "note":   "Preserve capital first. History shows patience is rewarded — best entries follow peak fear.",
            },
        }
        _sg = _SECTOR_GUIDANCE.get(_regime_label, {})
        _favor_tags = "".join(
            f'<span style="background:#1b5e2044;border:1px solid #66bb6a44;color:#a5d6a7;'
            f'border-radius:3px;padding:2px 6px;font-size:.65rem;margin:2px 2px 0 0;display:inline-block">'
            f'✓ {s}</span>' for s in _sg.get("favor", [])
        )
        _avoid_tags = "".join(
            f'<span style="background:#b71c1c22;border:1px solid #ef535044;color:#ef9a9a;'
            f'border-radius:3px;padding:2px 6px;font-size:.65rem;margin:2px 2px 0 0;display:inline-block">'
            f'✗ {s}</span>' for s in _sg.get("avoid", [])
        )
        _sector_block = (
            f'<div style="flex:2;min-width:220px;border-left:1px solid #ffffff11;padding-left:14px">'
            f'<div style="font-size:.60rem;font-weight:700;color:#78909c;text-transform:uppercase;'
            f'letter-spacing:.07em;margin-bottom:5px">Sectors in this regime</div>'
            f'<div style="margin-bottom:4px"><span style="font-size:.62rem;color:#78909c">TEND TO OUTPERFORM &nbsp;</span>'
            f'{_favor_tags}</div>'
            f'<div><span style="font-size:.62rem;color:#78909c">HEADWINDS &nbsp;</span>'
            f'{_avoid_tags}</div>'
            + (f'<div style="font-size:.65rem;color:#546e7a;margin-top:6px;font-style:italic">{_sg["note"]}</div>'
               if _sg.get("note") else "")
            + f'</div>'
        ) if _sg else ""

        _thumb_lines = [
            ("🔵 Low yields (<3%)",     "Growth / tech outperform"),
            ("🟢 Normal (3–4%)",        "Balanced — quality wins"),
            ("🟡 Elevated (4–5%)",      "Value, banks, energy edge"),
            ("🔴 High yields (>5%)",    "Defensives; compress multiples"),
        ]
        _thumb_html = "".join(
            f'<div style="display:flex;gap:6px;font-size:.66rem;color:#b0bec5;margin-bottom:3px">'
            f'<span>{lv}</span><span style="color:#546e7a">→</span><span>{rv}</span></div>'
            for lv, rv in _thumb_lines
        )

        _regime_bar = (
            f'<div style="margin:10px 0 8px;background:{_regime_bg};border-left:3px solid {_regime_color};'
            f'border-radius:4px;padding:10px 14px">'
            f'<div style="font-size:.62rem;font-weight:700;color:#78909c;text-transform:uppercase;'
            f'letter-spacing:.07em;margin-bottom:8px">📐 Market Regime &amp; Stock Style Signal</div>'
            f'<div style="display:flex;gap:16px;flex-wrap:wrap;align-items:flex-start">'
            # Col 1: regime label + description
            f'<div style="flex:2;min-width:190px">'
            f'<div style="font-size:.92rem;font-weight:700;color:{_regime_color};margin-bottom:4px">'
            f'{_regime_icon} {_regime_label}</div>'
            f'<div style="font-size:.72rem;color:#b0bec5;line-height:1.55">{_regime_desc}</div>'
            f'<div style="font-size:.65rem;color:#546e7a;margin-top:6px">'
            f'Current 10Y: <b style="color:{_regime_color}">{_dgs10_val}%</b> ({_rate_sig.lower()})</div>'
            f'</div>'
            # Col 2: sector guidance
            + _sector_block +
            # Col 3: rule of thumb cheat-sheet
            f'<div style="flex:1;min-width:170px;border-left:1px solid #ffffff11;padding-left:14px">'
            f'<div style="font-size:.60rem;font-weight:700;color:#78909c;text-transform:uppercase;'
            f'letter-spacing:.07em;margin-bottom:6px">Simple rule of thumb</div>'
            f'{_thumb_html}'
            f'</div>'
            f'</div>'
            f'</div>'
        )

        # ── Interpretation bar ─────────────────────────────────────────────
        parts = []
        vc = mc.get("yield_curve")
        cs = mc.get("curve_signal", "")
        if vc is not None:
            if cs == "INVERTED":
                parts.append(f"⚠️ Yield curve INVERTED ({vc:+.2f}%) — historically precedes recession 6-18mo")
            elif cs == "FLAT":
                parts.append(f"⚠️ Yield curve flat ({vc:+.2f}%) — slowdown risk")
            else:
                parts.append(f"✅ Yield curve {cs.lower()} ({vc:+.2f}%)")
        if mc.get("vix"):      parts.append(f"VIX {mc['vix']} ({mc.get('vix_signal','?').lower()})")
        if mc.get("cpi_yoy"):  parts.append(f"CPI {mc['cpi_yoy']}% YoY")
        if mc.get("fedfunds"): parts.append(f"Fed Funds {mc['fedfunds']}%")
        interp = "  ·  ".join(parts)

        # ── AI interpretation row ──────────────────────────────────────────
        ai_md = (ai or {}).get("macro_dashboard", {})
        ai_rows = ""
        if ai_md:
            rr  = (ai_md.get("recession_risk") or "").split()[0].upper()
            fp  = (ai_md.get("fed_policy") or "NEUTRAL").upper()
            re_ = ai_md.get("rate_environment", "")
            rr_cls = {"LOW": "badge-buy", "HIGH": "badge-bear"}.get(rr, "badge-elev")
            fp_cls = {"DOVISH": "badge-buy", "HAWKISH": "badge-bear"}.get(fp, "badge-neut")
            ai_rows = (
                f'<div style="margin-top:12px">'
                f'<span style="font-size:.75rem;color:#9e9e9e">AI INTERPRETATION &nbsp;</span>'
                f'<span class="badge {rr_cls}">RECESSION RISK: {rr}</span> &nbsp;'
                f'<span class="badge {fp_cls}">FED POLICY: {fp}</span>'
                + (f'<p class="interp" style="margin-top:8px">{re_}</p>' if re_ else "")
                + '</div>'
            )

        # ── AI market outlook block ────────────────────────────────────────
        _mo = (ai or {}).get("market_outlook", {})
        _outlook_block = ""
        if _mo:
            _nt  = _mo.get("near_term_bias", "NEUTRAL")
            _lt  = _mo.get("long_term_bias",  "NEUTRAL")
            _cr  = _mo.get("crash_risk",      "ELEVATED")
            _rat = _mo.get("rationale", "")
            _outlook_block = (
                '<div style="margin-bottom:14px">'
                '<h2 style="margin-bottom:8px;font-size:.75rem">AI MARKET OUTLOOK</h2>'
                '<div class="outlook-row">'
                f'<div class="outlook-tile" style="background:#1a237e22">'
                f'<div class="o-lbl">NEAR-TERM</div><div class="o-val">{_bias_badge(_nt)}</div></div>'
                f'<div class="outlook-tile" style="background:#1a237e22">'
                f'<div class="o-lbl">LONG-TERM</div><div class="o-val">{_bias_badge(_lt)}</div></div>'
                f'<div class="outlook-tile" style="background:#1a237e22">'
                f'<div class="o-lbl">CRASH RISK</div><div class="o-val">{_crash_badge(_cr)}</div></div>'
                '</div>'
                + (f'<p class="interp" style="margin-top:6px">{_rat}</p>' if _rat else "")
                + '</div>'
            )

        # ── AI synopsis block ──────────────────────────────────────────────
        _synopsis = (ai or {}).get("synopsis", "")
        _synopsis_block = (
            f'<div style="margin-bottom:14px">'
            f'<h2 style="margin-bottom:6px;font-size:.75rem">AI MARKET SYNOPSIS</h2>'
            f'<div class="synopsis">{_synopsis}</div>'
            f'</div>'
        ) if _synopsis else ""

        # ── Geo/macro context block ────────────────────────────────────────
        _macro_ctx_text = (ai or {}).get("macro_context", "")
        _geo_block = (
            f'<div style="margin-bottom:14px;background:#1a1f2e;border-left:3px solid #455a64;'
            f'border-radius:4px;padding:10px 14px">'
            f'<h2 style="margin-bottom:6px;font-size:.75rem;color:#78909c">'
            f'🌍 GEOPOLITICAL &amp; MACRO CONTEXT</h2>'
            f'<p style="color:#b0bec5;font-style:italic;margin:0;font-size:.82rem;line-height:1.6">'
            f'{_macro_ctx_text}</p>'
            f'</div>'
        ) if _macro_ctx_text else ""

        return f"""
<section id="macro">
  <div class="section-title">🌍 Macro Dashboard — FRED data as of {mc.get('as_of','?')}</div>
  {_outlook_block}
  {_synopsis_block}
  {_geo_block}
  <h2 style="margin-bottom:8px;margin-top:4px;font-size:.75rem">LIVE MACRO INDICATORS
    <span style="font-size:.65rem;font-weight:400;color:#666">— click any tile to see history &amp; context</span>
  </h2>
  <div class="macro-grid">{tiles_html}</div>
  {_regime_bar}
  {detail_panels}
  <p class="interp">{interp}</p>
  {ai_rows}
</section>"""

    # ── AI PICKS SECTION ──────────────────────────────────────────────────
    def _ai_section():
        """Merged tab: AI Master Manager top picks + all 11 specialist reports."""
        _sp = (ai or {}).get("_specialist_picks", {})
        picks = (ai or {}).get("picks", [])

        # ── Build spec-picks reverse-index: ticker → [agent keys] ────────────
        _spec_map = {}
        for _ak, _av in _sp.items():
            for _p2 in _av.get("picks", []):
                _tk2 = _p2.get("ticker", "")
                if _tk2:
                    _spec_map.setdefault(_tk2, []).append(_ak)

        _ATTR_ICONS = {
            "QualityGrowth": "🌱", "SpecialSit": "⚡", "CapAppreciation": "📈",
            "EmergingGrowth": "🚀", "TenBagger": "🎯",
            "LynchBWYK": "🛒",
            "CathieWood": "🚀", "Pabrai": "🎲",
            "HowardMarks": "🔄", "Burry": "🕳️",
            "InsiderTrack": "👁️",
        }
        _ATTR_LABELS = {
            "QualityGrowth": "Quality Growth", "SpecialSit": "Special Sit",
            "CapAppreciation": "Cap Appreciation", "EmergingGrowth": "Emerging Growth",
            "TenBagger": "10-Bagger Hunter",
            "LynchBWYK": "Lynch BWYK",
            "CathieWood": "Disruptive Innov.",
            "Pabrai": "Pabrai Asym.",
            "HowardMarks": "Marks 2nd-Level",
            "Burry": "Burry Deep Val.", "InsiderTrack": "Insider Track",
        }

        # ── MASTER MANAGER PICKS ─────────────────────────────────────────────
        mm_cards = []
        for i, p in enumerate(picks):
            t = p.get("ticker", "?")
            co = p.get("company", "")
            sec = p.get("sector", "")
            strat = p.get("strategy", "")
            hl = p.get("headline", "")
            story = p.get("story", "")
            catalyst = p.get("catalyst", "")
            watch = p.get("watch", "")
            conv = p.get("conviction", "MEDIUM")
            urg = p.get("urgency", "WATCH")
            endorsed = p.get("endorsed_by", "")
            s = stocks.get(t, {})
            price_str = _money(s.get("price"))
            peg_str   = _num(s.get("peg"))
            pe_str    = _num(s.get("pe"))
            roic_str  = _pct(s.get("roic"))
            cap_badge = _cap_badge(s.get("mktCap", 0))

            # Attribution chips
            _agents_for_t = _spec_map.get(t, [])
            if _agents_for_t:
                _chips = " ".join(
                    f'<span style="display:inline-block;background:#1a237e55;border:1px solid #1a237e99;'
                    f'border-radius:3px;padding:1px 5px;font-size:.66rem;color:#9fa8da;margin:1px">'
                    f'{_ATTR_ICONS.get(a,"")} {_ATTR_LABELS.get(a,a)}</span>'
                    for a in _agents_for_t
                )
                _attr_html = (f'<div class="mm-attr"><span style="font-size:.66rem;color:#546e7a">'
                              f'Agents: </span>{_chips}</div>')
            elif endorsed:
                _attr_html = (f'<div style="font-size:.68rem;color:#546e7a;margin-top:5px;font-style:italic">'
                              f'Suggested by: {endorsed}</div>')
            else:
                _attr_html = ""

            # A9/A10: Pull lynchCategory straight off the stock dict (set in assemble_stock_data)
            _lynch_raw = s.get("lynchCategory") or p.get("lynch_category", "")
            lynch_badge = _lynch_badge(_lynch_raw)

            _biz_syn    = p.get("business_synopsis", "")
            _biz_ind    = p.get("industry", "") or s.get("industry", "")
            _biz_comp   = p.get("key_competitors", "")
            _biz_block  = ""
            if _biz_syn or _biz_ind or _biz_comp:
                _ind_line  = (f'<span style="font-size:.68rem;color:#78909c"><b>Industry:</b> {_biz_ind}</span>'
                              if _biz_ind else "")
                _comp_line = (f'<span style="font-size:.68rem;color:#78909c"><b>vs:</b> {_biz_comp}</span>'
                              if _biz_comp else "")
                _meta_bits = " &nbsp;·&nbsp; ".join(x for x in [_ind_line, _comp_line] if x)
                _syn_p     = (f'<p style="font-size:.73rem;color:#b0bec5;line-height:1.55;margin:0 0 5px">'
                              f'{_biz_syn}</p>') if _biz_syn else ""
                _meta_div  = f'<div style="margin-top:3px">{_meta_bits}</div>' if _meta_bits else ""
                _biz_block = (
                    f'<div style="background:#12122088;border-left:3px solid #42a5f555;'
                    f'border-radius:4px;padding:7px 10px;margin-bottom:8px">'
                    f'<div style="font-size:.63rem;font-weight:700;color:#42a5f5;text-transform:uppercase;'
                    f'letter-spacing:.06em;margin-bottom:4px">🏢 About the Business</div>'
                    f'{_syn_p}{_meta_div}'
                    f'</div>'
                )

            # Sparkline + trend badge for MM judge picks
            _mm_trend  = _trend_badge(t)
            _mm_spark  = _sparkline_svg(t)
            if _mm_spark:
                _mm_prices = _sparklines_data.get(t, [])
                _mm_lo = f"${min(_mm_prices):.0f}" if _mm_prices else ""
                _mm_hi = f"${max(_mm_prices):.0f}" if _mm_prices else ""
                _mm_spark_block = (
                    f'<div class="sparkline-wrap">'
                    f'<div class="sparkline-label">'
                    f'<span>5Y price history</span>'
                    f'<span>{_mm_lo} → {_mm_hi}</span>'
                    f'</div>'
                    f'<div style="width:100%">{_mm_spark}</div>'
                    f'</div>'
                )
            else:
                _mm_spark_block = ""

            mm_cards.append(f"""
<div class="mm-card xcard" onclick="toggleExpand(this)">
  <div style="display:flex;align-items:center;flex-wrap:wrap;gap:6px">
    <span class="mm-rank">#{i+1}</span>
    <span class="mm-ticker">{t}</span>
    {cap_badge}
    <span class="mm-co">{co} · {sec}</span>
    {_conv_badge(conv)} {_urgency_badge(urg)}
    {_mm_trend}
    <span class="badge badge-hold" style="font-size:.63rem">{strat}</span>
    {lynch_badge}
    {_familiar_badge(s)}
    {_undercovered_badge(s)}
    {_ceo_badge(s)}
    {_divergence_badge(s)}
    <span class="xarrow">▼</span>
  </div>
  <div class="xbody" style="display:none">
    {_mm_spark_block}
    {_biz_block}
    {_capalloc_block(s)}
    <p class="mm-hl">"{hl}"</p>
    <p class="mm-story">{story}</p>
    <div class="mm-meta">
      <span><b>Price</b> {price_str}</span>
      <span><b>PEG</b> {peg_str}</span>
      <span><b>P/E</b> {pe_str}</span>
      <span><b>ROIC</b> {roic_str}</span>
      {f'<span><b>Catalyst</b> {catalyst}</span>' if catalyst else ''}
      {f'<span><b>Watch</b> {watch}</span>' if watch else ''}
    </div>
    {_attr_html}
  </div>
</div>""")

        attn = (ai or {}).get("attention", [])
        risks_html = ""
        if attn:
            risk_items = "".join(f"<li>{r}</li>" for r in attn)
            risks_html = (f'<div style="margin:10px 0 14px"><h2 style="margin-bottom:6px;font-size:.75rem">'
                          f'⚠ Key Risks</h2>'
                          f'<ul style="padding-left:1.2em;color:#ef9a9a;font-size:.78rem;line-height:1.8">'
                          f'{risk_items}</ul></div>')

        # A9/A10: Lynch category distribution bar for today's MM picks
        _lynch_dist = {}
        for p_l in picks:
            t_l = p_l.get("ticker","")
            s_l = stocks.get(t_l, {})
            raw_l = s_l.get("lynchCategory") or p_l.get("lynch_category","")
            if raw_l:
                for cat_l in str(raw_l).split("+"):
                    cat_l = cat_l.strip()
                    if cat_l:
                        _lynch_dist[cat_l] = _lynch_dist.get(cat_l, 0) + 1

        _LYNCH_DIST_CSS = {
            "FastGrower": "#1b5e20", "Stalwart": "#0d47a1", "SlowGrower": "#37474f",
            "Cyclical": "#e65100", "Turnaround": "#6a1b9a", "AssetPlay": "#4e342e",
        }
        _LYNCH_DIST_LABEL = {
            "FastGrower": "Fast Grower", "Stalwart": "Stalwart",
            "SlowGrower": "Slow Grower", "Cyclical": "Cyclical",
            "Turnaround": "Turnaround", "AssetPlay": "Asset Play",
        }
        _dist_chips = "".join(
            f'<span style="display:inline-flex;align-items:center;gap:4px;background:{_LYNCH_DIST_CSS.get(c,"#283593")};'
            f'color:#fff;border-radius:4px;padding:2px 7px;font-size:.68rem;font-weight:600">'
            f'{_LYNCH_DIST_LABEL.get(c,c)}&nbsp;<b style="font-size:.8em;opacity:.8">{n}</b></span>'
            for c, n in sorted(_lynch_dist.items(), key=lambda x: -x[1])
        ) if _lynch_dist else '<span style="font-size:.72rem;color:#546e7a">No Lynch category data yet</span>'

        # ── Sprint 2 A3: 🎯 HIGH-CONSENSUS PICKS panel (today's specialists) ──
        # Counts how many of today's active specialists picked each ticker.
        # Max = number of active specialists. No retired-agent noise.
        consensus_html = ""
        try:
            from collections import defaultdict
            _consensus_map = defaultdict(set)   # ticker → set(agent_name) from today's run
            _ticker_meta   = {}                 # ticker → {co, syn}
            # Pull directly from today's specialist results (already in memory as _sp)
            for _ag_name, _ag_data in _sp.items():
                for _pp in (_ag_data.get("picks") or []):
                    _ptk = (_pp.get("ticker") or "").strip().upper()
                    if not _ptk:
                        continue
                    _consensus_map[_ptk].add(_ag_name)
                    if _ptk not in _ticker_meta:
                        _s_tmp = stocks.get(_ptk, {})
                        _ticker_meta[_ptk] = {
                            "co":  (_pp.get("company") or _s_tmp.get("companyName",""))[:40],
                            "syn": (_pp.get("thesis") or _pp.get("synopsis") or "")[:90],
                        }
            # Build sorted consensus list (≥3 distinct specialists = meaningful agreement)
            _min_consensus = max(3, len(_sp) // 4)  # ~25% of active agents
            _hi_consensus = sorted(
                ((tk, len(srcs)) for tk, srcs in _consensus_map.items() if len(srcs) >= _min_consensus),
                key=lambda x: -x[1]
            )[:20]   # cap at 20 most-consensus
            if _hi_consensus:
                # Build lookup: ticker → best available AI thesis text
                # Priority: judge pick (full story) > any specialist thesis
                _judge_by_ticker = {p.get("ticker","").upper(): p for p in picks}
                _spec_by_ticker  = {}   # ticker → best specialist pick dict
                for _ag, _ag_data in _sp.items():
                    for _pp in (_ag_data.get("picks") or []):
                        _ptk = (_pp.get("ticker") or "").upper()
                        if _ptk and _ptk not in _spec_by_ticker:
                            _spec_by_ticker[_ptk] = _pp

                _cons_cards = []
                for _tk2, _n2 in _hi_consensus:
                    _meta = _ticker_meta.get(_tk2, {})
                    _co2  = _meta.get("co", "")
                    # Rich text: judge story > specialist thesis > CSV synopsis
                    _judge_p = _judge_by_ticker.get(_tk2)
                    _spec_p  = _spec_by_ticker.get(_tk2)
                    if _judge_p:
                        _hl2   = _judge_p.get("headline", "")
                        _story2 = _judge_p.get("story", "")
                        _cat2  = _judge_p.get("catalyst", "")
                        _watch2 = _judge_p.get("watch", "")
                    elif _spec_p:
                        _hl2   = _spec_p.get("headline", "")
                        _story2 = _spec_p.get("thesis", "") or _spec_p.get("story", "")
                        _cat2  = _spec_p.get("catalyst", "")
                        _watch2 = _spec_p.get("watch", "")
                    else:
                        _hl2   = ""
                        _story2 = _meta.get("syn", "")
                        _cat2  = ""
                        _watch2 = ""
                    _syn2 = _meta.get("syn", "")
                    _s2   = stocks.get(_tk2, {})
                    _fam_b = _familiar_badge(_s2)
                    _unc_b = _undercovered_badge(_s2)
                    _ceo_b = _ceo_badge(_s2)
                    _div_b = _divergence_badge(_s2)
                    _price2 = _s2.get("price")
                    _price_s = f"${_price2:.2f}" if _price2 else "—"
                    _sec2   = (_s2.get("sector") or "")
                    _ind2   = (_s2.get("industry") or "")
                    _mktcap2 = _s2.get("mktCap")
                    _cap_s  = (f"${_mktcap2/1e9:.1f}B" if _mktcap2 and _mktcap2 >= 1e9
                               else f"${_mktcap2/1e6:.0f}M" if _mktcap2 else "")
                    _peg2   = _s2.get("pegRatio")
                    _pe2    = _s2.get("peRatioTTM")
                    _roic2  = _s2.get("roic")
                    _fcfm2  = _s2.get("fcfMargin")
                    _rev2   = _s2.get("revenueGrowthYoy")
                    _peg_s  = f"{_peg2:.1f}" if _peg2 and abs(_peg2) < 99 else "—"
                    _pe_s   = f"{_pe2:.1f}x" if _pe2 and abs(_pe2) < 999 else "—"
                    _roic_s = f"{_roic2*100:.1f}%" if _roic2 else "—"
                    _fcfm_s = f"{_fcfm2*100:.1f}%" if _fcfm2 else "—"
                    _rev_s  = f"{_rev2*100:.1f}%" if _rev2 else "—"
                    _spark2 = _sparkline_svg(_tk2)
                    _lynch2 = _lynch_badge(_s2.get("lynchCategory",""))
                    # Color intensity by consensus count: 5-7 = blue, 8-12 = red/amber, 13+ = green
                    _bg = ("#1b5e20" if _n2 >= 13 else
                           "#bf360d" if _n2 >= 8 else
                           "#1565c0")
                    _cons_cards.append(
                        f'<div class="xcard" onclick="toggleExpand(this)" style="background:#1a1a2e;'
                        f'border-left:4px solid {_bg};border-radius:4px;padding:8px 10px;'
                        f'min-width:180px;flex:1 1 200px;max-width:280px;cursor:pointer">'
                        f'<div style="display:flex;align-items:center;flex-wrap:wrap;gap:5px;margin-bottom:3px">'
                        f'<span style="font-weight:700;font-size:.95rem;color:#90caf9">{_tk2}</span>'
                        f'<span style="background:{_bg};color:#fff;border-radius:3px;padding:1px 6px;'
                        f'font-size:.65rem;font-weight:700">{_n2} agents</span>'
                        f'{_fam_b}{_unc_b}{_ceo_b}{_div_b}{_lynch2}'
                        f'<span class="xarrow" style="margin-left:auto;font-size:.7rem;color:#546e7a">▼</span>'
                        f'</div>'
                        f'<div style="font-size:.7rem;color:#b0bec5;margin-bottom:2px">{_co2}</div>'
                        f'<div style="font-size:.65rem;color:#78909c">'
                        f'{_sec2}{(" · " + _price_s) if _price_s else ""}'
                        f'{(" · " + _cap_s) if _cap_s else ""}</div>'
                        f'<div class="xbody" style="display:none;margin-top:8px;border-top:1px solid #263238;padding-top:8px">'
                        + (f'<div style="margin-bottom:6px">{_spark2}</div>' if _spark2 else "")
                        + (f'<p class="mm-hl" style="font-size:.75rem;font-style:italic;color:#b0bec5;margin:0 0 6px">&ldquo;{_hl2}&rdquo;</p>' if _hl2 else "")
                        + (f'<p class="mm-story" style="font-size:.72rem;color:#cfd8dc;line-height:1.6;margin:0 0 8px">{_story2}</p>' if _story2 else (f'<p style="font-size:.68rem;color:#9e9e9e;font-style:italic;margin:0 0 8px">{_syn2}</p>' if _syn2 else ""))
                        + f'<div style="display:flex;flex-wrap:wrap;gap:8px;font-size:.7rem;color:#b0bec5;margin-bottom:6px">'
                        f'<span><b>Price</b> {_price_s}</span>'
                        f'<span><b>P/E</b> {_pe_s}</span>'
                        f'<span><b>PEG</b> {_peg_s}</span>'
                        f'<span><b>ROIC</b> {_roic_s}</span>'
                        f'<span><b>FCF Margin</b> {_fcfm_s}</span>'
                        f'<span><b>Rev Growth</b> {_rev_s}</span>'
                        f'{f"<span><b>Mkt Cap</b> {_cap_s}</span>" if _cap_s else ""}'
                        f'</div>'
                        + (f'<p style="font-size:.68rem;color:#80cbc4;margin:0 0 3px"><b>Catalyst:</b> {_cat2}</p>' if _cat2 else "")
                        + (f'<p style="font-size:.68rem;color:#ef9a9a;margin:0 0 3px"><b>Watch:</b> {_watch2}</p>' if _watch2 else "")
                        + f'<div style="font-size:.63rem;color:#546e7a;margin-top:4px">{_ind2}</div>'
                        f'</div>'
                        f'</div>'
                    )
                consensus_html = (
                    '<div style="background:#0d1117;border:1px solid #42a5f533;border-radius:8px;'
                    'padding:12px 14px;margin-bottom:14px">'
                    '<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px">'
                    '<span style="font-size:1rem;font-weight:700;color:#42a5f5">🎯 HIGH-CONSENSUS PICKS</span>'
                    f'<span style="font-size:.72rem;color:#78909c">— {len(_hi_consensus)} '
                    f'tickers nominated by {_min_consensus}+ of today\'s {len(_sp)} specialists</span>'
                    '</div>'
                    '<p style="font-size:.7rem;color:#546e7a;margin-bottom:10px;line-height:1.5">'
                    'When multiple independent lenses converge on the same name today, that is the strongest '
                    'signal in the system. Cross-check against your personal knowledge before sizing.'
                    '</p>'
                    f'<div style="display:flex;flex-wrap:wrap;gap:8px">{"".join(_cons_cards)}</div>'
                    '</div>'
                )
        except Exception:
            consensus_html = ""

        # ── 🛍️ MALL MANAGER SECTION (Lynch consumer-observable lens) ──────
        # Picks where there's a real-world consumer/observable edge data-driven
        # Wall Street is missing. Distinct mint/teal accent vs gold MM section.
        mall_section_html = ""
        try:
            _mall_picks = (mall or {}).get("picks", []) if mall else []
            if _mall_picks:
                _mall_synopsis = (mall or {}).get("synopsis", "")
                _mall_rejected = (mall or {}).get("rejected_examples", "")
                _mall_cards = []
                for _i, _mp in enumerate(_mall_picks):
                    _mt   = (_mp.get("ticker") or "?").upper()
                    _mco  = (_mp.get("company") or "")[:40]
                    _msec = _mp.get("sector", "")
                    _mhl  = _mp.get("headline", "")
                    _mstory = _mp.get("story", "")
                    _mthesis = _mp.get("consumer_thesis", "")
                    _mblind = _mp.get("wall_street_blindspot", "")
                    _mcat = _mp.get("catalyst", "")
                    _mwatch = _mp.get("watch", "")
                    _mconv = _mp.get("conviction", "MEDIUM")
                    _ms   = stocks.get(_mt, {})
                    _mprice = _ms.get("price")
                    _mpx_s = f"${_mprice:.2f}" if _mprice else "—"
                    _mmkt = _ms.get("mktCap")
                    _mmkt_s = (f"${_mmkt/1e9:.1f}B" if _mmkt and _mmkt >= 1e9
                               else f"${_mmkt/1e6:.0f}M" if _mmkt else "")
                    _mind = (_ms.get("industry") or "")
                    _mfam = _familiar_badge(_ms)
                    _munc = _undercovered_badge(_ms)
                    _mceo = _ceo_badge(_ms)
                    _mdiv = _divergence_badge(_ms)
                    _mlynch = _lynch_badge(_ms.get("lynchCategory",""))
                    _mspark = _sparkline_svg(_mt)
                    _mconv_color = "#26a69a" if _mconv == "HIGH" else "#80cbc4"

                    _mall_cards.append(
                        f'<div class="mm-card xcard" onclick="toggleExpand(this)" '
                        f'style="border-left:4px solid #26a69a">'
                        f'<div style="display:flex;align-items:center;flex-wrap:wrap;gap:6px">'
                        f'<span class="mm-rank" style="background:#004d40;color:#80cbc4">#{_i+1}</span>'
                        f'<span class="mm-ticker">{_mt}</span>'
                        f'<span class="mm-co">{_mco} · {_msec}</span>'
                        f'<span class="badge" style="background:{_mconv_color};color:#000;font-size:.62rem;'
                        f'padding:1px 6px;border-radius:3px;font-weight:700">{_mconv}</span>'
                        f'{_mlynch}{_mfam}{_munc}{_mceo}{_mdiv}'
                        f'<span class="xarrow">▼</span>'
                        f'</div>'
                        # Always-visible thesis line — the unique value of the Mall Manager
                        f'<p style="margin:6px 0 0;font-size:.78rem;color:#80cbc4;'
                        f'font-style:italic;line-height:1.5">'
                        f'🛍️ &ldquo;{_mthesis}&rdquo;</p>'
                        f'<div class="xbody" style="display:none">'
                        + (f'<div class="sparkline-wrap" style="margin-top:8px"><div style="width:100%">{_mspark}</div></div>' if _mspark else "")
                        + (f'<p class="mm-hl" style="margin-top:8px">&ldquo;{_mhl}&rdquo;</p>' if _mhl else "")
                        + (f'<p class="mm-story">{_mstory}</p>' if _mstory else "")
                        + (f'<p style="font-size:.72rem;color:#ffab91;margin:6px 0 0;'
                           f'background:#1a1a2e;padding:6px 8px;border-left:3px solid #ff7043;border-radius:3px">'
                           f'<b>🔭 Wall Street Blindspot:</b> {_mblind}</p>' if _mblind else "")
                        + f'<div class="mm-meta" style="margin-top:8px">'
                        f'<span><b>Price</b> {_mpx_s}</span>'
                        f'{f"<span><b>Mkt Cap</b> {_mmkt_s}</span>" if _mmkt_s else ""}'
                        + (f'<span><b>Catalyst</b> {_mcat}</span>' if _mcat else "")
                        + (f'<span><b>Watch</b> {_mwatch}</span>' if _mwatch else "")
                        + f'</div>'
                        + (f'<div style="font-size:.63rem;color:#546e7a;margin-top:5px">{_mind}</div>' if _mind else "")
                        + f'</div>'
                        f'</div>'
                    )

                _rejected_block = (
                    f'<p style="font-size:.7rem;color:#9e9e9e;margin:10px 0 4px;font-style:italic">'
                    f'<b>Rejected from Judge\'s list:</b> {_mall_rejected}</p>'
                ) if _mall_rejected else ""
                _synopsis_block = (
                    f'<p style="font-size:.74rem;color:#b0bec5;margin-bottom:8px;line-height:1.5">'
                    f'{_mall_synopsis}</p>'
                ) if _mall_synopsis else ""

                mall_section_html = (
                    '<div style="background:#0d1117;border:1px solid #26a69a44;border-radius:8px;'
                    'padding:12px 14px;margin-bottom:18px">'
                    '<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px">'
                    '<span style="font-size:1rem;font-weight:700;color:#26a69a">🛍️ AI Mall Manager</span>'
                    f'<span style="font-size:.72rem;color:#78909c">— {len(_mall_picks)} picks where '
                    'a real-world consumer-observable edge contradicts data-driven Wall Street consensus</span>'
                    '</div>'
                    '<p style="font-size:.72rem;color:#546e7a;margin-bottom:8px;line-height:1.5">'
                    'Peter Lynch \"walking through a shopping mall\" lens: each pick has a thesis you could '
                    'observe in daily life — products people are using more than the data shows yet. '
                    '<b style="color:#78909c">Click any card to expand the full thesis.</b>'
                    '</p>'
                    + _synopsis_block
                    + f'<div class="mm-grid">{"".join(_mall_cards)}</div>'
                    + _rejected_block
                    + '</div>'
                )
        except Exception as _emall:
            mall_section_html = f"<!-- mall_section error: {str(_emall)[:80]} -->"

        # ── SPECIALIST SECTIONS ───────────────────────────────────────────
        _ADESC = {
            "QualityGrowth":  ("🌱 Quality Growth",  "2E7D32",
                "Finds durable compounders with ROIC >15%, consistent multi-year revenue growth, "
                "and structural competitive moats. Prioritises FCF conversion and PEG <1.5."),
            "SpecialSit":     ("⚡ Special Situation", "6A1B9A",
                "Identifies event-driven, misunderstood, and inflection-point opportunities. Focuses on "
                "business model misclassification, regulatory catalysts, and hidden assets."),
            "CapAppreciation":("📈 Capital Appreciation", "1565C0",
                "Finds near-term re-rating candidates with specific catalysts in 1–6 months. Targets "
                "beaten-down quality (52wPos <65%), revenue re-acceleration, and cycle trough entries."),
            "EmergingGrowth": ("🚀 Emerging Growth",  "E65100",
                "Identifies smaller fast-growing companies ($100M–$15B) early in becoming compounders. "
                "Targets 20%+ revenue growth, rising ROIC, and large underserved TAMs."),
            "TenBagger":      ("🎯 10-Bagger Hunter", "BF360D",
                "Peter Lynch-style small-cap hunter. Looks for underfollowed companies ($50M–$2B) with "
                "15–40% EPS+revenue growth, at least 2 expansion levers not yet exhausted, and structural "
                "Wall Street under-coverage (orphans, post-restructuring, complex/sin-sector names)."),
            "LynchBWYK":      ("🛒 Lynch Buy What You Know","880E4F",
                "Lynch's 'buy what you know' — simple understandable businesses serving everyday needs, "
                "growing 15–25%/yr with real earnings at PEG <1.0."),
            "CathieWood":     ("🚀 Disruptive Innovation","0D47A1",
                "ARK Invest-style screen. Identifies pure-play companies in AI, robotics, genomics, "
                "or energy storage riding Wright's Law cost curves with network effects."),
            "Pabrai":         ("🎲 Pabrai Asymmetric Bet","33691E",
                "Mohnish Pabrai's framework: 3:1+ upside/downside asymmetry where downside is protected "
                "by real assets or essential business value — 'heads I win, tails I don't lose much'."),
            "HowardMarks":    ("🔄 Marks Second-Level","827717",
                "Howard Marks' contrarian analysis where market consensus is factually wrong. Targets "
                "oversold stocks where the negative narrative exceeds actual deterioration."),
            "Burry":          ("🕳️ Burry Deep Value",  "3E2723",
                "Michael Burry-style catalyst-driven deep value — hidden assets, temporary earnings "
                "distortions, and specific upcoming events that will force the market to reprice."),
            "InsiderTrack":   ("👁️ Insider & Smart Money","263238",
                "Tracks cluster insider buying (multiple executives buying simultaneously) and significant "
                "open-market purchases with personal capital — the highest-signal conviction indicator."),
        }
        _AGENT_ORDER = ["QualityGrowth", "SpecialSit", "CapAppreciation", "EmergingGrowth", "TenBagger",
                        "LynchBWYK", "CathieWood",
                        "Pabrai", "HowardMarks", "Burry", "InsiderTrack"]

        # ── Strategy Picks Summary table ──────────────────────────────────
        _strat_defs = [
            ("💎 IV Discount",        iv_rows,             "1A237E"),
            ("💎 Quality Comp.",       quality_compounders, "1B5E20"),
            ("📊 Stalwarts",           stalwarts,           "1565C0"),
            ("🚀 Fast Growers",        fast_growers,        "E65100"),
            ("🔄 Turnarounds",         turnarounds,         "6A1B9A"),
            ("📉 Slow Growers",        slow_growers,        "37474F"),
            ("🔁 Cyclicals",           cyclicals,           "827717"),
            ("🏗️ Asset Plays",         asset_plays,         "4E342E"),
            ("🎯 10-Baggers",          ten_baggers,         "BF360D"),
        ]
        _strat_rows_html = []
        for _sl, _sd, _sc in _strat_defs:
            if not _sd:
                continue
            _top3 = [r.get("Ticker","") for r in _sd[:3]]
            _top1  = _sd[0]
            _peg   = _top1.get("PEG")
            _mos   = _top1.get("MoS")
            _pio   = _top1.get("Piotroski")
            _co    = (_top1.get("Company") or "")[:28]
            _n     = len(_sd)
            _peg_s = f'{_peg:.2f}' if _peg else '—'
            _mos_s = f'{_mos*100:.0f}%' if _mos else '—'
            _pio_s = str(int(_pio)) if _pio else '—'
            _pio_color = ("#1b5e20" if _pio and _pio >= 7 else
                          "#e65100" if _pio and _pio <= 4 else "#263238")
            _top3_html = " ".join(
                f'<span style="background:#ffffff11;border-radius:3px;padding:1px 5px;'
                f'font-size:.68rem;font-family:monospace">{t}</span>' for t in _top3
            )
            _strat_rows_html.append(f"""<tr>
  <td style="padding:5px 8px;border-bottom:1px solid #ffffff0f">
    <span style="display:inline-block;width:8px;height:8px;border-radius:50%;
      background:#{_sc};margin-right:5px"></span>{_sl}
  </td>
  <td style="padding:5px 8px;border-bottom:1px solid #ffffff0f;color:#78909c;text-align:right">{_n}</td>
  <td style="padding:5px 8px;border-bottom:1px solid #ffffff0f">{_top3_html}</td>
  <td style="padding:5px 8px;border-bottom:1px solid #ffffff0f;color:#b0bec5;font-size:.72rem">{_co}</td>
  <td style="padding:5px 8px;border-bottom:1px solid #ffffff0f;color:#90caf9;text-align:right">{_peg_s}</td>
  <td style="padding:5px 8px;border-bottom:1px solid #ffffff0f;color:#a5d6a7;text-align:right">{_mos_s}</td>
  <td style="padding:5px 8px;border-bottom:1px solid #ffffff0f;background:{_pio_color};
    border-radius:3px;text-align:center;font-weight:700">{_pio_s}</td>
</tr>""")
        strat_summary_html = f"""
<details open style="margin-bottom:14px">
  <summary style="cursor:pointer;list-style:none;display:flex;align-items:center;gap:8px;
    padding:8px 12px;background:#161b22;border-radius:6px;border:1px solid #ffffff18;
    font-size:.75rem;font-weight:700;color:#90caf9;text-transform:uppercase;letter-spacing:.05em;
    user-select:none">
    <span style="font-size:.9rem">📊</span> Strategy Picks Summary
    <span style="font-size:.65rem;color:#546e7a;font-weight:400;text-transform:none;margin-left:4px">
      — top quantitative picks per strategy tab (rule-based, not AI)
    </span>
    <span style="margin-left:auto;font-size:.7rem;color:#546e7a">▼ click to collapse</span>
  </summary>
  <div style="margin-top:8px;background:#0d1117;border:1px solid #ffffff12;border-radius:6px;overflow:hidden">
    <table style="width:100%;border-collapse:collapse;font-size:.73rem">
      <thead>
        <tr style="background:#161b22">
          <th style="padding:6px 8px;text-align:left;color:#546e7a;font-weight:600">Strategy</th>
          <th style="padding:6px 8px;text-align:right;color:#546e7a;font-weight:600"># Qualifying</th>
          <th style="padding:6px 8px;text-align:left;color:#546e7a;font-weight:600">Top 3 Tickers</th>
          <th style="padding:6px 8px;text-align:left;color:#546e7a;font-weight:600">Top Pick</th>
          <th style="padding:6px 8px;text-align:right;color:#546e7a;font-weight:600">PEG</th>
          <th style="padding:6px 8px;text-align:right;color:#546e7a;font-weight:600">MoS</th>
          <th style="padding:6px 8px;text-align:center;color:#546e7a;font-weight:600">Piotroski</th>
        </tr>
      </thead>
      <tbody>{"".join(_strat_rows_html)}</tbody>
    </table>
  </div>
</details>""" if _strat_rows_html else ""

        # ── Wrap major AI blocks in collapsible <details> ─────────────────
        def _collapsible(title, subtitle, content, open_by_default=True, accent="#42a5f5"):
            _open = "open" if open_by_default else ""
            return f"""
<details {_open} style="margin-bottom:14px">
  <summary style="cursor:pointer;list-style:none;display:flex;align-items:center;gap:8px;
    padding:9px 14px;background:#161b22;border-radius:6px;border:1px solid {accent}33;
    font-size:.8rem;font-weight:700;color:{accent};user-select:none">
    {title}
    <span style="font-size:.65rem;color:#546e7a;font-weight:400;margin-left:4px">{subtitle}</span>
    <span style="margin-left:auto;font-size:.7rem;color:#546e7a">▼</span>
  </summary>
  <div style="margin-top:6px">{content}</div>
</details>"""

        # Wrap consensus + MM + Mall into collapsibles
        n_agents = sum(1 for ak in _AGENT_ORDER if ak in _sp)
        wrapped_mm = _collapsible("🧠 AI Master Manager",
                                  f"— synthesises {len(picks)} top picks from {n_agents} specialists",
                                  f"""<div style="background:#0d1117;border:1px solid #ffd54f33;border-radius:8px;padding:12px 14px">
  <p style="font-size:.73rem;color:#546e7a;margin-bottom:8px;line-height:1.5">
    The Master Manager conducts no independent research. It selects and ranks only from agent-nominated
    stocks, prioritising cross-specialist consensus, quality filters, and catalyst timing.
    <b style="color:#78909c">Click any card to expand.</b>
  </p>
  <div style="display:flex;flex-wrap:wrap;gap:5px;margin-bottom:10px;align-items:center">
    <span style="font-size:.66rem;color:#78909c;text-transform:uppercase;letter-spacing:.04em">Lynch Mix:</span>
    {_dist_chips}
  </div>
  <div class="mm-grid">{"".join(mm_cards)}</div>
  {risks_html}
</div>""",
                                  open_by_default=True, accent="#ffd54f")

        # Mall Manager collapsible
        mall_inner = mall_section_html.strip()
        wrapped_mall = (_collapsible("🛍️ AI Mall Manager",
                                     "— Peter Lynch consumer-observable lens on specialist picks",
                                     mall_inner,
                                     open_by_default=True, accent="#26a69a")
                        if mall_inner and not mall_inner.startswith("<!--") else "")

        consensus_wrapped = (_collapsible("🎯 High-Consensus Picks",
                                          f"— tickers nominated by 3+ of today's {n_agents} specialists",
                                          consensus_html.strip(),
                                          open_by_default=True, accent="#42a5f5")
                             if consensus_html.strip() else "")

        # Build specialist collapsibles directly from _sp iteration order
        spec_collapsibles = []
        for ak in _AGENT_ORDER:
            if ak not in _sp:
                continue
            sr_data   = _sp[ak]
            ap        = sr_data.get("picks", [])
            lbl, chx, dsc = _ADESC.get(ak, (ak, "37474F", ""))
            # Reuse already-built pick_cards from spec_sections loop
            # Re-build pick cards for this agent inline
            _pkc = []
            for pp in ap:
                tk = pp.get("ticker", "")
                s2 = stocks.get(tk, {})
                mc_b = s2.get("mktCapB")
                prc  = s2.get("price")
                conv2 = (pp.get("conviction") or "MEDIUM").upper()
                brief = pp.get("brief_case", "")
                key_m = pp.get("key_metric", "")
                price_disp  = f"${prc:.0f}" if prc else ""
                mktcap_disp = f"${mc_b:.1f}B" if mc_b else ""
                rationale3 = pp.get("rationale", "") or brief
                _lb3 = _lynch_badge(s2.get("lynchCategory"))
                _sp_syn   = pp.get("business_synopsis", "")
                _sp_ind   = pp.get("industry", "") or s2.get("industry", "")
                _sp_comp  = pp.get("key_competitors", "")
                _sp_biz   = ""
                if _sp_syn or _sp_ind or _sp_comp:
                    _sp_meta = "  ·  ".join(x for x in [
                        (f"<b>Industry:</b> {_sp_ind}" if _sp_ind else ""),
                        (f"<b>vs:</b> {_sp_comp}"      if _sp_comp else ""),
                    ] if x)
                    _sp_biz = (
                        f'<div style="background:#12122088;border-left:2px solid #42a5f555;'
                        f'border-radius:3px;padding:5px 8px;margin-top:6px">'
                        f'<div style="font-size:.60rem;font-weight:700;color:#42a5f5;text-transform:uppercase;'
                        f'letter-spacing:.05em;margin-bottom:3px">🏢 About</div>'
                        + (f'<div style="font-size:.68rem;color:#b0bec5;line-height:1.5">{_sp_syn}</div>' if _sp_syn else "")
                        + (f'<div style="font-size:.65rem;color:#78909c;margin-top:3px">{_sp_meta}</div>' if _sp_meta else "")
                        + f'</div>'
                    )
                _sp_co     = (pp.get("company") or s2.get("name") or tk)[:38]
                _sp_sector = s2.get("sector", "")
                _sp_cap    = _cap_badge(s2.get("mktCap", 0))
                _sp_trend  = _trend_badge(tk)
                _sp_spark  = _sparkline_svg(tk)
                _sp_price  = f"${s2['price']:.0f}" if s2.get("price") else ""
                _sp_pe     = _num(s2.get("pe"))   if s2.get("pe")   else ""
                _sp_peg    = _num(s2.get("peg"))  if s2.get("peg")  else ""
                _sp_roic   = _pct(s2.get("roic")) if s2.get("roic") else ""
                _sp_mc     = f"${mc_b:.1f}B"      if mc_b           else ""
                _sp_meta_items = " ".join(
                    f'<span><b>{k}</b> {v}</span>'
                    for k, v in [("Price", _sp_price), ("P/E", _sp_pe),
                                  ("PEG", _sp_peg), ("ROIC", _sp_roic), ("MktCap", _sp_mc)]
                    if v
                )
                if _sp_spark:
                    _sp_prices = _sparklines_data.get(tk, [])
                    _sp_lo = f"${min(_sp_prices):.0f}" if _sp_prices else ""
                    _sp_hi = f"${max(_sp_prices):.0f}" if _sp_prices else ""
                    _sp_spark_block = (
                        f'<div class="sparkline-wrap">'
                        f'<div class="sparkline-label"><span>5Y price</span>'
                        f'<span>{_sp_lo} → {_sp_hi}</span></div>'
                        f'<div style="width:100%">{_sp_spark}</div>'
                        f'</div>'
                    )
                else:
                    _sp_spark_block = ""
                _pkc.append(f"""
<div class="mm-card xcard" style="border-left-color:#{chx}" onclick="toggleExpand(this)">
  <div style="display:flex;align-items:center;flex-wrap:wrap;gap:5px">
    <span class="mm-ticker">{tk}</span>
    {_sp_cap}
    <span class="mm-co">{_sp_co}{(' · ' + _sp_sector) if _sp_sector else ''}</span>
    {_conv_badge(conv2)}
    {_sp_trend}
    {_lb3}
    {_familiar_badge(s2)}
    {_undercovered_badge(s2)}
    {_ceo_badge(s2)}
    {_divergence_badge(s2)}
    <span class="xarrow">▼</span>
  </div>
  {f'<div style="font-size:.72rem;color:#78909c;margin-top:4px;font-style:italic">{key_m[:90]}</div>' if key_m else ''}
  <div class="xbody" style="display:none">
    {_sp_spark_block}
    {_sp_biz}
    {_capalloc_block(s2)}
    <p class="mm-hl">"{rationale3}"</p>
    {f'<p class="mm-story">{brief}</p>' if brief else ''}
    <div class="mm-meta">{_sp_meta_items}</div>
  </div>
</div>""")

            inner_html = (
                f'<div class="agent-section" style="margin-bottom:0">'
                f'<div class="agent-hdr" style="background:#{chx}">'
                f'<div class="ag-title">{lbl}</div>'
                f'<div class="ag-desc">{dsc}</div>'
                f'</div>'
                f'<div class="agent-picks-grid">{"".join(_pkc)}</div>'
                f'</div>'
            )
            spec_collapsibles.append(
                _collapsible(lbl, f"— {len(ap)} picks", inner_html,
                             open_by_default=False, accent=f"#{chx}")
            )

        return f"""
<section id="ai" class="active">
  <div class="section-title">🤖 AI Analysis — Master Manager + {n_agents} Specialists</div>
  {strat_summary_html}
  {consensus_wrapped}
  {wrapped_mm}
  {wrapped_mall}
  <details style="margin-bottom:6px">
    <summary style="cursor:pointer;list-style:none;display:flex;align-items:center;gap:8px;
      padding:8px 12px;background:#161b22;border-radius:6px;border:1px solid #ffffff18;
      font-size:.75rem;font-weight:700;color:#78909c;text-transform:uppercase;letter-spacing:.05em;
      user-select:none">
      <span>📋 {n_agents} Specialist Reports</span>
      <span style="font-size:.65rem;font-weight:400;text-transform:none;color:#546e7a">
        — click any specialist to expand
      </span>
      <span style="margin-left:auto;font-size:.7rem;color:#546e7a">▼</span>
    </summary>
    <div style="margin-top:8px">{"".join(spec_collapsibles)}</div>
  </details>
</section>"""

    # ── MERGED SECTOR SECTION ─────────────────────────────────────────────
    def _sector_section():
        """Unified Sectors tab: ETF timing signals (top) + sector fundamentals (bottom)."""

        # ── ETF summary cards ──────────────────────────────────────────────
        def _etf_sig_badge(sig):
            s = (sig or "").upper()
            if "ROTATE IN" in s: return f'<span class="badge badge-buy">{sig}</span>'
            if "AVOID"     in s: return f'<span class="badge badge-avoid">{sig}</span>'
            if "TAKE PROF" in s: return f'<span class="badge badge-caut">{sig}</span>'
            return f'<span class="badge badge-hold">{sig}</span>'

        def _vs52h_bar(v):
            """Visual bar showing where price sits in its 52w range."""
            if v is None: return "—"
            pct = v * 100
            bar_w = max(4, min(100, int(pct)))
            color = ("#ef5350" if pct > 92 else "#ff9800" if pct > 80
                     else "#66bb6a" if pct < 70 else "#ffd54f")
            return (f'<div style="display:flex;align-items:center;gap:5px;min-width:90px">'
                    f'<div style="flex:1;background:#2a2a3e;border-radius:2px;height:6px">'
                    f'<div style="width:{bar_w}%;background:{color};height:6px;border-radius:2px"></div>'
                    f'</div><span style="font-size:.72rem;color:{color};width:34px;text-align:right">'
                    f'{pct:.0f}%</span></div>')

        # ── ETF ROTATION TABLE ─────────────────────────────────────────────
        etf_section_html = ""
        if etf_rows:
            # Sort: best score first
            sorted_etf = sorted(etf_rows, key=lambda r: -(r.get("score") or 0))
            etf_cards = []
            for r in sorted_etf:
                sc   = r.get("score", 0)
                sig  = r.get("signal", "HOLD")
                vs52 = r.get("vs52H")
                rsi  = r.get("rsi14")
                r1m  = r.get("1M")
                r3m  = r.get("3M")
                r1y  = r.get("1Y")
                a1m  = r.get("1M_alpha")
                a3m  = r.get("3M_alpha")
                ma50 = r.get("vs_ma50")
                ma200= r.get("vs_ma200")
                trend= r.get("trend","")
                fund = r.get("fund_sc", 0)
                price= r.get("price")
                etf  = r.get("etf","")
                sec  = r.get("sector","")
                cycle= r.get("cycle","—")

                # Row background tint based on signal
                bg = ("rgba(27,94,32,.18)" if "ROTATE IN" in sig.upper()
                      else "rgba(183,28,28,.18)" if "AVOID" in sig.upper()
                      else "rgba(230,81,0,.12)" if "TAKE PROF" in sig.upper()
                      else "transparent")

                def _ret(v):
                    if v is None: return "<td>—</td>"
                    c = "g" if v > 0 else ("r" if v < 0 else "")
                    return f'<td class="{c}">{v*100:+.1f}%</td>' if c else f"<td>{v*100:+.1f}%</td>"

                def _alpha(v):
                    if v is None: return "<td>—</td>"
                    c = "g" if v > 0.01 else ("r" if v < -0.01 else "")
                    return f'<td class="{c}">{v*100:+.1f}%</td>' if c else f"<td>{v*100:+.1f}%</td>"

                ma50_disp  = (f'<span style="color:{"#66bb6a" if ma50 and ma50>=1 else "#ef5350"}">'
                              f'{"▲" if ma50 and ma50>=1 else "▼"} MA50</span>') if ma50 else "—"
                ma200_disp = (f'<span style="color:{"#66bb6a" if ma200 and ma200>=1 else "#ef5350"}">'
                              f'{"▲" if ma200 and ma200>=1 else "▼"} MA200</span>') if ma200 else "—"
                rsi_cls = ("g" if rsi and rsi <= 35 else "r" if rsi and rsi >= 65 else "")
                sc_cls  = ("g" if sc >= 65 else "a" if sc >= 48 else "r")

                etf_cards.append(f"""
<tr style="background:{bg}">
  <td><b style="font-size:.82rem">{sec}</b><br><span style="color:#78909c;font-size:.7rem">{cycle}</span></td>
  <td><b style="color:#90caf9">{etf}</b></td>
  <td>{_etf_sig_badge(sig)}</td>
  <td class="{sc_cls}" style="text-align:center;font-weight:700">{sc}</td>
  <td>{_vs52h_bar(vs52)}</td>
  <td class="{rsi_cls}" style="text-align:center">{_num(rsi)}</td>
  <td style="text-align:center">{trend or "—"}</td>
  {_ret(r1m)}{_ret(r3m)}{_ret(r1y)}
  {_alpha(a1m)}{_alpha(a3m)}
  <td style="text-align:center">{ma50_disp}</td>
  <td style="text-align:center">{ma200_disp}</td>
  <td style="text-align:center;color:#78909c;font-size:.72rem">{fund}</td>
  <td>{_money(price)}</td>
</tr>""")

            etf_section_html = f"""
<h2 style="font-size:.78rem;margin:0 0 6px;color:#90caf9;text-transform:uppercase;letter-spacing:.05em">
  🔄 ETF Rotation — Buy when cheap, sell when extended
</h2>
<p style="background:#1a1a2e;padding:7px 12px;border-radius:4px;font-size:.73rem;
   color:#9e9e9e;margin-bottom:8px;line-height:1.6">
  Live daily signals. <b style="color:#e0e0e0">Score</b> = 52w positioning + RSI + MA50/200 + momentum + fundamentals.
  <b style="color:#e0e0e0">52w bar</b> shows where the ETF trades within its 1-year range — green &lt;70% = oversold opportunity.
  α = excess return vs SPY. Sort any column.
</p>
<div class="tbl-wrap">
<table>
<thead><tr>
  <th>Sector / Cycle</th><th>ETF</th><th>Signal</th><th>Score</th>
  <th>52w Range</th><th>RSI</th><th>Trend</th>
  <th>1M</th><th>3M</th><th>1Y</th><th>1M α</th><th>3M α</th>
  <th>vs MA50</th><th>vs MA200</th><th>Fund.Sc</th><th>Price</th>
</tr></thead>
<tbody>{"".join(etf_cards)}</tbody>
</table></div>"""

        # ── SECTOR FUNDAMENTALS TABLE ──────────────────────────────────────
        fund_section_html = ""
        if sector_rows:
            fund_rows_html = []
            for i, r in enumerate(sector_rows):
                alt = ' class="alt"' if i % 2 == 0 else ''
                sig = r.get("Signal", "HOLD")
                sig_cls = {"BUY": "g", "AVOID": "r"}.get(sig, "")
                peg  = r.get("Med PEG")
                pe   = r.get("Med P/E")
                fcf  = r.get("Avg FCF Yield")
                roe  = r.get("Avg ROE")
                rg   = r.get("Avg Rev Grw")
                sc   = r.get("Rot. Score", 0)
                sc_cls = "g" if sc >= 60 else ("a" if sc >= 45 else "r")
                fund_rows_html.append(f"""
<tr{alt}>
  <td><b>{r.get("Sector","")}</b></td>
  <td>{_signal_badge(sig)}</td>
  <td class="{sc_cls}" style="text-align:center">{sc}</td>
  <td style="text-align:center">{r.get("# Stocks","")}</td>
  <td class="{_peg_cls(peg) if peg else ''}">{_num(peg)}</td>
  <td style="text-align:center">{_num(pe)}</td>
  <td class="{"g" if fcf and fcf>0.05 else "a" if fcf and fcf>0.02 else ""}">{_pct(fcf)}</td>
  <td class="{"g" if roe and roe>0.15 else "a" if roe and roe>0.08 else ""}">{_pct(roe)}</td>
  <td class="{"g" if rg and rg>0.10 else "a" if rg and rg>0.05 else "r" if rg and rg<=0 else ""}">{_pct(rg)}</td>
  <td style="text-align:center;color:#78909c;font-size:.72rem">{r.get("ETF","—")}</td>
</tr>""")

            fund_section_html = f"""
<h2 style="font-size:.78rem;margin:18px 0 6px;color:#90caf9;text-transform:uppercase;letter-spacing:.05em">
  📊 Sector Fundamentals — Where to hunt for stocks
</h2>
<p style="background:#1a1a2e;padding:7px 12px;border-radius:4px;font-size:.73rem;
   color:#9e9e9e;margin-bottom:8px;line-height:1.6">
  Aggregated from {len(stocks):,} common stocks. <b style="color:#e0e0e0">Score</b> = PEG + FCF yield + ROE + revenue growth breadth.
  Fundamentals update with each earnings cycle (quarterly) — the ranking is intentionally stable; sector valuation
  regimes don't reverse daily. For timing within a sector use the ETF table above.
</p>
<div class="tbl-wrap">
<table>
<thead><tr>
  <th>Sector</th><th>Signal</th><th>Fund.Score</th><th># Stocks</th>
  <th>Med PEG</th><th>Med P/E</th><th>Avg FCF Yield</th><th>Avg ROE</th><th>Avg Rev Growth</th><th>ETF</th>
</tr></thead>
<tbody>{"".join(fund_rows_html)}</tbody>
</table></div>"""

        if not etf_section_html and not fund_section_html:
            return '<section id="sector"><p style="color:#666">No sector data available</p></section>'

        return f"""
<section id="sector">
  <div class="section-title">🗺 Sectors — ETF Rotation + Stock Hunting</div>
  {etf_section_html}
  {fund_section_html}
</section>"""

    # ── PERFORMANCE SECTION ───────────────────────────────────────────────
    def _perf_section():
        """Performance tab: agent scorecard + AI picks P&L + portfolio holdings."""
        import csv as _csv

        _AGENT_ICONS = {
            "AI-Judge":           "⚖️ Master Manager",
            "AI-MallManager":     "🛍️ Mall Manager",
            "AI-QualityGrowth":   "🌱 Qual.Growth",
            "AI-EmergingGrowth":  "🚀 Emerg.Growth",
            "AI-CapAppreciation": "📈 Cap.Apprecn",
            "AI-SpecialSit":      "⚡ Special Sit",
            "AI-TenBagger":       "🎯 10-Bagger",
            "AI-LynchBWYK":       "🛒 Lynch BWYK",
            "AI-CathieWood":      "🚀 Disruptive",
            "AI-Pabrai":          "🎲 Pabrai",
            "AI-HowardMarks":     "🔄 Contrarian",
            "AI-Burry":           "🕳️ Deep Value",
            "AI-InsiderTrack":    "👁️ Insider",
            # Retired agents — labels kept so historical CSV entries display correctly
            "AI-GoldmanSC":       "🏦 Goldman SC (retired)",
            "AI-SocialArb":       "📱 Social Arb (retired)",
            "AI-Mayer100x":       "💯 100-Bagger (retired)",
            "AI-MagicFormula":    "🔢 Magic Formula (retired)",
            "AI-NickSleep":       "🌀 Scale Econ (retired)",
            "AI-WallStBlind":     "🔍 WallStBlind (retired)",
        }
        _AGENT_ORDER = list(_AGENT_ICONS.keys())

        # ── Read AI picks log ─────────────────────────────────────────────
        ai_perf_rows = []
        if os.path.exists(AI_PICKS_LOG):
            try:
                with open(AI_PICKS_LOG, "r", encoding="utf-8") as _f:
                    for row in _csv.DictReader(_f):
                        t = row.get("ticker", "")
                        try: entry = float(row.get("entry_price") or 0)
                        except (ValueError, TypeError): entry = 0
                        s = stocks.get(t, {})
                        curr = s.get("price") or 0
                        ret  = ((curr - entry) / entry) if entry > 0 and curr > 0 else None
                        try:
                            days = (datetime.date.today()
                                    - datetime.date.fromisoformat(row["date"])).days
                        except Exception:
                            days = None
                        src = row.get("source", "")
                        ai_perf_rows.append({
                            "_src":      src,
                            "Date":      row.get("date",""),
                            "Agent":     _AGENT_ICONS.get(src, src),
                            "Ticker":    t,
                            "Company":   row.get("company","")[:28],
                            "Strategy":  row.get("strategy","")[:22],
                            "Entry $":   entry or None,
                            "Current $": curr or None,
                            "Return":    ret,
                            "Days":      days,
                            "Conviction":row.get("conviction",""),
                            "Headline":  row.get("headline","")[:60],
                        })
            except Exception:
                pass

        # ── Build per-agent scorecard ─────────────────────────────────────
        _agent_stats = {}  # src → {picks, rets, days_list, wins, best_t, best_r, worst_t, worst_r}
        for r in ai_perf_rows:
            src = r["_src"]
            if src not in _agent_stats:
                _agent_stats[src] = {"picks": 0, "rets": [], "days_list": [],
                                     "wins": 0, "best_t": "—", "best_r": None,
                                     "worst_t": "—", "worst_r": None}
            st = _agent_stats[src]
            st["picks"] += 1
            ret = r.get("Return")
            days_val = r.get("Days")
            if ret is not None:
                st["rets"].append(ret)
                if days_val:
                    try: st["days_list"].append(float(days_val))
                    except Exception: pass
                if ret > 0:
                    st["wins"] += 1
                if st["best_r"] is None or ret > st["best_r"]:
                    st["best_r"] = ret
                    st["best_t"] = r["Ticker"]
                if st["worst_r"] is None or ret < st["worst_r"]:
                    st["worst_r"] = ret
                    st["worst_t"] = r["Ticker"]

        scorecard_tiles = []
        # Sort: judge first, then specialists in fixed order, then any unknown
        _sorted_srcs = (
            [s for s in _AGENT_ORDER if s in _agent_stats] +
            [s for s in _agent_stats if s not in _AGENT_ORDER]
        )
        for src in _sorted_srcs:
            st = _agent_stats[src]
            n          = st["picks"]
            wins       = st["wins"]
            rets       = st["rets"]
            days_list  = st["days_list"]
            avg_r      = sum(rets)/len(rets) if rets else None
            win_r      = wins/n if n else None
            best_r     = st["best_r"]
            best_t     = st["best_t"]
            worst_r    = st["worst_r"]
            worst_t    = st["worst_t"]
            name       = _AGENT_ICONS.get(src, src)
            # Sharpe: annualised on raw returns (no SPY subtraction here — relative perf within scorecard)
            sharpe_val = None
            if len(rets) >= 3 and avg_r is not None:
                _var = sum((x - avg_r)**2 for x in rets) / (len(rets) - 1)
                _std = _var ** 0.5
                if _std > 0:
                    _avg_days = sum(days_list) / len(days_list) if days_list else 30
                    _periods  = max(1.0, 252.0 / max(_avg_days, 1.0))
                    sharpe_val = (avg_r / _std) * (_periods ** 0.5)
            # Card colour: green if Sharpe > 0, red if < 0, neutral otherwise
            card_cls = ("win" if sharpe_val and sharpe_val > 0 else
                        "loss" if sharpe_val is not None and sharpe_val <= 0 else
                        ("win" if win_r and win_r > 0.5 else ""))
            avg_color   = "#a5d6a7" if avg_r and avg_r > 0 else "#ef9a9a"
            sharpe_str  = f"{sharpe_val:.2f}" if sharpe_val is not None else "—"
            sharpe_color = "#a5d6a7" if sharpe_val and sharpe_val > 0 else "#ef9a9a"
            worst_html = (f'<br>Worst: <b>{worst_t}</b> '
                          f'<span style="color:#ef9a9a">{_pct(worst_r)}</span>'
                          if worst_r is not None and worst_r < 0 else "")
            scorecard_tiles.append(f"""
<div class="agent-card {card_cls}">
  <div class="ag-name">{name}</div>
  <div class="ag-stat">
    <b>{n}</b> picks &nbsp;·&nbsp;
    <b style="color:{'#a5d6a7' if win_r and win_r>0.5 else '#ef9a9a'}">{f"{win_r*100:.0f}%" if win_r is not None else "—"}</b> win rate<br>
    Avg: <b style="color:{avg_color}">{_pct(avg_r) if avg_r is not None else "—"}</b>
    &nbsp;·&nbsp; Sharpe: <b style="color:{sharpe_color}">{sharpe_str}</b><br>
    Best: <b>{best_t}</b> {f'<span style="color:#a5d6a7">{_pct(best_r)}</span>' if best_r else ''}
    {worst_html}
  </div>
</div>""")

        scorecard_html = ""
        if scorecard_tiles:
            scorecard_html = f"""
<h2 style="margin:0 0 8px;font-size:.75rem">AGENT SCORECARD</h2>
<div class="agent-grid">{"".join(scorecard_tiles)}</div>"""

        # ── B8: Agent Leaderboard (SPY-adjusted attribution, uses compute_agent_performance) ──
        leaderboard_html = ""
        if agent_perf:
            _lb_hdrs = ["Agent", "Picks", "Avg Ret", "SPY Ret", "Alpha", "Sharpe",
                        "Win %", "90d Hit %", "Med Hold", "Best Pick", "Worst Pick"]
            _lb_rows = []
            _lb_src_order = (
                [s for s in _AGENT_ORDER if s in agent_perf] +
                [s for s in agent_perf if s not in _AGENT_ORDER]
            )
            for _src in _lb_src_order:
                _st = agent_perf[_src]
                _n      = _st.get("n_picks", 0)
                _alpha  = _st.get("alpha")
                _sharpe = _st.get("sharpe")
                _avg_r  = _st.get("avg_ret")
                _avg_spy = _st.get("avg_spy")
                _win    = _st.get("win_rate")
                _h90    = _st.get("hit_90d")
                _med    = _st.get("med_hold_days")
                _b_t    = _st.get("best_ticker", "—")
                _b_r    = _st.get("best_ret")
                _w_t    = _st.get("worst_ticker", "—")
                _w_r    = _st.get("worst_ret")
                _name   = _AGENT_ICONS.get(_src, _src)
                # Row colour: alpha-driven
                if _alpha is not None and _alpha > 0.005:
                    _row_cls = ' class="g"'
                elif _alpha is not None and _alpha < -0.005:
                    _row_cls = ' class="r"'
                else:
                    _row_cls = ""
                def _lb_pct(v):
                    return f"{v*100:+.1f}%" if v is not None else "—"
                def _lb_f(v, fmt="{:.2f}"):
                    return fmt.format(v) if v is not None else "—"
                _best_str  = (f"{_b_t} {_lb_pct(_b_r)}" if _b_r is not None else "—")
                _worst_str = (f"{_w_t} {_lb_pct(_w_r)}" if _w_r is not None else "—")
                _alpha_color  = "#a5d6a7" if (_alpha or 0) > 0 else "#ef9a9a"
                _sharpe_color = "#a5d6a7" if (_sharpe or 0) > 0 else "#ef9a9a"
                _cells = [
                    f'<td><b>{_name}</b></td>',
                    f'<td style="text-align:center">{_n}</td>',
                    f'<td>{_lb_pct(_avg_r)}</td>',
                    f'<td>{_lb_pct(_avg_spy)}</td>',
                    f'<td style="color:{_alpha_color};font-weight:700">{_lb_pct(_alpha)}</td>',
                    f'<td style="color:{_sharpe_color};font-weight:700">{_lb_f(_sharpe)}</td>',
                    f'<td>{_lb_pct(_win)}</td>',
                    f'<td>{_lb_pct(_h90) if _h90 is not None else "—"}</td>',
                    f'<td>{f"{int(_med)}d" if _med else "—"}</td>',
                    f'<td style="color:#a5d6a7;font-size:.72rem">{_best_str}</td>',
                    f'<td style="color:#ef9a9a;font-size:.72rem">{_worst_str}</td>',
                ]
                _lb_rows.append(f"<tr{_row_cls}>{''.join(_cells)}</tr>")
            if _lb_rows:
                _th = "".join(f'<th onclick="sortTable(this)">{h}</th>' for h in _lb_hdrs)
                leaderboard_html = f"""
<h2 style="margin:14px 0 6px;font-size:.75rem">🏆 AGENT LEADERBOARD (SPY-adjusted, all-time)</h2>
<p style="font-size:.7rem;color:#666;margin-bottom:8px">Alpha = agent avg return minus SPY avg return over same hold periods.
Sharpe on alpha series. Click any column header to sort.</p>
<div class="tbl-wrap">
  <table id="lb-table">
    <thead><tr>{_th}</tr></thead>
    <tbody>{"".join(_lb_rows)}</tbody>
  </table>
</div>"""

        # ── Picks table (sorted: judge first, then by agent, then date) ───
        def _sort_key(r):
            src = r["_src"]
            try: order = _AGENT_ORDER.index(src)
            except ValueError: order = 99
            return (order, r.get("Date",""))

        ai_perf_rows.sort(key=_sort_key)
        ai_hdr = ["Date","Agent","Ticker","Company","Strategy",
                  "Entry $","Current $","Return","Days","Conviction","Headline"]
        ai_body = []
        for i, r in enumerate(ai_perf_rows):
            ret = r.get("Return")
            ret_cls = "g" if ret and ret > 0 else ("r" if ret and ret < 0 else "")
            alt = ' class="alt"' if i % 2 == 0 else ''
            cells = []
            for c in ai_hdr:
                v = r.get(c)
                if c == "Return":
                    disp = _pct(v) if v is not None else "—"
                    cells.append(f'<td class="{ret_cls}">{disp}</td>' if ret_cls
                                 else f"<td>{disp}</td>")
                elif c in ("Entry $", "Current $"):
                    cells.append(f"<td>{_money(v)}</td>")
                elif c == "Conviction":
                    cells.append(f"<td>{_conv_badge(v) if v else '—'}</td>")
                else:
                    cells.append(f"<td>{v if v is not None else '—'}</td>")
            ai_body.append(f"<tr{alt}>{''.join(cells)}</tr>")

        _n_hist = len(ai_perf_rows)
        ai_table = f"""
<h2 style="margin:14px 0 8px;font-size:.75rem;cursor:pointer;user-select:none"
    onclick="var d=this.nextElementSibling;var a=this.querySelector('.xarrow');var open=d.style.display!=='none';d.style.display=open?'none':'block';a.textContent=open?'▶':'▼';">
  <span class="xarrow">▶</span> Purchase History ({_n_hist} entries)
</h2>
<div style="display:none">
<div class="tbl-wrap">
  <table>
    <thead><tr>{"".join(f"<th>{c}</th>" for c in ai_hdr)}</tr></thead>
    <tbody>{"".join(ai_body)}</tbody>
  </table>
</div>
</div>""" if ai_perf_rows else ""

        # ── Portfolio holdings ────────────────────────────────────────────
        port_rows = []
        total_mkt = 0.0
        port_total_ret = None
        if portfolio and isinstance(portfolio.get("holdings"), list):
            init_cash = portfolio.get("initial_cash", 100000)
            cash      = portfolio.get("cash", 0)
            for h in portfolio["holdings"]:
                t      = h.get("ticker","")
                entry  = float(h.get("entry_price") or 0)
                shares = int(h.get("shares") or 0)
                curr   = (stocks.get(t, {}).get("price") or 0)
                ret    = ((curr - entry) / entry) if entry > 0 and curr > 0 else None
                mktval = curr * shares
                total_mkt += mktval
                port_rows.append({
                    "Entry Date":  h.get("entry_date",""),
                    "Ticker":      t,
                    "Company":     h.get("company","")[:26],
                    "Shares":      shares,
                    "Entry $":     entry or None,
                    "Current $":   curr or None,
                    "Cost Basis":  entry * shares or None,
                    "Mkt Value":   mktval or None,
                    "Return":      ret,
                    "Conviction":  h.get("conviction",""),
                })
            port_total_ret = ((total_mkt + cash - init_cash) / init_cash) if init_cash else None

        if not ai_perf_rows and not port_rows:
            return ('<section id="perf"><p style="color:#666;padding:16px">'
                    'No performance data yet — will populate after first logged picks.'
                    '</p></section>')

        # Portfolio summary bar
        summary_bar = ""
        if port_rows and portfolio:
            cash      = portfolio.get("cash", 0)
            init_cash = portfolio.get("initial_cash", 100000)
            tot_val   = total_mkt + cash
            ret_color = "#a5d6a7" if tot_val >= init_cash else "#ef9a9a"
            summary_bar = f"""
<div style="background:#1a1a2e;padding:10px 14px;border-radius:4px;margin-bottom:12px;
            display:flex;gap:20px;flex-wrap:wrap;font-size:.8rem">
  <span>💰 <b>Cash:</b> {_money(cash)}</span>
  <span>📊 <b>Holdings:</b> {_money(total_mkt)}</span>
  <span>📊 <b>Total value:</b> {_money(tot_val)}</span>
  <span>📈 <b>Total return:</b>
    <b style="color:{ret_color}">{_pct(port_total_ret) if port_total_ret is not None else "—"}</b>
  </span>
  <span style="color:#666">Started: {portfolio.get('started','?')}</span>
</div>"""

        port_hdr = ["Entry Date","Ticker","Company","Shares","Entry $","Current $",
                    "Cost Basis","Mkt Value","Return","Conviction"]
        port_body = []
        for i, r in enumerate(port_rows):
            ret = r.get("Return")
            ret_cls = "g" if ret and ret > 0 else ("r" if ret and ret < 0 else "")
            alt = ' class="alt"' if i % 2 == 0 else ''
            cells = []
            for c in port_hdr:
                v = r.get(c)
                if c == "Return":
                    disp = _pct(v) if v is not None else "—"
                    cells.append(f'<td class="{ret_cls}">{disp}</td>' if ret_cls
                                 else f"<td>{disp}</td>")
                elif c in ("Entry $","Current $","Cost Basis","Mkt Value"):
                    cells.append(f"<td>{_money(v)}</td>")
                elif c == "Conviction":
                    cells.append(f"<td>{_conv_badge(v) if v else '—'}</td>")
                else:
                    cells.append(f"<td>{v if v is not None else '—'}</td>")
            port_body.append(f"<tr{alt}>{''.join(cells)}</tr>")

        # ── C6: Portfolio risk dashboard ───────────────────────────────────
        risk_dash_html = ""
        if portfolio and isinstance(portfolio.get("holdings"), list) and portfolio["holdings"]:
            import math as _math
            _risk_metrics = []  # per-holding (ticker, weight, vol_60d, beta, curr_dd)

            # Ensure SPY in _hist_price_cache (for beta computation)
            if "SPY" not in _hist_price_cache:
                fetch_price_on_date("SPY",
                                    (datetime.date.today() - datetime.timedelta(days=1)).isoformat())
            # Ensure all holding tickers are cached (for vol/beta)
            for _ch in portfolio["holdings"]:
                _ct = _ch.get("ticker","")
                if _ct and _ct not in _hist_price_cache:
                    fetch_price_on_date(_ct,
                                        (datetime.date.today() - datetime.timedelta(days=1)).isoformat())

            # Build SPY daily return series from _hist_price_cache (if available)
            _spy_prices_60 = _hist_price_cache.get("SPY", {})
            _spy_sorted = sorted(_spy_prices_60.items())[-65:] if _spy_prices_60 else []
            _spy_rets = []
            for _i in range(1, len(_spy_sorted)):
                _p0, _p1 = _spy_sorted[_i-1][1], _spy_sorted[_i][1]
                if _p0 > 0 and _p1 > 0:
                    _spy_rets.append((_p1 - _p0) / _p0)
            _spy_var  = (sum((_r - (sum(_spy_rets)/len(_spy_rets)))**2 for _r in _spy_rets)
                         / (len(_spy_rets) - 1)) if len(_spy_rets) > 2 else None

            total_w = max(total_mkt, 1)
            for h in portfolio["holdings"]:
                _ht   = h.get("ticker","")
                _hcurr = stocks.get(_ht, {}).get("price") or h.get("entry_price") or 0
                _hentry = float(h.get("entry_price") or 0)
                _wt   = (_hcurr * int(h.get("shares", 0))) / total_w

                # 60d price history for this holding
                _phist = _hist_price_cache.get(_ht, {})
                _psorted = sorted(_phist.items())[-65:] if _phist else []
                _prets = []
                for _i in range(1, len(_psorted)):
                    _p0, _p1 = _psorted[_i-1][1], _psorted[_i][1]
                    if _p0 > 0 and _p1 > 0:
                        _prets.append((_p1 - _p0) / _p0)

                _vol_60 = None
                _beta   = None
                if len(_prets) > 5:
                    _pm = sum(_prets) / len(_prets)
                    _vol_60 = (_math.sqrt(sum((_r - _pm)**2 for _r in _prets) /
                                          (len(_prets) - 1)) * _math.sqrt(252))
                    # Beta = cov(stock, spy) / var(spy)
                    if _spy_var and len(_spy_rets) >= len(_prets):
                        _n = min(len(_prets), len(_spy_rets))
                        _sp = _spy_rets[-_n:]
                        _st = _prets[-_n:]
                        _spm = sum(_sp) / _n
                        _stm = sum(_st) / _n
                        _cov = sum((_sp[_i2] - _spm) * (_st[_i2] - _stm) for _i2 in range(_n)) / (_n - 1)
                        _beta = _cov / _spy_var if _spy_var else None

                _curr_dd = ((_hcurr - _hentry) / _hentry) if _hentry > 0 and _hcurr > 0 else None
                _risk_metrics.append({"t": _ht, "w": _wt, "vol": _vol_60, "beta": _beta, "dd": _curr_dd})

            # Portfolio-level aggregations (weight-averaged)
            _port_vol_items  = [(_m["vol"],  _m["w"]) for _m in _risk_metrics if _m["vol"]  is not None]
            _port_beta_items = [(_m["beta"], _m["w"]) for _m in _risk_metrics if _m["beta"] is not None]
            _wt_sum_v = sum(_w for _, _w in _port_vol_items)
            _wt_sum_b = sum(_w for _, _w in _port_beta_items)
            _port_vol  = (sum(_v * _w for _v, _w in _port_vol_items)  / _wt_sum_v) if _wt_sum_v > 0 else None
            _port_beta = (sum(_b * _w for _b, _w in _port_beta_items) / _wt_sum_b) if _wt_sum_b > 0 else None

            # Max drawdown: worst current holding return (proxy — true NAV DD needs full history)
            _all_dds = [_m["dd"] for _m in _risk_metrics if _m["dd"] is not None]
            _max_dd_proxy = min(_all_dds) if _all_dds else None  # most negative = worst DD

            def _risk_color(v, good_thresh, bad_thresh, invert=False):
                if v is None: return "#888"
                ok = v <= good_thresh if not invert else v >= good_thresh
                bad = v >= bad_thresh if not invert else v <= bad_thresh
                return "#a5d6a7" if ok else ("#ef9a9a" if bad else "#ffe082")

            _vol_str  = f"{_port_vol*100:.1f}%" if _port_vol is not None else "—"
            _beta_str = f"{_port_beta:.2f}" if _port_beta is not None else "—"
            _dd_str   = f"{_max_dd_proxy:+.1%}" if _max_dd_proxy is not None else "—"
            _vol_col  = _risk_color(_port_vol, 0.15, 0.25)  # green <15%, red >25%
            _beta_col = _risk_color(_port_beta, 1.0, 1.4)   # green <1.0, red >1.4
            _dd_col   = _risk_color(_max_dd_proxy, -0.15, -0.30)  # green >-15%, red <-30%

            risk_dash_html = f"""
<div style="background:#1a1a2e;border:1px solid #2a2a3e;border-radius:6px;
            padding:12px 16px;margin-bottom:12px">
  <div style="font-size:.72rem;font-weight:700;color:#90caf9;
              text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px">
    ⚠️ C6 Portfolio Risk Dashboard (60-day, proxy)
  </div>
  <div style="display:flex;gap:24px;flex-wrap:wrap;font-size:.8rem">
    <span>📉 <b>Worst Holding DD:</b>
      <b style="color:{_dd_col}">{_dd_str}</b>
      <small style="color:#666">(proxy for max-DD)</small>
    </span>
    <span>〰️ <b>Port. Vol (ann.):</b>
      <b style="color:{_vol_col}">{_vol_str}</b>
      <small style="color:#666">wt-avg 60d</small>
    </span>
    <span>📊 <b>Beta vs SPY:</b>
      <b style="color:{_beta_col}">{_beta_str}</b>
      <small style="color:#666">wt-avg 60d</small>
    </span>
  </div>
</div>"""

        port_table = f"""
<h2 style="margin:14px 0 8px;font-size:.75rem">AI PORTFOLIO — PAPER TRADING</h2>
{risk_dash_html}
{summary_bar}
<div class="tbl-wrap">
  <table>
    <thead><tr>{"".join(f"<th>{c}</th>" for c in port_hdr)}</tr></thead>
    <tbody>{"".join(port_body)}</tbody>
  </table>
</div>""" if port_rows else ""

        return f"""
<section id="perf">
  <div class="section-title">📈 Performance Tracking</div>
  {scorecard_html}
  {leaderboard_html}
  {port_table}
  {ai_table}
</section>"""

    # ── STRATEGY TABLE COLS ───────────────────────────────────────────────
    STRAT_COLS = ["Rank","Ticker","Company","Sector","Price","Score",
                  "PEG","Fwd PEG","P/E","ROIC","ROE","FCF Yield",
                  "MoS","Rev Growth","Piotroski","MktCap ($B)"]

    # ── ASSEMBLE ──────────────────────────────────────────────────────────
    body = f"""
<div class="header">
  <div>
    <h1>📈 FMP Stock Screener</h1>
    <small>{now} · {len(stocks):,} stocks · {fmp_call_count} API calls</small>
  </div>
  <small style="color:#9fa8da">17 Specialists · Master Manager</small>
</div>
<nav class="nav">{nav_html}</nav>
{_ai_section()}
{_macro_section()}
{_strategy_table(iv_rows,   STRAT_COLS + ["IV","D/E","EV/EBITDA"],   "iv",       "📊 IV Discount Picks",
    "DCF intrinsic value discount ≥5% · Piotroski ≥6 (value-trap gate only) · Altman Z ≥1.5. "
    "Ranked on: MoS + Lynch quality (rev consistency, FCF conversion, EPS growth, buybacks) "
    "+ FCF yield + ROIC + ROE + multi-metric cheapness. Top 50.")}
{_strategy_table(stalwarts,  STRAT_COLS + ["CEO Score","FCF/Sh 5Y","Divergence"],    "stalwart", "🏛 Stalwarts",
    "Revenue growth 5–25% · MktCap >$2B · P/E <50 · FCF positive · Piotroski ≥5 · "
    "Rev consistency ≥60% · Excl. Basic Materials. Lynch 'boring but reliable' category.")}
{_strategy_table(fast_growers, STRAT_COLS + ["Rev Growth 5Y","EPS Growth 5Y"], "fastg",    "🚀 Fast Growers",
    "Revenue growth >20% · PEG <1.5 · ROIC >10% · FCF positive or high-growth exception. "
    "Lynch's highest-return category — find the next 10-bagger before it's obvious.")}
{_strategy_table(ten_baggers or [], STRAT_COLS + ["Gross Margin","Oper Margin","Net Debt/EBITDA"], "tenb", "🎯 Lynch 10-Baggers",
    "Small-cap $50M–$2B · PEG<2 · Gross margin>20% · Operating margin>0. "
    "Lynch's real criteria: pricing power + growth at a reasonable price. No FCF kill — early Amazon had negative FCF too.")}
{_strategy_table(slow_growers, STRAT_COLS + ["Div Yield"],              "slowg",    "🐢 Slow Growers",
    "Dividend yield ≥2% · Stable multi-year earnings · Large established companies. "
    "Lynch category: own for income + capital preservation, not growth.")}
{_strategy_table(cyclicals,  STRAT_COLS,                               "cycl",     "🔄 Cyclicals",
    "Cyclical sectors (Industrials, Energy, Materials, Consumer Cyclical) at earnings trough. "
    "Lynch: P/E is INVERTED for cyclicals — HIGH or missing P/E = depressed earnings = BUY. "
    "LOW P/E = peak earnings = SELL. Best entry: decline decelerating + FCF positive + net debt &lt; 3× EBITDA.")}
{_strategy_table(turnarounds, STRAT_COLS,                              "turn",     "🔁 Turnarounds",
    "Down ≥40% from 52W high · Revenue recovering (positive growth trend) · Piotroski ≥4. "
    "Lynch: near-bankrupt companies that survive can be 10x — but require highest conviction.")}
{_strategy_table(asset_plays, STRAT_COLS + ["P/B","Graham NN"],        "asset",    "🏗 Asset Plays",
    "P/B <1 · Hidden tangible asset value (real estate, cash, IP) · FCF positive. "
    "Lynch/Graham: buy $1 of assets for <$1. Best in Financial Services, Real Estate, Industrials.")}
{_strategy_table(quality_compounders, STRAT_COLS + ["P/FCF","EV/EBITDA","CEO Score","FCF/Sh 5Y","Divergence"], "qual",  "🏆 Quality Compounders",
    "ROIC >15% · PEG <2 · FCF positive · Operating margin >20% · Revenue growth >8%. "
    "Buffett category: wonderful companies at fair prices — hold forever, let compounding work. "
    "(absorbs IV Discount + Stalwarts cuts — same compounding thesis, different ranking lenses)")}
{_strategy_table(hold_forever or [], STRAT_COLS + ["FCF Margin","CEO Score","FCF/Sh 5Y","Divergence"], "hold",  "💎 Hold Forever — Buy-and-Forget",
    "Strict cuts: ROIC >15% · 80%+ rev consistency · 5–25% steady CAGR · moat (gross margin >40% or oper margin >15%) · "
    "no dilution · low debt · Piotroski ≥6. Top 25 names. The user's natural buy-and-hold candidate pool — "
    "scan for 🛒 (Familiar Brand) and 🔍 (Wall St under-coverage) tags; cross-check against personal knowledge before sizing.")}
{_sector_section()}
{_perf_section()}
<footer style="padding:16px;color:#555;font-size:.72rem;text-align:center">
  FMP Screener · {now} · Not investment advice.
</footer>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FMP Screener — {today}</title>
<style>{css}</style>
</head>
<body>
{body}
<script>{js}</script>
</body>
</html>"""


# ─────────────────────────────────────────────
# B5: BACKTEST CLI
# ─────────────────────────────────────────────

def backtest_picks(cutoff_date_str: str) -> None:
    """B5: Walk-forward backtest.
    For every AI pick logged BEFORE cutoff_date_str, fetch 30/90/180d returns
    using FMP historical prices. Emit per-pick and per-agent CSV report.

    Usage: python FMP_stock_screener.py --backtest 2026-01-01
    """
    import csv as _csv
    try:
        cutoff_dt = datetime.date.fromisoformat(cutoff_date_str)
    except ValueError:
        print(f"  ❌ Invalid date: {cutoff_date_str} (use YYYY-MM-DD)")
        return

    if not os.path.exists(AI_PICKS_LOG):
        print(f"  ❌ AI picks log not found: {AI_PICKS_LOG}")
        return

    print(f"\n{'='*65}")
    print(f"  📊 BACKTEST — picks before {cutoff_date_str}")
    print(f"{'='*65}")

    # ── Read picks ──────────────────────────────────────────────────────
    picks = []
    with open(AI_PICKS_LOG, "r", encoding="utf-8") as _f:
        for row in _csv.DictReader(_f):
            try:
                d = datetime.date.fromisoformat(row["date"])
            except Exception:
                continue
            if d >= cutoff_dt:
                continue
            try:
                entry = float(row.get("entry_price") or 0)
            except (ValueError, TypeError):
                entry = 0
            if entry <= 0:
                continue
            picks.append({
                "date":     row["date"],
                "source":   row.get("source", ""),
                "ticker":   row.get("ticker", ""),
                "company":  row.get("company", ""),
                "entry":    entry,
                "prompt_v": row.get("prompt_version", ""),
            })

    if not picks:
        print(f"  ⚠️  No picks with entry price before {cutoff_date_str}")
        return

    print(f"  📦 {len(picks)} picks to evaluate (with valid entry price)")

    # ── Fetch SPY history for benchmark ─────────────────────────────────
    earliest = min(p["date"] for p in picks)
    spy_hist = fetch_spy_history(earliest)
    spy_vals = sorted(spy_hist.items())  # [(date_str, price), ...]

    def _spy_ret_at(entry_date_str: str, target_date_str: str) -> "float | None":
        entry_p = spy_hist.get(entry_date_str)
        if not entry_p:
            # forward-search entry
            try:
                ed = datetime.date.fromisoformat(entry_date_str)
            except ValueError:
                return None
            for delta in range(6):
                c = (ed + datetime.timedelta(days=delta)).isoformat()
                entry_p = spy_hist.get(c)
                if entry_p:
                    break
        if not entry_p:
            return None
        target_p = None
        try:
            td = datetime.date.fromisoformat(target_date_str)
        except ValueError:
            return None
        for delta in range(6):
            c = (td + datetime.timedelta(days=delta)).isoformat()
            target_p = spy_hist.get(c)
            if target_p:
                break
        if not target_p:
            return None
        return (target_p - entry_p) / entry_p

    # ── Evaluate each pick ───────────────────────────────────────────────
    print(f"  ⏱  Fetching historical prices (may take a while for large logs)...")
    result_rows = []
    for i, p in enumerate(picks):
        t        = p["ticker"]
        entry    = p["entry"]
        edate    = p["date"]

        if i > 0 and i % 50 == 0:
            print(f"    ... {i}/{len(picks)} evaluated")

        rets = {}
        for horizon in (30, 90, 180):
            ret = _checkpoint_ret_b1(t, entry, edate, horizon, horizon + 1)
            spy_ret = None
            if ret is not None:
                try:
                    tgt = (datetime.date.fromisoformat(edate) +
                           datetime.timedelta(days=horizon)).isoformat()
                    spy_ret = _spy_ret_at(edate, tgt)
                except Exception:
                    pass
            rets[f"ret_{horizon}d"]    = ret
            rets[f"spy_{horizon}d"]    = spy_ret
            rets[f"alpha_{horizon}d"]  = (ret - spy_ret) if ret is not None and spy_ret is not None else None

        result_rows.append({
            **p,
            **rets,
        })

    # ── Per-agent summary ────────────────────────────────────────────────
    _agent_bt: dict = {}
    for r in result_rows:
        src = r["source"]
        if src not in _agent_bt:
            _agent_bt[src] = {"n": 0, "alpha_30": [], "alpha_90": [], "alpha_180": []}
        _agent_bt[src]["n"] += 1
        for h in (30, 90, 180):
            a = r.get(f"alpha_{h}d")
            if a is not None:
                _agent_bt[src][f"alpha_{h}"].append(a)

    # ── Write per-pick CSV ───────────────────────────────────────────────
    pick_csv = f"fmp_backtest_{cutoff_date_str}.csv"
    pick_fields = ["date", "source", "ticker", "company", "entry", "prompt_v",
                   "ret_30d", "spy_30d", "alpha_30d",
                   "ret_90d", "spy_90d", "alpha_90d",
                   "ret_180d", "spy_180d", "alpha_180d"]
    with open(pick_csv, "w", newline="", encoding="utf-8") as _f:
        w = _csv.DictWriter(_f, fieldnames=pick_fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(result_rows)
    print(f"\n  📄 Per-pick results → {pick_csv}")

    # ── Write per-agent summary CSV ──────────────────────────────────────
    agent_csv = f"fmp_backtest_{cutoff_date_str}_by_agent.csv"
    agent_fields = ["source", "n_picks",
                    "avg_alpha_30d", "avg_alpha_90d", "avg_alpha_180d",
                    "n_with_30d", "n_with_90d", "n_with_180d"]
    agent_rows = []
    for src, st in sorted(_agent_bt.items()):
        def _mavg(lst):
            return round(sum(lst) / len(lst) * 100, 2) if lst else None
        agent_rows.append({
            "source":       src,
            "n_picks":      st["n"],
            "avg_alpha_30d":  _mavg(st["alpha_30"]),
            "avg_alpha_90d":  _mavg(st["alpha_90"]),
            "avg_alpha_180d": _mavg(st["alpha_180"]),
            "n_with_30d":    len(st["alpha_30"]),
            "n_with_90d":    len(st["alpha_90"]),
            "n_with_180d":   len(st["alpha_180"]),
        })
    with open(agent_csv, "w", newline="", encoding="utf-8") as _f:
        w = _csv.DictWriter(_f, fieldnames=agent_fields)
        w.writeheader()
        w.writerows(agent_rows)
    print(f"  📄 Per-agent summary → {agent_csv}")

    # ── Print quick console summary ──────────────────────────────────────
    print(f"\n  📊 PER-AGENT ALPHA (vs SPY, % basis):")
    print(f"  {'Agent':<30} {'N':>4}  {'30d α':>7}  {'90d α':>7}  {'180d α':>7}")
    print(f"  {'-'*60}")
    for r2 in sorted(agent_rows, key=lambda x: (x["avg_alpha_90d"] or -999), reverse=True):
        def _fs(v):
            return f"{v:+.2f}%" if v is not None else "  —"
        print(f"  {r2['source']:<30} {r2['n_picks']:>4}  "
              f"{_fs(r2['avg_alpha_30d']):>7}  {_fs(r2['avg_alpha_90d']):>7}  "
              f"{_fs(r2['avg_alpha_180d']):>7}")
    print(f"\n  ✅ Backtest complete — {len(result_rows)} picks evaluated")
    print(f"{'='*65}\n")


# ─────────────────────────────────────────────
# C4: BACKTEST REPLAY ENGINE
# ─────────────────────────────────────────────

def backtest_replay(from_date_str: str = None) -> None:
    """C4: Full walk-forward portfolio replay.

    For each day in the AI picks log, simulates the portfolio as if the PM
    ran on that date using the picks available at that time and historical prices.
    Emits a cumulative NAV curve, Sharpe vs SPY, and an HTML report.

    Usage: python FMP_stock_screener.py --replay [YYYY-MM-DD]
    (from_date_str defaults to the earliest date in the AI picks log)
    """
    import csv as _csv, math as _math

    print(f"\n{'='*65}")
    print(f"  🔄 C4 BACKTEST REPLAY ENGINE")
    print(f"{'='*65}")

    if not os.path.exists(AI_PICKS_LOG):
        print(f"  ❌ AI picks log not found: {AI_PICKS_LOG}")
        return

    # ── 1. Read picks grouped by date ──────────────────────────────────
    picks_by_date: dict = {}   # date_str → list of {ticker, source, entry_price}
    all_tickers = set()
    with open(AI_PICKS_LOG, "r", encoding="utf-8") as _f:
        for row in _csv.DictReader(_f):
            d   = row.get("date", "")
            t   = row.get("ticker", "")
            src = row.get("source", "")
            try:
                ep = float(row.get("entry_price") or 0)
            except (ValueError, TypeError):
                ep = 0
            if not d or not t or ep <= 0:
                continue
            picks_by_date.setdefault(d, []).append({
                "ticker": t, "source": src, "entry_price": ep,
                "company": row.get("company", ""),
            })
            all_tickers.add(t)

    if not picks_by_date:
        print("  ⚠️  No picks with valid entry prices found — nothing to replay.")
        return

    all_dates = sorted(picks_by_date.keys())
    start_date_str = from_date_str or all_dates[0]
    try:
        start_dt = datetime.date.fromisoformat(start_date_str)
    except ValueError:
        print(f"  ❌ Invalid from-date: {start_date_str}")
        return

    replay_dates = [d for d in all_dates if d >= start_date_str]
    if not replay_dates:
        print(f"  ⚠️  No picks on or after {start_date_str}")
        return

    print(f"  📦 {len(replay_dates)} pick dates from {replay_dates[0]} to {replay_dates[-1]}")
    print(f"  🌐 Pre-fetching price history for {len(all_tickers)} tickers + SPY...")

    # ── 2. Pre-fetch price histories ────────────────────────────────────
    # SPY first
    if "SPY" not in _hist_price_cache:
        fetch_price_on_date("SPY", (datetime.date.today() - datetime.timedelta(days=1)).isoformat())
    # All tickers (batch with small throttle to avoid rate limits)
    for i, t in enumerate(sorted(all_tickers)):
        if t not in _hist_price_cache:
            fetch_price_on_date(t, (datetime.date.today() - datetime.timedelta(days=1)).isoformat())
        if i > 0 and i % 50 == 0:
            print(f"    ... {i}/{len(all_tickers)} tickers cached")

    spy_hist = _hist_price_cache.get("SPY", {})

    # ── 3. Simulate portfolio ────────────────────────────────────────────
    # Simple equal-weight, rebalance on each pick date by adding up to 10 positions
    # (new picks replace old ones if max positions reached)
    REPLAY_INITIAL_CASH = 100_000.0
    REPLAY_MAX_POS      = 10
    REPLAY_POS_SIZE     = REPLAY_INITIAL_CASH / REPLAY_MAX_POS   # $10K per slot

    # holdings: ticker → {shares, entry_price, entry_date}
    holdings: dict = {}
    cash = REPLAY_INITIAL_CASH

    # NAV timeline: list of (date_str, nav, spy_indexed)
    nav_timeline = []
    spy_start_price = None

    def _price_on(ticker: str, date_str: str) -> "float | None":
        prices = _hist_price_cache.get(ticker, {})
        try:
            td = datetime.date.fromisoformat(date_str)
        except ValueError:
            return None
        for delta in range(6):
            c = (td + datetime.timedelta(days=delta)).isoformat()
            p = prices.get(c)
            if p and p > 0:
                return p
        return None

    def _nav_on(date_str: str) -> float:
        total = cash
        for t_h, h in holdings.items():
            p = _price_on(t_h, date_str) or h["entry_price"]
            total += h["shares"] * p
        return total

    def _spy_idx(date_str: str) -> "float | None":
        nonlocal spy_start_price
        p = _price_on("SPY", date_str)
        if p is None:
            return None
        if spy_start_price is None:
            spy_start_price = p
        return (p / spy_start_price) * REPLAY_INITIAL_CASH

    today_str = datetime.date.today().isoformat()

    for rdate in replay_dates:
        # ─ Add new picks from this date as new positions ─────────────────
        new_picks = picks_by_date.get(rdate, [])
        # Deduplicate: only add tickers not already held
        for pick in new_picks:
            t = pick["ticker"]
            if t in holdings:
                continue
            if len(holdings) >= REPLAY_MAX_POS:
                break
            entry_price = _price_on(t, rdate) or pick["entry_price"]
            if entry_price <= 0:
                continue
            invest = min(REPLAY_POS_SIZE, cash)
            if invest < entry_price:
                continue
            shares = max(1, int(invest / entry_price))
            cost   = shares * entry_price
            cash  -= cost
            holdings[t] = {
                "shares":       shares,
                "entry_price":  entry_price,
                "entry_date":   rdate,
                "source":       pick["source"],
            }

        # ─ Record NAV and SPY on this date ───────────────────────────────
        nav = _nav_on(rdate)
        spy = _spy_idx(rdate)
        nav_timeline.append({"date": rdate, "nav": round(nav, 2),
                              "spy_indexed": round(spy, 2) if spy else None})

    # Final snapshot (today)
    nav_final = _nav_on(today_str)
    spy_final = _spy_idx(today_str)
    nav_timeline.append({"date": today_str, "nav": round(nav_final, 2),
                          "spy_indexed": round(spy_final, 2) if spy_final else None})

    # ── 4. Compute metrics ───────────────────────────────────────────────
    navs = [r["nav"] for r in nav_timeline]

    # Daily returns
    port_rets = [(navs[i] - navs[i-1]) / navs[i-1] for i in range(1, len(navs))]
    spy_navs  = [r["spy_indexed"] or REPLAY_INITIAL_CASH for r in nav_timeline]
    spy_rets  = [(spy_navs[i] - spy_navs[i-1]) / spy_navs[i-1] for i in range(1, len(spy_navs))]

    def _annualised_sharpe(rets: list) -> "float | None":
        if len(rets) < 5:
            return None
        avg = sum(rets) / len(rets)
        std = (_math.sqrt(sum((r - avg)**2 for r in rets) / (len(rets) - 1))
               if len(rets) > 1 else 0)
        return (avg / std * _math.sqrt(252)) if std > 0 else None

    def _max_drawdown(navs: list) -> float:
        peak = navs[0]
        max_dd = 0.0
        for n in navs:
            if n > peak:
                peak = n
            dd = (n - peak) / peak
            if dd < max_dd:
                max_dd = dd
        return max_dd

    total_ret    = (navs[-1] - navs[0]) / navs[0] if navs[0] > 0 else 0
    spy_total    = (spy_navs[-1] - spy_navs[0]) / spy_navs[0] if spy_navs[0] > 0 else 0
    alpha_total  = total_ret - spy_total
    port_sharpe  = _annualised_sharpe(port_rets)
    spy_sharpe   = _annualised_sharpe(spy_rets)
    max_dd       = _max_drawdown(navs)
    n_days       = len(nav_timeline)
    ann_factor   = 252 / max(n_days, 1)
    port_ann     = ((navs[-1] / navs[0]) ** ann_factor - 1) if navs[0] > 0 else 0
    spy_ann      = ((spy_navs[-1] / spy_navs[0]) ** ann_factor - 1) if spy_navs[0] > 0 else 0

    # ── 5. Write NAV CSV ─────────────────────────────────────────────────
    nav_csv = f"fmp_replay_{start_date_str}_nav.csv"
    with open(nav_csv, "w", newline="", encoding="utf-8") as _f:
        w = _csv.DictWriter(_f, fieldnames=["date", "nav", "spy_indexed"])
        w.writeheader()
        w.writerows(nav_timeline)
    print(f"\n  📄 NAV curve → {nav_csv}")

    # ── 6. Write HTML report ─────────────────────────────────────────────
    def _pct_s(v):
        return f"{v*100:+.1f}%" if v is not None else "—"

    chart_labels  = [r["date"]        for r in nav_timeline]
    chart_port    = [r["nav"]          for r in nav_timeline]
    chart_spy     = [r["spy_indexed"]  for r in nav_timeline]

    import json as _json
    labels_json = _json.dumps(chart_labels)
    port_json   = _json.dumps(chart_port)
    spy_json    = _json.dumps([x if x else "null" for x in chart_spy])

    # Current holdings table
    hold_rows = ""
    for t_h, h in sorted(holdings.items()):
        curr_p = _price_on(t_h, today_str) or h["entry_price"]
        ret_h  = (curr_p - h["entry_price"]) / h["entry_price"] if h["entry_price"] > 0 else 0
        clr    = "#a5d6a7" if ret_h > 0 else "#ef9a9a"
        hold_rows += (f'<tr><td>{t_h}</td><td>{h["source"]}</td>'
                      f'<td>{h["entry_date"]}</td>'
                      f'<td>${h["entry_price"]:.2f}</td><td>${curr_p:.2f}</td>'
                      f'<td style="color:{clr};font-weight:700">{ret_h:+.1%}</td>'
                      f'<td>{h["shares"]}</td></tr>\n')

    html_report = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FMP Replay — {start_date_str}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
       background: #0f1117; color: #e0e0e0; font-size: 14px; padding: 16px; }}
h1 {{ color: #90caf9; margin-bottom: 4px; font-size: 1.1rem; }}
h2 {{ color: #90caf9; font-size: .85rem; text-transform: uppercase;
     letter-spacing: .05em; margin: 18px 0 8px; }}
.stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
              gap: 10px; margin-bottom: 20px; }}
.stat-card {{ background: #1a1a2e; border-radius: 6px; padding: 12px 14px; }}
.stat-label {{ font-size: .7rem; color: #666; text-transform: uppercase; letter-spacing: .04em; }}
.stat-value {{ font-size: 1.1rem; font-weight: 700; margin-top: 3px; }}
.g {{ color: #a5d6a7; }} .r {{ color: #ef9a9a; }} .n {{ color: #e0e0e0; }}
canvas {{ max-height: 380px; }}
table {{ border-collapse: collapse; width: 100%; font-size: .78rem; white-space: nowrap; }}
th {{ background: #1a237e; color: #fff; padding: 6px 8px; text-align: left; }}
td {{ padding: 5px 8px; border-bottom: 1px solid #2a2a2a; }}
.tbl-wrap {{ overflow-x: auto; }}
</style>
</head>
<body>
<h1>🔄 FMP Backtest Replay — Portfolio NAV Curve</h1>
<p style="color:#666;font-size:.78rem;margin-bottom:16px">
  Walk-forward simulation · From {start_date_str} to {today_str}
  · {len(replay_dates)} pick dates · {len(holdings)} current positions
</p>

<div class="stats-grid">
  <div class="stat-card">
    <div class="stat-label">Portfolio Total Return</div>
    <div class="stat-value {'g' if total_ret > 0 else 'r'}">{_pct_s(total_ret)}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">SPY Total Return</div>
    <div class="stat-value {'g' if spy_total > 0 else 'r'}">{_pct_s(spy_total)}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">Alpha vs SPY</div>
    <div class="stat-value {'g' if alpha_total > 0 else 'r'}">{_pct_s(alpha_total)}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">Port. Ann. Return</div>
    <div class="stat-value {'g' if port_ann > 0 else 'r'}">{_pct_s(port_ann)}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">SPY Ann. Return</div>
    <div class="stat-value {'g' if spy_ann > 0 else 'r'}">{_pct_s(spy_ann)}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">Port. Sharpe</div>
    <div class="stat-value {'g' if (port_sharpe or 0) > 0 else 'r'}">{f"{port_sharpe:.2f}" if port_sharpe else "—"}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">SPY Sharpe</div>
    <div class="stat-value n">{f"{spy_sharpe:.2f}" if spy_sharpe else "—"}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">Max Drawdown</div>
    <div class="stat-value {'g' if max_dd > -0.10 else 'r'}">{_pct_s(max_dd)}</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">Final NAV</div>
    <div class="stat-value n">${navs[-1]:,.0f}</div>
  </div>
</div>

<h2>📈 NAV Curve</h2>
<canvas id="navChart"></canvas>

<h2>📋 Current Holdings</h2>
<div class="tbl-wrap">
<table>
<thead><tr><th>Ticker</th><th>Source Agent</th><th>Entry Date</th>
<th>Entry $</th><th>Current $</th><th>Return</th><th>Shares</th></tr></thead>
<tbody>{hold_rows or '<tr><td colspan="7" style="color:#666">No open positions</td></tr>'}</tbody>
</table>
</div>

<p style="color:#444;font-size:.7rem;margin-top:20px">
  ⚠️ This is a simplified replay: equal-weight, no transaction costs, no rebalancing on exits.
  Full walk-forward with PM decisions is not simulated. Use as directional signal only.
</p>

<script>
const ctx = document.getElementById('navChart').getContext('2d');
new Chart(ctx, {{
  type: 'line',
  data: {{
    labels: {labels_json},
    datasets: [
      {{
        label: 'Portfolio NAV ($)',
        data: {port_json},
        borderColor: '#42a5f5',
        backgroundColor: 'rgba(66,165,245,0.08)',
        borderWidth: 2,
        pointRadius: 0,
        tension: 0.2,
      }},
      {{
        label: 'SPY (indexed to $100K)',
        data: {spy_json},
        borderColor: '#66bb6a',
        backgroundColor: 'rgba(102,187,106,0.06)',
        borderWidth: 2,
        pointRadius: 0,
        tension: 0.2,
      }},
    ]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: true,
    plugins: {{
      legend: {{ labels: {{ color: '#ccc' }} }},
      tooltip: {{ mode: 'index', intersect: false }},
    }},
    scales: {{
      x: {{ ticks: {{ color: '#666', maxTicksLimit: 12 }}, grid: {{ color: '#1a1a2e' }} }},
      y: {{ ticks: {{ color: '#ccc', callback: v => '$'+v.toLocaleString() }},
            grid: {{ color: '#1a1a2e' }} }},
    }},
  }}
}});
</script>
</body>
</html>"""

    replay_html = f"fmp_replay_{start_date_str}.html"
    with open(replay_html, "w", encoding="utf-8") as _f:
        _f.write(html_report)
    print(f"  📄 HTML report  → {replay_html}")

    # Console summary
    print(f"\n  📊 REPLAY SUMMARY:")
    print(f"  {'Period:':<22} {start_date_str} → {today_str} ({n_days} days)")
    print(f"  {'Portfolio total ret:':<22} {_pct_s(total_ret)}")
    print(f"  {'SPY total ret:':<22} {_pct_s(spy_total)}")
    print(f"  {'Alpha:':<22} {_pct_s(alpha_total)}")
    print(f"  {'Port. ann. return:':<22} {_pct_s(port_ann)}")
    print(f"  {'SPY ann. return:':<22} {_pct_s(spy_ann)}")
    print(f"  {'Portfolio Sharpe:':<22} {f'{port_sharpe:.2f}' if port_sharpe else '—'}")
    print(f"  {'SPY Sharpe:':<22} {f'{spy_sharpe:.2f}' if spy_sharpe else '—'}")
    print(f"  {'Max drawdown:':<22} {_pct_s(max_dd)}")
    print(f"\n  ✅ Replay complete")
    print(f"{'='*65}\n")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    global _fmp_call_count, _run_start

    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", action="store_true", help="Debug mode: 20 stocks only, fast iteration")
    parser.add_argument("--log-picks", action="store_true", help="Log today's top picks to CSV for performance tracking")
    parser.add_argument("--backtest", metavar="YYYY-MM-DD",
                        help="B5: Run backtest on picks before this date and exit (e.g. 2026-01-01)")
    parser.add_argument("--replay", metavar="YYYY-MM-DD", nargs="?", const="all",
                        help="C4: Walk-forward portfolio replay from date (or all-time if omitted)")
    parser.add_argument("--force-fresh-ai", action="store_true",
                        help="Bypass the daily AI cache and re-run all specialists + judge for this run")
    args = parser.parse_args()

    # B5: Backtest mode — run and exit without building the full report
    if args.backtest:
        backtest_picks(args.backtest)
        return

    # C4: Replay mode — walk-forward portfolio simulation
    if args.replay:
        from_d = None if args.replay == "all" else args.replay
        backtest_replay(from_d)
        return

    DEBUG = args.debug
    _run_start = time.time()

    print("=" * 65)
    print("  📈 FMP STOCK SCREENER — Professional Fundamentals Edition")
    if DEBUG:
        print("  ⚡ DEBUG MODE — 20 stocks only for fast testing")
    if NTFY_TOPIC:
        print(f"  🔔 Phone notifications: ntfy.sh/{NTFY_TOPIC}")
    print("=" * 65)

    # ── API key status ──
    fmp_status  = "✅ loaded" if FMP_KEY  else "❌ NOT FOUND"
    ai_status   = "✅ loaded" if ANTHROPIC_KEY else "⚠️  not set  (AI analysis disabled)"
    print(f"  🔑 FMP_API_KEY:        {fmp_status}")
    print(f"  🔑 ANTHROPIC_API_KEY:  {ai_status}")
    if not FMP_KEY or not ANTHROPIC_KEY:
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
        if not os.path.exists(env_path):
            print(f"\n  💡 Tip: create {env_path} with:")
            print("     FMP_API_KEY=your_fmp_key")
            print("     ANTHROPIC_API_KEY=sk-ant-api03-...")
    print()

    if not FMP_KEY:
        print("\n  ❌ FMP_API_KEY not set! Add it to .env or environment variables.")
        sys.exit(1)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    output_file = os.path.join(OUTPUT_DIR, f"fmp_screener_{ts}.xlsx")

    # Load cache
    load_cache()

    # Start notification — sent before the heavy lifting begins
    if NTFY_TOPIC:
        _cache_warm = bool(_cache.get("universe"))
        _eta_start = "~10–15 min (cache warm)" if _cache_warm else "~2.5–3 h (full fetch)"
        _purl_start = (f"https://{GITHUB_REPO.split('/', 1)[0]}.github.io/"
                       f"{GITHUB_REPO.split('/', 1)[1]}/"
                       if GITHUB_REPO and '/' in GITHUB_REPO else "")
        notify_phone(
            title="FMP Screener starting 🚀",
            message=(f"Run started {datetime.datetime.now().strftime('%H:%M')} · ETA {_eta_start}"
                     + (f"\n📱 When done: {_purl_start}" if _purl_start else "")),
            tags="rocket",
        )

    # Fetch macro indicators early (fast, cached, free via FRED)
    phase_start("macro", "Fetching macro indicators (FRED)")
    macro_data = fetch_macro_indicators()
    market_intel = fetch_market_intelligence()   # FMP news for Social Arb / Disruptive / Insider agents

    # Phase 1: Universe from screener (already has price, mktCap, sector)
    phase_start("universe", "Fetching US stock universe (live prices)")
    universe = fetch_us_universe()

    # Phase 2: Skip bulk profiles — screener data is sufficient for filtering
    # Use screener data directly as our "profiles"
    profiles = {}
    for t, u in universe.items():
        if u.get("price") and u.get("mktCap") and u["mktCap"] > 0:
            profiles[t] = u  # screener data IS the profile

    # Filter: stocks with positive price and market cap > $50M for enrichment
    candidates = [t for t in profiles if profiles[t].get("mktCap", 0) > 50_000_000
                  and profiles[t].get("price", 0) > 1]
    # Sort by market cap, take top N for detailed analysis
    candidates.sort(key=lambda t: -(profiles.get(t, {}).get("mktCap") or 0))
    enrich_count = 20 if DEBUG else 4000
    top_for_enrichment = candidates[:enrich_count]
    print(f"\n  📊 {len(candidates)} stocks > $50M mktcap, enriching top {len(top_for_enrichment)}")
    phase_start("enrichment", f"Enriching top {len(top_for_enrichment)} stocks (fundamentals)")

    key_metrics = fetch_key_metrics(top_for_enrichment)
    ratios_ttm = fetch_ratios_ttm(top_for_enrichment)
    dcf_data = fetch_dcf_bulk(top_for_enrichment)
    estimates = fetch_growth_estimates(top_for_enrichment)
    scores = fetch_financial_scores(top_for_enrichment)
    ratings = fetch_ratings(top_for_enrichment)
    growth_data = fetch_financial_growth(top_for_enrichment)
    balance_sheet = fetch_balance_sheet(top_for_enrichment)
    earnings_surp = fetch_earnings_surprises(top_for_enrichment)
    insider_data = fetch_insider_trading(tickers=top_for_enrichment)

    # ── Capital allocator data: CEO tenure + 5Y financials ─────────
    # Lighter universe: only top 2000 by mktcap (where capital-allocation
    # analysis matters most — covers all strategy tab universes)
    capalloc_universe = top_for_enrichment[:2000]
    bs_5y_data        = fetch_balance_sheet_5y(capalloc_universe)
    cfs_5y_data       = fetch_cash_flow_5y(capalloc_universe)
    executives_data   = fetch_key_executives(capalloc_universe)

    # Fetch 52-week high/low (not returned by screener on Starter plan; batched via /quote)
    phase_start("52w_ranges", "Fetching 52-week ranges (parallel)")
    w52 = fetch_52w_ranges(top_for_enrichment)
    for _t, _rng in w52.items():
        if _t in universe:
            universe[_t]["yearHigh"] = _rng.get("yearHigh")
            universe[_t]["yearLow"]  = _rng.get("yearLow")

    # Save cache after all fetches
    save_cache()

    # ── Small-cap supplemental enrichment for Lynch 10-Baggers tab ────────────
    # Main enrichment covers top 4000 by market cap (roughly $2.5B+).
    # Small-caps ($50M–$2B) need a targeted pass for the 2 critical fields:
    #   grossMargin / operatingMargin (ratios-ttm) + revGrowth (financial-growth)
    # Uses separate cache keys to avoid disrupting the main cache.
    phase_start("sc_enrichment", "Small-cap supplemental enrichment (10-Baggers tab)")
    _top_enrich_set = set(top_for_enrichment)
    _sc_candidates = [
        t for t, u in profiles.items()
        if 50e6 < (u.get("mktCap") or 0) < 2e9
        and t not in _top_enrich_set
        and (u.get("sector") or "") not in ("Real Estate", "Basic Materials")
        and (u.get("price") or 0) > 1
    ]
    _sc_candidates.sort(key=lambda t: -(profiles.get(t, {}).get("mktCap") or 0))
    _sc_candidates = _sc_candidates[:3500]  # top 3500 small-caps by mktCap (covers ~$200M+)
    print(f"  🎯 {len(_sc_candidates)} small-cap candidates for 10-Baggers enrichment")

    # Invalidate SC cache if the cached entries are sparse relative to candidates.
    # If the cache was built from a smaller candidate set (e.g. 1500) and we now have
    # 3500 candidates, the entry density will be low (<30%) because ~2000 new candidates
    # were never fetched. Density check is more reliable than tracking candidate counts
    # because the count key can get out of sync (stored as 3500 even if only 688 fetched).
    _sc_cached_ratios_count = len(_cache.get("ratios_ttm_sc") or {})
    _sc_density = _sc_cached_ratios_count / max(len(_sc_candidates), 1)
    if _sc_cached_ratios_count > 0 and _sc_density < 0.28:
        print(f"  🔄 SC cache sparse ({_sc_cached_ratios_count} entries / {len(_sc_candidates)} candidates = {_sc_density:.0%}) — clearing for full refresh")
        _cache.pop("ratios_ttm_sc", None)
        _cache.pop("growth_sc", None)

    if _sc_candidates:
        # Fetch ratios (gross/oper margin) — uses separate cache key to avoid
        # overwriting the main ratios_ttm cache which covers large-caps only
        _sc_ratios = _parallel_fetch(_sc_candidates, "ratios-ttm",
                                     "SC financial ratios", "ratios_ttm_sc")

        # Fetch growth (revenue growth) — separate cache key
        _sc_growth_cache_key = "growth_sc"
        if _cache.get(_sc_growth_cache_key):
            _sc_growth = _cache[_sc_growth_cache_key]
            print(f"  📦 Using cached SC growth ({len(_sc_growth)} stocks)")
        else:
            from concurrent.futures import ThreadPoolExecutor as _TpeSc, as_completed as _ascSc
            import threading as _thr_sc
            _sc_growth = {}
            _lock_sc = _thr_sc.Lock()
            _thr_sc_sem = _thr_sc.Semaphore(4)

            def _fetch_sc_growth(t):
                with _thr_sc_sem:
                    d = fmp_get("financial-growth",
                                {"symbol": t, "period": "annual", "limit": "5"})
                    time.sleep(0.25)
                    return t, (d if d and isinstance(d, list) and d else None)

            with _TpeSc(max_workers=12) as _pool_sc:
                _futs_sc = {_pool_sc.submit(_fetch_sc_growth, t): t for t in _sc_candidates}
                _done_sc = 0
                for _fut_sc in _ascSc(_futs_sc):
                    _t_sc, _d_sc = _fut_sc.result()
                    if _d_sc:
                        with _lock_sc: _sc_growth[_t_sc] = _d_sc
                    _done_sc += 1
                    if _done_sc % 300 == 0:
                        print(f"    [{_done_sc}/{len(_sc_candidates)}] SC growth fetched...")

            _cache[_sc_growth_cache_key] = _sc_growth
            print(f"  ✅ SC growth loaded: {len(_sc_growth)} stocks")

        # Merge into main dicts (don't overwrite existing large-cap data)
        _sc_ratios_added = sum(1 for t, d in _sc_ratios.items() if t not in ratios_ttm)
        _sc_growth_added = sum(1 for t, d in _sc_growth.items() if t not in growth_data)
        for t, d in _sc_ratios.items():
            if t not in ratios_ttm:
                ratios_ttm[t] = d
        for t, d in _sc_growth.items():
            if t not in growth_data:
                growth_data[t] = d
        print(f"  ✅ SC merge: +{_sc_ratios_added} ratios, +{_sc_growth_added} growth entries")

        # ── SC 52w range fetch ────────────────────────────────────────────────
        # The FMP company screener does NOT return yearHigh/yearLow for small-caps
        # (these are null in the universe dict for all stocks below top-4000 by mktCap).
        # Without yearHigh, priceVs52H = None for every Lynch 10-Baggers stock.
        # Fix: fetch /quote for SC candidates; update universe[t] with yearHigh/yearLow.
        _sc_ranges_cache_key = "52w_ranges_sc"
        _sc_cached_ranges = _cache.get(_sc_ranges_cache_key) or {}
        _sc_ranges_density = len(_sc_cached_ranges) / max(len(_sc_candidates), 1)
        _sc_needing_ranges = [t for t in _sc_candidates if not _sc_cached_ranges.get(t)
                              and not universe.get(t, {}).get("yearHigh")]
        if len(_sc_cached_ranges) > 0 and _sc_ranges_density < 0.28:
            print(f"  🔄 SC ranges cache sparse ({len(_sc_cached_ranges)}/{len(_sc_candidates)}) — clearing for full refresh")
            _sc_cached_ranges = {}
            _cache.pop(_sc_ranges_cache_key, None)
            _sc_needing_ranges = _sc_candidates[:]

        if _sc_needing_ranges:
            print(f"  📊 Fetching 52w ranges for {len(_sc_needing_ranges)} SC stocks (parallel)...")
            from concurrent.futures import ThreadPoolExecutor as _TpeRng, as_completed as _ascRng
            import threading as _thr_rng
            _rng_results = dict(_sc_cached_ranges)
            _rng_lock = _thr_rng.Lock()
            _rng_sem  = _thr_rng.Semaphore(4)

            def _fetch_sc_range(t):
                with _rng_sem:
                    try:
                        r = requests.get(f"{FMP_BASE}/quote",
                                         params={"symbol": t, "apikey": FMP_KEY}, timeout=10)
                        time.sleep(0.2)
                        if r.status_code == 200:
                            d = r.json()
                            item = d[0] if isinstance(d, list) and d else (d if isinstance(d, dict) else {})
                            yh = item.get("yearHigh"); yl = item.get("yearLow")
                            if yh and float(yh) > 0:
                                return t, {"yearHigh": float(yh), "yearLow": float(yl or 0)}
                    except Exception:
                        pass
                    return t, None

            with _TpeRng(max_workers=12) as _rng_pool:
                _rng_futs = {_rng_pool.submit(_fetch_sc_range, t): t for t in _sc_needing_ranges}
                _rng_done = 0
                for _fut in _ascRng(_rng_futs):
                    _t2, _r2 = _fut.result()
                    if _r2:
                        with _rng_lock: _rng_results[_t2] = _r2
                    _rng_done += 1
                    if _rng_done % 500 == 0:
                        print(f"    [{_rng_done}/{len(_sc_needing_ranges)}] SC ranges fetched...")

            _cache[_sc_ranges_cache_key] = _rng_results
            print(f"  ✅ SC 52w ranges fetched: {len(_rng_results)} stocks")
        else:
            _rng_results = _sc_cached_ranges
            print(f"  📦 Using cached SC 52w ranges ({len(_rng_results)} stocks)")

        # Apply SC 52w ranges to universe (only where main enrichment didn't supply data)
        _rng_applied = 0
        for _t3, _rng3 in _rng_results.items():
            if _t3 in universe and not universe[_t3].get("yearHigh") and _rng3:
                universe[_t3]["yearHigh"] = _rng3.get("yearHigh")
                universe[_t3]["yearLow"]  = _rng3.get("yearLow")
                _rng_applied += 1
        print(f"  ✅ SC 52w applied: +{_rng_applied} stocks now have yearHigh")

        save_cache()   # persist SC cache keys for next run

    # Assemble
    stocks = assemble_stock_data(universe, profiles, key_metrics, ratios_ttm, dcf_data,
                                 estimates, scores, ratings, growth_data, insider_data,
                                 balance_sheet, earnings_surp,
                                 bs_5y=bs_5y_data, cfs_5y=cfs_5y_data,
                                 executives=executives_data)

    print(f"\n  📊 FMP API calls this run: {_fmp_call_count}")

    # Build Excel
    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "Overview"
    # Pre-create AI picks sheet in position 2, portfolio sheet in position 3, agent reports in position 4
    ws_ai_picks   = wb.create_sheet("1b. AI Top Picks")
    ws_ai_portf   = wb.create_sheet("1c. AI Portfolio")
    # 1d. Agent Reports is created inside build_agent_reports_tab()

    phase_start("tabs", "Building strategy tabs (IV Discount, Lynch, Sector, AI)")
    # Tab 2: IV Discount
    iv_rows = build_iv_discount(wb, stocks)

    # Lynch tabs
    # ─── Stalwarts: consistent 5-25% revenue growers, large cap, quality gate ───
    def _stalwart_filter(s):
        if not _is_common_stock(s): return False
        rg  = s.get("revGrowth")
        pe  = s.get("pe")
        pio = s.get("piotroski")
        fcf = s.get("fcfYield")
        # Kill criteria: FCF must be positive (real business, not accounting mirage)
        if fcf is not None and fcf <= 0:
            return False
        # Kill: Basic Materials excluded — commodity cycle drives revenue, not business quality
        if (s.get("sector") or "") == "Basic Materials":
            return False
        # Kill: RevConsistency < 0.60 — Stalwarts must show steady multi-year growth
        rc = s.get("revConsistency")
        if rc is not None and rc < 0.60:
            return False
        return (rg and 0.05 <= rg <= 0.25
                and s.get("mktCap", 0) > 2e9
                and (pe is None or (0 < pe < 50))
                and (pio is None or pio >= 5))

    def _stalwart_score(s):
        sc = 0
        _roic_s  = s.get("roic")
        _fcf_s   = s.get("fcfYield")
        _rg_s    = s.get("revGrowth") or 0
        _rg_p_s  = s.get("revGrowthPrev")
        _fin_s       = "financial" in (s.get("sector") or "").lower()
        _basic_mat_s = "basic materials" in (s.get("sector") or "").lower()
        _energy_s    = (s.get("sector") or "") == "Energy"

        # ── ROIC first: is this a quality business worth owning? ─────────────
        if _roic_s and _roic_s > 0.25:   sc += 18
        elif _roic_s and _roic_s > 0.20: sc += 14
        elif _roic_s and _roic_s > 0.15: sc += 10
        elif _roic_s and _roic_s > 0.10: sc += 5
        elif _roic_s is not None and _roic_s < 0.08:
            sc -= 4  # low ROIC Stalwart = growing revenue but not compounding capital

        # ── PEG second: is the quality priced fairly? ────────────────────────
        # Max 12pts (ROIC leads; PEG confirms entry price)
        best_peg = s.get("fwdPEG") or s.get("peg")
        peg_sc = 0
        if best_peg and 0 < best_peg < 1:    peg_sc = 12
        elif best_peg and 0 < best_peg < 1.5: peg_sc = 8
        elif best_peg and 0 < best_peg < 2:   peg_sc = 4
        # Quality bonus: ROIC validates the PEG is moat-backed
        if peg_sc and _roic_s and _roic_s > 0.15 and _fcf_s and _fcf_s > 0.04:
            peg_sc = min(12, round(peg_sc * 1.25))
        # Deceleration penalty
        if peg_sc and _rg_p_s is not None and (_rg_s - _rg_p_s) < -0.10:
            peg_sc = round(peg_sc * 0.60)
        # Cyclical sector penalty: financials, commodity miners, E&P energy
        if peg_sc and _fin_s and best_peg and best_peg < 1.5:
            peg_sc = round(peg_sc * 0.50)  # insurance float / cycle distorts PEG
        if peg_sc and _basic_mat_s and best_peg and best_peg < 1.5:
            peg_sc = round(peg_sc * 0.60)  # gold/commodity miners: trough-earnings PEG unreliable
        if peg_sc and _energy_s and best_peg and best_peg < 1.5:
            peg_sc = round(peg_sc * 0.60)  # E&P revenue/earnings move with oil/gas price, not business quality
        sc += peg_sc
        roe = s.get("roe")
        if roe and roe > 0.20:   sc += 8
        elif roe and roe > 0.15: sc += 5
        elif roe and roe > 0.08: sc += 2
        fcf = s.get("fcfYield")
        if fcf and fcf > 0.05: sc += 5
        elif fcf and fcf > 0.02: sc += 2
        pio = s.get("piotroski")
        if pio and pio >= 8: sc += 8
        elif pio and pio >= 7: sc += 5
        elif pio and pio >= 5: sc += 2
        if s.get("mos") and s.get("mos") > 0.2: sc += 5
        if s.get("pe") and 0 < s.get("pe") < 20: sc += 3
        roic = s.get("roic")
        if roic and roic > 0.15: sc += 4
        # Earnings beat rate — consistent executors are the best Stalwarts
        br = s.get("beatRate")
        if br and br >= 0.875: sc += 7
        elif br and br >= 0.75: sc += 4
        elif br and br >= 0.625: sc += 2
        # Revenue acceleration — growth improving YoY = re-accelerating Stalwart
        rg = s.get("revGrowth") or 0
        rg_prev = s.get("revGrowthPrev")
        if rg_prev is not None and rg > 0.05:
            rg_delta = rg - rg_prev
            if rg_delta > 0.05:   sc += 5   # re-accelerating = high conviction
            elif rg_delta < -0.08: sc -= 3  # decelerating toward slow grower zone
        # Revenue consistency — steady multi-year growers carry lower risk
        rc = s.get("revConsistency")
        if rc is not None:
            if rc >= 0.80:   sc += 7   # 4+ of 5 years positive
            elif rc >= 0.60: sc += 3
            elif rc < 0.40:  sc -= 4   # majority of years declining = structural issue
        return sc

    _st_headers = [
        "Rank", "Ticker", "Company", "Sector", "Price",
        "Fwd PEG", "PEG", "Fwd P/E", "P/E", "ROIC",
        "ROE", "FCF Yield", "FCF Conv.", "Oper Margin", "Rev Growth", "Rev Consist.", "Grwth Gap",
        "EPS Growth 5Y", "Piotroski", "52w vs High", "Div Yield",
        "CEO Score", "FCF/Sh 5Y", "Divergence",
        "MktCap ($B)", "Score", "🏦 Insider",
    ]
    _st_widths = [5, 8, 22, 15, 8, 7, 6, 7, 7, 8, 7, 8, 8, 8, 9, 9, 8, 11, 7, 9, 7, 12, 9, 14, 10, 6, 14]

    stalwarts = build_lynch_tab(wb, stocks, "Stalwarts", "3", _stalwart_filter, _stalwart_score,
                                "4A148C", "Consistent 5-25% revenue growers, >$2B. Buy on dips.",
                                custom_headers=_st_headers, custom_widths=_st_widths)

    # ─── Fast Growers: Peter Lynch 10-bagger style ─────────────────────────
    # Lynch: "Find a company growing 20-25%/yr with a low PEG — that's a 10-bagger."
    # Key: consistent multi-year growth + reasonable PEG + profitability improving
    # NOT primarily Piotroski — Lynch focused on growth quality, not balance sheet scores
    def _fast_grower_filter(s):
        if not _is_common_stock(s): return False
        # Real Estate excluded: ROIC is inflated by land-at-cost book values; revenue growth
        # is asset-cycle driven, not business quality compounding
        if (s.get("sector") or "") == "Real Estate": return False
        # Basic Materials excluded: commodity prices (gold, copper, fertilizer) drive revenue
        # growth, not underlying business quality — miners are not Lynch fast growers
        if (s.get("sector") or "") == "Basic Materials": return False
        rg  = s.get("revGrowth")       # trailing 1-yr revenue growth
        rg5 = s.get("revGrowth5y")     # forward 5-yr revenue CAGR (analyst estimates)
        eg5 = s.get("epsGrowth5y")     # forward 5-yr EPS CAGR
        rg5h = s.get("fiveYRevGrowth") # historical 5-yr rev/share growth
        pe  = s.get("pe")
        mktcap = s.get("mktCap", 0)

        # Size: Lynch preferred small-mid caps — lots of runway to grow
        if mktcap < 100e6 or mktcap > 150e9:
            return False

        # Primary qualifier — at least ONE strong growth signal must exist:
        # A) Current year revenue growing >20% (classic fast grower)
        # B) Multi-year revenue growth >15% (steadier compounder trajectory)
        # C) Forward EPS CAGR >15% + any positive current growth (earnings-led growth)
        strong_current = (rg and rg > 0.20)
        strong_5yr     = (rg5 and rg5 > 0.15 and (rg is None or rg > 0.10))
        strong_5yr_h   = (rg5h and rg5h > 0.15 and (rg is None or rg > 0.10))
        eps_led_growth = (eg5 and eg5 > 0.15 and (rg is None or rg > 0.08))
        if not (strong_current or strong_5yr or strong_5yr_h or eps_led_growth):
            return False

        # A1: Kill isolated single-year spike — >50% growth needs prior-year support
        # Prevents one-time events (pharma milestones, asset sales) from inflating rankings
        if rg is not None and rg > 0.50:
            _rg_prev_a1 = s.get("revGrowthPrev"); _rg_yr2_a1 = s.get("revGrowthYr2")
            if not ((_rg_prev_a1 and _rg_prev_a1 > 0.15) or (_rg_yr2_a1 and _rg_yr2_a1 > 0.15)):
                return False  # isolated spike — not a structural fast grower

        # ── KILL CRITERIA: hard rejections — mediocre businesses filtered out ──
        # Kill 1: FCF must be positive — with one exception for genuine high-growth small caps.
        # Lynch would not have excluded early Amazon/Salesforce for reinvesting aggressively.
        # Exception: mktcap < $2B AND rev growth > 25% AND FCF yield > -5% (not deeply burning).
        fcf = s.get("fcfYield")
        if fcf is not None and fcf <= 0:
            _small = mktcap < FG_SMALL_CAP_MAX
            _fast  = (rg is not None and rg > FG_RELAXED_FCF_REV_GROWTH)
            _mild  = fcf > -0.05  # FCF yield better than -5% — not deeply destroying cash
            if not (_small and _fast and _mild):
                return False
            # Exception granted — scoring still penalises: 0 FCF yield pts,
            # -3 for FCF margin < 2%, likely -5 for low FCF conversion → net -8 to -11 pts.

        # Kill 2: Revenue consistency — 3 of 5 years must show positive growth
        rc = s.get("revConsistency")
        if rc is not None and rc < 0.40:
            return False  # majority of years declining = structural issue, not a fast grower

        # Kill 3: Profitability floor — must have SOME evidence of earnings quality
        has_pe       = (pe and 0 < pe < 300)
        has_eps_gr   = (eg5 and eg5 > 0)
        if not (has_pe or has_eps_gr):
            return False

        # Kill 4: Exclude distress — Altman Z < 1 is bankruptcy danger zone
        az = s.get("altmanZ")
        if az is not None and az < 1.0:
            return False

        # A2: Leverage kill — too much debt makes growth fragile and funding uncertain
        _nde_a2 = s.get("netDebtEbitda")
        if _nde_a2 is not None and _nde_a2 > 5.0:
            return False  # net debt > 5× EBITDA — too leveraged to sustain growth

        # Kill 5: Financial Services with low ROIC — PEG cheap but capital-inefficient
        # Insurance/bank PEGs look cheap at cycle peaks; require ROIC ≥ 10% as minimum quality floor
        _sector_fg = (s.get("sector") or "")
        if "financial" in _sector_fg.lower() or "insurance" in _sector_fg.lower():
            roic_fg = s.get("roic")
            if roic_fg is not None and roic_fg < 0.10:
                return False

        # Kill 6: Extreme share dilution — >15%/yr destroys shareholder value faster than growth creates it
        sg = s.get("sharesGrowth")
        if sg is not None and sg > 0.15:
            return False

        return True

    def _fast_grower_score(s):
        rg      = s.get("revGrowth") or 0
        rg_prev = s.get("revGrowthPrev")          # prior year — for acceleration detection
        rg5     = s.get("revGrowth5y") or 0
        rg5h    = s.get("fiveYRevGrowth") or 0
        eg5     = s.get("epsGrowth5y") or 0
        eg      = s.get("epsGrowth") or 0
        eg_prev = s.get("epsGrowthPrev")

        # Core 1: Current revenue growth (primary signal, capped to avoid SPAC distortions)
        sc = min(rg * 28, 18)  # 20% growth = 5.6 pts, 60% = 18 pts (cap)

        # Core 2: Multi-year revenue growth consistency — Lynch wanted steady compounders
        best_5yr = max(rg5, rg5h)   # take the better of forward/historical 5yr
        if best_5yr > 0.25:  sc += 14   # exceptional compounder (25%+/yr for 5 years)
        elif best_5yr > 0.20: sc += 10
        elif best_5yr > 0.15: sc += 7
        elif best_5yr > 0.10: sc += 4
        elif best_5yr > 0.05: sc += 1

        # Core 3: EPS growth trajectory — Lynch: earnings should grow with revenue
        if eg5 > 0.25:  sc += 12
        elif eg5 > 0.20: sc += 9
        elif eg5 > 0.15: sc += 6
        elif eg5 > 0.10: sc += 3
        elif eg5 > 0.05: sc += 1

        # Consistency bonus: both revenue AND earnings growing fast = high conviction
        if best_5yr > 0.15 and eg5 > 0.15:
            sc += 6

        # ── Growth acceleration / deceleration vs prior year ────────────────
        # Lynch's key insight: you want to catch fast growers BEFORE they decelerate
        # Accelerating = current growth BETTER than last year (momentum building)
        # Decelerating = current growth WORSE than last year (late cycle, avoid)
        if rg_prev is not None and rg > 0.10:
            rg_delta = rg - rg_prev   # positive = accelerating
            if rg_delta > 0.10:       sc += 10  # meaningful acceleration (+10pp+)
            elif rg_delta > 0.05:     sc += 6   # moderate acceleration
            elif rg_delta > 0:        sc += 3   # slight acceleration
            elif rg_delta < -0.10:    sc -= 8   # meaningful deceleration (red flag)
            elif rg_delta < -0.05:    sc -= 4   # moderate deceleration (caution)

        # EPS acceleration (even stronger signal — earnings leverage on revenue)
        if eg_prev is not None and eg > 0:
            eg_delta = eg - eg_prev
            if eg_delta > 0.10:  sc += 6   # EPS accelerating strongly
            elif eg_delta > 0:   sc += 3   # EPS acceleration
            elif eg_delta < -0.10: sc -= 4  # EPS decelerating

        # Positive growth 3 consecutive years (all 3 years we have show growth)
        rg_yr2 = s.get("revGrowthYr2")
        if rg > 0.10 and rg_prev and rg_prev > 0.10 and rg_yr2 and rg_yr2 > 0.05:
            sc += 8  # 3-year consistent growth = high conviction fast grower

        # Revenue consistency 5-year: penalise structurally declining businesses
        rc_fg = s.get("revConsistency")
        if rc_fg is not None:
            if rc_fg >= 0.80:   sc += 7
            elif rc_fg >= 0.60: sc += 3
            elif rc_fg < 0.40:  sc -= 5  # majority of years negative = not a fast grower

        # Share dilution: buybacks = capital discipline, heavy dilution = value destruction
        sg_fg = s.get("sharesGrowth")
        if sg_fg is not None:
            if sg_fg < -0.03:   sc += 5
            elif sg_fg < 0:     sc += 2
            elif sg_fg > 0.05:  sc -= 4

        # PEG — Lynch's defining metric: growth at a reasonable price
        # Prefer forward PEG (consistent fwdPE/fwdGrowth) to avoid trailing-vs-forward mismatch
        best_peg_fg = s.get("fwdPEG") or s.get("peg")
        _roic_fg = s.get("roic"); _fcf_fg = s.get("fcfYield")
        _fin_fg       = "financial" in (s.get("sector") or "").lower()
        _basic_mat_fg = "basic materials" in (s.get("sector") or "").lower()
        _energy_fg    = (s.get("sector") or "") == "Energy"
        # PEG: secondary valuation check — confirms whether the ROIC is priced fairly
        # Max 15pts (was 22) — ROIC now leads, PEG validates the entry price
        peg_sc_fg = 0
        if best_peg_fg and 0 < best_peg_fg < 0.75:   peg_sc_fg = 15
        elif best_peg_fg and 0 < best_peg_fg < 1.0:  peg_sc_fg = 11
        elif best_peg_fg and 0 < best_peg_fg < 1.5:  peg_sc_fg = 6
        elif best_peg_fg and 0 < best_peg_fg < 2.0:  peg_sc_fg = 3
        # Quality bonus: ROIC+FCF confirms the low PEG is moat-backed, not an earnings spike
        if peg_sc_fg and _roic_fg and _roic_fg > 0.15 and _fcf_fg and _fcf_fg > 0.04:
            peg_sc_fg = min(22, round(peg_sc_fg * 1.25))
        # Deceleration penalty: fast-falling growth makes a low PEG misleading
        if peg_sc_fg and rg_prev is not None and (rg - rg_prev) < -0.10:
            peg_sc_fg = round(peg_sc_fg * 0.60)
        # Cyclical penalty: insurance/financial PEGs are structurally unreliable
        if peg_sc_fg and _fin_fg and best_peg_fg and best_peg_fg < 1.5:
            peg_sc_fg = round(peg_sc_fg * 0.50)
        # Cyclical penalty: commodity miners — PEG based on trough-earnings is misleading
        if peg_sc_fg and _basic_mat_fg and best_peg_fg and best_peg_fg < 1.5:
            peg_sc_fg = round(peg_sc_fg * 0.60)
        # Cyclical penalty: E&P energy — revenue growth driven by commodity prices, not business quality
        if peg_sc_fg and _energy_fg and best_peg_fg and best_peg_fg < 1.5:
            peg_sc_fg = round(peg_sc_fg * 0.60)
        # Filter 4: Growth optimism penalty — analysts significantly more bullish than history
        _go_fg = s.get("growthOptimism")
        _ego_fg = s.get("epsGrowthOptimism")
        if peg_sc_fg and _go_fg is not None and _go_fg > 0.50:
            peg_sc_fg = round(peg_sc_fg * 0.70)  # revenue estimates too rosy
        # Double penalty if BOTH revenue AND EPS estimates far exceed history
        if peg_sc_fg and _go_fg is not None and _go_fg > 0.50 and _ego_fg is not None and _ego_fg > 0.50:
            peg_sc_fg = round(peg_sc_fg * 0.80)  # combined red flag — analyst euphoria
        sc += peg_sc_fg

        # FCF data missing entirely = uncertainty penalty (different from FCF ≤ 0 which is a kill)
        # A legitimate fast grower should have measurable FCF; None means FMP has no data.
        # Micro-caps frequently have incomplete FMP coverage despite real FCF — reduce penalty.
        fcf_fg = s.get("fcfYield")
        if fcf_fg is None:
            if s.get("mktCap", 0) < FG_MICRO_CAP_MAX:
                sc -= 2   # reduced: data gap reflects thin coverage, not poor quality
            else:
                sc -= 5   # full penalty for mid/large caps — no excuse for missing FCF data

        # Negative forward EPS growth = earnings expected to shrink — contradicts "fast grower" thesis
        eg5_fg = s.get("epsGrowth5y") or 0
        if eg5_fg < 0:
            sc -= 8  # declining projected earnings is a fundamental contradiction for this category

        # Filter 3: FCF conversion — cash earnings vs reported earnings quality check
        fcc_fg = s.get("fcfConversion")
        if fcc_fg is not None:
            if fcc_fg >= 0.80:   sc += 6   # FCF quality confirmed
            elif fcc_fg >= 0.60: sc += 3
            elif fcc_fg < 0.40:  sc -= 5   # earnings not converting to cash

        # FCF Margin: how much of each revenue dollar becomes FCF (capital efficiency of growth)
        fcm_fg = s.get("fcfMargin")
        if fcm_fg is not None:
            if fcm_fg >= 0.20:   sc += 8   # >20% FCF margin = elite cash generator
            elif fcm_fg >= 0.12: sc += 5
            elif fcm_fg >= 0.06: sc += 2
            elif fcm_fg < 0.02:  sc -= 3   # thin FCF margin means growth eats all cash

        # FCF growth consistency: is the FCF trend stable across years?
        fgc_fg = s.get("fcfGrowthConsistency")
        if fgc_fg is not None:
            if fgc_fg >= 0.80:   sc += 5
            elif fgc_fg >= 0.60: sc += 2
            elif fgc_fg < 0.40:  sc -= 3

        # ── ROIC: PRIMARY quality signal — evaluated BEFORE PEG ─────────────
        # High ROIC = durable moat. The biggest winners sustain ROIC > 15% for years.
        # PEG tells you the price; ROIC tells you if the business deserves a high price.
        roic = s.get("roic")
        if roic and roic > 0.30:   sc += 20  # elite compounder — ROIC sustains itself
        elif roic and roic > 0.25: sc += 16
        elif roic and roic > 0.20: sc += 12
        elif roic and roic > 0.15: sc += 8   # moat confirmed — growth is worth paying for
        elif roic and roic > 0.10: sc += 3
        elif roic and roic is not None and roic < 0.08:
            sc -= 5  # low ROIC growth = burning capital to grow, not compounding

        # ROE — secondary; can be lever-inflated, so lower weight than ROIC
        roe = s.get("roe")
        if roe and roe > 0.25:   sc += 4
        elif roe and roe > 0.15: sc += 2

        # Gross margin — pricing power / scalability
        gm = s.get("grossMargin")
        if gm and gm > 0.60:   sc += 6   # software/IP-like margins
        elif gm and gm > 0.40: sc += 4
        elif gm and gm > 0.25: sc += 2

        # FCF quality (positive = bonus, negative = neutral — many fast growers reinvest heavily)
        fcf = s.get("fcfYield")
        if fcf and fcf > 0.08:  sc += 7
        elif fcf and fcf > 0.04: sc += 4
        elif fcf and fcf > 0.01: sc += 2

        # P/FCF — cash-flow based value check (lower = more reasonable vs growth)
        pfcf = s.get("pFcf")
        if pfcf and 0 < pfcf < 20:    sc += 4   # cheap on FCF basis
        elif pfcf and 0 < pfcf < 35:  sc += 2

        # Earnings beat rate — management consistently delivers vs expectations
        br = s.get("beatRate")
        if br and br >= 0.875:  sc += 8   # beat 7+ of 8 quarters = very reliable compounder
        elif br and br >= 0.75: sc += 5
        elif br and br >= 0.625: sc += 2

        # A4: EPS beat streak — consecutive quarters beating = execution momentum building
        ebs = s.get("epsBeatStreak")
        if ebs is not None:
            if ebs >= 4:   sc += 8   # 4+ consecutive beats — strong execution momentum
            elif ebs >= 2: sc += 4

        # Valuation sanity
        pe = s.get("pe")
        if pe and 0 < pe < 25:   sc += 3
        elif pe and 0 < pe < 40: sc += 1

        # Insider buying = management confidence in their own growth story
        if s.get("insiderBuys", 0) >= 3:   sc += 6
        elif s.get("insiderBuys", 0) >= 1: sc += 3

        # ── Market cap size bonus: surfaces underfollowed small caps ─────────
        # Smaller companies have more analytical blind spots = more mispricing = more alpha.
        # Additive to quality score — a junk small-cap still scores low overall.
        _mc_fg = s.get("mktCap", 0)
        if _mc_fg < FG_MICRO_CAP_MAX:
            sc += FG_MICRO_CAP_BONUS    # $100M–$500M: most neglected
        elif _mc_fg < FG_SMALL_CAP_MAX:
            sc += FG_SMALL_CAP_BONUS    # $500M–$2B: still under-covered
        elif _mc_fg < FG_SMALL_MID_MAX:
            sc += FG_SMALL_MID_BONUS    # $2B–$5B: marginal nudge

        return sc

    _fg_headers = [
        "Rank", "Ticker", "Company", "Sector", "Price",
        "Fwd PEG", "PEG", "Fwd P/E", "P/E", "ROIC",
        "FCF Yield", "FCF Margin", "FCF Conv.", "FCF Consist.",
        "Rev Growth", "Rev Gr Prev", "Rev Growth 5Y", "EPS Growth 5Y",
        "Rev Consist.", "Grwth Gap", "Shares Δ", "Beat Rate", "52w vs High", "ROE", "Gross Margin",
        "P/FCF", "MktCap ($B)", "Cap Size", "Score", "🏦 Insider",
    ]
    _fg_widths = [5, 8, 22, 15, 8, 7, 6, 7, 7, 8, 8, 9, 8, 9, 9, 10, 11, 11, 9, 8, 8, 8, 9, 7, 11, 7, 10, 7, 6, 14]

    fast_growers = build_lynch_tab(wb, stocks, "Fast Growers", "4", _fast_grower_filter, _fast_grower_score,
                                   "1B5E20", "Revenue >20% or multi-year 15%+ CAGR. Lynch 10-baggers. PEG is king.",
                                   custom_headers=_fg_headers, custom_widths=_fg_widths)

    # ─── Slow Growers: income focus — dividend sustainability is key ───
    def _slow_grower_filter(s):
        if not _is_common_stock(s): return False
        dy = s.get("divYield")
        rg = s.get("revGrowth")
        pio = s.get("piotroski")
        fcf = s.get("fcfYield")
        # Require: >2% yield, low growth, quality floor, FCF must cover dividend
        return (dy and 0.02 < dy < 0.15          # cap at 15% — above that is usually distressed
                and (not rg or rg < 0.10)
                and (pio is None or pio >= 5)     # quality gate
                and (fcf is None or fcf > dy * 0.5))  # FCF covers at least half the dividend

    def _slow_grower_score(s):
        dy = s.get("divYield") or 0
        sc = dy * 80                              # yield weighted, but capped by filter
        pio = s.get("piotroski")
        if pio and pio >= 8: sc += 10
        elif pio and pio >= 7: sc += 6
        elif pio and pio >= 5: sc += 2
        fcf = s.get("fcfYield")
        if fcf and fcf > 0.07: sc += 5
        elif fcf and fcf > 0.04: sc += 3
        de = s.get("de")
        if de is not None and de < 0.5: sc += 4
        elif de is not None and de < 1.0: sc += 2
        if s.get("pe") and 0 < s.get("pe") < 18: sc += 3
        roe = s.get("roe")
        if roe and roe > 0.15: sc += 3
        if s.get("mos") and s.get("mos") > 0.1: sc += 3
        return sc

    slow_growers = build_lynch_tab(wb, stocks, "Slow Growers", "5", _slow_grower_filter, _slow_grower_score,
                                   "455A64", "Dividend >2%, low growth. Buy for income, not appreciation.")

    # ─── Cyclicals: Lynch BUY list — identify companies at the TROUGH ────────
    # Lynch's key insight: cyclicals are the OPPOSITE of normal stocks.
    # BUY when P/E is HIGH (earnings depressed at trough) → sell when P/E is LOW (peak earnings).
    # Also BUY when P/E is not available (reporting losses = maximum trough opportunity).
    # The goal is to catch cyclicals before earnings recover, not after.
    def _cyclical_filter(s):
        """Lynch trough-buying filter for cyclicals.

        KEY INSIGHT: For cyclicals, the P/E signal is INVERTED vs normal stocks.
          - Low P/E  → earnings at PEAK  → Lynch would SELL
          - High P/E → earnings DEPRESSED → Lynch would BUY
          - No P/E   → maximum trough   → best entry (if balance sheet survives)

        We require at least one confirmed trough signal AND balance-sheet survival.
        """
        if not _is_common_stock(s): return False
        pio = s.get("piotroski")
        pb  = s.get("pb")
        pe  = s.get("pe")
        az  = s.get("altmanZ")
        nde = s.get("netDebtEbitda")  # net debt / EBITDA — survival metric
        cr  = s.get("currentRatio")   # liquidity

        # Must be in a recognised cyclical sector
        if s.get("sector") not in CYCLICAL_SECTORS:
            return False

        # Size: large enough to survive the trough
        if s.get("mktCap", 0) < 500e6:
            return False

        # Quality floor — must still be financially alive
        if pio is not None and pio < 4:
            return False

        # Exclude companies in imminent-bankruptcy territory
        if az is not None and az < 0.8:
            return False

        # Hard debt survival gate: net debt > 6× EBITDA at trough = bankruptcy risk
        # (Cyclicals with heavy leverage go bust before the cycle turns)
        if nde is not None and nde > 6.0:
            return False

        # Liquidity gate: current ratio < 0.8 = can't pay near-term obligations
        if cr is not None and cr < 0.8:
            return False

        # ── Trough signal: need at least one of three ───────────────────
        # A) Elevated P/E (≥15) — earnings depressed, NOT at cycle peak
        #    Note: P/E 10–14 is excluded — that's more likely MID/PEAK for cyclicals
        trough_pe    = (pe is not None and pe >= 15)

        # B) No P/E / reporting losses + asset backing (P/B < 1.5)
        #    Maximum trough — company unprofitable but assets > liabilities
        loss_trough  = (pe is None and pb is not None and 0 < pb < 1.5)

        # C) P/B < 0.9 — extreme asset discount regardless of earnings
        #    Sector so beaten down that stock trades below liquidation value
        asset_trough = (pb is not None and 0 < pb < 0.9)

        return trough_pe or loss_trough or asset_trough

    def _cyclical_score(s):
        """Score cyclicals for trough depth, survival quality, and early recovery signals.

        Scoring philosophy (Lynch):
          1. Deepest trough = best entry (high/no P/E preferred)
          2. Balance sheet must survive 2+ more years of downcycle
          3. Deceleration of decline signals approaching bottom (buy BEFORE recovery)
          4. Price near multi-year lows means more of the pain is priced in
          5. Insider buying = management knows something is improving
        """
        sc  = 0
        pe  = s.get("pe")
        pb  = s.get("pb")
        fcf = s.get("fcfYield")
        de  = s.get("de")
        nde = s.get("netDebtEbitda")
        rg  = s.get("revGrowth")
        rg_prev = s.get("revGrowthPrev")   # prior-year revenue growth (for deceleration signal)
        cr  = s.get("currentRatio")

        # ── 1. P/E TROUGH SIGNAL (INVERTED vs normal stocks) ───────────
        # Lynch: high P/E at trough = BUY;  low P/E at peak = SELL
        if pe is None:
            sc += 22   # no earnings / loss = maximum trough = best entry
        elif pe > 50:
            sc += 15   # extreme trough or heavy losses — still a buy if assets OK
        elif 25 < pe <= 50:
            sc += 22   # classic cyclical trough band — primary buy zone
        elif 15 < pe <= 25:
            sc += 14   # elevated — early/mid trough, still attractive
        elif 10 < pe <= 15:
            sc -= 4    # mid-to-peak territory — Lynch would be cautious
        elif pe <= 10:
            sc -= 10   # low P/E = peak earnings — Lynch would SELL, not buy

        # ── 2. ASSET FLOOR (P/B) ────────────────────────────────────────
        # Book value is the downside floor when earnings are depressed
        if pb is not None and 0 < pb < 0.5:    sc += 18  # extreme asset discount
        elif pb is not None and 0 < pb < 0.8:  sc += 13
        elif pb is not None and 0 < pb < 1.2:  sc += 8
        elif pb is not None and 0 < pb < 2.0:  sc += 3

        # ── 3. FCF: cash generation at trough = quality signal ──────────
        # A cyclical generating positive FCF while earnings are depressed
        # is NOT in distress — it will survive and recover strongly
        if fcf is not None and fcf > 0.10:    sc += 18
        elif fcf is not None and fcf > 0.07:  sc += 13
        elif fcf is not None and fcf > 0.04:  sc += 8
        elif fcf is not None and fcf > 0.01:  sc += 3
        elif fcf is not None and fcf < -0.05: sc -= 8   # burning cash = survival risk

        # ── 4. REVENUE TREND + DECELERATION (approaching-bottom signal) ─
        # Lynch: buy BEFORE revenue recovers, not after.
        # Best signal: decline decelerating (e.g. -20% → -8%) = bottom approaching
        decline_decelerating = (
            rg is not None and rg_prev is not None
            and rg < 0 and rg_prev < 0          # both declining years
            and rg > rg_prev                     # decline is getting smaller
        )
        if decline_decelerating:
            sc += 12  # strongest signal — approaching the trough inflection point

        # Current revenue trend score
        if rg is None or rg == 0:
            sc += 5    # flat / no data — possibly at the floor
        elif -0.20 <= rg < 0:
            sc += 8    # mild decline — near trough
        elif -0.40 <= rg < -0.20:
            sc += 5    # severe decline — maybe near bottom
        elif rg < -0.40:
            sc += 2    # catastrophic decline — deep trough but survival risk rises
        elif 0 < rg <= 0.05:
            sc -= 2    # slight recovery — might be late
        elif 0.05 < rg <= 0.15:
            sc -= 6    # recovery underway — Lynch would be reducing, not buying
        elif rg > 0.15:
            sc -= 12   # strong recovery = peak approaching — time to SELL not BUY

        # ── 5. DEBT SURVIVAL: outlast the downcycle ─────────────────────
        # Net debt/EBITDA is the primary survival metric at trough
        if nde is not None:
            if nde < 0:           sc += 10  # net cash = can't go bankrupt
            elif nde < 1.0:       sc += 8
            elif nde < 2.0:       sc += 5
            elif nde < 3.5:       sc += 2
            elif nde < 5.0:       sc -= 3   # stretched — one more bad year is dangerous
            # > 6x already blocked by filter

        # Supplementary: D/E ratio
        if de is not None and de < 0.3:   sc += 5
        elif de is not None and de < 0.7: sc += 3
        elif de is not None and de < 1.5: sc += 1

        # ── 6. LIQUIDITY: can pay bills during the trough ───────────────
        if cr is not None and cr > 2.0:   sc += 6
        elif cr is not None and cr > 1.5: sc += 4
        elif cr is not None and cr > 1.2: sc += 2

        # ── 7. PIOTROSKI: fundamental health score ──────────────────────
        pio = s.get("piotroski")
        if pio is not None and pio >= 8:   sc += 8
        elif pio is not None and pio >= 6: sc += 5
        elif pio is not None and pio >= 4: sc += 2

        # ── 8. DIVIDEND MAINTAINED: financial strength at trough ────────
        div = s.get("divYield")
        if div is not None and div > 0.06:   sc += 7
        elif div is not None and div > 0.03: sc += 4
        elif div is not None and div > 0.01: sc += 2

        # ── 9. EV/REVENUE: useful when earnings near zero ───────────────
        # EV/Revenue < 0.5 means you're paying < 50c per $1 of revenue — deep value
        ev_rev = s.get("evRevenue")
        if ev_rev is not None and 0 < ev_rev < 0.3:   sc += 10  # extreme cheapness on sales
        elif ev_rev is not None and 0 < ev_rev < 0.6: sc += 7
        elif ev_rev is not None and 0 < ev_rev < 1.0: sc += 4

        # ── 10. EV/EBITDA: elevated = earnings depressed (same inversion) ─
        ev_eb = s.get("evEbitda")
        if ev_eb is not None and ev_eb > 25:    sc += 5
        elif ev_eb is not None and ev_eb > 15:  sc += 3

        # ── 11. 52-WEEK POSITION: near lows = pain is priced in ─────────
        pvs52h = s.get("priceVs52H")
        if pvs52h is not None and pvs52h < 0.45:   sc += 12  # >55% off 52wk high
        elif pvs52h is not None and pvs52h < 0.60: sc += 8
        elif pvs52h is not None and pvs52h < 0.75: sc += 4

        # ── 12. INSIDER BUYING: management knows cycle is turning ────────
        if s.get("insiderBuys", 0) >= 3:   sc += 12
        elif s.get("insiderBuys", 0) >= 1: sc += 6

        return sc

    _cy_headers = [
        "Rank", "Ticker", "Company", "Sector", "Price",
        "P/E", "EV/EBITDA", "EV/Rev", "P/B", "FCF Yield",
        "Net Debt/EBITDA", "Curr Ratio", "Rev Growth", "Rev Gr Prev",
        "Div Yield", "52w Pos", "Piotroski", "Score", "🏦 Insider",
    ]
    _cy_widths = [5, 8, 22, 15, 8, 7, 9, 7, 6, 8, 11, 9, 9, 10, 8, 7, 8, 6, 14]

    cyclicals = build_lynch_tab(wb, stocks, "Cyclicals", "6", _cyclical_filter, _cyclical_score,
                                "E65100",
                                "Lynch trough-buying: BUY at HIGH/NO P/E (depressed earnings). "
                                "Low P/E = peak earnings = SELL signal for cyclicals. "
                                "Best entry: decline decelerating + P/B < 1.2 + FCF positive + net debt < 3× EBITDA.",
                                custom_headers=_cy_headers, custom_widths=_cy_widths)

    # ─── Turnarounds: recovering businesses — distinct from IV Discount ────────
    # Lynch: "Turnarounds are companies that have had serious problems and are now recovering."
    # Different from IV Discount (always-good business temporarily cheap):
    # Turnarounds had real problems (low Piotroski, earnings trough) and are now showing ACTIVE recovery.
    # Key: must show concrete recovery signals — not just cheap, but measurably improving.
    def _turnaround_filter(s):
        if not _is_common_stock(s): return False
        mos   = s.get("mos")
        pb    = s.get("pb")
        pio   = s.get("piotroski")
        rg    = s.get("revGrowth")
        eg    = s.get("epsGrowth")
        nig   = s.get("netIncomeGrowth")
        fcf   = s.get("fcfYield")
        az    = s.get("altmanZ")

        if s.get("mktCap", 0) < 300e6:  # must be substantial enough to recover
            return False

        # Turnarounds are NOT already-great businesses (those belong in IV Discount)
        # If Piotroski is already >= 8, it's a quality business — not a true turnaround
        if pio is not None and pio >= 8:
            return False

        # Fast growers masquerading as turnarounds: high ROIC + strong revenue growth = not distressed
        # These belong in Fast Growers, not here. A genuine turnaround has weak/recovering ROIC.
        roic_ta = s.get("roic"); rg_ta = s.get("revGrowth")
        if (roic_ta is not None and roic_ta > 0.20
                and rg_ta is not None and rg_ta > 0.25):
            return False

        # Not in active bankruptcy — some financial life remaining
        if az is not None and az < 0.5:
            return False

        # Valuation: beaten-down price (cheap relative to DCF or book)
        deep_discount = (mos and mos > 0.20)    # DCF says 20%+ undervalued
        asset_cheap   = (pb and 0 < pb < 0.8)   # trading well below book

        if not (deep_discount or asset_cheap):
            return False

        # B1: Liquidity gate — tight liquidity needs stronger evidence before qualifying
        _cr_b1 = s.get("currentRatio"); _ncr_b1 = s.get("netCashRatio")
        _required_signals = 3 if (_cr_b1 is not None and _cr_b1 < 1.0
                                   and (_ncr_b1 is None or _ncr_b1 < 0)) else 2

        # RECOVERY SIGNAL — must show at least two concrete improvement signals:
        # Without real recovery signals, it's just a cheap deteriorating business (value trap)
        recovery_signals = 0
        if rg  and rg  > 0.03:   recovery_signals += 1  # revenue growing again
        if eg  and eg  > 0:      recovery_signals += 1  # EPS improving
        if nig and nig > 0:      recovery_signals += 1  # net income improving
        if fcf and fcf > 0:      recovery_signals += 1  # cash flow positive
        if s.get("revGrowth5y") and s.get("revGrowth5y") > 0.05:
            recovery_signals += 1  # analysts expect sustained growth ahead
        if pio and pio >= 5:     recovery_signals += 1  # improving financial health
        # Beat rate >= 0.625 (5 of 8 quarters beating) = execution is consistently improving
        if s.get("beatRate") and s.get("beatRate") >= 0.625: recovery_signals += 1
        # B5: EPS beat streak — 3+ consecutive beats = management executing on recovery
        if s.get("epsBeatStreak") and s.get("epsBeatStreak") >= 3: recovery_signals += 1

        return recovery_signals >= _required_signals

    def _turnaround_score(s):
        sc = 0

        # ── Valuation cheapness (how big is the opportunity) ────────────
        mos = s.get("mos")
        if mos and mos > 0.50:    sc += 20
        elif mos and mos > 0.35:  sc += 15
        elif mos and mos > 0.20:  sc += 10

        pb = s.get("pb")
        if pb and 0 < pb < 0.4:   sc += 12
        elif pb and 0 < pb < 0.7: sc += 8
        elif pb and 0 < pb < 1.0: sc += 4

        # ── Revenue recovery momentum ────────────────────────────────────
        rg = s.get("revGrowth")
        if rg and rg > 0.20:   sc += 14   # strong recovery = high conviction
        elif rg and rg > 0.10: sc += 10
        elif rg and rg > 0.03: sc += 6

        # B2: Multi-year consecutive revenue recovery — distinguishes real turnarounds from blips
        _rg_prev_b2 = s.get("revGrowthPrev"); _rg_yr2_b2 = s.get("revGrowthYr2")
        if rg and rg > 0.03 and _rg_prev_b2 and _rg_prev_b2 > 0 and _rg_yr2_b2 and _rg_yr2_b2 > 0:
            sc += 10  # 3 consecutive recovery years — real turnaround, not a blip
        elif rg and rg > 0.03 and _rg_prev_b2 and _rg_prev_b2 > 0:
            sc += 5   # 2 consecutive recovery years — early confirmation

        # ── EPS / earnings recovery ──────────────────────────────────────
        eg = s.get("epsGrowth")
        if eg and eg > 0.30:   sc += 12
        elif eg and eg > 0.10: sc += 8
        elif eg and eg > 0:    sc += 4

        nig = s.get("netIncomeGrowth")
        if nig and nig > 0.30:   sc += 8
        elif nig and nig > 0.10: sc += 5
        elif nig and nig > 0:    sc += 2

        # ── Cash flow turning around ─────────────────────────────────────
        fcf = s.get("fcfYield")
        if fcf and fcf > 0.08:   sc += 12
        elif fcf and fcf > 0.04: sc += 8
        elif fcf and fcf > 0:    sc += 4

        # ── Financial health improving (Piotroski 5-7 = recovering zone) ─
        pio = s.get("piotroski")
        if pio and pio >= 7:     sc += 10  # well into recovery
        elif pio and pio >= 5:   sc += 6
        elif pio and pio >= 3:   sc += 2

        # ── Altman Z: grey zone = turnaround in progress ─────────────────
        az = s.get("altmanZ")
        if az and az > 3.0:       sc += 5   # back in safe zone = recovery confirmed
        elif az and az > 1.8:     sc += 8   # grey zone, improving = best turnaround signal
        elif az and az > 0.8:     sc += 3   # still distressed but surviving

        # ── Analyst growth outlook confirms the turnaround ───────────────
        rg5 = s.get("revGrowth5y")
        if rg5 and rg5 > 0.15:   sc += 7
        elif rg5 and rg5 > 0.08: sc += 4
        elif rg5 and rg5 > 0.03: sc += 2

        # ── Earnings beat rate — consistently beating = execution improving ─
        br = s.get("beatRate")
        if br and br >= 0.875:  sc += 10  # beat 7/8 quarters = high-conviction recovery
        elif br and br >= 0.75: sc += 7
        elif br and br >= 0.625: sc += 4

        # ── 52-week positioning: deep drawdown = more of the pain is priced in ─
        pvs52h = s.get("priceVs52H")
        if pvs52h and pvs52h < 0.45:    sc += 8   # >55% off high = deep distress priced in
        elif pvs52h and pvs52h < 0.65:  sc += 5

        # ── Insider buying = management backs the recovery ───────────────
        if s.get("insiderBuys", 0) >= 3:   sc += 12  # big signal for turnarounds
        elif s.get("insiderBuys", 0) >= 1: sc += 6

        return sc

    _ta_headers = [
        "Rank", "Ticker", "Company", "Sector", "Price",
        "MoS", "P/B", "Rev Growth", "EPS Growth", "FCF Yield",
        "Beat Rate", "52w Pos", "52w vs Low", "Net D/E", "Piotroski", "Altman Z", "Rev Growth 5Y",
        "MktCap ($B)", "Score", "🏦 Insider",
    ]
    _ta_widths = [5, 8, 22, 15, 8, 7, 6, 9, 9, 9, 8, 7, 9, 7, 8, 8, 11, 10, 6, 14]

    turnarounds = build_lynch_tab(wb, stocks, "Turnarounds", "7", _turnaround_filter, _turnaround_score,
                                  "B71C1C",
                                  "Recovery plays: beaten-down price + ≥2 active recovery signals. NOT already-great businesses.",
                                  custom_headers=_ta_headers, custom_widths=_ta_widths)

    # ─── Asset Plays: Hidden value vs asset prices — Lynch style ───────────
    # Lynch: "Find companies where specific assets are worth more than the whole."
    # Five qualifying paths:
    #   1. Classic: P/B < 1 (market undervalues the book)
    #   2. Tangible book > price (ignoring goodwill/intangibles reveals real asset discount)
    #   3. Cash fortress: very high liquidity + minimal debt + moderate valuation
    #   4. Net cash positive: cash + investments > total debt (company has net cash per share)
    #   5. Graham net-net: trading below net working capital — Graham's deepest value
    _ASSET_HEAVY_SECTORS = {
        "Financial Services", "Real Estate", "Energy", "Basic Materials",
        "Utilities", "Industrials",
    }

    def _asset_play_filter(s):
        if not _is_common_stock(s): return False
        pb     = s.get("pb")
        tb     = s.get("tangibleBook")
        price  = s.get("price", 0)
        pio    = s.get("piotroski")
        mktcap = s.get("mktCap", 0)
        cr     = s.get("currentRatio")
        de     = s.get("de")
        ncr    = s.get("netCashRatio")   # net cash / price
        gnn    = s.get("grahamNetNet")   # Graham net-net per share

        if mktcap < 150e6:  # no micro-cap junk
            return False
        # Quality floor — not a financial basket case
        if pio is not None and pio < 4:
            return False

        # Path 1: Classic P/B < 1 (book value > market cap)
        below_book = (pb and 0 < pb < 1.0)

        # Path 2: Trading below TANGIBLE book — strips goodwill/intangibles
        # (a company with P/B=1.2 but tangible book = 1.5× price is a real asset play)
        below_tangible = (tb and tb > 0 and price > 0 and (tb / price) >= 0.95)

        # Path 3: Cash fortress — so much net cash it distorts value
        # Current ratio >3.0 + D/E < 0.3 + P/B < 2.0 = assets >> liabilities >> market cap
        cash_fortress = (cr and cr > 3.0
                         and de is not None and de < 0.3
                         and pb and 0 < pb < 2.0)

        # Path 4: Net cash positive — company has more cash than debt
        # ncr > 0.25 means net cash > 25% of price (significant "freebie")
        net_cash_play = (ncr is not None and ncr > 0.25)

        # Path 5: Graham net-net — trading below net working capital per share
        # This is the deepest possible value — Graham called it the "margin of safety"
        graham_net_net = (gnn is not None and price > 0 and gnn > price * 0.5)

        if not (below_book or below_tangible or cash_fortress or net_cash_play or graham_net_net):
            return False

        # Must have some financial activity — not a zombie shell
        has_activity = (
            s.get("fcfYield") is not None or
            (s.get("pe") and s.get("pe") > 0) or
            s.get("revGrowth") is not None or
            s.get("roe") is not None
        )
        return has_activity

    def _asset_play_score(s):
        pb    = s.get("pb")
        tb    = s.get("tangibleBook")
        price = s.get("price", 0)
        sc    = 0

        # ── P/B discount ────────────────────────────────────────────
        if pb and 0 < pb < 1.0:
            sc += (1.0 - pb) * 32       # 0.5× P/B → 16 pts, 0.2× → 25.6 pts
        elif pb and 1.0 <= pb < 1.5:
            sc += (1.5 - pb) * 6        # partial credit 1.0–1.5×

        # ── TANGIBLE BOOK vs price — the gold standard for Lynch asset plays ──
        # When tangible book > price, you're buying real assets at a discount
        if tb and tb > 0 and price > 0:
            tb_ratio = tb / price
            if tb_ratio >= 3.0:   sc += 24  # tangible book is 3× price — extraordinary
            elif tb_ratio >= 2.0: sc += 18
            elif tb_ratio >= 1.5: sc += 13
            elif tb_ratio >= 1.2: sc += 8
            elif tb_ratio >= 1.0: sc += 4   # at or below tangible book
            elif tb_ratio >= 0.8: sc += 1

        # ── Net cash per share — Lynch's "freebie" metric ───────────────────
        # If a company has more cash than debt, you're being paid to hold the business
        ncr = s.get("netCashRatio")
        if ncr is not None:
            if ncr >= 0.75:   sc += 22  # net cash > 75% of price = extraordinary freebie
            elif ncr >= 0.50: sc += 17
            elif ncr >= 0.30: sc += 12
            elif ncr >= 0.15: sc += 7
            elif ncr >= 0.05: sc += 3

        # ── Graham Net-Net — deepest possible value signal ──────────────────
        # Trading below net working capital is Graham's ultimate cheap stock
        gnn = s.get("grahamNetNet")
        if gnn is not None and price > 0:
            gnn_ratio = gnn / price
            if gnn_ratio >= 2.0:   sc += 18  # trading at less than half net working capital
            elif gnn_ratio >= 1.5: sc += 13
            elif gnn_ratio >= 1.0: sc += 8   # at net working capital per share
            elif gnn_ratio >= 0.7: sc += 4

        # ── Cash / liquidity quality ─────────────────────────────────
        cr = s.get("currentRatio")
        if cr and cr > 4.0:   sc += 8   # cash-rich relative to near-term liabilities
        elif cr and cr > 3.0: sc += 5
        elif cr and cr > 2.0: sc += 3
        elif cr and cr > 1.5: sc += 1

        # ── Debt load — key for asset plays (low debt = cleaner asset story) ──
        de = s.get("de")
        if de is not None and de < 0.2:   sc += 8
        elif de is not None and de < 0.5: sc += 5
        elif de is not None and de < 1.0: sc += 2

        # ── FCF yield — the asset base is generating cash ───────────
        fcf = s.get("fcfYield")
        if fcf and fcf > 0.12:   sc += 10
        elif fcf and fcf > 0.08: sc += 7
        elif fcf and fcf > 0.05: sc += 4
        elif fcf and fcf > 0.02: sc += 1

        # ── Piotroski — financial health of the asset base ──────────
        pio = s.get("piotroski")
        if pio and pio >= 8:   sc += 8
        elif pio and pio >= 7: sc += 5
        elif pio and pio >= 5: sc += 2

        # ── ROE — the assets are actually working ───────────────────
        roe = s.get("roe")
        if roe and roe > 0.15:   sc += 6
        elif roe and roe > 0.08: sc += 3
        elif roe and roe > 0.02: sc += 1

        # ── DCF confirms cheapness ───────────────────────────────────
        mos = s.get("mos")
        if mos and mos > 0.40:   sc += 6
        elif mos and mos > 0.25: sc += 4
        elif mos and mos > 0.10: sc += 2

        # ── Asset-heavy sector bonus (book values are more tangible) ──
        if s.get("sector") in _ASSET_HEAVY_SECTORS:
            sc += 3

        # ── Insider buying — insiders see the hidden value ───────────
        if s.get("insiderBuys", 0) >= 3:   sc += 7
        elif s.get("insiderBuys", 0) >= 1: sc += 3

        return sc

    _ap_headers = [
        "Rank", "Ticker", "Company", "Sector", "Price",
        "P/B", "Tangible Book", "Net Cash/Sh", "Graham NN",
        "Curr Ratio", "D/E", "FCF Yield", "ROE", "MoS", "Piotroski",
        "MktCap ($B)", "Score", "🏦 Insider",
    ]
    _ap_widths = [5, 8, 22, 15, 8, 6, 12, 11, 10, 9, 7, 9, 7, 7, 8, 10, 6, 14]

    asset_plays = build_lynch_tab(wb, stocks, "Asset Plays", "8", _asset_play_filter, _asset_play_score,
                                  "0D47A1", "P/B<1, tangible book, net cash, or Graham net-net. Lynch hidden value.",
                                  custom_headers=_ap_headers, custom_widths=_ap_widths)

    # ─── Lynch 10-Baggers: pure small-cap opportunity list ─────────────────
    # Lynch's real insight: PEG + gross margin + operating margin are the quality gates.
    # FCF is NOT required — early Amazon, early Starbucks, early Home Depot all had negative FCF.
    # The filters here replace FCF with profitability (gross + operating margin > 0).
    # This is a dedicated small-cap tab ($50M–$2B) — Fast Growers goes to $150B.
    def _ten_bagger_filter(s):
        if not _is_common_stock(s): return False
        # Exclude sectors where revenue growth is commodity-price driven
        if (s.get("sector") or "") in ("Real Estate", "Basic Materials"): return False
        # Exclude Financial Services: banks, BDCs, and investment vehicles dominate
        # FMP's gross margin metric (net interest margin ≈ 80-100% → false quality signal)
        # and have structurally low PEGs that are not comparable to operating businesses.
        # Lynch's 10-baggers were operating companies — Dunkin Donuts, Home Depot, Chrysler —
        # not capital allocators. Fintech/payment companies appear under Technology.
        if (s.get("sector") or "") == "Financial Services": return False

        mktcap = s.get("mktCap", 0)
        # Pure small-cap focus: $50M floor (nano excluded — too illiquid), $2B ceiling
        if mktcap < 50e6 or mktcap > 2e9:
            return False

        rg  = s.get("revGrowth")
        rg5 = s.get("revGrowth5y")
        eg5 = s.get("epsGrowth5y")
        rg5h = s.get("fiveYRevGrowth")
        gm  = s.get("grossMargin")
        om  = s.get("operatingMargin")
        az  = s.get("altmanZ")
        nde = s.get("netDebtEbitda")

        # Growth qualifier — at least one strong signal (slightly lower bar than Fast Growers)
        # Cap rev growth at 200%: spikes above that are almost always one-time events
        # (pharma milestone payments, licence fees, COVID-vaccine base-effects, etc.)
        # — they don't represent the durable operating growth Lynch looked for.
        _rg_clean = rg if (rg is not None and rg < 2.0) else None
        strong_current = (_rg_clean is not None and _rg_clean > 0.15)
        strong_5yr_f   = (rg5  is not None and rg5  > 0.12)
        strong_5yr_h   = (rg5h is not None and rg5h > 0.12)
        eps_led        = (eg5  is not None and eg5  > 0.15 and (rg is None or rg > 0.05))
        if not (strong_current or strong_5yr_f or strong_5yr_h or eps_led):
            return False

        # Kill 1: Gross margin must exist and show real pricing power
        # This is Lynch's proxy for business quality — replaces FCF requirement
        if gm is None or gm < 0.20:
            return False

        # Kill 2: Business model must work — operating margin must be positive,
        # and must be sane (>80% op margin is physically impossible for real operating
        # businesses and signals property revaluations, licensing one-offs, etc.)
        if om is not None and om < 0:
            return False
        if om is not None and om > 0.80:
            return False

        # Kill 3: Not in distress — Altman Z > 1.0 (slightly lower than Fast Growers' 1.0 kill)
        if az is not None and az < 1.0:
            return False

        # Kill 4: Not over-leveraged (allow some debt — small caps need growth capital)
        if nde is not None and nde > 6:
            return False

        # Kill 5: Share dilution ≤ 20% (slightly more lenient than FG — small caps issue equity)
        sd = s.get("sharesGrowth")
        if sd is not None and sd > 0.20:
            return False

        # Kill 6: PEG valuation — Lynch's anchor metric. PEG > 2 means paying too much
        # for the growth you're getting. Allow missing PEG (small-caps often lack consensus).
        # Only exclude if BOTH trailing and forward PEG are above the threshold.
        peg  = s.get("peg")
        fpeg = s.get("fwdPEG")
        if peg is not None and peg > 2.0:
            if fpeg is None or fpeg > 2.0:
                return False

        return True

    def _ten_bagger_score(s):
        sc = 0
        gm  = s.get("grossMargin") or 0
        om  = s.get("operatingMargin") or 0
        rg  = s.get("revGrowth") or 0
        rg5 = s.get("revGrowth5y")
        peg = s.get("peg")
        fpeg = s.get("fwdPEG")
        nde = s.get("netDebtEbitda")
        az  = s.get("altmanZ")
        rc  = s.get("revConsistency") or 0

        # ── PEG ratio: Lynch's anchor metric — most important single signal ──
        # Low PEG = paying reasonable price for genuine growth = 10-bagger setup
        best_peg = min(p for p in [peg, fpeg] if p and 0 < p < 5) if any(p and 0 < p < 5 for p in [peg, fpeg]) else None
        if best_peg is not None:
            if best_peg < 0.5:  sc += 25
            elif best_peg < 0.75: sc += 20
            elif best_peg < 1.0: sc += 16
            elif best_peg < 1.25: sc += 12
            elif best_peg < 1.5: sc += 8
            elif best_peg < 2.0: sc += 4
        else:
            sc -= 4  # missing PEG — can't verify valuation vs growth

        # ── Gross margin: pricing power = durable competitive advantage ──
        if gm > 0.70:   sc += 18
        elif gm > 0.55: sc += 14
        elif gm > 0.40: sc += 10
        elif gm > 0.30: sc += 6
        elif gm > 0.20: sc += 2

        # ── Operating margin: business model efficiency ──
        if om > 0.25:   sc += 12
        elif om > 0.15: sc += 9
        elif om > 0.08: sc += 6
        elif om > 0.03: sc += 3
        elif om > 0:    sc += 1

        # ── Revenue growth: the engine ──
        if rg > 0.50:   sc += 16
        elif rg > 0.35: sc += 12
        elif rg > 0.25: sc += 9
        elif rg > 0.15: sc += 6
        elif rg > 0.08: sc += 3
        if rg5 and rg5 > 0.20: sc += 6
        elif rg5 and rg5 > 0.12: sc += 3

        # ── Revenue consistency: sustained growth > one-year spike ──
        if rc > 0.80:   sc += 8
        elif rc > 0.65: sc += 5
        elif rc > 0.50: sc += 2

        # ── Balance sheet: net debt / EBITDA ──
        # Net cash is special — Lynch loved "Fort Knox" balance sheets on small caps
        if nde is not None:
            if nde < -0.5:  sc += 12  # net cash position
            elif nde < 0.5: sc += 8   # essentially debt-free
            elif nde < 1.5: sc += 5
            elif nde < 3.0: sc += 2
        # else: no data — no penalty (small caps often missing)

        # ── Altman Z: financial health ──
        if az is not None:
            if az > 4.0:    sc += 6
            elif az > 2.5:  sc += 4
            elif az > 1.5:  sc += 2

        # ── 52-week position: buy beaten-down small caps, not momentum ──
        pvs52h = s.get("priceVs52H")
        if pvs52h is not None:
            if pvs52h < 0.70:   sc += 8   # >30% off high — maximum opportunity
            elif pvs52h < 0.80: sc += 5
            elif pvs52h < 0.90: sc += 2

        # ── Insider buying: management eats their own cooking ──
        if s.get("insiderBuys", 0) >= 3:   sc += 10
        elif s.get("insiderBuys", 0) >= 1: sc += 5

        # ── Cap size bonus: smaller = more neglected = more mispricing ──
        _mc = s.get("mktCap", 0)
        if _mc < 150e6:    sc += 15   # sub-$150M: maximum neglect
        elif _mc < 350e6:  sc += 11   # $150M–$350M: very underfollowed
        elif _mc < FG_SMALL_CAP_MAX:  sc += 7   # $350M–$2B: small cap territory

        return sc

    _tb_headers = [
        "Rank", "Ticker", "Company", "Sector", "Price",
        "PEG", "Fwd PEG", "P/E", "Fwd P/E",
        "Gross Margin", "Oper Margin", "Net D/EBITDA",
        "Rev Growth", "Rev Growth 5Y", "Rev Consist.",
        "52w vs High", "Altman Z", "Shares Δ",
        "MktCap ($B)", "Cap Size", "Score", "🏦 Insider",
    ]
    _tb_widths = [5, 8, 22, 15, 8, 6, 8, 6, 8, 12, 11, 12, 10, 12, 11, 11, 9, 9, 10, 7, 6, 14]

    ten_baggers = build_lynch_tab(wb, stocks, "Lynch 10-Baggers", "9", _ten_bagger_filter, _ten_bagger_score,
                                  "BF360D", "Small-cap $50M–$2B · PEG<2 · Gross margin>20% · Oper margin>0. "
                                  "Lynch's real criteria: pricing power + growth + reasonable price. No FCF kill.",
                                  custom_headers=_tb_headers, custom_widths=_tb_widths)

    quality_compounders = build_quality_compounders(wb, stocks)
    # Sprint 3 A4: 💎 Hold Forever — strict long-term shortlist (~25 names)
    hold_forever = build_hold_forever_tab(wb, stocks)

    # Sector Valuations (ETF Rotation tab retired in v4 — pure trader signal, not aligned with long-term picking)
    _sector_rows, _etf_rets = build_sector_valuations(wb, stocks)
    _etf_rows = None

    # Insider Buying
    build_insider_tab(wb, stocks, insider_data)

    build_picks_tracking(wb, stocks)

    # B1: Compute per-agent attribution (SPY prices fetched inside build_picks_tracking
    # already; re-fetch here using the same helper — cached so no extra API calls).
    _b1_spy_prices = {}; _b1_spy_today = None
    if os.path.exists(AI_PICKS_LOG):
        try:
            import csv as _csv_tmp
            with open(AI_PICKS_LOG, "r", encoding="utf-8") as _f:
                _dates = [r["date"] for r in _csv_tmp.DictReader(_f) if r.get("date")]
            if _dates:
                _earliest_ai = min(_dates)
                _b1_spy_prices = fetch_spy_history(_earliest_ai)
                _b1_spy_today  = fetch_live_price("SPY")
        except Exception:
            pass
    agent_perf = compute_agent_performance(_b1_spy_prices, _b1_spy_today)

    # Auto-log strategy picks every run (deduped by date+ticker+strategy)
    current_prices = {t: s["price"] for t, s in stocks.items() if s.get("price")}
    log_picks({
        "IV Discount": iv_rows, "Quality Compounders": quality_compounders,
        "Stalwarts": stalwarts, "Fast Growers": fast_growers,
        "Turnarounds": turnarounds, "Slow Growers": slow_growers,
        "Cyclicals": cyclicals, "Asset Plays": asset_plays,
        "Lynch 10-Baggers": ten_baggers,
    }, current_prices)

    # ── AI analysis (single call — result shared across Overview + AI tab) ──
    picks_data = {
        "IV Discount (Buffett/DCF)": iv_rows,
        "Quality Compounders (Buffett)": quality_compounders,
        "Stalwarts (Lynch)": stalwarts,
        "Fast Growers (Lynch)": fast_growers,
        "Turnarounds (Lynch)": turnarounds,
        "Slow Growers / Income (Lynch)": slow_growers,
        "Cyclicals (Lynch)": cyclicals,
        "Asset Plays (Lynch)": asset_plays,
        "Lynch 10-Baggers": ten_baggers,
    }
    # ── Daily AI result cache — skip Claude calls if already run today ──
    _AI_CACHE_KEY = "_ai_result"
    _today_str    = datetime.date.today().isoformat()
    _cached_ai    = _cache.get(_AI_CACHE_KEY, {})
    _ai_from_cache = (
        not args.force_fresh_ai      # C2: --force-fresh-ai bypasses the daily cache
        and isinstance(_cached_ai, dict)
        and _cached_ai.get("date") == _today_str
        and isinstance(_cached_ai.get("result"), dict)
        and _cached_ai["result"].get("picks")
    )

    if _ai_from_cache:
        ai_result   = _cached_ai["result"]
        mall_result = _cached_ai.get("mall_result", {}) or {}
        n_picks = len(ai_result.get("picks", []))
        n_specs = len(ai_result.get("_specialist_picks", {}))
        n_mall  = len(mall_result.get("picks", []))
        mall_note = f" + {n_mall} mall" if n_mall else ""
        print(f"\n  📦 Using cached AI analysis ({n_picks} judge picks, {n_specs} specialists{mall_note} — already run today)")
    else:
        if args.force_fresh_ai and _cached_ai.get("date") == _today_str:
            print("\n  🔄 --force-fresh-ai: bypassing today's AI cache, re-running all specialists + judge + mall manager")
        phase_start("ai_analysis", "Running multi-agent AI analysis (11 specialists + judge)")
        ai_result = call_claude_analysis(picks_data, stocks, macro=macro_data,
                                         market_intel=market_intel,
                                         agent_perf=agent_perf)   # B4: performance feedback
        mall_result = {}
        if ai_result and ai_result.get("picks"):
            # Mall Manager runs after the judge, takes specialist pool + strategy tabs
            mall_result = call_mall_manager(ai_result, stocks,
                                            macro=macro_data,
                                            market_intel=market_intel,
                                            fast_growers=fast_growers,
                                            quality_compounders=quality_compounders) or {}
            # Cache result for remainder of today
            _cache[_AI_CACHE_KEY] = {"date": _today_str,
                                     "result": ai_result,
                                     "mall_result": mall_result}
            save_cache()

    # If cache was loaded but mall_result is missing (older cache schema), run only mall manager
    if ai_result and ai_result.get("picks") and not mall_result and _ai_from_cache:
        print("  🛍️ Cache pre-dates Mall Manager — running it now to backfill")
        mall_result = call_mall_manager(ai_result, stocks,
                                        macro=macro_data,
                                        market_intel=market_intel,
                                        fast_growers=fast_growers,
                                        quality_compounders=quality_compounders) or {}
        if mall_result:
            _cache[_AI_CACHE_KEY] = {"date": _today_str,
                                     "result": ai_result,
                                     "mall_result": mall_result}
            save_cache()

    # Auto-log AI picks every run (no flag needed)
    if ai_result:
        log_ai_picks(ai_result, stocks, mall_result=mall_result)

    # Tab 1b: AI Top Picks (uses pre-created sheet so it stays in position 2)
    build_ai_picks_tab(wb, ai_result, stocks, ws=ws_ai_picks)

    # Tab 1d: Agent Reports (individual specialist picks)
    build_agent_reports_tab(wb, ai_result, stocks)

    # Fetch 5Y sparklines for all AI picks (judge + all specialists, deduplicated)
    _sparkline_tickers = set()
    if ai_result:
        for _sp in ai_result.get("picks", []):
            _sparkline_tickers.add(_sp.get("ticker", ""))
        for _av in ai_result.get("_specialist_picks", {}).values():
            for _sp in _av.get("picks", []):
                _sparkline_tickers.add(_sp.get("ticker", ""))
    _sparkline_tickers.discard("")
    _sparklines = fetch_sparkline_data(list(_sparkline_tickers)) if _sparkline_tickers else {}
    if _sparklines:
        save_cache()   # persist sparkline cache (7-day TTL) for next run

    # ── AI Portfolio Manager ─────────────────────────────────────────
    phase_start("portfolio", "Running AI Portfolio Manager")
    print("\n  💼 Running AI Portfolio Manager...")
    portfolio = load_portfolio()

    # Reuse the candidate pool + sector context built inside call_claude_analysis
    # by re-building the same meta-score pool here for the PM context
    _pm_meta = {}
    _pm_strat_short = {
        "IV Discount (Buffett/DCF)": "IV Discount",
        "Quality Compounders (Buffett)": "Quality Compounder",
        "Stalwarts (Lynch)": "Stalwart",
        "Fast Growers (Lynch)": "Fast Grower",
        "Turnarounds (Lynch)": "Turnaround",
        "Slow Growers / Income (Lynch)": "Slow Grower",
        "Cyclicals (Lynch)": "Cyclical",
        "Asset Plays (Lynch)": "Asset Play",
        "Lynch 10-Baggers": "10-Bagger",
    }
    for _tab_name, _rows in picks_data.items():
        _short = _pm_strat_short.get(_tab_name, _tab_name)
        for _ri, _row in enumerate(_rows[:15]):
            _t = _row.get("Ticker", "?")
            _sc = _row.get("Score", 0) or 0
            if _t not in _pm_meta:
                _pm_meta[_t] = {"strategies": [], "best_rank": _ri+1, "max_score": _sc, "row": _row}
            _pm_meta[_t]["strategies"].append(_short)
            if _sc > _pm_meta[_t]["max_score"]: _pm_meta[_t]["max_score"] = _sc
            if _ri+1 < _pm_meta[_t]["best_rank"]: _pm_meta[_t]["best_rank"] = _ri+1

    # Build compact candidates block for PM (top 40 by meta-score)
    _pm_top = sorted(_pm_meta.values(),
                     key=lambda x: -(len(x["strategies"])*25 + x["max_score"]*0.4))[:40]

    def _pm_fmt(m):
        r2 = m["row"]; t2 = r2.get("Ticker","?"); s2 = stocks.get(t2,{})
        strats2 = "+".join(m["strategies"])
        parts2 = [f"{t2}({r2.get('Company','')[:18]}) [{strats2}]"]
        roic2 = s2.get("roic"); fcf2 = s2.get("fcfYield")
        peg2 = s2.get("fwdPEG") or s2.get("peg"); rg2 = s2.get("revGrowth")
        if roic2: parts2.append(f"ROIC={roic2*100:.0f}%")
        if fcf2:  parts2.append(f"FCF={fcf2:.1%}")
        if peg2:  parts2.append(f"PEG={peg2:.2f}")
        if rg2:   parts2.append(f"RG={rg2:+.0%}")
        mos2 = s2.get("mos")
        if mos2:  parts2.append(f"MoS={mos2:.0%}")
        return "  " + " | ".join(parts2)

    _pm_candidates = "\n".join(_pm_fmt(m) for m in _pm_top)

    # Simple sector context for PM
    _pm_sec_lines = []
    _pm_sec = {}
    for _sv in stocks.values():
        _sec = _sv.get("sector","Unknown")
        if _sec not in _pm_sec: _pm_sec[_sec] = []
        _peg = _sv.get("peg")
        if _peg and 0 < _peg < 15: _pm_sec[_sec].append(_peg)
    for _sec, _pegs in sorted(_pm_sec.items(), key=lambda x: (sum(x[1])/len(x[1])) if x[1] else 99):
        if len(_pegs) >= 5:
            _pm_sec_lines.append(f"  {_sec}: med PEG={sorted(_pegs)[len(_pegs)//2]:.1f} ({len(_pegs)} stocks)")
    _pm_sector_block = "\n".join(_pm_sec_lines[:12])

    # ── Auto-exit holdings in sectors excluded from quality portfolio ────
    _EXCLUDED_PTF_SECTORS = {"Basic Materials", "Real Estate"}
    _today_auto = datetime.date.today().isoformat()
    _auto_sell_tickers = []
    for _ah in list(portfolio.get("holdings", [])):
        _at = _ah.get("ticker", "")
        _asector = stocks.get(_at, {}).get("sector") or _ah.get("sector") or ""
        if _asector in _EXCLUDED_PTF_SECTORS:
            _alive_p = fetch_live_price(_at) or _ah.get("entry_price", 0) or 0
            _ashares = _ah.get("shares", 0)
            _aproceeds = round(_alive_p * _ashares, 2)
            _aret = ((_alive_p / _ah["entry_price"]) - 1) if (_ah.get("entry_price") and _ah["entry_price"] > 0) else 0
            portfolio["cash"] = round(portfolio.get("cash", 0) + _aproceeds, 2)
            portfolio["transactions"].append({
                "date": _today_auto,
                "action": "SELL",
                "ticker": _at,
                "company": _ah.get("company", ""),
                "shares": _ashares,
                "price": _alive_p,
                "proceeds": _aproceeds,
                "return_pct": round(_aret * 100, 2),
                "rationale": f"Auto-exit: {_asector} excluded from quality growth portfolio",
            })
            _auto_sell_tickers.append(_at)
            print(f"    🚫 Auto-sell {_at} ({_asector}) @ ${_alive_p:.2f} ({_aret:+.1%}) — excluded sector")
    if _auto_sell_tickers:
        portfolio["holdings"] = [_h for _h in portfolio["holdings"] if _h["ticker"] not in _auto_sell_tickers]
        save_portfolio(portfolio)

    pm_decisions = run_portfolio_manager(portfolio, _pm_candidates, _pm_sector_block, stocks)
    if pm_decisions:
        portfolio = apply_portfolio_decisions(portfolio, pm_decisions, stocks)
        save_portfolio(portfolio)
    portfolio["_last_thesis"] = pm_decisions.get("portfolio_thesis", "")

    # Fetch SPY data for portfolio tab (reuse from picks tracking if already fetched)
    _ptf_spy_prices = {}; _ptf_spy_today = None
    if portfolio.get("started"):
        _ptf_spy_prices = fetch_spy_history(portfolio["started"])
        _ptf_spy_today  = fetch_live_price("SPY")

    # Tab 1c: AI Portfolio (uses pre-created sheet so it stays in position 3)
    _ptf_nav, _ptf_ret, _ptf_spy_ret = build_portfolio_tab(
        wb, portfolio, stocks, _ptf_spy_prices, _ptf_spy_today,
        ws=ws_ai_portf)

    # Overview (last — needs data from all tabs)
    build_overview_tab(ws_overview, stocks, iv_rows, stalwarts, fast_growers, turnarounds,
                       slow_growers, cyclicals, asset_plays, quality_compounders,
                       _fmp_call_count, ai=ai_result,
                       portfolio=portfolio, portfolio_nav=_ptf_nav,
                       portfolio_ret=_ptf_ret, portfolio_spy_ret=_ptf_spy_ret,
                       macro=macro_data, ten_baggers=ten_baggers)

    # Save
    phase_start("save", "Saving Excel + HTML dashboard")
    wb.save(output_file)

    # Generate HTML dashboard (same data, no extra fetching)
    html_content = build_html_report(
        stocks, iv_rows, stalwarts, fast_growers, turnarounds,
        slow_growers, cyclicals, asset_plays, quality_compounders,
        sector_rows=_sector_rows, etf_rows=_etf_rows,
        ai=ai_result, macro=macro_data,
        portfolio=portfolio, fmp_call_count=_fmp_call_count,
        ten_baggers=ten_baggers,
        agent_perf=agent_perf,   # B1: per-agent attribution for leaderboard (B8)
        sparklines=_sparklines,  # 5Y price history for AI pick mini-charts
        hold_forever=hold_forever,  # Sprint 3 A4: long-term shortlist tab
        mall=mall_result,        # 🛍️ Mall Manager picks (Lynch consumer-observable)
    )
    html_file = output_file.replace(".xlsx", ".html")
    with open(html_file, "w", encoding="utf-8") as _hf:
        _hf.write(html_content)
    print(f"  🌐 HTML dashboard: {html_file}")

    # Push to GitHub Pages (if configured)
    pages_url = push_html_to_github(html_content)
    if pages_url:
        print(f"  🌐 Live at: {pages_url}")

    # Commit data files (portfolio + picks logs) for local runs
    commit_data_files()

    total_elapsed = time.time() - _run_start
    n_picks = len(ai_result.get("picks", [])) if ai_result else 0

    print(f"\n{'=' * 65}")
    print(f"  ✅ DONE! Saved to: {output_file}")
    print(f"  ⏱  Total run time: {_fmt_elapsed(total_elapsed)}")
    print(f"  📊 FMP API calls: {_fmp_call_count}")
    print(f"  📦 Cache: {CACHE_FILE} (valid {CACHE_DAYS} days)")
    if pages_url:
        print(f"  🌐 Dashboard: {pages_url}")
    if NTFY_TOPIC:
        print(f"  🔔 Sending phone notification to ntfy.sh/{NTFY_TOPIC}...")
    print(f"{'=' * 65}")

    # Phone notification — includes clickable dashboard link
    notify_phone(
        title="FMP Screener Done ✅",
        message=(
            f"Run complete in {_fmt_elapsed(total_elapsed)}\n"
            f"{n_picks} AI picks | {_fmp_call_count} FMP calls\n"
            + (f"📱 View: {pages_url}" if pages_url else f"File: {os.path.basename(output_file)}")
        ),
        tags="chart_with_upwards_trend",
    )


if __name__ == "__main__":
    main()
